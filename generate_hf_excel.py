"""
Hedge Fund Level DCF Upgrade
Adds 5 new sheets to the existing generate_dcf_excel.py model:
  Sheet 9:  Reverse DCF (market-implied expectations)
  Sheet 10: ROIC-Driven Growth Model
  Sheet 11: Monte Carlo Simulation (1000 runs)
  Sheet 12: Downside Protection & Upside Asymmetry
  Sheet 13: Final Investment Framework (hedge fund conclusion)

Usage:
    python generate_hf_excel.py --ticker TCS.NS
    python generate_hf_excel.py --ticker ITC.NS
"""

import sys, argparse, io, warnings
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Reuse the institutional palette ────────────────────────────
C_INPUT_BG  = "EBF3FB"; C_INPUT_FG  = "0000FF"
C_FORMULA   = "000000"; C_LINK_FG   = "008000"
C_HEADER_BG = "1F3864"; C_HEADER_FG = "FFFFFF"
C_SUBHDR    = "2E75B6"; C_SECTION   = "D6E4F0"
C_ALT       = "F2F7FB"; C_WHITE     = "FFFFFF"
C_YELLOW    = "FFFF00"; C_GREEN     = "00B050"
C_RED       = "C00000"; C_AMBER     = "ED7D31"
C_BEAR_BG   = "FCE4D6"; C_BULL_BG   = "E2EFDA"
C_NAVY      = "1F3864"; C_DARK_NAVY = "172035"

def _c(s):
    if not s: return "FFFFFFFF"
    s = str(s).lstrip("#").upper()
    if len(s) == 6: return "FF" + s
    if len(s) == 8: return s
    if len(s) >= 10: return s[:8]
    return "FFFFFFFF"

def hf(c): return PatternFill("solid", fgColor=_c(c))
def bdr(style="thin", color="BFBFBF"):
    s = Side(style=style, color=_c(color))
    return Border(left=s, right=s, top=s, bottom=s)

def wc(ws, row, col, val, bg=C_WHITE, fg=C_FORMULA, bold=False, sz=10,
       nf=None, align="right", italic=False, wrap=False, height=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = hf(bg)
    c.font = Font(name="Calibri", bold=bold, size=sz, color=_c(fg), italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = bdr()
    if nf: c.number_format = nf
    if height: ws.row_dimensions[row].height = height
    return c

def inp(ws, row, col, val, nf=None, comment=None):
    """Blue = hardcoded input"""
    c = wc(ws, row, col, val, bg=C_INPUT_BG, fg=C_INPUT_FG, nf=nf)
    return c

def hdr(ws, row, labels, bg=C_HEADER_BG, fg=C_HEADER_FG, sz=10, height=20):
    ws.row_dimensions[row].height = height
    for col, lbl in enumerate(labels, 1):
        c = ws.cell(row=row, column=col, value=lbl)
        c.fill = hf(bg); c.font = Font(name="Calibri", bold=True, size=sz, color=_c(fg))
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr()

def sec(ws, row, text, ncols, bg=C_SECTION, fg=C_NAVY, sz=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=f"  {text}")
    c.fill = hf(bg); c.font = Font(name="Calibri", bold=True, size=sz, color=_c(fg))
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = Border(bottom=Side(style="medium", color=_c(C_NAVY)))
    ws.row_dimensions[row].height = 18

def title(ws, row, text, sub, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = hf(C_NAVY); c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=ncols)
    c2 = ws.cell(row=row+1, column=1, value=sub)
    c2.fill = hf(C_SUBHDR); c2.font = Font(name="Calibri", size=9, color="FFFFFFFF", italic=True)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row+1].height = 15


def build_hedge_fund_sheets(wb, ticker, enriched, dcf_res, forecast_result,
                             scenarios, wacc_data, wacc, terminal_g,
                             forecast_yrs, sym, fx):
    """Add hedge fund sheets to existing workbook."""

    price       = enriched.get("price", 0) * fx
    shares      = enriched.get("shares", 1)
    op_margin   = enriched.get("op_margin", 0)
    rev_growth  = enriched.get("revenue_growth", 0)
    latest_rev  = enriched.get("latest_revenue", 0) * fx / 1e9
    latest_fcf  = enriched.get("latest_fcf",  0)  * fx / 1e9
    total_debt  = enriched.get("total_debt",  0)  * fx / 1e9
    total_cash  = enriched.get("total_cash",  0)  * fx / 1e9
    tax_rate    = wacc_data.get("tax_rate", 0.25)
    sector      = enriched.get("sector_name", enriched.get("sector", "General"))
    moat        = enriched.get("moat_grade", "None")
    iv_base     = dcf_res.get("intrinsic_value_per_share", 0) * fx
    mktcap      = price * shares / 1e9

    bear_iv  = scenarios.get("Bear 🐻", {}).get("iv", iv_base * 0.6) * fx
    base_iv  = scenarios.get("Base 📊", {}).get("iv", iv_base)       * fx
    bull_iv  = scenarios.get("Bull 🐂", {}).get("iv", iv_base * 1.4) * fx
    pw_iv    = bear_iv * 0.25 + base_iv * 0.50 + bull_iv * 0.25

    # ══════════════════════════════════════════════════════════
    # SHEET 9 — REVERSE DCF (Market-Implied Expectations)
    # ══════════════════════════════════════════════════════════
    ws9 = wb.create_sheet("9. Reverse DCF")
    ws9.sheet_view.showGridLines = False
    ws9.column_dimensions["A"].width = 36
    ws9.column_dimensions["B"].width = 18
    ws9.column_dimensions["C"].width = 18
    ws9.column_dimensions["D"].width = 18
    ws9.column_dimensions["E"].width = 35

    title(ws9, 1, f"REVERSE DCF — {ticker}",
          "What growth/margin assumptions is the market currently pricing in?", 5)

    # Reverse-engineer market-implied growth
    # Formula: Price = FCF_base × (1+g_impl)^n / (WACC-g_impl) × (1/(1+WACC)^n)
    # Solve numerically for g_impl that makes IV = Price
    fcf_base_raw = forecast_result.get("fcf_base", 0) * fx / 1e9

    def compute_iv_for_growth(g_imp, fcf_b, w, tg, n_yrs, debt, cash, sh):
        """Compute IV/share for a given growth rate assumption."""
        if w <= tg: return 0
        fcf = fcf_b
        pv_sum = 0
        for yr in range(1, n_yrs + 1):
            fade_g = tg + (g_imp - tg) * np.exp(-0.25 * yr)
            fcf   *= (1 + fade_g)
            pv_sum += fcf / (1 + w) ** yr
        term_norm = fcf
        tv  = term_norm * (1 + tg) / (w - tg)
        pv_tv = tv / (1 + w) ** n_yrs
        ev = pv_sum + pv_tv
        eq = max(ev - debt + cash, 0)
        return eq * 1e9 / sh * fx if sh > 0 else 0

    # Binary search for implied growth
    implied_g = rev_growth  # default
    if fcf_base_raw > 0 and shares > 0 and price > 0:
        lo, hi = -0.10, 0.50
        for _ in range(60):
            mid = (lo + hi) / 2
            iv_mid = compute_iv_for_growth(mid, fcf_base_raw, wacc, terminal_g,
                                           forecast_yrs, total_debt, total_cash, shares)
            if iv_mid < price: lo = mid
            else: hi = mid
        implied_g = (lo + hi) / 2

    # Implied margin (holding growth constant at model rate, solve for margin)
    implied_margin = op_margin
    if latest_rev > 0 and shares > 0 and price > 0:
        lo_m, hi_m = 0.0, 0.80
        for _ in range(60):
            mid_m = (lo_m + hi_m) / 2
            fcf_test = latest_rev * mid_m * (1 - tax_rate) * 0.75
            iv_test = compute_iv_for_growth(rev_growth, fcf_test, wacc, terminal_g,
                                            forecast_yrs, total_debt, total_cash, shares)
            if iv_test < price: lo_m = mid_m
            else: hi_m = mid_m
        implied_margin = (lo_m + hi_m) / 2

    # EV-implied reinvestment rate
    ev_current = mktcap + total_debt - total_cash
    nopat_latest = latest_rev * op_margin * (1 - tax_rate)
    implied_reinv_rate = 1 - (latest_fcf / nopat_latest) if nopat_latest > 0 else 0.5

    sec(ws9, 3, "STEP 1 — MARKET-IMPLIED ASSUMPTIONS (What is the market pricing in?)", 5)
    hdr(ws9, 4, ["Metric", "Market-Implied", "My Base Case", "Difference", "Verdict"], bg=C_SUBHDR)

    impl_rows = [
        ("Revenue / FCF Growth Rate",
         implied_g, rev_growth, implied_g - rev_growth,
         "Market pricing HIGHER growth" if implied_g > rev_growth else "Market pricing LOWER growth"),
        ("Operating Margin",
         implied_margin, op_margin, implied_margin - op_margin,
         "Market expects MARGIN EXPANSION" if implied_margin > op_margin else "Market expects MARGIN COMPRESSION"),
        ("Reinvestment Rate (implied)",
         implied_reinv_rate, 0.35, implied_reinv_rate - 0.35,
         "Market assumes HIGH reinvestment" if implied_reinv_rate > 0.4 else "Market assumes CAPITAL-LIGHT model"),
        ("EV/Revenue Multiple (current)",
         ev_current / latest_rev if latest_rev > 0 else 0, None, None,
         "Premium to fundamentals" if (ev_current/latest_rev if latest_rev>0 else 0) > 3 else "Reasonable multiple"),
        ("FCF Yield (current price)",
         latest_fcf * 1e9 / shares / price if price > 0 else 0, None, None,
         "LOW yield = growth premium" if (latest_fcf * 1e9 / shares / price if price > 0 else 0) < 0.03 else "REASONABLE yield"),
    ]

    for i, row_data in enumerate(impl_rows):
        label, impl_val, base_val, diff, verdict = row_data
        r = 5 + i
        ws9.row_dimensions[r].height = 18
        bg = C_ALT if i % 2 == 0 else C_WHITE
        wc(ws9, r, 1, label, bg=bg, fg=C_FORMULA, sz=10, align="left")

        if impl_val is not None:
            nf = "0.0%" if i < 3 else "0.00x" if i == 3 else "0.0%"
            fg = C_AMBER if diff and abs(diff) > 0.05 else C_FORMULA
            c = wc(ws9, r, 2, impl_val, bg=C_YELLOW if i < 2 else bg, fg=fg, bold=i<2, sz=10, nf=nf)

        if base_val is not None:
            nf = "0.0%"
            wc(ws9, r, 3, base_val, bg=bg, fg=C_INPUT_FG, bold=False, sz=10, nf=nf)
        else:
            wc(ws9, r, 3, "—", bg=bg, fg="595959", sz=10, align="center")

        if diff is not None:
            fg = C_RED if diff > 0.05 else (C_GREEN if diff < -0.05 else C_FORMULA)
            wc(ws9, r, 4, diff, bg=bg, fg=fg, bold=abs(diff)>0.05, sz=10, nf="+0.0%;(0.0%);-")
        else:
            wc(ws9, r, 4, "—", bg=bg, fg="595959", sz=10, align="center")

        verdict_bg = C_BEAR_BG if "HIGHER" in verdict or "EXPANSION" in verdict else (C_BULL_BG if "LOWER" in verdict else bg)
        wc(ws9, r, 5, verdict, bg=verdict_bg, fg=C_FORMULA, sz=9, align="left", italic=True)

    # Market-implied conclusions
    sec(ws9, 11, "REVERSE DCF INTERPRETATION", 5)
    ws9.row_dimensions[12].height = 60
    ws9.merge_cells("A12:E12")
    conclusion_text = (
        f"At {sym}{price:,.0f}/share, the market is pricing in: "
        f"FCF growth of {implied_g:.1%} (your base: {rev_growth:.1%}), "
        f"operating margin of {implied_margin:.1%} (current: {op_margin:.1%}). "
        f"{'The market appears OVERLY OPTIMISTIC — implied growth exceeds your base case significantly. This is a HIGH BAR to clear.' if implied_g > rev_growth * 1.3 else ''}"
        f"{'The market appears PESSIMISTIC relative to fundamentals — implied growth is well below historical.' if implied_g < rev_growth * 0.7 else ''}"
        f"{'The market appears FAIRLY RATIONAL — implied assumptions close to your base case.' if abs(implied_g - rev_growth) < rev_growth * 0.3 else ''}"
    )
    c = ws9.cell(12, 1, value=conclusion_text)
    c.fill = hf("FFF2CC"); c.font = Font(name="Calibri", size=10, color=_c(C_FORMULA), italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = bdr()

    # ══════════════════════════════════════════════════════════
    # SHEET 10 — ROIC-DRIVEN GROWTH MODEL
    # ══════════════════════════════════════════════════════════
    ws10 = wb.create_sheet("10. ROIC Growth Model")
    ws10.sheet_view.showGridLines = False
    ws10.column_dimensions["A"].width = 32
    for col in "BCDEFGHIJKL":
        ws10.column_dimensions[col].width = 14

    title(ws10, 1, f"ROIC-DRIVEN GROWTH — {ticker}",
          "Growth = ROIC × Reinvestment Rate  |  Enforcing Fundamental Consistency", 12)

    # ROIC calculation
    ic_proxy = max(total_debt + latest_rev * 0.4, latest_rev * 0.1)
    nopat    = latest_rev * op_margin * (1 - tax_rate)
    roic     = nopat / ic_proxy if ic_proxy > 0 else 0
    roic_wacc_spread = roic - wacc

    # Reinvestment schedule — declining as company matures
    reinv_schedule = []
    roic_schedule  = []
    growth_derived = []
    implied_fcf_conv = []
    base_g = forecast_result.get("base_growth", rev_growth)

    for yr in range(1, forecast_yrs + 1):
        # ROIC decays toward WACC as competition erodes moat
        # Wide moat: slower decay; No moat: faster decay
        moat_factor = {"Wide": 0.05, "Narrow": 0.10, "None": 0.18}.get(moat, 0.10)
        yr_roic = max(wacc, roic * np.exp(-moat_factor * yr) + wacc * (1 - np.exp(-moat_factor * yr)))

        # Growth derived from ROIC × Reinvestment Rate
        # Fade growth from base toward terminal
        yr_g = terminal_g + (base_g - terminal_g) * np.exp(-0.25 * yr)
        yr_reinv = yr_g / yr_roic if yr_roic > 0 else 0.5
        yr_reinv = min(max(yr_reinv, 0), 1.0)  # cap 0-100%
        yr_fcf_conv = 1 - yr_reinv

        roic_schedule.append(round(yr_roic, 4))
        reinv_schedule.append(round(yr_reinv, 4))
        growth_derived.append(round(yr_g, 4))
        implied_fcf_conv.append(round(yr_fcf_conv, 4))

    sec(ws10, 3, "STEP 2 — ROIC vs WACC SPREAD (Value Creation Test)", 12)
    hdr(ws10, 4, ["Metric", "Value", "", "Interpretation"], bg=C_SUBHDR)
    roic_rows = [
        ("NOPAT (latest)",            nopat,           "#,##0.0",  f"= {op_margin:.1%} margin × {latest_rev:.1f}B rev × (1-tax)"),
        ("Invested Capital (proxy)",  ic_proxy,        "#,##0.0",  "Debt + 40% Revenue (tangible asset proxy)"),
        ("ROIC (latest)",             roic,            "0.0%",     "= NOPAT / Invested Capital"),
        ("WACC",                      wacc,            "0.0%",     "Minimum required return"),
        ("ROIC − WACC Spread",        roic_wacc_spread,"0.0%",     "POSITIVE = value creation  |  NEGATIVE = destruction"),
        ("Value Creation Status",
         "✓ CREATES VALUE" if roic_wacc_spread > 0.02 else ("≈ NEUTRAL" if roic_wacc_spread > -0.02 else "✗ DESTROYS VALUE"),
         "@", "Critical: growth only adds value if ROIC > WACC"),
    ]
    for i, (label, val, nf, note) in enumerate(roic_rows):
        r = 5 + i
        ws10.row_dimensions[r].height = 17
        bg = C_ALT if i%2==0 else C_WHITE
        is_spread = "Spread" in label
        is_status = "Status" in label
        fg = (C_GREEN if roic_wacc_spread > 0.02 else (C_AMBER if roic_wacc_spread > -0.02 else C_RED)) if (is_spread or is_status) else C_FORMULA
        wc(ws10, r, 1, label, bg=bg, fg=C_FORMULA, sz=10, align="left")
        c = wc(ws10, r, 2, val if isinstance(val,str) else round(float(val),4), bg=C_YELLOW if is_spread else bg, fg=fg, bold=is_spread or is_status, sz=10)
        if nf != "@" and not isinstance(val,str): c.number_format = nf
        wc(ws10, r, 3, "", bg=bg)
        wc(ws10, r, 4, note, bg=C_ALT, fg="595959", sz=9, align="left", italic=True)

    sec(ws10, 12, "ROIC-DERIVED GROWTH SCHEDULE (Growth = ROIC × Reinvestment Rate)", 12)
    yr_labels = [f"Y{i+1}" for i in range(forecast_yrs)]
    hdr(ws10, 13, ["Metric"] + yr_labels + ["Terminal"], bg=C_SUBHDR)

    schedule_rows = [
        ("ROIC (year-specific)",          roic_schedule,   "0.0%",   C_WHITE,   C_FORMULA),
        ("Reinvestment Rate",              reinv_schedule,  "0.0%",   C_INPUT_BG,C_INPUT_FG),
        ("Derived Growth (ROIC×Reinv)",   growth_derived,  "0.0%",   C_ALT,     C_FORMULA),
        ("FCF Conversion (1−Reinv Rate)", implied_fcf_conv,"0.0%",   C_WHITE,   C_GREEN),
        ("Moat Decay Factor (annual)",
         [round(moat_factor,3)]*forecast_yrs,              "0.000",  C_ALT,     "595959"),
    ]
    for k, (label, vals, nf, bg, fg) in enumerate(schedule_rows):
        r = 14 + k
        ws10.row_dimensions[r].height = 17
        wc(ws10, r, 1, label, bg=bg, fg=C_FORMULA, sz=10, align="left")
        for j, v in enumerate(vals):
            c = wc(ws10, r, j+2, v, bg=bg, fg=fg, sz=10)
            c.number_format = nf
        wc(ws10, r, forecast_yrs+2, terminal_g if "Growth" in label else (terminal_g/wacc if "Reinv" in label else wacc),
           bg="FFF2CC", fg=C_FORMULA, bold=True, sz=10, nf=nf)

    # Flag value destruction years
    sec(ws10, 20, "CAPITAL ALLOCATION QUALITY CHECK", 12)
    hdr(ws10, 21, ["Year", "ROIC", "WACC", "Spread", "Growth", "Value Created?", "Assessment"], bg=C_SUBHDR)
    for i in range(forecast_yrs):
        r = 22 + i
        ws10.row_dimensions[r].height = 16
        spread = roic_schedule[i] - wacc
        creates = spread > 0
        bg = C_BULL_BG if creates else C_BEAR_BG
        wc(ws10, r, 1, f"Year {i+1}",         bg=bg, fg=C_FORMULA, sz=9, align="center")
        wc(ws10, r, 2, roic_schedule[i],       bg=bg, fg=C_FORMULA, sz=9, nf="0.0%")
        wc(ws10, r, 3, wacc,                   bg=bg, fg=C_FORMULA, sz=9, nf="0.0%")
        wc(ws10, r, 4, spread,                 bg=bg, fg=C_GREEN if creates else C_RED, bold=True, sz=9, nf="+0.0%;(0.0%);-")
        wc(ws10, r, 5, growth_derived[i],      bg=bg, fg=C_FORMULA, sz=9, nf="0.0%")
        wc(ws10, r, 6, "✓ Creates" if creates else "✗ Destroys", bg=bg, fg=C_GREEN if creates else C_RED, bold=True, sz=9, align="center")
        assessment = "Moat intact" if roic_schedule[i] > wacc*1.5 else ("Moat eroding" if creates else "Value destruction — monitor closely")
        wc(ws10, r, 7, assessment, bg=bg, fg="595959", sz=9, align="left", italic=True)

    # ══════════════════════════════════════════════════════════
    # SHEET 11 — MONTE CARLO SIMULATION
    # ══════════════════════════════════════════════════════════
    ws11 = wb.create_sheet("11. Monte Carlo")
    ws11.sheet_view.showGridLines = False
    ws11.column_dimensions["A"].width = 30
    ws11.column_dimensions["B"].width = 18
    ws11.column_dimensions["C"].width = 18
    ws11.column_dimensions["D"].width = 18
    ws11.column_dimensions["E"].width = 18
    ws11.column_dimensions["F"].width = 20

    title(ws11, 1, f"MONTE CARLO SIMULATION — {ticker}",
          "1,000 Probabilistic Valuation Runs  |  Normal Distributions on Key Inputs", 6)

    # Run Monte Carlo
    np.random.seed(42)
    N = 1000

    # Input distributions
    g_mean  = base_g;          g_std   = max(base_g * 0.30, 0.03)
    m_mean  = op_margin;       m_std   = max(op_margin * 0.15, 0.02)
    w_mean  = wacc;            w_std   = 0.015
    tg_mean = terminal_g;      tg_std  = 0.005

    g_sims  = np.clip(np.random.normal(g_mean,  g_std,  N), -0.10, 0.40)
    m_sims  = np.clip(np.random.normal(m_mean,  m_std,  N),  0.02, 0.60)
    w_sims  = np.clip(np.random.normal(w_mean,  w_std,  N),  0.08, 0.22)
    tg_sims = np.clip(np.random.normal(tg_mean, tg_std, N),  0.01, 0.04)
    tg_sims = np.minimum(tg_sims, w_sims - 0.01)  # enforce tg < wacc

    iv_sims = []
    for i in range(N):
        g0 = g_sims[i]; m = m_sims[i]; w = w_sims[i]; tg = tg_sims[i]
        if w <= tg: tg = w - 0.01
        fcf_b = latest_rev * m * (1 - tax_rate) * 0.75
        if fcf_b <= 0: continue
        fcf = fcf_b; pv_sum = 0
        for yr in range(1, forecast_yrs + 1):
            yr_g = tg + (g0 - tg) * np.exp(-0.25 * yr)
            fcf *= (1 + yr_g)
            pv_sum += fcf / (1 + w) ** yr
        tv     = fcf * (1 + tg) / (w - tg)
        pv_tv  = tv / (1 + w) ** forecast_yrs
        ev     = pv_sum + pv_tv
        eq     = max(ev - total_debt + total_cash, 0)
        iv_ps  = eq * 1e9 / shares * fx if shares > 0 else 0
        if 0 < iv_ps < price * 15:  # filter outliers
            iv_sims.append(iv_ps)

    iv_arr  = np.array(iv_sims)
    mc_mean = float(np.mean(iv_arr))
    mc_med  = float(np.median(iv_arr))
    mc_std  = float(np.std(iv_arr))
    mc_p5   = float(np.percentile(iv_arr, 5))
    mc_p25  = float(np.percentile(iv_arr, 25))
    mc_p75  = float(np.percentile(iv_arr, 75))
    mc_p95  = float(np.percentile(iv_arr, 95))
    prob_under = float(np.mean(iv_arr > price))

    sec(ws11, 3, "STEP 3 — INPUT DISTRIBUTIONS (What we're simulating)", 6)
    hdr(ws11, 4, ["Input Variable", "Mean", "Std Dev", "Min (sim)", "Max (sim)", "Distribution"], bg=C_SUBHDR)
    dist_rows = [
        ("FCF / Revenue Growth Rate",  g_mean, g_std,  float(np.min(g_sims)),  float(np.max(g_sims)),  "Normal"),
        ("Operating Margin",           m_mean, m_std,  float(np.min(m_sims)),  float(np.max(m_sims)),  "Normal"),
        ("WACC",                       w_mean, w_std,  float(np.min(w_sims)),  float(np.max(w_sims)),  "Normal"),
        ("Terminal Growth Rate",       tg_mean,tg_std, float(np.min(tg_sims)), float(np.max(tg_sims)), "Normal, capped < WACC"),
    ]
    for i, (label, mean, std, mn, mx, dist) in enumerate(dist_rows):
        r = 5 + i
        ws11.row_dimensions[r].height = 17
        bg = C_ALT if i%2==0 else C_WHITE
        wc(ws11, r, 1, label, bg=bg, fg=C_FORMULA, sz=10, align="left")
        inp(ws11, r, 2, mean).number_format = "0.0%"
        wc(ws11, r, 3, std,  bg=bg, fg=C_FORMULA, sz=10, nf="0.0%")
        wc(ws11, r, 4, mn,   bg=bg, fg=C_FORMULA, sz=10, nf="0.0%")
        wc(ws11, r, 5, mx,   bg=bg, fg=C_FORMULA, sz=10, nf="0.0%")
        wc(ws11, r, 6, dist, bg=bg, fg="595959",  sz=9, align="left", italic=True)

    sec(ws11, 10, f"STEP 3 — SIMULATION RESULTS ({len(iv_sims):,} valid runs of {N:,} total)", 6)
    hdr(ws11, 11, ["Statistic", "Value", "", "Interpretation", "", ""], bg=C_SUBHDR)

    mc_results = [
        ("Mean Intrinsic Value",        mc_mean,  f'{sym}#,##0',  C_WHITE,      C_FORMULA,
         "Expected value across all scenarios"),
        ("Median Intrinsic Value",      mc_med,   f'{sym}#,##0',  C_WHITE,      C_FORMULA,
         "50th percentile — central tendency"),
        ("Standard Deviation",          mc_std,   f'{sym}#,##0',  C_ALT,        C_FORMULA,
         f"Uncertainty spread — {mc_std/mc_mean:.0%} of mean"),
        ("5th Percentile (Downside)",   mc_p5,    f'{sym}#,##0',  C_BEAR_BG,    C_RED,
         "Worst-case realistic scenario (1-in-20)"),
        ("25th Percentile",             mc_p25,   f'{sym}#,##0',  C_ALT,        C_FORMULA,
         "Conservative scenario"),
        ("75th Percentile",             mc_p75,   f'{sym}#,##0',  C_ALT,        C_FORMULA,
         "Optimistic scenario"),
        ("95th Percentile (Upside)",    mc_p95,   f'{sym}#,##0',  C_BULL_BG,    C_GREEN,
         "Best-case realistic scenario (1-in-20)"),
        ("Current Market Price",        price,    f'{sym}#,##0',  C_YELLOW,     C_INPUT_FG,
         "Reference — what market is paying today"),
        ("P(IV > Price) — Upside Prob.",prob_under,"0.0%",        C_YELLOW,     C_INPUT_FG if prob_under>0.5 else C_RED,
         f"{'Majority of scenarios show upside' if prob_under>0.5 else 'Majority of scenarios show downside'}"),
    ]
    for i, row_data in enumerate(mc_results):
        label, val, nf, bg, fg, note = row_data
        r = 12 + i
        ws11.row_dimensions[r].height = 18
        wc(ws11, r, 1, label, bg=bg, fg=fg, bold="Percentile" in label or "Price" in label, sz=10, align="left")
        c = wc(ws11, r, 2, round(val, 2), bg=bg, fg=fg, bold=True, sz=11)
        if nf: c.number_format = nf
        wc(ws11, r, 3, "", bg=bg)
        ws11.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        wc(ws11, r, 4, note, bg=C_ALT, fg="595959", sz=9, align="left", italic=True)

    # Distribution histogram (bucket the IV values)
    sec(ws11, 22, "IV DISTRIBUTION — HISTOGRAM DATA", 6)
    buckets = np.linspace(mc_p5 * 0.8, mc_p95 * 1.1, 20)
    counts  = np.histogram(iv_arr, bins=buckets)[0]
    hdr(ws11, 23, ["IV Bucket (midpoint)", "Count", "% of Sims", "vs Price", "", ""], bg=C_SUBHDR)
    for i, (cnt, lo, hi) in enumerate(zip(counts, buckets[:-1], buckets[1:])):
        r = 24 + i
        mid = (lo + hi) / 2
        pct = cnt / len(iv_sims)
        above = mid > price
        bg = C_BULL_BG if above else C_BEAR_BG
        ws11.row_dimensions[r].height = 15
        c = wc(ws11, r, 1, round(mid, 0), bg=bg, fg=C_FORMULA, sz=9, nf=f'"{sym}"#,##0')
        wc(ws11, r, 2, int(cnt),           bg=bg, fg=C_FORMULA, sz=9, nf="#,##0")
        wc(ws11, r, 3, pct,                bg=bg, fg=C_FORMULA, sz=9, nf="0.0%")
        wc(ws11, r, 4, "↑ UPSIDE" if above else "↓ BELOW",
           bg=bg, fg=C_GREEN if above else C_RED, bold=True, sz=9, align="center")

    # Add bar chart for distribution
    chart = BarChart()
    chart.type = "col"; chart.grouping = "clustered"
    chart.title = f"Monte Carlo IV Distribution — {ticker}"
    chart.y_axis.title = "Frequency (# simulations)"
    chart.x_axis.title = f"Intrinsic Value ({sym})"
    chart.style = 2; chart.width = 22; chart.height = 14
    data_ref = Reference(ws11, min_col=2, min_row=23, max_row=23+len(counts))
    cat_ref  = Reference(ws11, min_col=1, min_row=24, max_row=23+len(counts))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    ws11.add_chart(chart, "H3")

    # ══════════════════════════════════════════════════════════
    # SHEET 12 — DOWNSIDE & UPSIDE ASYMMETRY
    # ══════════════════════════════════════════════════════════
    ws12 = wb.create_sheet("12. Risk-Reward")
    ws12.sheet_view.showGridLines = False
    ws12.column_dimensions["A"].width = 36
    ws12.column_dimensions["B"].width = 20
    ws12.column_dimensions["C"].width = 20
    ws12.column_dimensions["D"].width = 35

    title(ws12, 1, f"DOWNSIDE PROTECTION & UPSIDE ASYMMETRY — {ticker}",
          "Step 7 & 8: Hedge Fund Risk/Reward Framework", 4)

    # Step 7 — Downside
    sec(ws12, 3, "STEP 7 — DOWNSIDE PROTECTION (What can go wrong?)", 4)
    hdr(ws12, 4, ["Risk Scenario", "IV Estimate", "% vs Price", "Trigger"], bg=C_SUBHDR, fg=C_HEADER_FG)

    downside_scenarios = [
        ("MC 5th Percentile (1-in-20 bad)",        mc_p5,     "Historical tail-risk materialises"),
        ("Bear DCF Scenario",                      bear_iv,   "Growth misses + margin compression"),
        ("Zero Terminal Value (no going concern)", max(dcf_res.get("sum_pv_fcfs",0)*fx/1e9 * 1e9/shares*fx if shares>0 else 0, 0), "Regulatory/structural disruption"),
        ("P/E Bear (sector trough multiple)",      enriched.get("latest_fcf",0)*fx/shares * 8 if shares>0 else 0, "Re-rating to trough valuation"),
        ("Liquidation Value (assets - debt)",      max((total_cash - total_debt) * 1e9 / shares * fx if shares>0 else 0, 0), "Complete business failure"),
    ]
    for i, (label, iv_est, trigger) in enumerate(downside_scenarios):
        r = 5 + i
        ws12.row_dimensions[r].height = 18
        pct_vs = (iv_est - price) / price if price > 0 else 0
        fg  = C_RED if pct_vs < -0.20 else (C_AMBER if pct_vs < 0 else C_GREEN)
        wc(ws12, r, 1, label,   bg=C_BEAR_BG, fg=C_FORMULA, sz=10, align="left")
        c = wc(ws12, r, 2, round(iv_est, 0), bg=C_BEAR_BG, fg=fg, bold=True, sz=10, nf=f'"{sym}"#,##0')
        wc(ws12, r, 3, pct_vs, bg=C_BEAR_BG, fg=fg, bold=True, sz=10, nf="+0.0%;(0.0%);-")
        wc(ws12, r, 4, trigger, bg=C_ALT,     fg="595959",  sz=9, align="left", italic=True)

    worst_case = min(mc_p5, bear_iv) if bear_iv > 0 else mc_p5
    max_loss   = (worst_case - price) / price if price > 0 else 0
    ws12.row_dimensions[11].height = 20
    wc(ws12, 11, 1, "MAX REALISTIC DOWNSIDE",   bg=C_BEAR_BG, fg=C_RED, bold=True, sz=11, align="left")
    c = wc(ws12, 11, 2, worst_case, bg=C_BEAR_BG, fg=C_RED, bold=True, sz=12, nf=f'"{sym}"#,##0')
    wc(ws12, 11, 3, max_loss, bg=C_BEAR_BG, fg=C_RED, bold=True, sz=12, nf="0.0%;(0.0%);-")
    wc(ws12, 11, 4, "Model worst-case floor estimate", bg=C_BEAR_BG, fg="595959", sz=9, italic=True, align="left")

    # Step 8 — Upside
    sec(ws12, 13, "STEP 8 — UPSIDE ASYMMETRY (Is the upside worth the risk?)", 4)
    hdr(ws12, 14, ["Upside Scenario", "IV Estimate", "% vs Price", "Catalyst"], bg=C_SUBHDR)

    upside_scenarios = [
        ("MC 95th Percentile (1-in-20 great)",     mc_p95,  "Everything goes right"),
        ("Bull DCF Scenario",                      bull_iv, "Moat widens, margins expand"),
        ("MC Mean (probability-weighted centre)",  mc_mean, "Expected value of all 1000 sims"),
        ("P/E Bull (sector peak multiple)",        enriched.get("latest_fcf",0)*fx/shares * 25 if shares>0 else 0, "Re-rating to peak sector multiple"),
    ]
    for i, (label, iv_est, catalyst) in enumerate(upside_scenarios):
        r = 15 + i
        ws12.row_dimensions[r].height = 18
        pct_vs = (iv_est - price) / price if price > 0 else 0
        fg = C_GREEN if pct_vs > 0.20 else (C_AMBER if pct_vs > 0 else C_RED)
        wc(ws12, r, 1, label,    bg=C_BULL_BG, fg=C_FORMULA, sz=10, align="left")
        wc(ws12, r, 2, round(iv_est,0), bg=C_BULL_BG, fg=fg, bold=True, sz=10, nf=f'"{sym}"#,##0')
        wc(ws12, r, 3, pct_vs,  bg=C_BULL_BG, fg=fg, bold=True, sz=10, nf="+0.0%;(0.0%);-")
        wc(ws12, r, 4, catalyst, bg=C_ALT,     fg="595959",  sz=9, align="left", italic=True)

    best_case   = max(mc_p95, bull_iv)
    max_gain    = (best_case - price) / price if price > 0 else 0
    ws12.row_dimensions[20].height = 20
    wc(ws12, 20, 1, "MAX REALISTIC UPSIDE",  bg=C_BULL_BG, fg=C_GREEN, bold=True, sz=11, align="left")
    wc(ws12, 20, 2, best_case, bg=C_BULL_BG, fg=C_GREEN, bold=True, sz=12, nf=f'"{sym}"#,##0')
    wc(ws12, 20, 3, max_gain,  bg=C_BULL_BG, fg=C_GREEN, bold=True, sz=12, nf="+0.0%;(0.0%);-")
    wc(ws12, 20, 4, "Best-case realisation if moat expands + catalyst", bg=C_BULL_BG, fg="595959", sz=9, italic=True, align="left")

    # Asymmetry ratio
    asymmetry = abs(max_gain / max_loss) if max_loss < 0 else 999
    sec(ws12, 22, "RISK/REWARD ASYMMETRY RATIO", 4)
    ws12.row_dimensions[23].height = 28
    wc(ws12, 23, 1, "Upside / Downside Ratio",  bg=C_YELLOW, fg=C_FORMULA, bold=True, sz=12, align="left")
    c = wc(ws12, 23, 2, round(asymmetry, 2),   bg=C_YELLOW, fg=C_GREEN if asymmetry > 2 else (C_AMBER if asymmetry > 1 else C_RED), bold=True, sz=14, nf="0.0x")
    wc(ws12, 23, 3, "Target: >2x for institutional investment", bg=C_YELLOW, fg="595959", sz=9, italic=True, align="left")
    verdict_asym = ("FAVOURABLE — Asymmetric upside. Risk/reward supports position." if asymmetry > 2
                    else "BALANCED — Symmetric risk/reward. Conviction required." if asymmetry > 1
                    else "UNFAVOURABLE — Downside exceeds upside. Caution advised.")
    ws12.row_dimensions[24].height = 22
    ws12.merge_cells("A24:D24")
    c24 = ws12.cell(24, 1, value=f"  {verdict_asym}")
    c24.fill = hf(C_BULL_BG if asymmetry > 2 else (C_YELLOW if asymmetry > 1 else C_BEAR_BG))
    c24.font = Font(name="Calibri", bold=True, size=11, color=_c(C_GREEN if asymmetry > 2 else (C_AMBER if asymmetry > 1 else C_RED)))
    c24.alignment = Alignment(horizontal="left", vertical="center")
    c24.border = bdr()

    # ══════════════════════════════════════════════════════════
    # SHEET 13 — FINAL INVESTMENT FRAMEWORK
    # ══════════════════════════════════════════════════════════
    ws13 = wb.create_sheet("13. HF Conclusion")
    ws13.sheet_view.showGridLines = False
    ws13.column_dimensions["A"].width = 34
    ws13.column_dimensions["B"].width = 22
    ws13.column_dimensions["C"].width = 38

    title(ws13, 1, f"HEDGE FUND INVESTMENT CONCLUSION — {ticker}",
          f"Sector: {sector}  |  Moat: {moat}  |  {datetime.now().strftime('%d %b %Y')}", 3)

    # Final call logic
    mos_base   = (base_iv - price) / price if price > 0 else 0
    mos_pw     = (pw_iv   - price) / price if price > 0 else 0
    mos_mc     = (mc_mean - price) / price if price > 0 else 0

    if mos_pw > 0.30 and asymmetry > 2.5 and prob_under > 0.65:
        final_call = "STRONG UNDERVALUED"
        call_color = C_GREEN
        call_bg    = C_BULL_BG
    elif mos_pw > 0.15 and asymmetry > 1.5:
        final_call = "UNDERVALUED"
        call_color = "00B050"
        call_bg    = "CCFFCC"
    elif mos_pw > -0.10:
        final_call = "NEAR FAIR VALUE"
        call_color = C_AMBER
        call_bg    = "FFF2CC"
    elif mos_pw > -0.25 and asymmetry < 1:
        final_call = "OVERVALUED"
        call_color = C_RED
        call_bg    = C_BEAR_BG
    else:
        final_call = "SIGNIFICANTLY OVERVALUED"
        call_color = C_RED
        call_bg    = "FCE4D6"

    # Final call banner
    ws13.merge_cells("A3:C5")
    c_call = ws13.cell(3, 1, value=f"  {final_call}")
    c_call.fill = hf(call_bg)
    c_call.font = Font(name="Calibri", bold=True, size=32, color=_c(call_color))
    c_call.alignment = Alignment(horizontal="center", vertical="center")
    ws13.row_dimensions[3].height = 70

    sec(ws13, 6, "STEP 9 — COMPLETE HEDGE FUND INVESTMENT FRAMEWORK", 3)

    # Section 1: Market-implied
    sec(ws13, 7, "① MARKET-IMPLIED EXPECTATIONS", 3)
    hdr(ws13, 8, ["What market is pricing in", "Implied Value", "My Base Case"], bg=C_SUBHDR)
    mkt_rows = [
        ("Implied FCF Growth",    implied_g,      rev_growth),
        ("Implied Op Margin",     implied_margin, op_margin),
        ("Current P/FCF",         price / (latest_fcf*1e9/shares*fx) if latest_fcf*shares>0 else 0, None),
    ]
    for i, (lbl, impl, base) in enumerate(mkt_rows):
        r = 9 + i
        ws13.row_dimensions[r].height = 17
        bg = C_ALT if i%2==0 else C_WHITE
        wc(ws13, r, 1, lbl,  bg=bg, fg=C_FORMULA, sz=10, align="left")
        nf = "0.0%" if i < 2 else "0.0x"
        wc(ws13, r, 2, impl, bg=C_YELLOW, fg=C_AMBER, bold=True, sz=10, nf=nf)
        if base: wc(ws13, r, 3, base, bg=bg, fg=C_INPUT_FG, sz=10, nf=nf)
        else: wc(ws13, r, 3, "—", bg=bg, fg="595959", sz=10, align="center")

    # Section 2: Valuation range
    sec(ws13, 13, "② INTRINSIC VALUATION RANGE", 3)
    hdr(ws13, 14, ["Methodology", "Intrinsic Value", "vs Current Price"], bg=C_SUBHDR)
    val_rows = [
        ("DCF Base Case",               base_iv,  (base_iv-price)/price if price>0 else 0),
        ("Probability-Weighted (B/B/B)",pw_iv,    (pw_iv-price)/price if price>0 else 0),
        ("Monte Carlo Mean",            mc_mean,  (mc_mean-price)/price if price>0 else 0),
        ("MC 5th Pct (Downside floor)", mc_p5,    (mc_p5-price)/price if price>0 else 0),
        ("MC 95th Pct (Upside ceiling)",mc_p95,   (mc_p95-price)/price if price>0 else 0),
    ]
    for i, (meth, iv_v, pct) in enumerate(val_rows):
        r = 15 + i
        ws13.row_dimensions[r].height = 17
        is_pw = "Prob" in meth
        bg = C_YELLOW if is_pw else (C_ALT if i%2==0 else C_WHITE)
        fg_pct = C_GREEN if pct > 0.10 else (C_AMBER if pct > -0.10 else C_RED)
        wc(ws13, r, 1, meth,           bg=bg, fg=C_FORMULA, bold=is_pw, sz=10, align="left")
        wc(ws13, r, 2, round(iv_v,0),  bg=bg, fg=C_INPUT_FG if is_pw else C_FORMULA, bold=is_pw, sz=10, nf=f'"{sym}"#,##0')
        wc(ws13, r, 3, pct,            bg=bg, fg=fg_pct, bold=is_pw, sz=10, nf="+0.0%;(0.0%);-")

    # Section 3: Risk/reward summary
    sec(ws13, 21, "③ DOWNSIDE vs UPSIDE", 3)
    rr_rows = [
        ("Max Realistic Downside",  worst_case, max_loss,  C_BEAR_BG, C_RED),
        ("Max Realistic Upside",    best_case,  max_gain,  C_BULL_BG, C_GREEN),
        ("Risk/Reward Ratio",       asymmetry,  None,      C_YELLOW,  C_AMBER),
        ("Prob(Undervalued) — MC",  prob_under, None,      C_YELLOW,  C_GREEN if prob_under>0.5 else C_RED),
    ]
    for i, (lbl, val1, val2, bg, fg) in enumerate(rr_rows):
        r = 22 + i
        ws13.row_dimensions[r].height = 18
        wc(ws13, r, 1, lbl, bg=bg, fg=C_FORMULA, bold=True, sz=10, align="left")
        nf1 = f'"{sym}"#,##0' if "Downside" in lbl or "Upside" in lbl else ("0.0x" if "Ratio" in lbl else "0.0%")
        wc(ws13, r, 2, round(float(val1),2), bg=bg, fg=fg, bold=True, sz=11, nf=nf1)
        if val2 is not None:
            wc(ws13, r, 3, val2, bg=bg, fg=fg, bold=True, sz=11, nf="+0.0%;(0.0%);-")
        else:
            wc(ws13, r, 3, "", bg=bg)

    # Section 4: Final written conclusion
    sec(ws13, 27, "④ HEDGE FUND INVESTMENT MEMO", 3)
    ws13.row_dimensions[28].height = 120
    ws13.merge_cells("A28:C28")

    memo = (
        f"TICKER: {ticker}  |  SECTOR: {sector}  |  MOAT: {moat}  |  DATE: {datetime.now().strftime('%d %b %Y')}\n\n"
        f"MARKET-IMPLIED EXPECTATIONS: At {sym}{price:,.0f}, the market prices in FCF growth of {implied_g:.1%} "
        f"(our base: {rev_growth:.1%}) and operating margin of {implied_margin:.1%} (current: {op_margin:.1%}). "
        f"{'The bar is HIGH — market optimism exceeds our base case.' if implied_g > rev_growth*1.2 else 'The market appears conservative relative to our base case.'}\n\n"
        f"VALUATION RANGE: Bear {sym}{mc_p5:,.0f} — Base {sym}{pw_iv:,.0f} — Bull {sym}{mc_p95:,.0f}. "
        f"Monte Carlo (N=1,000) gives mean IV of {sym}{mc_mean:,.0f} with {prob_under:.0%} probability of upside.\n\n"
        f"ROIC QUALITY: ROIC of {roic:.1%} vs WACC of {wacc:.1%} = spread of {roic_wacc_spread:+.1%}. "
        f"{'Growth CREATES value — reinvestment earns above cost of capital.' if roic_wacc_spread > 0 else 'WARNING: Growth DESTROYS value — ROIC below WACC. Prioritise capital return over growth.'}\n\n"
        f"RISK/REWARD: Downside {max_loss:.0%} / Upside {max_gain:+.0%} = {asymmetry:.1f}x asymmetry ratio. "
        f"{'Favourable asymmetry supports position sizing.' if asymmetry > 2 else 'Limited asymmetry — require wider margin of safety before commitment.'}\n\n"
        f"FINAL CALL: {final_call}"
    )
    c_memo = ws13.cell(28, 1, value=memo)
    c_memo.fill = hf("F8FBFF")
    c_memo.font = Font(name="Calibri", size=10, color=_c(C_FORMULA))
    c_memo.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c_memo.border = bdr("medium", C_NAVY)

    for ws in [ws9, ws10, ws11, ws12, ws13]:
        ws.freeze_panes = "B5"

    return wb


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--ticker", required=True)
    args = parser.parse_args()
    ticker = args.ticker.upper().strip()

    print(f"[1/6] Fetching data for {ticker}...")
    from data.collector import StockDataCollector
    from data.processor import compute_metrics
    from models.forecaster import FCFForecaster, compute_wacc
    from models.industry_wacc import get_industry_wacc
    from screener.dcf_engine import DCFEngine, margin_of_safety, assign_signal
    from screener.scenarios import run_scenarios
    from screener.valuation_model import generate_valuation_summary as generate_investment_plan
    from screener.moat_engine import compute_moat_score, apply_moat_adjustments
    from generate_dcf_excel import generate_institutional_dcf

    collector = StockDataCollector(ticker)
    raw = collector.get_all()
    if not raw:
        print(f"Could not fetch {ticker}"); return

    print(f"[2/6] Running DCF engine...")
    enriched      = compute_metrics(raw)
    forecaster    = FCFForecaster()
    wacc_data_raw = compute_wacc(collector._ticker_obj, is_indian=ticker.endswith((".NS",".BO")))
    ind_info      = get_industry_wacc(ticker=ticker, capm_wacc=wacc_data_raw.get("wacc",0.11))
    final_wacc    = ind_info["wacc"]
    terminal_g    = ind_info["terminal_growth"]
    forecast_yrs  = 10

    dcf_engine      = DCFEngine(discount_rate=final_wacc, terminal_growth=terminal_g)
    forecast_result = forecaster.predict(enriched, years=forecast_yrs)
    dcf_res         = dcf_engine.intrinsic_value_per_share(
        projected_fcfs=forecast_result["projections"],
        terminal_fcf_norm=forecast_result["terminal_fcf_norm"],
        total_debt=enriched["total_debt"], total_cash=enriched["total_cash"],
        shares_outstanding=enriched["shares"], current_price=enriched["price"],
        ticker=ticker,
    )

    print(f"[3/6] Computing moat & scenarios...")
    moat_result = compute_moat_score(enriched, final_wacc)
    enriched["moat_grade"]        = moat_result.get("grade", "None")
    enriched["fundamental_grade"] = "N/A"
    enriched["fundamental_score"] = 0

    iv_n     = dcf_res.get("intrinsic_value_per_share", 0)
    price_n  = enriched["price"]
    mos      = margin_of_safety(iv_n, price_n)
    inv_plan = generate_investment_plan(enriched, price_n, iv_n, mos)
    sig      = assign_signal(mos, dcf_res.get("suspicious",False), forecast_result.get("reliable",True))
    scenarios = run_scenarios(
        enriched=enriched, fcf_base=forecast_result.get("fcf_base",1e9),
        base_growth=forecast_result.get("base_growth",0.08),
        base_wacc=final_wacc, base_terminal_g=terminal_g,
        total_debt=enriched["total_debt"], total_cash=enriched["total_cash"],
        shares=enriched["shares"], current_price=price_n, years=forecast_yrs,
    )
    report_data = {
        "price":   price_n, "iv": iv_n, "mos_pct": mos*100, "signal": sig,
        "bear_iv": scenarios.get("Bear 🐻",{}).get("iv", iv_n*0.7),
        "bull_iv": scenarios.get("Bull 🐂",{}).get("iv", iv_n*1.3),
    }

    print(f"[4/6] Building institutional DCF sheets...")
    excel_bytes = generate_institutional_dcf(
        ticker=ticker, enriched=enriched, dcf_res=dcf_res,
        forecast_result=forecast_result, scenarios=scenarios,
        wacc_data=wacc_data_raw, wacc=final_wacc, terminal_g=terminal_g,
        forecast_yrs=forecast_yrs, sym="₹", to_code="INR", fx=1.0,
    )

    print(f"[5/6] Adding hedge fund sheets (Reverse DCF, Monte Carlo, Risk/Reward)...")
    from openpyxl import load_workbook
    wb = load_workbook(filename=__import__("io").BytesIO(excel_bytes))
    wb = build_hedge_fund_sheets(
        wb=wb, ticker=ticker, enriched=enriched, dcf_res=dcf_res,
        forecast_result=forecast_result, scenarios=scenarios,
        wacc_data=wacc_data_raw, wacc=final_wacc, terminal_g=terminal_g,
        forecast_yrs=forecast_yrs, sym="₹", fx=1.0,
    )

    print(f"[6/6] Saving file...")
    out_path = f"{ticker.replace('.','_')}_HedgeFund_DCF_{datetime.now().strftime('%Y%m%d')}.xlsx"
    buf = __import__("io").BytesIO()
    wb.save(buf)
    with open(out_path, "wb") as f:
        f.write(buf.getvalue())

    print(f"\n✅ Saved: {out_path}")
    print(f"\nSheets included:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"  {name}")


if __name__ == "__main__":
    main()
