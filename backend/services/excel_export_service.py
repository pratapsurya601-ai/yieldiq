"""
Excel export service for the /analysis page.

Builds a multi-sheet workbook where the DCF projection is *formula-driven*
from a small Inputs sheet (WACC / terminal growth / forecast years / base
FCF / shares / net debt). The user can change any blue cell on the
Inputs tab and Excel re-runs the projection live without re-hitting the
backend.

Why a separate, slimmer builder than ``generate_dcf_excel.py``:
  * That script reuses a full ``enriched`` payload (income/CF/BS pandas
    frames) which is only available on the local fundamentals worker.
  * The hot API path only has access to the cached AnalysisResponse
    (Pydantic). Re-fetching pandas frames per request would defeat the
    whole point of the analysis cache.
  * The institutional formatting in generate_dcf_excel.py is preserved
    here as a thin reuse of its colour palette + helper functions —
    we share the *visual* convention, not the data path.

Sheets:
  1. Inputs    — editable WACC / g / forecast years / base FCF / shares
                 / net debt. Blue = input convention.
  2. DCF       — formula-driven FCF projection, PV table, terminal value,
                 EV → Equity → IV/share, MoS. Every cell references
                 the Inputs sheet by name so flipping g from 3% to 4% on
                 the Inputs tab updates intrinsic value live.
  3. Scenarios — Bear / Base / Bull cases pulled from
                 ``ValuationOutput.bear_case / base_case / bull_case``
                 plus the per-scenario WACC + growth (when present in
                 ``ScenariosOutput``).
  4. Source    — raw inputs the model used (company info, valuation
                 metadata, freshness stamps) for audit / sanity-check.

Pro-tier gated: the route handler in ``backend/routers/analysis.py``
returns 402 ``Payment Required`` for free-tier users; this module is
caller-trusted (it does not re-check the tier).
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# Reuse the institutional colour palette from generate_dcf_excel.py so
# the two exports look like they came from the same shop. Only the
# constants are imported — the script's heavy pandas-driven builder is
# not used here.
C_INPUT_BG    = "EBF3FB"
C_INPUT_FG    = "0000FF"
C_FORMULA_FG  = "000000"
C_LINK_FG     = "008000"
C_HEADER_BG   = "1F3864"
C_HEADER_FG   = "FFFFFF"
C_SUBHDR_BG   = "2E75B6"
C_SECTION_BG  = "D6E4F0"
C_ALT_ROW     = "F2F7FB"
C_WHITE       = "FFFFFF"
C_YELLOW_FLAG = "FFFF00"
C_GREEN_POS   = "00B050"
C_RED_NEG     = "C00000"


# ── tiny helpers ────────────────────────────────────────────────
def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor="FF" + hex6)


def _font(bold: bool = False, sz: int = 10, color: str = "000000",
          italic: bool = False) -> Font:
    return Font(name="Calibri", bold=bold, size=sz,
                color="FF" + color, italic=italic)


def _border(weight: str = "thin") -> Border:
    s = Side(style=weight, color="FFBFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _set(ws, row, col, value, *, bg=None, fg="000000", bold=False, sz=10,
         nf: str | None = None, h: str = "right", italic: bool = False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, sz=sz, color=fg, italic=italic)
    c.alignment = Alignment(horizontal=h, vertical="center")
    c.border = _border()
    if bg:
        c.fill = _fill(bg)
    if nf:
        c.number_format = nf
    return c


def _section(ws, row, text, ncols, bg=C_SECTION_BG, fg=C_HEADER_BG):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=f"  {text}")
    c.fill = _fill(bg)
    c.font = _font(bold=True, sz=10, color=fg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _border("medium")
    ws.row_dimensions[row].height = 18


def _title(ws, row, ticker: str, subtitle: str, ncols: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1,
                value=f"YieldIQ DCF Model — {ticker}")
    c.fill = _fill(C_HEADER_BG)
    c.font = _font(bold=True, sz=14, color=C_HEADER_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

    ws.merge_cells(start_row=row + 1, start_column=1,
                   end_row=row + 1, end_column=ncols)
    c2 = ws.cell(row=row + 1, column=1, value=subtitle)
    c2.fill = _fill(C_SUBHDR_BG)
    c2.font = _font(sz=9, color=C_HEADER_FG, italic=True)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row + 1].height = 16


# ── public entry point ──────────────────────────────────────────
def build_workbook(analysis: Any) -> bytes:
    """Return an XLSX file as bytes for the given AnalysisResponse.

    ``analysis`` is the Pydantic ``AnalysisResponse`` model (or a dict
    matching its shape — see backend/models/responses.py).
    """
    # Tolerate both Pydantic and dict input so cached payloads (which
    # arrive as dicts on the fast-path) can also be exported without an
    # extra model_validate step.
    def _g(obj, key, default=None):
        if isinstance(obj, dict):
            return obj.get(key, default)
        return getattr(obj, key, default)

    ticker      = _g(analysis, "ticker", "TICKER")
    company     = _g(analysis, "company", {}) or {}
    valuation   = _g(analysis, "valuation", {}) or {}
    scenarios   = _g(analysis, "scenarios", {}) or {}
    quality     = _g(analysis, "quality", {}) or {}

    company_name = _g(company, "name", ticker)
    sector       = _g(company, "sector", "—")
    currency     = _g(company, "currency", "INR") or "INR"
    market_cap   = float(_g(company, "market_cap", 0) or 0)
    shares_out   = float(_g(company, "shares_outstanding", 0) or 0)

    current_price   = float(_g(valuation, "current_price", 0) or 0)
    fair_value      = float(_g(valuation, "fair_value", 0) or 0)
    wacc            = float(_g(valuation, "wacc", 0.12) or 0.12)
    terminal_g      = float(_g(valuation, "terminal_growth", 0.03) or 0.03)
    fcf_growth      = float(_g(valuation, "fcf_growth_rate", 0.08) or 0.08)
    enterprise_val  = float(_g(valuation, "enterprise_value", 0) or 0)
    equity_val      = float(_g(valuation, "equity_value", 0) or 0)
    pv_fcfs_total   = float(_g(valuation, "pv_fcfs", 0) or 0)
    pv_terminal     = float(_g(valuation, "pv_terminal", 0) or 0)
    valuation_model = _g(valuation, "valuation_model", "dcf")

    # Net debt = EV - Equity (when both are present); else 0 fallback.
    net_debt = max(enterprise_val - equity_val, 0.0) if (
        enterprise_val and equity_val
    ) else 0.0

    # Base FCF: when ev = sum(pv_fcf) + pv_terminal, the base FCF that
    # produced that PV stream is approximately pv_fcfs_total / sum of
    # discount factors over the forecast horizon. We invert that here so
    # the user's editable Inputs.base_fcf matches the model's working
    # assumption. Fallback to a conservative estimate when DCF metadata
    # is missing (e.g. financials valued via P/BV).
    forecast_yrs = 10
    if pv_fcfs_total > 0 and wacc > 0 and fcf_growth is not None:
        # Approximate base FCF assuming a flat growth-rate projection.
        # discount_factor_sum = sum 1/(1+wacc)^t for t=1..N when growth=g
        # PV(FCF) ≈ FCF0 * sum((1+g)^t / (1+wacc)^t)
        ratio = (1 + fcf_growth) / (1 + wacc) if wacc != fcf_growth else 1.0
        if abs(ratio - 1.0) < 1e-9:
            denom = forecast_yrs
        else:
            denom = ratio * (1 - ratio ** forecast_yrs) / (1 - ratio)
        base_fcf = pv_fcfs_total / denom if denom > 0 else 0.0
    else:
        # Fallback for non-DCF (P/BV financials path) — set base FCF to 0
        # so the user can plug a real number in. The Inputs sheet text
        # explicitly notes when the model used a non-DCF path.
        base_fcf = 0.0

    wb = Workbook()

    # ════════════════════════════════════════════════════════════
    # SHEET 1 — INPUTS
    # ════════════════════════════════════════════════════════════
    ws_in = wb.active
    ws_in.title = "Inputs"
    ws_in.sheet_view.showGridLines = False
    ws_in.column_dimensions["A"].width = 36
    ws_in.column_dimensions["B"].width = 18
    ws_in.column_dimensions["C"].width = 12
    ws_in.column_dimensions["D"].width = 50

    _title(
        ws_in, 1, ticker,
        f"{company_name}  |  Sector: {sector}  |  Currency: {currency}  "
        f"|  Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}",
        4,
    )

    _section(ws_in, 4, "EDITABLE ASSUMPTIONS  (blue cells — change these)", 4)
    # row → (label, value, number_format, note)  ; blue = input
    input_rows: list[tuple[str, Any, str, str]] = [
        ("WACC",                        wacc,       "0.00%",
         "Discount rate. Industry-band capped at the analysis layer."),
        ("Terminal growth (g)",         terminal_g, "0.00%",
         "Long-run nominal GDP. India hard-cap 4%."),
        ("Forecast horizon (years)",    forecast_yrs, "0",
         "Number of explicit projection years. Default 10."),
        ("Base FCF",                    base_fcf,   "#,##0.00",
         f"Year-0 free cash flow ({currency}). Inverted from model PV."),
        ("FCF growth rate (flat)",      fcf_growth, "0.00%",
         "Applied flat across the forecast horizon for simplicity."),
        ("Shares outstanding",          shares_out, "#,##0",
         "From company info — used for IV/share."),
        ("Net debt",                    net_debt,   "#,##0.00",
         f"EV − Equity ({currency}). 0 when model used a non-DCF path."),
        ("Current price",               current_price, "#,##0.00",
         f"Live price ({currency}) at compute time."),
    ]
    # Defined names (workbook-scoped) so the DCF sheet can reference
    # `WACC`, `Term_g`, etc. instead of `Inputs!$B$6`.
    name_map = {
        0: "WACC",
        1: "Term_g",
        2: "Yrs",
        3: "FCF0",
        4: "G_FCF",
        5: "Shares",
        6: "NetDebt",
        7: "CurrPrice",
    }
    start = 5
    for i, (label, val, nf, note) in enumerate(input_rows):
        r = start + i
        _set(ws_in, r, 1, label, bg=C_WHITE, fg=C_FORMULA_FG, h="left")
        c = _set(ws_in, r, 2, val, bg=C_INPUT_BG, fg=C_INPUT_FG,
                 bold=True, nf=nf)
        _set(ws_in, r, 3, "input", bg=C_ALT_ROW, fg="595959", sz=9, h="center",
             italic=True)
        _set(ws_in, r, 4, note, bg=C_ALT_ROW, fg="595959", sz=9, h="left",
             italic=True)
        if i in name_map:
            ref = f"Inputs!${get_column_letter(2)}${r}"
            wb.defined_names[name_map[i]] = DefinedName(
                name=name_map[i], attr_text=ref,
            )

    _section(ws_in, start + len(input_rows) + 1,
             "MODEL METADATA (read-only)", 4)
    meta_start = start + len(input_rows) + 2
    meta_rows = [
        ("Model used",
         valuation_model,
         "@",
         "dcf = full DCF, pb_ratio = peer-median P/BV (financials)."),
        ("Backend fair value",          fair_value,    "#,##0.00",
         f"What the YieldIQ engine returned ({currency})."),
        ("Backend MoS",
         float(_g(valuation, "margin_of_safety", 0) or 0) / 100.0,
         "0.0%",
         "(FV − price) / price."),
        ("Backend WACC band",
         f"{float(_g(valuation, 'wacc_industry_min', 0) or 0):.2%} – "
         f"{float(_g(valuation, 'wacc_industry_max', 0) or 0):.2%}",
         "@",
         "Industry-floor / ceiling clamp."),
        ("Confidence score",
         int(_g(valuation, "confidence_score", 0) or 0),
         "0",
         "0–100. Higher = more model agreement."),
        ("TV % of EV",
         float(_g(valuation, "tv_pct_of_ev", 0) or 0),
         "0.0%",
         "Flagged when > 75% (terminal-heavy)."),
    ]
    for i, (label, val, nf, note) in enumerate(meta_rows):
        r = meta_start + i
        _set(ws_in, r, 1, label, bg=C_WHITE, fg=C_FORMULA_FG, h="left")
        c = _set(ws_in, r, 2, val, bg=C_WHITE, fg=C_FORMULA_FG)
        if nf and nf != "@":
            c.number_format = nf
        _set(ws_in, r, 3, "meta", bg=C_ALT_ROW, fg="595959", sz=9,
             h="center", italic=True)
        _set(ws_in, r, 4, note, bg=C_ALT_ROW, fg="595959", sz=9, h="left",
             italic=True)

    # ════════════════════════════════════════════════════════════
    # SHEET 2 — DCF (formula-driven)
    # ════════════════════════════════════════════════════════════
    ws_dcf = wb.create_sheet("DCF")
    ws_dcf.sheet_view.showGridLines = False
    ws_dcf.column_dimensions["A"].width = 36
    # Pre-allocate enough columns for a 10-year forecast + terminal.
    for i in range(2, 14):
        ws_dcf.column_dimensions[get_column_letter(i)].width = 14

    _title(
        ws_dcf, 1, ticker,
        f"Formula-driven projection  |  WACC = =WACC  |  g = =Term_g  "
        f"|  All cells reference the Inputs sheet — change Inputs to re-run.",
        12,
    )

    _section(ws_dcf, 4, "FREE CASH FLOW PROJECTION", 12)
    # Year header row.
    _set(ws_dcf, 5, 1, "Year", bg=C_SUBHDR_BG, fg=C_HEADER_FG, bold=True,
         h="center")
    for t in range(1, forecast_yrs + 1):
        _set(ws_dcf, 5, 1 + t, f"Year {t}", bg=C_SUBHDR_BG, fg=C_HEADER_FG,
             bold=True, h="center")
    _set(ws_dcf, 5, 2 + forecast_yrs, "Terminal", bg=C_SUBHDR_BG,
         fg=C_HEADER_FG, bold=True, h="center")

    # Row 6 — projected FCF: FCF0 * (1+G_FCF)^t
    _set(ws_dcf, 6, 1, "Projected FCF", bg=C_WHITE, fg=C_FORMULA_FG, h="left")
    for t in range(1, forecast_yrs + 1):
        _set(ws_dcf, 6, 1 + t, f"=FCF0*(1+G_FCF)^{t}",
             bg=C_WHITE, fg=C_FORMULA_FG, nf="#,##0.00")
    # Terminal FCF column = FCF in last forecast year × (1+g)
    _set(ws_dcf, 6, 2 + forecast_yrs,
         f"=FCF0*(1+G_FCF)^{forecast_yrs}*(1+Term_g)",
         bg=C_WHITE, fg=C_FORMULA_FG, nf="#,##0.00")

    # Row 7 — discount factor: 1/(1+WACC)^t
    _set(ws_dcf, 7, 1, "Discount factor",
         bg=C_ALT_ROW, fg=C_FORMULA_FG, h="left")
    for t in range(1, forecast_yrs + 1):
        _set(ws_dcf, 7, 1 + t, f"=1/(1+WACC)^{t}",
             bg=C_ALT_ROW, fg=C_FORMULA_FG, nf="0.0000")
    _set(ws_dcf, 7, 2 + forecast_yrs, f"=1/(1+WACC)^{forecast_yrs}",
         bg=C_ALT_ROW, fg=C_FORMULA_FG, nf="0.0000")

    # Row 8 — PV(FCF) = projected × discount
    _set(ws_dcf, 8, 1, "PV of FCF", bg=C_WHITE, fg=C_FORMULA_FG, bold=True,
         h="left")
    for t in range(1, forecast_yrs + 1):
        col = get_column_letter(1 + t)
        _set(ws_dcf, 8, 1 + t, f"={col}6*{col}7",
             bg=C_WHITE, fg=C_FORMULA_FG, bold=True, nf="#,##0.00")

    # ── Terminal value block ────────────────────────────────────
    _section(ws_dcf, 10, "TERMINAL VALUE (Gordon growth)", 12)
    last_col = get_column_letter(2 + forecast_yrs)        # terminal col
    last_fc_col = get_column_letter(1 + forecast_yrs)     # last forecast col
    _set(ws_dcf, 11, 1, "Terminal FCF (Yr N+1)",
         bg=C_WHITE, fg=C_FORMULA_FG, h="left")
    _set(ws_dcf, 11, 2, f"={last_col}6",
         bg=C_WHITE, fg=C_FORMULA_FG, nf="#,##0.00")

    _set(ws_dcf, 12, 1, "WACC − g (spread)",
         bg=C_ALT_ROW, fg=C_FORMULA_FG, h="left")
    _set(ws_dcf, 12, 2, "=WACC-Term_g",
         bg=C_ALT_ROW, fg=C_FORMULA_FG, nf="0.00%")

    _set(ws_dcf, 13, 1, "Undiscounted Terminal Value",
         bg=C_WHITE, fg=C_FORMULA_FG, h="left")
    _set(ws_dcf, 13, 2, "=IFERROR(B11/B12,0)",
         bg=C_WHITE, fg=C_FORMULA_FG, nf="#,##0.00")

    _set(ws_dcf, 14, 1, "PV of Terminal Value",
         bg=C_YELLOW_FLAG, fg=C_FORMULA_FG, bold=True, h="left")
    _set(ws_dcf, 14, 2, f"=B13/(1+WACC)^{forecast_yrs}",
         bg=C_YELLOW_FLAG, fg=C_FORMULA_FG, bold=True, nf="#,##0.00")

    # ── Valuation bridge ────────────────────────────────────────
    _section(ws_dcf, 16, "VALUATION BRIDGE  →  Intrinsic Value / Share", 12)
    sum_pv_range = f"B8:{last_fc_col}8"
    _set(ws_dcf, 17, 1, "Σ PV of forecast FCFs",
         bg=C_WHITE, fg=C_FORMULA_FG, h="left")
    _set(ws_dcf, 17, 2, f"=SUM({sum_pv_range})",
         bg=C_WHITE, fg=C_FORMULA_FG, nf="#,##0.00")

    _set(ws_dcf, 18, 1, "+ PV of Terminal Value",
         bg=C_WHITE, fg=C_FORMULA_FG, h="left")
    _set(ws_dcf, 18, 2, "=B14",
         bg=C_WHITE, fg=C_FORMULA_FG, nf="#,##0.00")

    _set(ws_dcf, 19, 1, "= Enterprise Value",
         bg=C_SECTION_BG, fg=C_FORMULA_FG, bold=True, h="left")
    _set(ws_dcf, 19, 2, "=B17+B18",
         bg=C_SECTION_BG, fg=C_FORMULA_FG, bold=True, nf="#,##0.00")

    _set(ws_dcf, 20, 1, "− Net Debt",
         bg=C_WHITE, fg=C_RED_NEG, h="left")
    _set(ws_dcf, 20, 2, "=NetDebt",
         bg=C_WHITE, fg=C_RED_NEG, nf="#,##0.00")

    _set(ws_dcf, 21, 1, "= Equity Value",
         bg=C_SECTION_BG, fg=C_FORMULA_FG, bold=True, h="left")
    _set(ws_dcf, 21, 2, "=B19-B20",
         bg=C_SECTION_BG, fg=C_FORMULA_FG, bold=True, nf="#,##0.00")

    _set(ws_dcf, 22, 1, "÷ Shares Outstanding",
         bg=C_WHITE, fg=C_FORMULA_FG, h="left")
    _set(ws_dcf, 22, 2, "=Shares",
         bg=C_WHITE, fg=C_FORMULA_FG, nf="#,##0")

    _set(ws_dcf, 23, 1, "= Intrinsic Value / Share",
         bg=C_YELLOW_FLAG, fg=C_GREEN_POS, bold=True, h="left")
    _set(ws_dcf, 23, 2, "=IFERROR(B21/B22,0)",
         bg=C_YELLOW_FLAG, fg=C_GREEN_POS, bold=True, nf="#,##0.00")

    _set(ws_dcf, 24, 1, "Current Price",
         bg=C_WHITE, fg=C_INPUT_FG, h="left")
    _set(ws_dcf, 24, 2, "=CurrPrice",
         bg=C_WHITE, fg=C_INPUT_FG, nf="#,##0.00")

    _set(ws_dcf, 25, 1, "Margin of Safety",
         bg=C_YELLOW_FLAG, fg=C_FORMULA_FG, bold=True, h="left")
    _set(ws_dcf, 25, 2, "=IFERROR((B23-B24)/B24,0)",
         bg=C_YELLOW_FLAG, fg=C_FORMULA_FG, bold=True, nf="0.0%")

    # ════════════════════════════════════════════════════════════
    # SHEET 3 — SCENARIOS (Bear / Base / Bull)
    # ════════════════════════════════════════════════════════════
    ws_sc = wb.create_sheet("Scenarios")
    ws_sc.sheet_view.showGridLines = False
    ws_sc.column_dimensions["A"].width = 22
    for col in range(2, 6):
        ws_sc.column_dimensions[get_column_letter(col)].width = 16

    _title(ws_sc, 1, ticker,
           "Bear / Base / Bull — pulled from the YieldIQ analysis cache.",
           5)

    _section(ws_sc, 4, "SCENARIO COMPARISON", 5)
    headers = ["Scenario", "Intrinsic Value", "WACC", "Growth", "MoS"]
    for col, lbl in enumerate(headers, 1):
        _set(ws_sc, 5, col, lbl, bg=C_SUBHDR_BG, fg=C_HEADER_FG, bold=True,
             h="center")

    bear = _g(scenarios, "bear", {}) or {}
    base = _g(scenarios, "base", {}) or {}
    bull = _g(scenarios, "bull", {}) or {}

    def _scenario_row(r: int, label: str, case: Any, fallback_iv: float,
                      colour: str):
        iv = float(_g(case, "iv", 0) or 0) or fallback_iv
        wacc_s = float(_g(case, "wacc", 0) or 0) or wacc
        g_s = float(_g(case, "growth", 0) or 0) or fcf_growth
        mos_s = float(_g(case, "mos_pct", 0) or 0) / 100.0
        # If mos_pct missing on the scenario, derive from iv / current price.
        if not mos_s and current_price > 0:
            mos_s = (iv - current_price) / current_price if iv else 0.0
        _set(ws_sc, r, 1, label, bg=C_WHITE, fg=colour, bold=True, h="left")
        _set(ws_sc, r, 2, iv, bg=C_WHITE, fg=C_FORMULA_FG, nf="#,##0.00")
        _set(ws_sc, r, 3, wacc_s, bg=C_WHITE, fg=C_FORMULA_FG, nf="0.00%")
        _set(ws_sc, r, 4, g_s, bg=C_WHITE, fg=C_FORMULA_FG, nf="0.00%")
        _set(ws_sc, r, 5, mos_s, bg=C_WHITE,
             fg=C_GREEN_POS if mos_s > 0 else C_RED_NEG, nf="0.0%")

    _scenario_row(6, "Bear", bear,
                  float(_g(valuation, "bear_case", 0) or 0), C_RED_NEG)
    _scenario_row(7, "Base", base,
                  float(_g(valuation, "base_case", 0) or fair_value), C_FORMULA_FG)
    _scenario_row(8, "Bull", bull,
                  float(_g(valuation, "bull_case", 0) or 0), C_GREEN_POS)

    _set(ws_sc, 10, 1,
         "Tip: change WACC / Term_g / G_FCF on the Inputs sheet to re-run "
         "the DCF tab live. Scenario rows above are static snapshots from "
         "the backend at compute time.",
         bg=C_ALT_ROW, fg="595959", italic=True, h="left", sz=9)
    ws_sc.merge_cells(start_row=10, start_column=1, end_row=10, end_column=5)

    # ════════════════════════════════════════════════════════════
    # SHEET 4 — SOURCE DATA
    # ════════════════════════════════════════════════════════════
    ws_src = wb.create_sheet("Source Data")
    ws_src.sheet_view.showGridLines = False
    ws_src.column_dimensions["A"].width = 32
    ws_src.column_dimensions["B"].width = 38

    _title(ws_src, 1, ticker,
           "Raw inputs the model used. Audit / sanity-check.", 2)

    _section(ws_src, 4, "COMPANY", 2)
    company_rows = [
        ("Name",                _g(company, "name", "")),
        ("Ticker",              ticker),
        ("Sector",              _g(company, "sector", "")),
        ("Industry",            _g(company, "industry", "")),
        ("Currency",            currency),
        ("Market cap",          market_cap),
        ("Shares outstanding",  shares_out),
        ("Market-cap source",   _g(company, "market_cap_source", "")),
        ("Market-cap as-of",    _g(company, "market_cap_as_of", "")),
        ("Shares source",       _g(company, "shares_outstanding_source", "")),
    ]
    for i, (label, val) in enumerate(company_rows):
        r = 5 + i
        _set(ws_src, r, 1, label, bg=C_WHITE, fg=C_FORMULA_FG, h="left")
        _set(ws_src, r, 2, val if val is not None else "—",
             bg=C_ALT_ROW, fg=C_FORMULA_FG, h="left")

    next_section = 5 + len(company_rows) + 1
    _section(ws_src, next_section, "VALUATION (backend snapshot)", 2)
    val_rows = [
        ("Model used",                  valuation_model),
        ("Engine",                      _g(valuation, "valuation_engine_used", "")),
        ("Fair value",                  fair_value),
        ("Current price",               current_price),
        ("Margin of safety (%)",        _g(valuation, "margin_of_safety", 0)),
        ("Verdict",                     _g(valuation, "verdict", "")),
        ("WACC",                        wacc),
        ("Terminal growth",             terminal_g),
        ("FCF growth (modelled)",       fcf_growth),
        ("Enterprise value",            enterprise_val),
        ("Equity value",                equity_val),
        ("Σ PV(FCF)",                   pv_fcfs_total),
        ("PV(Terminal)",                pv_terminal),
        ("TV % of EV",                  _g(valuation, "tv_pct_of_ev", 0)),
        ("Confidence score",            _g(valuation, "confidence_score", 0)),
        ("FV computed at",              _g(valuation, "fair_value_computed_at", "")),
        ("Price source",                _g(valuation, "current_price_source", "")),
        ("Price as-of",                 _g(valuation, "current_price_as_of", "")),
    ]
    for i, (label, val) in enumerate(val_rows):
        r = next_section + 1 + i
        _set(ws_src, r, 1, label, bg=C_WHITE, fg=C_FORMULA_FG, h="left")
        _set(ws_src, r, 2, val if val is not None else "—",
             bg=C_ALT_ROW, fg=C_FORMULA_FG, h="left")

    next_section = next_section + 1 + len(val_rows) + 1
    _section(ws_src, next_section, "QUALITY (snapshot)", 2)
    quality_rows = [
        ("YieldIQ score",               _g(quality, "score", "")),
        ("Grade",                       _g(quality, "grade", "")),
        ("Piotroski",                   _g(quality, "piotroski_score", "")),
        ("Moat grade",                  _g(quality, "moat_grade", "")),
        ("Latest filing period",        _g(quality, "latest_filing_period_end", "")),
        ("Revenue CAGR window",         _g(quality, "revenue_cagr_window", "")),
    ]
    for i, (label, val) in enumerate(quality_rows):
        r = next_section + 1 + i
        _set(ws_src, r, 1, label, bg=C_WHITE, fg=C_FORMULA_FG, h="left")
        _set(ws_src, r, 2, val if val is not None else "—",
             bg=C_ALT_ROW, fg=C_FORMULA_FG, h="left")

    # If the analysis used a non-DCF model (financials path), make that
    # loud so the user doesn't think the formula-driven DCF tab is the
    # source-of-truth for this name.
    if valuation_model and valuation_model != "dcf":
        warn_row = next_section + 1 + len(quality_rows) + 2
        _set(ws_src, warn_row, 1,
             "NOTE: backend used non-DCF model",
             bg=C_YELLOW_FLAG, fg=C_RED_NEG, bold=True, h="left")
        _set(ws_src, warn_row, 2,
             f"Model = {valuation_model}. DCF tab is provided for "
             "scenario exploration only — backend FV came from the "
             "peer-multiple path.",
             bg=C_YELLOW_FLAG, fg=C_RED_NEG, h="left")

    # ── Serialise ───────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["build_workbook"]
