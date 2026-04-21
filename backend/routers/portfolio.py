# backend/routers/portfolio.py
from __future__ import annotations
import sys, os
import csv
import io
import logging
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel

_PROJECT_ROOT = str(Path(__file__).resolve().parent.parent.parent)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)
_DASHBOARD_ROOT = os.path.join(_PROJECT_ROOT, "dashboard")
if _DASHBOARD_ROOT not in sys.path:
    sys.path.insert(0, _DASHBOARD_ROOT)

from backend.models.requests import AddHoldingRequest
from backend.models.responses import (
    HoldingResponse, PortfolioHealthResponse, SuccessResponse,
)
from backend.middleware.auth import get_current_user

logger = logging.getLogger("yieldiq.portfolio")

router = APIRouter(prefix="/api/v1/portfolio", tags=["portfolio"])


class ImportCSVRequest(BaseModel):
    csv_text: str
    broker: str = "zerodha"  # zerodha, groww, upstox, icici, custom


@router.get("/holdings", response_model=list[HoldingResponse])
async def get_holdings(user: dict = Depends(get_current_user)):
    """Get all portfolio holdings for the authenticated user (raw, no live data)."""
    from backend.services.portfolio_service import get_holdings as _get
    email = user.get("email", "")
    if not email:
        return []
    holdings = _get(email)
    return [
        HoldingResponse(
            ticker=h.get("ticker", ""),
            company_name=h.get("company_name", ""),
            entry_price=h.get("entry_price", 0) or 0,
            iv=h.get("iv", 0) or 0,
            mos_pct=h.get("mos_pct", 0) or 0,
            signal=h.get("signal", "") or "",
            sector=h.get("sector", "") or "",
            notes=h.get("notes", "") or "",
            saved_at=str(h.get("saved_at", "") or ""),
        )
        for h in holdings
    ]


@router.get("/holdings-live")
async def get_holdings_live(user: dict = Depends(get_current_user)):
    """
    Get holdings enriched with live prices, P&L, fair value, and verdict.
    Returns: { holdings: [...], summary: {...} }
    """
    from backend.services.portfolio_service import get_holdings_with_live_data
    email = user.get("email", "")
    if not email:
        return {"holdings": [], "summary": {}}
    return get_holdings_with_live_data(email)


@router.post("/holdings", response_model=SuccessResponse)
async def add_holding(req: AddHoldingRequest, user: dict = Depends(get_current_user)):
    """Add a single stock to user's portfolio."""
    from backend.services.portfolio_service import save_holding
    email = user.get("email", "")
    if not email:
        raise HTTPException(status_code=401, detail="Email required")

    ok, err = save_holding(
        user_email=email,
        ticker=req.ticker,
        entry_price=req.entry_price,
        iv=req.iv,
        mos_pct=req.mos_pct,
        signal=req.signal,
        wacc=req.wacc,
        sector=req.sector,
        notes=req.notes,
    )
    if not ok:
        raise HTTPException(status_code=500, detail=f"Failed to save holding: {err}")
    return SuccessResponse(message=f"{req.ticker} added to portfolio")


@router.delete("/holdings/{ticker}", response_model=SuccessResponse)
async def remove_holding(ticker: str, user: dict = Depends(get_current_user)):
    """Remove a stock from user's portfolio."""
    from backend.services.portfolio_service import remove_holding as _remove
    email = user.get("email", "")
    if not email:
        raise HTTPException(status_code=401, detail="Email required")

    ok = _remove(email, ticker)
    if not ok:
        raise HTTPException(status_code=404, detail=f"{ticker} not found in portfolio")
    return SuccessResponse(message=f"{ticker} removed from portfolio")


def _do_import(parsed: list[dict], broker: str, user: dict) -> dict:
    """Shared import logic — accepts pre-parsed holdings list."""
    from backend.services.portfolio_service import save_holding
    from backend.services import analysis_service as svc
    from backend.services.cache_service import cache as _c

    email = user.get("email", "")
    tier = user.get("tier", "free")

    if not email:
        raise HTTPException(status_code=401, detail="Email required")

    if not parsed:
        raise HTTPException(status_code=400, detail="No valid holdings found")

    if tier == "free" and len(parsed) > 5:
        raise HTTPException(
            status_code=402,
            detail=f"Free tier limited to 5 holdings. Upgrade to Pro for unlimited import ({len(parsed)} holdings detected).",
        )

    imported = 0
    skipped = 0
    errors: list[str] = []

    for row in parsed:
        raw_ticker = row["ticker"]
        qty = row.get("quantity", 0)
        avg_cost = row.get("avg_cost", 0)
        if not raw_ticker or qty <= 0 or avg_cost <= 0:
            skipped += 1
            continue

        clean = raw_ticker.replace(".NS", "").replace(".BO", "").upper()
        full_ticker = f"{clean}.NS"

        # PERF: don't run DCF synchronously during import -- a user with
        # 20 holdings would wait 20 x ~8s = ~3 min, breaking the
        # frontend's 30s timeout with "Network error". Instead:
        #   - Use cache if already warm (common case)
        #   - Otherwise save with placeholders; portfolio page triggers
        #     analysis on first view, cache warms, subsequent views fast
        try:
            cached = _c.get(f"analysis:{full_ticker}")
            if cached and hasattr(cached, "valuation"):
                iv = cached.valuation.fair_value
                wacc_val = cached.valuation.wacc
                sector = cached.company.sector
                verdict = cached.valuation.verdict
                company_name = cached.company.company_name
                mos = (iv - avg_cost) / avg_cost * 100 if avg_cost > 0 else 0.0
            else:
                # Not cached -> save placeholders, defer analysis to
                # next portfolio-page view. import stays fast.
                iv = 0
                wacc_val = 0.12
                sector = ""
                verdict = ""
                company_name = clean
                mos = 0
        except Exception as e:
            logger.warning(f"Import: enrichment failed for {full_ticker}: {e}")
            iv = 0
            wacc_val = 0.12
            sector = ""
            verdict = ""
            company_name = clean
            mos = 0

        try:
            ok, err = save_holding(
                user_email=email,
                ticker=full_ticker,
                entry_price=avg_cost,
                iv=iv,
                mos_pct=mos,
                signal=verdict,
                wacc=wacc_val,
                company_name=company_name,
                sector=sector,
                notes=f"Imported from {broker} ({qty} shares)",
                # 2026-04-21 multi-account: tag each row with the broker
                # so two CSVs from different demats stay separate. Old
                # single-broker users have account_label='zerodha' (or
                # whatever they uploaded as) instead of merging.
                account_label=broker or "default",
                quantity=float(qty) if qty else None,
            )
            if ok:
                imported += 1
            else:
                skipped += 1
                errors.append(f"{clean}: {err}" if err else f"{clean}: save failed")
        except Exception as e:
            skipped += 1
            errors.append(f"{clean}: {type(e).__name__}: {e}")

    return {
        "imported": imported,
        "skipped": skipped,
        "total_parsed": len(parsed),
        "errors": errors[:10],
        "tier": tier,
    }


@router.post("/import")
async def import_holdings(req: ImportCSVRequest, user: dict = Depends(get_current_user)):
    """
    Bulk import holdings from a broker CSV (text).
    Supports: Zerodha Console, Groww, Upstox, ICICI Direct, Custom.
    """
    try:
        parsed = _parse_broker_csv(req.csv_text, req.broker)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse CSV: {e}")

    return _do_import(parsed, req.broker, user)


@router.post("/import-file")
async def import_holdings_file(
    file: UploadFile = File(...),
    broker: str = Form(default="zerodha"),
    user: dict = Depends(get_current_user),
):
    """
    Bulk import holdings from an uploaded file (.csv or .xlsx).
    Auto-detects header row (handles Zerodha's metadata-prefixed xlsx).
    """
    filename = (file.filename or "").lower()
    content_type = (file.content_type or "").lower()
    contents = await file.read()
    logger.info(f"import-file: filename={filename!r}, ct={content_type}, size={len(contents)}, broker={broker}")

    # Detect xlsx by filename OR by magic bytes (PK header for ZIP-based formats)
    is_xlsx = (
        filename.endswith(".xlsx")
        or filename.endswith(".xlsm")
        or filename.endswith(".xls")
        or "sheet" in content_type
        or "excel" in content_type
        or "officedocument" in content_type
        or (len(contents) >= 4 and contents[:4] == b"PK\x03\x04")  # ZIP/OOXML magic
    )

    try:
        if is_xlsx:
            logger.info(f"import-file: parsing as XLSX")
            parsed = _parse_xlsx_bytes(contents, broker)
        else:
            # Treat as CSV/text
            logger.info(f"import-file: parsing as CSV/text")
            try:
                csv_text = contents.decode("utf-8")
            except UnicodeDecodeError:
                csv_text = contents.decode("latin-1", errors="ignore")
            parsed = _parse_broker_csv(csv_text, broker)
        logger.info(f"import-file: parsed {len(parsed)} holdings")
    except ValueError as e:
        logger.warning(f"import-file ValueError for {filename}: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.warning(f"File import parse failed: {e}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"Could not parse file: {type(e).__name__}: {e}")

    return _do_import(parsed, broker, user)


# Broker-specific field maps (lowercase key → list of possible header names)
BROKER_FIELD_MAPS = {
    "zerodha": {
        "ticker": ["symbol", "tradingsymbol", "stock"],
        "quantity": ["qty", "quantity", "qty.", "quantity available"],
        "avg_cost": ["avg.cost", "avg cost", "avgcost", "avg. cost", "avg price", "average price"],
    },
    "groww": {
        "ticker": ["stock name", "symbol", "ticker"],
        "quantity": ["quantity", "qty"],
        "avg_cost": ["average buy price", "avg buy price", "avg price", "avg. cost"],
    },
    "upstox": {
        "ticker": ["company name", "tradingsymbol", "symbol"],
        "quantity": ["quantity", "qty"],
        "avg_cost": ["avg price", "average price", "avg. cost"],
    },
    "icici": {
        "ticker": ["stock", "symbol", "company"],
        "quantity": ["qty", "quantity"],
        "avg_cost": ["avg price", "average price", "avg. cost"],
    },
    "custom": {
        "ticker": ["ticker", "symbol", "stock", "company"],
        "quantity": ["quantity", "qty", "shares", "units"],
        "avg_cost": ["avg_price", "avg_cost", "price", "average price", "avg price", "buy price"],
    },
}


def _find_header_row(rows: list[list], field_map: dict) -> int:
    """
    Find the row index containing the header. Zerodha's xlsx has
    metadata rows at the top, so we scan for the first row that
    has all 3 required field names.
    """
    required_tokens = []
    for key in ("ticker", "quantity", "avg_cost"):
        required_tokens.append([t.lower() for t in field_map[key]])

    for idx, row in enumerate(rows):
        cells_lower = [str(c or "").lower().strip() for c in row]
        found_all = True
        for token_list in required_tokens:
            if not any(t in cells_lower for t in token_list):
                found_all = False
                break
        if found_all:
            return idx
    return -1


def _rows_to_dicts(rows: list[list], header_row_idx: int) -> tuple[list[dict], list[str]]:
    """Convert rows[header_row_idx+1:] to dicts using header row as keys."""
    headers = [str(c or "").strip() for c in rows[header_row_idx]]
    dicts = []
    for row in rows[header_row_idx + 1:]:
        if not any(c for c in row if c is not None and str(c).strip()):
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            row_dict[h] = row[i] if i < len(row) else None
        dicts.append(row_dict)
    return dicts, headers


def _parse_rows_to_holdings(rows: list[list], broker: str) -> list[dict]:
    """
    Given a 2D list of rows (from CSV or XLSX), find the header row and
    extract holdings.
    """
    field_map = BROKER_FIELD_MAPS.get(broker, BROKER_FIELD_MAPS["custom"])

    # Find header row (might not be row 0 for Zerodha xlsx)
    header_idx = _find_header_row(rows, field_map)
    if header_idx < 0:
        raise ValueError(
            f"Required columns not found for {broker}. "
            f"Need ticker, quantity, and average cost columns."
        )

    dicts, headers = _rows_to_dicts(rows, header_idx)
    headers_lower = {h.lower().strip(): h for h in headers}

    def _find_header(candidates: list[str]) -> str | None:
        for c in candidates:
            if c.lower() in headers_lower:
                return headers_lower[c.lower()]
        return None

    ticker_h = _find_header(field_map["ticker"])
    qty_h = _find_header(field_map["quantity"])
    cost_h = _find_header(field_map["avg_cost"])

    if not ticker_h or not qty_h or not cost_h:
        raise ValueError(
            f"Required columns not found for {broker}. "
            f"Need ticker, quantity, and average cost columns."
        )

    parsed: list[dict] = []
    for row_dict in dicts:
        try:
            raw_ticker = str(row_dict.get(ticker_h) or "").strip().upper()
            if not raw_ticker or raw_ticker in ("SYMBOL", "TICKER", "-"):
                continue
            if ":" in raw_ticker:
                raw_ticker = raw_ticker.split(":")[-1]
            raw_ticker = raw_ticker.strip()

            qty_str = str(row_dict.get(qty_h, "") or "").replace(",", "").strip()
            cost_str = str(row_dict.get(cost_h, "") or "").replace(",", "").replace("\u20B9", "").strip()

            qty = float(qty_str) if qty_str and qty_str not in ("-", "None") else 0
            cost = float(cost_str) if cost_str and cost_str not in ("-", "None") else 0

            if qty > 0 and cost > 0:
                parsed.append({
                    "ticker": raw_ticker,
                    "quantity": qty,
                    "avg_cost": cost,
                })
        except (ValueError, KeyError, TypeError):
            continue

    return parsed


def _parse_xlsx_bytes(xlsx_bytes: bytes, broker: str) -> list[dict]:
    """Parse an uploaded .xlsx file to holdings."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)

    # For Zerodha holdings file, prefer "Equity" sheet if available
    preferred_order = ["Equity", "Combined", "Holdings", "equity", "Sheet1"]
    sheet_name = None
    for pref in preferred_order:
        if pref in wb.sheetnames:
            sheet_name = pref
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    return _parse_rows_to_holdings(rows, broker)


def _parse_broker_csv(csv_text: str, broker: str) -> list[dict]:
    """
    Parse broker-specific CSV text. Returns list of dicts with:
        ticker (str), quantity (float), avg_cost (float)
    """
    csv_text = csv_text.strip()
    if not csv_text:
        return []

    # Auto-detect delimiter
    first_line = csv_text.split("\n")[0]
    delim = "\t" if "\t" in first_line else ","

    # Parse as 2D list so we can find header row in Zerodha files
    reader = csv.reader(io.StringIO(csv_text), delimiter=delim)
    rows = [list(row) for row in reader]
    if not rows:
        return []

    return _parse_rows_to_holdings(rows, broker)


def _parse_broker_csv_legacy(csv_text: str, broker: str) -> list[dict]:
    """Deprecated — kept as reference, not used."""
    csv_text = csv_text.strip()
    if not csv_text:
        return []
    first_line = csv_text.split("\n")[0]
    delim = "\t" if "\t" in first_line else ","
    reader = csv.DictReader(io.StringIO(csv_text), delimiter=delim)
    if not reader.fieldnames:
        return []
    headers_lower = {h.lower().strip(): h for h in reader.fieldnames}
    field_map = BROKER_FIELD_MAPS.get(broker, BROKER_FIELD_MAPS["custom"])
    def _find_header(candidates: list[str]) -> str | None:
        for c in candidates:
            if c.lower() in headers_lower:
                return headers_lower[c.lower()]
        return None
    ticker_h = _find_header(field_map["ticker"])
    qty_h = _find_header(field_map["quantity"])
    cost_h = _find_header(field_map["avg_cost"])
    if not ticker_h or not qty_h or not cost_h:
        raise ValueError(
            f"Required columns not found for {broker}. "
            f"Need ticker, quantity, and average cost columns."
        )

    parsed: list[dict] = []
    for row in reader:
        try:
            raw_ticker = (row.get(ticker_h) or "").strip().upper()
            if not raw_ticker:
                continue
            # Strip exchange prefixes (e.g. "NSE:RELIANCE" → "RELIANCE")
            if ":" in raw_ticker:
                raw_ticker = raw_ticker.split(":")[-1]
            # Strip whitespace and non-printable chars
            raw_ticker = raw_ticker.strip()

            qty_str = str(row.get(qty_h, "")).replace(",", "").strip()
            cost_str = str(row.get(cost_h, "")).replace(",", "").replace("\u20B9", "").strip()

            qty = float(qty_str) if qty_str else 0
            cost = float(cost_str) if cost_str else 0

            if qty > 0 and cost > 0:
                parsed.append({
                    "ticker": raw_ticker,
                    "quantity": qty,
                    "avg_cost": cost,
                })
        except (ValueError, KeyError):
            continue

    return parsed


@router.get("/health", response_model=PortfolioHealthResponse)
async def get_portfolio_health(user: dict = Depends(get_current_user)):
    """Portfolio health score (0-100) for the authenticated user.

    Uses live-enriched holdings (current price, fair value, score from
    cache) so the score reflects the real portfolio state.
    """
    from backend.services.portfolio_service import get_holdings_with_live_data
    from backend.services.cache_service import cache as _c
    from dashboard.utils.portfolio_health import calculate_portfolio_health

    email = user.get("email", "")
    if not email:
        return PortfolioHealthResponse(score=0, grade="F", summary="No holdings", issues=[], strengths=[], overvalued_count=0, undervalued_count=0, danger_positions=[], concentration_warning=None)

    enriched = get_holdings_with_live_data(email)
    live_holdings = enriched.get("holdings", [])

    # Map to the field names that calculate_portfolio_health expects
    mapped = []
    for h in live_holdings:
        ticker = h.get("ticker", "")
        # Pull additional context from analysis cache (red flags, moat)
        red_flag_count = 0
        moat = "None"
        try:
            cached = _c.get(f"analysis:{ticker}")
            if cached and hasattr(cached, "insights"):
                rf = getattr(cached.insights, "red_flags", []) or []
                red_flag_count = len(rf) if isinstance(rf, list) else 0
            if cached and hasattr(cached, "quality"):
                moat = cached.quality.moat or "None"
        except Exception:
            pass

        mapped.append({
            "ticker": h.get("ticker", ""),
            "shares": h.get("quantity", 1),
            "avg_buy_price": h.get("entry_price", 0),
            "current_price": h.get("current_price", 0),
            "yieldiq_score": h.get("score") if h.get("score") is not None else 50,
            "mos": h.get("mos_pct", 0),  # field rename: mos_pct -> mos
            "moat": moat,
            "red_flags": red_flag_count,
            "sector": h.get("sector", "Unknown") or "Unknown",
        })

    health = calculate_portfolio_health(mapped)
    return PortfolioHealthResponse(**health)
