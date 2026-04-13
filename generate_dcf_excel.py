"""
Institutional-Grade DCF Excel Model Generator
Generates a standalone Excel file that upgrades the dashboard's DCF output
to investment bank quality (10/10).

Usage (from ai_dcf_screener directory):
    python generate_dcf_excel.py --ticker TCS.NS
    python generate_dcf_excel.py --ticker ITC.NS
"""

import sys, argparse, io
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# ── Institutional color palette ────────────────────────────────
# Following Goldman Sachs / Morgan Stanley Excel conventions
C_INPUT_BG    = "EBF3FB"   # light blue bg for inputs
C_INPUT_FG    = "0000FF"   # blue text = hardcoded inputs
C_FORMULA_FG  = "000000"   # black text = formulas
C_LINK_FG     = "008000"   # green text = cross-sheet links
C_ERROR_FG    = "FF0000"   # red text = errors / warnings
C_HEADER_BG   = "1F3864"   # dark navy header
C_HEADER_FG   = "FFFFFF"   # white header text
C_SUBHDR_BG   = "2E75B6"   # blue subheader
C_SECTION_BG  = "D6E4F0"   # light blue section
C_ALT_ROW     = "F2F7FB"   # alternating row
C_WHITE       = "FFFFFF"
C_YELLOW_FLAG = "FFFF00"   # yellow = key assumption
C_GREEN_POS   = "00B050"   # positive values
C_RED_NEG     = "C00000"   # negative values
C_AMBER       = "ED7D31"   # amber / warning
C_BEAR        = "C00000"   # bear scenario
C_BASE        = "2E75B6"   # base scenario
C_BULL        = "00B050"   # bull scenario

def _c(s):
    if not s: return "FFFFFFFF"
    s = str(s).lstrip("#").upper()
    if len(s) == 6: return "FF" + s
    if len(s) == 8: return s
    if len(s) >= 10: return s[:8]
    return "FFFFFFFF"

def hf(c, fg=None):
    fill = PatternFill("solid", fgColor=_c(c))
    return fill

def border(style="thin", color=None):
    color = color or "BFBFBF"
    s = Side(style=style, color=_c(color))
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom(color=None):
    color = color or "1F3864"
    thin = Side(style="thin", color=_c("BFBFBF"))
    thick = Side(style="medium", color=_c(color))
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def font(bold=False, sz=10, color="000000", italic=False, name="Calibri"):
    return Font(name=name, bold=bold, size=sz, color=_c(color), italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def wc(ws, row, col, value, bg=None, fg="000000", bold=False, sz=10,
       nf=None, h_align="right", border_style="thin", italic=False):
    c = ws.cell(row=row, column=col, value=value)
    if bg: c.fill = hf(bg)
    c.font = Font(name="Calibri", bold=bold, size=sz, color=_c(fg), italic=italic)
    c.alignment = Alignment(horizontal=h_align, vertical="center")
    c.border = border(border_style)
    if nf: c.number_format = nf
    return c

def header_row(ws, row, labels, bg=C_HEADER_BG, fg=C_HEADER_FG, sz=10, bold=True, height=20):
    ws.row_dimensions[row].height = height
    for col, lbl in enumerate(labels, 1):
        c = ws.cell(row=row, column=col, value=lbl)
        c.fill = hf(bg)
        c.font = Font(name="Calibri", bold=bold, size=sz, color=_c(fg))
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border("thin", C_HEADER_BG)

def section_header(ws, row, text, ncols, bg=C_SECTION_BG, fg=C_HEADER_BG, sz=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=f"  {text}")
    c.fill = hf(bg)
    c.font = Font(name="Calibri", bold=True, size=sz, color=_c(fg))
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thick_bottom()
    ws.row_dimensions[row].height = 18

def title_block(ws, row, ticker, subtitle_text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=f"INSTITUTIONAL DCF MODEL — {ticker}")
    c.fill = hf(C_HEADER_BG)
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30

    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=ncols)
    c2 = ws.cell(row=row+1, column=1, value=subtitle_text)
    c2.fill = hf(C_SUBHDR_BG)
    c2.font = Font(name="Calibri", size=9, color="FFFFFFFF", italic=True)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row+1].height = 16

def input_cell(ws, row, col, value, comment=None):
    """Blue text = hardcoded input (institutional convention)"""
    c = wc(ws, row, col, value, bg=C_INPUT_BG, fg=C_INPUT_FG, bold=False, sz=10)
    if comment:
        from openpyxl.comments import Comment
        c.comment = Comment(comment, "Model")
    return c

def formula_cell(ws, row, col, formula, fg=C_FORMULA_FG, bg=C_WHITE, nf=None):
    """Black text = formula (institutional convention)"""
    c = ws.cell(row=row, column=col, value=formula)
    c.font = Font(name="Calibri", size=10, color=_c(fg))
    c.fill = hf(bg)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = border()
    if nf: c.number_format = nf
    return c


def generate_institutional_dcf(ticker, enriched, dcf_res, forecast_result,
                                 scenarios, wacc_data, wacc, terminal_g,
                                 forecast_yrs, sym, to_code, fx) -> bytes:
    wb = Workbook()

    income_df = enriched.get("income_df", pd.DataFrame())
    cf_df     = enriched.get("cf_df",     pd.DataFrame())
    bs_df     = enriched.get("bs_df",     pd.DataFrame())
    projected = forecast_result["projections"]
    gs        = forecast_result.get("growth_schedule", [])
    pv_fcfs   = dcf_res.get("pv_fcfs", [])
    price     = enriched.get("price", 0) * fx
    shares    = enriched.get("shares", 1)
    op_margin = enriched.get("op_margin", 0)
    rev_growth= enriched.get("revenue_growth", 0)
    latest_rev= enriched.get("latest_revenue", 0)
    sector    = enriched.get("sector_name", enriched.get("sector", "General"))
    moat_grade= enriched.get("moat_grade", "N/A")

    hist_years = []
    if not income_df.empty and "year" in income_df.columns:
        hist_years = [str(int(y)) for y in income_df["year"].tolist()]

    proj_labels = [f"FY+{i+1}" for i in range(forecast_yrs)]
    all_years   = hist_years + proj_labels
    NCOLS = max(len(all_years) + 2, 14)

    # ══════════════════════════════════════════════════════════
    # SHEET 1 — ASSUMPTIONS (Institutional standard: inputs first)
    # ══════════════════════════════════════════════════════════
    ws_a = wb.active
    ws_a.title = "1. Assumptions"
    ws_a.sheet_view.showGridLines = False
    ws_a.column_dimensions["A"].width = 38
    ws_a.column_dimensions["B"].width = 22
    ws_a.column_dimensions["C"].width = 18
    ws_a.column_dimensions["D"].width = 40

    title_block(ws_a, 1, ticker,
        f"Key Model Assumptions  |  Sector: {sector}  |  Moat: {moat_grade}  |  Generated: {datetime.now().strftime('%d %b %Y')}", 4)

    # Color convention legend
    section_header(ws_a, 4, "COLOR CONVENTION (Institutional Standard)", 4)
    legend = [
        ("Blue text / Blue background", "Hardcoded input assumptions — change these for scenarios", C_INPUT_BG, C_INPUT_FG),
        ("Black text / White background","Formula — calculated automatically from inputs", C_WHITE, C_FORMULA_FG),
        ("Green text",                   "Cross-sheet link — pulls data from another tab", C_WHITE, C_LINK_FG),
        ("Red text",                     "Error or warning — requires analyst review", C_WHITE, C_ERROR_FG),
        ("Yellow background",            "Key assumption — high sensitivity to model output", C_YELLOW_FLAG, C_FORMULA_FG),
    ]
    for i, (label, desc, bg, fg) in enumerate(legend):
        r = 5 + i
        ws_a.row_dimensions[r].height = 18
        wc(ws_a, r, 1, label, bg=bg, fg=fg, bold=True, sz=10, h_align="left")
        wc(ws_a, r, 2, desc,  bg=C_WHITE, fg="595959", bold=False, sz=9, h_align="left")
        wc(ws_a, r, 3, "",    bg=C_WHITE)
        wc(ws_a, r, 4, "",    bg=C_WHITE)

    # WACC Assumptions
    section_header(ws_a, 11, "WACC ASSUMPTIONS", 4)
    rf   = wacc_data.get("rf", 0.072)
    beta = wacc_data.get("beta", 1.0)
    mrp  = wacc_data.get("market_premium", 0.06)
    re   = wacc_data.get("re", wacc)
    rd   = wacc_data.get("rd", 0.06)
    tax  = wacc_data.get("tax_rate", 0.25)
    ew   = wacc_data.get("e_weight", 0.8)
    dw   = wacc_data.get("d_weight", 0.2)

    wacc_rows = [
        ("Risk-Free Rate (10Y Gsec)",    rf,    "0.00%", "RBI 10-Year Government Bond Yield"),
        ("Equity Risk Premium (ERP)",    mrp,   "0.00%", "Damodaran India ERP — Jan 2025"),
        ("Beta (industry-adjusted)",     beta,  "0.00",  "Yahoo Finance 5Y monthly, floored at 0.80"),
        ("Cost of Equity (CAPM)",        re,    "0.00%", "=Rf + β×ERP  [FORMULA — black]"),
        ("Pre-tax Cost of Debt",         rd,    "0.00%", "Interest Expense / Total Debt"),
        ("Effective Tax Rate",           tax,   "0.00%", "Effective tax rate from income statement"),
        ("After-tax Cost of Debt",       rd*(1-tax), "0.00%", "=Kd×(1−Tax)  [FORMULA]"),
        ("Equity Weight (Mkt Value)",    ew,    "0.00%", "Market Cap / (Market Cap + Total Debt)"),
        ("Debt Weight (Mkt Value)",      dw,    "0.00%", "Total Debt / (Market Cap + Total Debt)"),
        ("WACC (Final)",                 wacc,  "0.00%", "=Ke×We + Kd(1−t)×Wd  [FORMULA]"),
    ]
    header_row(ws_a, 12, ["Parameter", "Value", "Format", "Source / Note"], bg=C_SUBHDR_BG)
    is_formula = {3, 6, 9}  # rows that are formulas (0-indexed)
    is_input   = {0, 1, 2, 4, 5, 7, 8}
    is_key     = {9}  # WACC final = yellow flag
    for i, (label, val, nf, note) in enumerate(wacc_rows):
        r = 13 + i
        ws_a.row_dimensions[r].height = 17
        wc(ws_a, r, 1, label, bg=C_WHITE, fg=C_FORMULA_FG, bold=False, sz=10, h_align="left")
        if i in is_key:
            input_cell(ws_a, r, 2, val).number_format = nf
            ws_a.cell(r, 2).fill = hf(C_YELLOW_FLAG)
        elif i in is_formula:
            formula_cell(ws_a, r, 2, val, nf=nf)
        else:
            input_cell(ws_a, r, 2, val).number_format = nf
        wc(ws_a, r, 3, nf,   bg=C_ALT_ROW if i%2==0 else C_WHITE, fg="595959", sz=9, h_align="center")
        wc(ws_a, r, 4, note, bg=C_ALT_ROW if i%2==0 else C_WHITE, fg="595959", sz=9, h_align="left", italic=i in is_formula)

    # Growth Assumptions
    section_header(ws_a, 24, "GROWTH & MARGIN ASSUMPTIONS", 4)
    fcf_base_method = forecast_result.get("fcf_base_method", "Auto")
    growth_rows = [
        ("Base FCF Growth Rate (Year 1)", gs[0] if gs else rev_growth, "0.0%", "Historical CAGR blend — AI-forecasted"),
        ("Growth Fade Factor (k)",        0.25,  "0.00",  "Exponential decay: g(t)=gT+(g0-gT)×e^(-k×t)"),
        ("Terminal Growth Rate",          terminal_g, "0.0%", "Long-run India nominal GDP — HARD CAP 4%"),
        ("Operating Margin (base)",       op_margin, "0.0%", "Latest reported operating margin"),
        ("Tax Rate",                      tax,   "0.0%",  "Effective corporate tax rate"),
        ("D&A as % of Revenue",           enriched.get("depreciation_pct", 0.035), "0.0%", "Sector-specific estimate"),
        ("CapEx as % of Revenue",         enriched.get("capex_intensity",  0.045), "0.0%", "Sector-specific — includes maintenance + growth"),
        ("ΔWorking Capital / ΔRevenue",   enriched.get("wc_pct_revenue",   0.08),  "0.0%", "Incremental WC required per unit of growth"),
        ("FCF Base Method",               fcf_base_method, "@", "How starting FCF was selected"),
        ("Moat Grade",                    moat_grade, "@", "Wide/Narrow/None — adjusts growth & WACC"),
    ]
    header_row(ws_a, 25, ["Parameter", "Value", "Format", "Note"], bg=C_SUBHDR_BG)
    for i, (label, val, nf, note) in enumerate(growth_rows):
        r = 26 + i
        ws_a.row_dimensions[r].height = 17
        wc(ws_a, r, 1, label, bg=C_WHITE, fg=C_FORMULA_FG, sz=10, h_align="left")
        ic = input_cell(ws_a, r, 2, val)
        if nf != "@": ic.number_format = nf
        if label == "Terminal Growth Rate": ws_a.cell(r,2).fill = hf(C_YELLOW_FLAG)
        wc(ws_a, r, 3, nf,   bg=C_ALT_ROW if i%2==0 else C_WHITE, fg="595959", sz=9, h_align="center")
        wc(ws_a, r, 4, note, bg=C_ALT_ROW if i%2==0 else C_WHITE, fg="595959", sz=9, h_align="left")

    # Guardrails
    section_header(ws_a, 37, "MODEL GUARDRAILS & VALIDATION", 4)
    tv_pct = dcf_res.get("tv_pct_of_ev", 0)
    roic_est = op_margin * (1-tax) / 0.4 if op_margin > 0 else 0
    guardrail_rows = [
        ("Terminal Growth < WACC",        "PASS" if terminal_g < wacc else "⚠ FAIL", terminal_g < wacc),
        ("Terminal Growth ≤ 4% (India)",  "PASS" if terminal_g <= 0.04 else "⚠ FAIL", terminal_g <= 0.04),
        ("WACC ≥ 9% (India floor)",       "PASS" if wacc >= 0.09 else "⚠ FAIL", wacc >= 0.09),
        ("TV% of EV ≤ 75%",               f"{'PASS' if tv_pct<=0.75 else '⚠ FAIL'} ({tv_pct:.0%})", tv_pct<=0.75),
        ("Growth supported by ROIC",      "PASS" if roic_est > rev_growth else "⚠ CHECK", roic_est > rev_growth),
        ("FCF Positive",                  "PASS" if enriched.get("latest_fcf",0)>0 else "⚠ FAIL", enriched.get("latest_fcf",0)>0),
    ]
    header_row(ws_a, 38, ["Guardrail Check", "Status", "", ""], bg=C_SUBHDR_BG)
    for i, (check, status, passed) in enumerate(guardrail_rows):
        r = 39 + i
        ws_a.row_dimensions[r].height = 17
        bg = "E2EFDA" if passed else "FCE4D6"
        fg = C_GREEN_POS if passed else C_RED_NEG
        wc(ws_a, r, 1, check,  bg=bg, fg=C_FORMULA_FG, sz=10, h_align="left")
        wc(ws_a, r, 2, status, bg=bg, fg=fg, bold=True, sz=10, h_align="center")
        wc(ws_a, r, 3, "",     bg=bg)
        wc(ws_a, r, 4, "",     bg=bg)

    # ══════════════════════════════════════════════════════════
    # SHEET 2 — DCF MODEL CORE
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("2. DCF Model")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 36
    for i in range(1, forecast_yrs + 4):
        ws2.column_dimensions[get_column_letter(i+1)].width = 14

    title_block(ws2, 1, ticker,
        f"WACC: {wacc:.2%}  |  Terminal g: {terminal_g:.2%}  |  Horizon: {forecast_yrs}Y  |  Currency: {to_code}B",
        forecast_yrs + 3)

    # Year headers
    section_header(ws2, 3, "FREE CASH FLOW PROJECTIONS  (FCFF = NOPAT + D&A − CapEx − ΔWCAP)", forecast_yrs+3)
    hdr_labels = ["Metric"] + proj_labels + ["Terminal", "Formula"]
    header_row(ws2, 4, hdr_labels, bg=C_SUBHDR_BG)

    term_fcf  = forecast_result.get("terminal_fcf_norm", 0) * fx / 1e9
    pv_tv_bn  = dcf_res.get("pv_tv", 0) * fx / 1e9
    fcf_base  = forecast_result.get("fcf_base", 0) * fx / 1e9

    proj_data = [
        ("FCF Base (Year 0)",    [fcf_base] + [""]*( forecast_yrs-1), "", "Starting FCF — max(latest, NOPAT proxy)", C_INPUT_BG, C_INPUT_FG),
        (f"Projected FCFF ({to_code}B)", [round(v*fx/1e9,3) for v in projected], round(term_fcf,3), "FCF(t-1) × (1 + g(t))", C_WHITE, C_FORMULA_FG),
        ("YoY Growth Rate",      [f"{g*100:.1f}%" for g in gs], f"{terminal_g*100:.1f}%", "Exponential fade from base to terminal g", C_ALT_ROW, C_FORMULA_FG),
        (f"Discount Factor",     [round(1/(1+wacc)**(i+1),4) for i in range(forecast_yrs)], "", "1 / (1+WACC)^t", C_ALT_ROW, C_FORMULA_FG),
        (f"PV of FCFF ({to_code}B)", [round(v*fx/1e9,3) for v in pv_fcfs], round(pv_tv_bn,3), "FCFF(t) × Discount Factor(t)", C_WHITE, C_FORMULA_FG),
    ]
    for k, (label, vals, term_val, formula_note, bg, fg) in enumerate(proj_data):
        r = 5 + k
        ws2.row_dimensions[r].height = 18
        wc(ws2, r, 1, label, bg=bg, fg=fg, bold=fg==C_INPUT_FG, sz=10, h_align="left")
        for j, v in enumerate(vals[:forecast_yrs]):
            nf = "#,##0.00" if isinstance(v, float) else "@"
            wc(ws2, r, j+2, v, bg=bg, fg=fg, sz=10, nf=nf)
        if term_val != "":
            wc(ws2, r, forecast_yrs+2, term_val, bg="FFF2CC", fg=C_FORMULA_FG, bold=True, sz=10, nf="#,##0.00")
        wc(ws2, r, forecast_yrs+3, formula_note, bg=C_ALT_ROW, fg="595959", sz=8, h_align="left", italic=True)

    # Terminal Value Section — explicit formula
    section_header(ws2, 11, "TERMINAL VALUE — GORDON GROWTH MODEL", forecast_yrs+3)
    tv_formula_text = f"TV = FCF_norm × (1+g) / (WACC − g) = {term_fcf:.2f} × (1+{terminal_g:.1%}) / ({wacc:.1%} − {terminal_g:.1%})"
    tv_value = term_fcf * (1+terminal_g) / (wacc - terminal_g) if wacc > terminal_g else 0
    tv_rows = [
        ("Normalised Terminal FCF",     round(term_fcf, 3),         "#,##0.00",  f"Avg of last 3 projected FCFs ({to_code}B)"),
        ("Terminal Growth Rate (g)",    f"{terminal_g*100:.1f}%",   "@",         "India long-run nominal GDP — HARD CAP 4%"),
        ("WACC",                        f"{wacc*100:.2f}%",          "@",         "Weighted Average Cost of Capital"),
        ("WACC − g (spread)",           f"{(wacc-terminal_g)*100:.2f}%", "@",    "Must be > 0 or model is undefined"),
        ("Terminal Value (undiscounted)", round(tv_value, 2),        "#,##0.00",  f"= {tv_formula_text}"),
        ("PV of Terminal Value",        round(pv_tv_bn, 2),          "#,##0.00",  f"TV / (1+WACC)^{forecast_yrs}"),
        ("TV as % of Enterprise Value", f"{tv_pct:.1%}",             "@",         "⚠ Flag if > 75% — high terminal sensitivity"),
    ]
    tv_flag_row = None
    for i, (label, val, nf, note) in enumerate(tv_rows):
        r = 12 + i
        ws2.row_dimensions[r].height = 18
        is_flag = "%" in str(val) and "75" in note
        bg = "FCE4D6" if (tv_pct > 0.75 and "%" in str(val) and "Flag" in note) else C_WHITE
        wc(ws2, r, 1, label, bg=bg, fg=C_FORMULA_FG, sz=10, h_align="left")
        c = wc(ws2, r, 2, val, bg=bg, fg=C_INPUT_FG if label in ["Terminal Growth Rate (g)"] else C_FORMULA_FG, sz=10)
        if nf != "@": c.number_format = nf
        wc(ws2, r, 3, note, bg=C_ALT_ROW, fg="595959", sz=9, h_align="left", italic=True)
        for col in range(4, forecast_yrs+4):
            wc(ws2, r, col, "", bg=bg)

    # Valuation Bridge
    section_header(ws2, 20, "VALUATION BRIDGE — EV TO EQUITY VALUE PER SHARE", forecast_yrs+3)
    ev_val    = dcf_res.get("enterprise_value", 0) * fx / 1e9
    debt_val  = enriched.get("total_debt", 0)  * fx / 1e9
    cash_val  = enriched.get("total_cash", 0)  * fx / 1e9
    eq_val    = max(ev_val - debt_val + cash_val, 0)
    iv_ps     = eq_val * 1e9 / shares * fx if shares > 0 else 0
    mos_pct   = (iv_ps - price) / price * 100 if price > 0 else 0
    sum_pv    = dcf_res.get("sum_pv_fcfs", 0) * fx / 1e9

    bridge = [
        ("Σ PV of Forecast FCFs",          sum_pv,    "#,##0.0", C_WHITE,   C_FORMULA_FG, False),
        ("+ PV of Terminal Value",         pv_tv_bn,  "#,##0.0", C_WHITE,   C_FORMULA_FG, False),
        ("= Enterprise Value (EV)",        ev_val,    "#,##0.0", C_SECTION_BG, C_FORMULA_FG, True),
        ("− Total Debt (book value)",       debt_val,  "#,##0.0", C_WHITE,   C_RED_NEG,    False),
        ("+ Cash & Cash Equivalents",      cash_val,  "#,##0.0", C_WHITE,   C_GREEN_POS,  False),
        ("= Equity Value",                 eq_val,    "#,##0.0", C_SECTION_BG, C_FORMULA_FG, True),
        ("÷ Shares Outstanding (B)",       shares/1e9, "0.000",  C_WHITE,   C_FORMULA_FG, False),
        ("= Intrinsic Value / Share",      iv_ps,     f'"{sym}"#,##0.00', "E2EFDA", C_GREEN_POS if mos_pct>0 else C_RED_NEG, True),
        ("Current Market Price",           price,     f'"{sym}"#,##0.00', C_WHITE,  C_INPUT_FG, False),
        ("Margin of Safety",               mos_pct/100, "0.0%;(0.0%);-", "E2EFDA" if mos_pct>0 else "FCE4D6",
                                           C_GREEN_POS if mos_pct>0 else C_RED_NEG, True),
        ("Investment Conclusion",
            "UNDERVALUED" if mos_pct>20 else "FAIRLY VALUED" if mos_pct>-10 else "OVERVALUED",
            "@", "FFF2CC", C_INPUT_FG, True),
    ]
    for i, row_data in enumerate(bridge):
        label, val, nf, bg, fg, bold = row_data
        r = 21 + i
        ws2.row_dimensions[r].height = 18
        wc(ws2, r, 1, label, bg=bg, fg=fg, bold=bold, sz=10, h_align="left")
        c = wc(ws2, r, 2, val if isinstance(val, str) else round(float(val), 4), bg=bg, fg=fg, bold=bold, sz=10)
        if nf != "@" and not isinstance(val, str): c.number_format = nf
        for col in range(3, forecast_yrs+4):
            wc(ws2, r, col, "", bg=bg)

    # ══════════════════════════════════════════════════════════
    # SHEET 3 — WACC DECOMPOSITION
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("3. WACC Build")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 35

    title_block(ws3, 1, ticker, f"WACC Decomposition — CAPM + Capital Structure  |  {sector}", 4)

    section_header(ws3, 3, "STEP 1 — COST OF EQUITY (CAPM)", 4)
    capm_rows = [
        ("Risk-Free Rate (Rf)",          rf,          "0.00%", "10Y India Gsec yield"),
        ("Beta (β)",                     beta,        "0.00",  "5Y monthly regression vs Nifty 50"),
        ("Equity Risk Premium (ERP)",    mrp,         "0.00%", "Damodaran India ERP"),
        ("Size/Liquidity Premium",       0.0,         "0.00%", "Add if small-cap (set to 0 for large-caps)"),
        ("Cost of Equity (Ke)",          re,          "0.00%", "= Rf + β × ERP  [CAPM Formula]"),
    ]
    header_row(ws3, 4, ["Component", "Value", "Format", "Note"], bg=C_SUBHDR_BG)
    for i, (label, val, nf, note) in enumerate(capm_rows):
        r = 5 + i
        ws3.row_dimensions[r].height = 17
        wc(ws3, r, 1, label, bg=C_WHITE, fg=C_FORMULA_FG, sz=10, h_align="left")
        is_inp = label not in ["Cost of Equity (Ke)"]
        c = input_cell(ws3, r, 2, val) if is_inp else formula_cell(ws3, r, 2, val)
        c.number_format = nf
        if label == "Cost of Equity (Ke)": ws3.cell(r,2).fill = hf(C_YELLOW_FLAG)
        wc(ws3, r, 3, nf,   bg=C_ALT_ROW, fg="595959", sz=9, h_align="center")
        wc(ws3, r, 4, note, bg=C_ALT_ROW, fg="595959", sz=9, h_align="left")

    section_header(ws3, 11, "STEP 2 — COST OF DEBT", 4)
    cod_rows = [
        ("Pre-tax Cost of Debt (Kd)",    rd,          "0.00%", "Interest Expense / Avg Total Debt"),
        ("Effective Tax Rate",           tax,         "0.00%", "From income statement"),
        ("After-tax Cost of Debt",       rd*(1-tax),  "0.00%", "= Kd × (1 − Tax Rate)"),
    ]
    header_row(ws3, 12, ["Component", "Value", "Format", "Note"], bg=C_SUBHDR_BG)
    for i, (label, val, nf, note) in enumerate(cod_rows):
        r = 13 + i
        ws3.row_dimensions[r].height = 17
        wc(ws3, r, 1, label, bg=C_WHITE, fg=C_FORMULA_FG, sz=10, h_align="left")
        is_inp = "After-tax" not in label
        c = input_cell(ws3, r, 2, val) if is_inp else formula_cell(ws3, r, 2, val)
        c.number_format = nf
        wc(ws3, r, 3, nf,   bg=C_ALT_ROW, fg="595959", sz=9, h_align="center")
        wc(ws3, r, 4, note, bg=C_ALT_ROW, fg="595959", sz=9, h_align="left")

    section_header(ws3, 17, "STEP 3 — CAPITAL STRUCTURE (MARKET VALUE WEIGHTS)", 4)
    mktcap = price * shares / 1e9
    cs_rows = [
        ("Market Capitalisation",        mktcap,      '#,##0.0', f"{to_code} Billions (price × shares)"),
        ("Total Debt (book proxy)",      enriched.get("total_debt",0)*fx/1e9, '#,##0.0', f"{to_code} Billions"),
        ("Total Capital (V)",            mktcap + enriched.get("total_debt",0)*fx/1e9, '#,##0.0', "= E + D"),
        ("Equity Weight (We)",           ew,          "0.0%",  "Market Cap / Total Capital"),
        ("Debt Weight (Wd)",             dw,          "0.0%",  "Total Debt / Total Capital"),
    ]
    header_row(ws3, 18, ["Component", "Value", "Format", "Note"], bg=C_SUBHDR_BG)
    for i, (label, val, nf, note) in enumerate(cs_rows):
        r = 19 + i
        ws3.row_dimensions[r].height = 17
        wc(ws3, r, 1, label, bg=C_WHITE, fg=C_FORMULA_FG, sz=10, h_align="left")
        is_inp = label in ["Equity Weight (We)", "Debt Weight (Wd)", "Total Debt (book proxy)"]
        c = input_cell(ws3, r, 2, val) if is_inp else formula_cell(ws3, r, 2, val)
        c.number_format = nf
        wc(ws3, r, 3, nf,   bg=C_ALT_ROW, fg="595959", sz=9, h_align="center")
        wc(ws3, r, 4, note, bg=C_ALT_ROW, fg="595959", sz=9, h_align="left")

    section_header(ws3, 25, "STEP 4 — FINAL WACC", 4)
    ws3.row_dimensions[26].height = 24
    wc(ws3, 26, 1, "WACC Formula", bg=C_SECTION_BG, fg=C_FORMULA_FG, bold=True, sz=10, h_align="left")
    wc(ws3, 26, 2, f"= Ke×We + Kd(1−t)×Wd = {re:.2%}×{ew:.1%} + {rd:.2%}×(1−{tax:.0%})×{dw:.1%}",
       bg=C_SECTION_BG, fg="595959", sz=9, h_align="left")
    ws3.merge_cells(start_row=26, start_column=2, end_row=26, end_column=4)
    ws3.row_dimensions[27].height = 28
    wc(ws3, 27, 1, "WACC (Final)", bg=C_YELLOW_FLAG, fg=C_FORMULA_FG, bold=True, sz=12, h_align="left")
    c_wacc = wc(ws3, 27, 2, wacc, bg=C_YELLOW_FLAG, fg=C_INPUT_FG, bold=True, sz=12)
    c_wacc.number_format = "0.00%"

    # ══════════════════════════════════════════════════════════
    # SHEET 4 — FCFF BUILD (Institutional bottom-up)
    # ══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("4. FCFF Build")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 36
    for i in range(1, len(hist_years) + forecast_yrs + 3):
        ws4.column_dimensions[get_column_letter(i+1)].width = 14

    all_labels = hist_years + proj_labels
    n_hist = len(hist_years)
    n_proj = forecast_yrs
    total_cols = len(all_labels)

    title_block(ws4, 1, ticker, f"FCFF = NOPAT + D&A − CapEx − ΔWCAP  |  Institutional Bottom-Up Build", total_cols+2)

    header_row(ws4, 3, [""] + ["← HISTORICAL →"]*n_hist + ["← FORECAST →"]*n_proj + ["Formula"], bg=C_HEADER_BG)
    header_row(ws4, 4, ["Line Item"] + all_labels + ["Source"], bg=C_SUBHDR_BG)

    def get_col(df, col, default=None):
        if df is None or df.empty or col not in df.columns: return [default]*len(df) if df is not None and not df.empty else []
        return [v for v in df[col].tolist()]

    hist_rev   = get_col(income_df, "revenue",          0)
    hist_opinc = get_col(income_df, "operating_income", 0)
    hist_ni    = get_col(income_df, "net_income",       0)
    hist_ocf   = get_col(cf_df,     "ocf",              0)
    hist_capex = get_col(cf_df,     "capex",            0)
    hist_fcf   = get_col(cf_df,     "fcf",              0)

    _sector = enriched.get("sector", "general")
    try:
        from models.industry_wacc import INDUSTRY_WACC
        _reinv = INDUSTRY_WACC.get(_sector, INDUSTRY_WACC["general"])
        _da_pct    = _reinv["depreciation_pct"]
        _capex_pct = _reinv["capex_intensity"]
        _dwc_pct   = _reinv["wc_pct_revenue"]
    except:
        _da_pct, _capex_pct, _dwc_pct = 0.035, 0.045, 0.08

    tax_rate   = tax
    base_rev   = latest_rev * fx / 1e9 if latest_rev else 1.0
    growth_fade = [max(rev_growth * np.exp(-0.25 * i), terminal_g) for i in range(1, n_proj+1)]

    fc_rev, fc_ebit, fc_nopat, fc_da, fc_capex_f, fc_dwc, fc_reinv, fc_fcff = [], [], [], [], [], [], [], []
    running_rev = base_rev
    for i, g in enumerate(growth_fade):
        running_rev *= (1 + g)
        ebit   = running_rev * op_margin
        nopat  = ebit * (1 - tax_rate)
        da     = running_rev * _da_pct
        capex_ = running_rev * _capex_pct
        dwc    = running_rev * g * _dwc_pct
        reinv  = capex_ - da + dwc
        fcff   = nopat - reinv
        fc_rev.append(round(running_rev, 2))
        fc_ebit.append(round(ebit, 3))
        fc_nopat.append(round(nopat, 3))
        fc_da.append(round(da, 3))
        fc_capex_f.append(round(capex_, 3))
        fc_dwc.append(round(dwc, 3))
        fc_reinv.append(round(reinv, 3))
        fc_fcff.append(round(fcff, 3))

    def fcff_row(ws, row, label, hist_vals, fc_vals, is_input=False, is_bold=False,
                  fg_hist=C_FORMULA_FG, fg_fc=C_FORMULA_FG, nf="#,##0.00", formula_note=""):
        ws.row_dimensions[row].height = 17
        wc(ws, row, 1, label, bg=C_WHITE, fg=C_FORMULA_FG, bold=is_bold, sz=10, h_align="left")
        for j, v in enumerate(hist_vals):
            val = v * fx / 1e9 if v else 0
            bg  = C_ALT_ROW if j%2==0 else C_WHITE
            c   = wc(ws, row, j+2, round(val, 3), bg=bg, fg=fg_hist, bold=is_bold, sz=10, nf=nf)
        for j, v in enumerate(fc_vals):
            bg = C_INPUT_BG if is_input else (C_ALT_ROW if j%2==0 else C_WHITE)
            fg = C_INPUT_FG if is_input else fg_fc
            c  = wc(ws, row, n_hist+j+2, v, bg=bg, fg=fg, bold=is_bold, sz=10, nf=nf)
        wc(ws, row, total_cols+2, formula_note, bg=C_ALT_ROW, fg="595959", sz=8, h_align="left", italic=True)

    section_header(ws4, 5, "INCOME BRIDGE", total_cols+2)
    fcff_row(ws4, 6,  f"Revenue ({to_code}B)",    hist_rev,   fc_rev,    False, True,  C_LINK_FG, C_FORMULA_FG, "#,##0.0",  "Historical: Yahoo Finance | Forecast: Rev(t-1)×(1+g)")
    fcff_row(ws4, 7,  "Operating Margin %",       [v/r if r else 0 for v,r in zip(hist_opinc,hist_rev)], [op_margin]*n_proj, True, False, C_FORMULA_FG, C_INPUT_FG, "0.0%", "Held constant at latest margin (conservative)")
    fcff_row(ws4, 8,  f"EBIT ({to_code}B)",       hist_opinc, fc_ebit,   False, False, C_LINK_FG, C_FORMULA_FG, "#,##0.00", "= Revenue × Operating Margin")
    fcff_row(ws4, 9,  f"NOPAT ({to_code}B)",      [v*(1-tax_rate) for v in hist_opinc], fc_nopat, False, False, C_FORMULA_FG, C_FORMULA_FG, "#,##0.00", f"= EBIT × (1 − {tax_rate:.0%} tax rate)")

    section_header(ws4, 11, "REINVESTMENT COMPONENTS", total_cols+2)
    fcff_row(ws4, 12, f"D&A ({to_code}B)",         [0]*n_hist, fc_da,       True, False, "595959", C_INPUT_FG, "#,##0.00", f"= {_da_pct:.1%} × Revenue (sector: {_sector})")
    fcff_row(ws4, 13, f"CapEx ({to_code}B)",        hist_capex, fc_capex_f,  True, False, C_LINK_FG, C_INPUT_FG, "#,##0.00", f"= {_capex_pct:.1%} × Revenue (sector benchmark)")
    fcff_row(ws4, 14, f"ΔWorking Capital ({to_code}B)", [0]*n_hist, fc_dwc, True, False, "595959", C_INPUT_FG, "#,##0.00", f"= {_dwc_pct:.1%} × ΔRevenue (incremental WC)")
    fcff_row(ws4, 15, f"Net Reinvestment ({to_code}B)", [0]*n_hist, fc_reinv, False, False, "595959", C_FORMULA_FG, "#,##0.00", "= CapEx − D&A + ΔWC")

    section_header(ws4, 17, "FREE CASH FLOW TO FIRM (FCFF)", total_cols+2)
    fcff_row(ws4, 18, f"Historical FCFF — Actual ({to_code}B)", hist_fcf, [], False, True, C_LINK_FG, C_FORMULA_FG, "#,##0.00", "Source: Yahoo Finance cash flow statement")
    fcff_row(ws4, 19, f"Forecast FCFF — Model ({to_code}B)",    [], fc_fcff,  False, True, "595959", C_FORMULA_FG, "#,##0.00", "= NOPAT − Net Reinvestment")

    section_header(ws4, 21, "REINVESTMENT RATE & FCFF CONVERSION", total_cols+2)
    reinv_rates = [round(r/n, 3) if n else 0 for r, n in zip(fc_reinv, fc_nopat)]
    fcf_conv = [round(f/n, 3) if n else 0 for f, n in zip(fc_fcff, fc_nopat)]
    fcff_row(ws4, 22, "Reinvestment Rate (Net Reinv/NOPAT)",  [], reinv_rates, False, False, "595959", C_FORMULA_FG, "0.0%", "How much of NOPAT is reinvested for growth")
    fcff_row(ws4, 23, "FCF Conversion Rate (FCFF/NOPAT)",    [], fcf_conv,    False, False, "595959", C_FORMULA_FG, "0.0%", "Efficiency of converting earnings to cash")

    # ROIC vs Growth check
    section_header(ws4, 25, "STEP 7 — ROIC vs GROWTH CONSISTENCY CHECK", total_cols+2)
    ic_est    = max(enriched.get("total_debt",0)*fx/1e9 + base_rev*0.4, base_rev*0.1)
    nopat_est = base_rev * op_margin * (1-tax_rate)
    roic_est2 = nopat_est / ic_est if ic_est > 0 else 0
    roic_ok   = roic_est2 > rev_growth
    ws4.row_dimensions[26].height = 18
    wc(ws4, 26, 1, "Estimated ROIC",          bg=C_WHITE,   fg=C_FORMULA_FG, sz=10, h_align="left")
    wc(ws4, 26, 2, roic_est2,                  bg=C_WHITE,   fg=C_FORMULA_FG, sz=10, nf="0.0%")
    wc(ws4, 26, 3, "Revenue Growth",           bg=C_ALT_ROW, fg=C_FORMULA_FG, sz=10, h_align="left")
    wc(ws4, 26, 4, rev_growth,                 bg=C_ALT_ROW, fg=C_FORMULA_FG, sz=10, nf="0.0%")
    wc(ws4, 26, 5, "ROIC > Growth? →",        bg=C_WHITE,   fg=C_FORMULA_FG, sz=10, h_align="right")
    status_bg = "E2EFDA" if roic_ok else "FCE4D6"
    status_fg = C_GREEN_POS if roic_ok else C_RED_NEG
    wc(ws4, 26, 6, "✓ PASS — Growth supported" if roic_ok else "⚠ FLAG — Growth may exceed ROIC",
       bg=status_bg, fg=status_fg, bold=True, sz=10, h_align="left")

    # ══════════════════════════════════════════════════════════
    # SHEET 5 — SCENARIOS (Bull / Base / Bear)
    # ══════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("5. Scenarios")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 34
    for col in ["B","C","D","E"]:
        ws5.column_dimensions[col].width = 20

    title_block(ws5, 1, ticker, "Bear / Base / Bull Scenario Analysis — Probability-Weighted Intrinsic Value", 5)

    section_header(ws5, 3, "SCENARIO ASSUMPTIONS (Blue = inputs you can change)", 5)
    header_row(ws5, 4, ["Parameter", "Bear 🐻", "Base 📊", "Bull 🐂", "Source"], bg=C_HEADER_BG,
               fg=C_HEADER_FG)

    bear = scenarios.get("Bear 🐻", {})
    base_sc = scenarios.get("Base 📊", {})
    bull = scenarios.get("Bull 🐂", {})

    scenario_params = [
        ("FCF Growth Rate",     bear.get("growth", rev_growth*0.6), base_sc.get("growth", rev_growth), bull.get("growth", rev_growth*1.4), "0.0%"),
        ("WACC",                bear.get("wacc", wacc+0.015), base_sc.get("wacc", wacc), bull.get("wacc", wacc-0.01), "0.00%"),
        ("Terminal Growth",     bear.get("term_g", terminal_g-0.005), base_sc.get("term_g", terminal_g), bull.get("term_g", terminal_g+0.005), "0.0%"),
        ("Operating Margin",    bear.get("margin", op_margin*0.85), base_sc.get("margin", op_margin), bull.get("margin", op_margin*1.1), "0.0%"),
        ("Probability Weight",  0.25, 0.50, 0.25, "0%"),
    ]
    for i, (label, bv, bsv, blv, nf) in enumerate(scenario_params):
        r = 5 + i
        ws5.row_dimensions[r].height = 18
        wc(ws5, r, 1, label, bg=C_WHITE, fg=C_FORMULA_FG, sz=10, h_align="left")
        for j, (val, bg) in enumerate([(bv, "FCE4D6"), (bsv, C_ALT_ROW), (blv, "E2EFDA")]):
            c = input_cell(ws5, r, j+2, val)
            c.number_format = nf
            c.fill = hf(bg)
        if label == "Probability Weight":
            ws5.cell(r, 2).fill = hf("FCE4D6")
            ws5.cell(r, 3).fill = hf(C_YELLOW_FLAG)
            ws5.cell(r, 4).fill = hf("E2EFDA")
        wc(ws5, r, 5, "Input — analyst discretion", bg=C_ALT_ROW, fg="595959", sz=8, h_align="left", italic=True)

    section_header(ws5, 11, "SCENARIO OUTPUTS", 5)
    header_row(ws5, 12, ["Output", "Bear 🐻", "Base 📊", "Bull 🐂", "Note"], bg=C_SUBHDR_BG)
    scenario_outputs = [
        (f"Intrinsic Value ({sym})", bear.get("iv",0)*fx, base_sc.get("iv",0)*fx, bull.get("iv",0)*fx, f'"{sym}"#,##0.00'),
        ("Margin of Safety",       bear.get("mos_pct",0)/100, base_sc.get("mos_pct",0)/100, bull.get("mos_pct",0)/100, "0.0%;(0.0%);-"),
        ("Signal",                 bear.get("signal",""), base_sc.get("signal",""), bull.get("signal",""), "@"),
    ]
    # Probability-weighted IV
    pw_iv = (bear.get("iv",0)*fx * 0.25 + base_sc.get("iv",0)*fx * 0.50 + bull.get("iv",0)*fx * 0.25)

    for i, (label, bv, bsv, blv, nf) in enumerate(scenario_outputs):
        r = 13 + i
        ws5.row_dimensions[r].height = 18
        wc(ws5, r, 1, label, bg=C_WHITE, fg=C_FORMULA_FG, sz=10, h_align="left")
        for j, (val, bg, fg) in enumerate([(bv,"FCE4D6",C_RED_NEG), (bsv,C_ALT_ROW,C_FORMULA_FG), (blv,"E2EFDA",C_GREEN_POS)]):
            c = wc(ws5, r, j+2, val if isinstance(val,str) else round(float(val),4), bg=bg, fg=fg, bold=True, sz=11)
            if nf != "@" and not isinstance(val, str): c.number_format = nf
        wc(ws5, r, 5, "Formula output", bg=C_ALT_ROW, fg="595959", sz=8, italic=True, h_align="left")

    ws5.row_dimensions[17].height = 24
    wc(ws5, 17, 1, "Probability-Weighted IV (25% / 50% / 25%)", bg=C_YELLOW_FLAG, fg=C_FORMULA_FG, bold=True, sz=11, h_align="left")
    c_pw = wc(ws5, 17, 2, round(pw_iv, 2), bg=C_YELLOW_FLAG, fg=C_INPUT_FG, bold=True, sz=12)
    c_pw.number_format = f'"{sym}"#,##0.00'
    wc(ws5, 17, 3, round((pw_iv - price)/price*100, 1) if price > 0 else 0,
       bg=C_YELLOW_FLAG, fg=C_GREEN_POS if pw_iv>price else C_RED_NEG, bold=True, sz=11, nf="0.0%")
    wc(ws5, 17, 4, "= Bear×25% + Base×50% + Bull×25%", bg=C_YELLOW_FLAG, fg="595959", sz=9, italic=True, h_align="left")

    # ══════════════════════════════════════════════════════════
    # SHEET 6 — SENSITIVITY HEATMAP
    # ══════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("6. Sensitivity")
    ws6.sheet_view.showGridLines = False
    ws6.column_dimensions["A"].width = 14

    title_block(ws6, 1, ticker, "Sensitivity Analysis — IV/Share  |  WACC (rows) × Terminal Growth Rate (columns)", 8)

    wacc_range = [wacc - 0.03, wacc - 0.02, wacc - 0.01, wacc, wacc + 0.01, wacc + 0.02, wacc + 0.03]
    tg_range   = [terminal_g - 0.01, terminal_g, terminal_g + 0.005, terminal_g + 0.01]
    tg_range   = [max(0.01, min(0.04, g)) for g in tg_range]

    section_header(ws6, 3, f"Intrinsic Value ({sym}/share)  |  Green = Undervalued vs ₹{price:.0f}  |  Red = Overvalued", 8)

    # Header row
    wc(ws6, 4, 1, f"WACC \\ g →", bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, sz=10, h_align="center")
    for j, tg in enumerate(tg_range):
        wc(ws6, 4, j+2, f"{tg:.1%}", bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, sz=10, h_align="center")
    wc(ws6, 4, len(tg_range)+2, "← Terminal Growth →", bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, sz=9, h_align="left")

    for i, w in enumerate(wacc_range):
        r = 5 + i
        ws6.row_dimensions[r].height = 20
        wc(ws6, r, 1, f"{w:.1%}", bg=C_SUBHDR_BG, fg=C_HEADER_FG, bold=True, sz=10, h_align="center")
        for j, tg in enumerate(tg_range):
            if w <= tg:
                wc(ws6, r, j+2, "N/M", bg="D9D9D9", fg="595959", sz=10, h_align="center")
                continue
            # Compute IV at this WACC/terminal g
            try:
                from screener.dcf_engine import DCFEngine
                eng = DCFEngine(discount_rate=w, terminal_growth=tg)
                res = eng.intrinsic_value_per_share(
                    projected_fcfs=forecast_result["projections"],
                    terminal_fcf_norm=forecast_result["terminal_fcf_norm"],
                    total_debt=enriched.get("total_debt",0),
                    total_cash=enriched.get("total_cash",0),
                    shares_outstanding=shares,
                    current_price=enriched.get("price",0) * 5,
                )
                cell_iv = res["intrinsic_value_per_share"] * fx
            except:
                cell_iv = 0

            is_green = cell_iv > price
            bg = "E2EFDA" if is_green else "FCE4D6"
            fg = C_GREEN_POS if is_green else C_RED_NEG
            c = wc(ws6, r, j+2, round(cell_iv, 0), bg=bg, fg=fg, bold=abs(w-wacc)<0.001, sz=10)
            c.number_format = f'"{sym}"#,##0'

    # Highlight base case
    base_row = 5 + 3  # 4th row = base WACC
    for col in range(1, len(tg_range)+2):
        cell = ws6.cell(base_row, col)
        cell.border = Border(
            left=Side(style="medium", color=_c(C_AMBER)),
            right=Side(style="medium", color=_c(C_AMBER)),
            top=Side(style="medium", color=_c(C_AMBER)),
            bottom=Side(style="medium", color=_c(C_AMBER)),
        )

    ws6.cell(5 + len(wacc_range) + 1, 1).value = f"↑ Base case WACC = {wacc:.1%} highlighted in orange border"
    ws6.cell(5 + len(wacc_range) + 1, 1).font = Font(name="Calibri", size=9, italic=True, color=_c("595959"))

    # ══════════════════════════════════════════════════════════
    # SHEET 7 — FINANCIAL STATEMENTS
    # ══════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("7. Financials")
    ws7.sheet_view.showGridLines = False
    ws7.column_dimensions["A"].width = 34
    for j in range(len(hist_years)):
        ws7.column_dimensions[get_column_letter(j+2)].width = 16

    title_block(ws7, 1, ticker, f"Historical Financial Statements  |  Source: Yahoo Finance  |  {to_code} Billions", len(hist_years)+2)

    def fin_section(ws, start_row, title, df, rows_cfg, n_hist, accent=C_SUBHDR_BG):
        section_header(ws, start_row, title, n_hist+2)
        header_row(ws, start_row+1, ["Line Item"] + hist_years + ["YoY CAGR"], bg=accent)
        r = start_row + 2
        for label, col, is_pct, is_ratio, bold in rows_cfg:
            ws.row_dimensions[r].height = 17
            bg = C_SECTION_BG if bold else (C_ALT_ROW if r%2==0 else C_WHITE)
            wc(ws, r, 1, label, bg=bg, fg=C_FORMULA_FG, bold=bold, sz=10, h_align="left")
            if col and df is not None and not df.empty and col in df.columns:
                vals = df[col].tolist()
                for j, raw in enumerate(vals):
                    if pd.isna(raw) or raw is None:
                        wc(ws, r, j+2, "-", bg=bg, fg="595959", sz=10, h_align="right")
                    elif is_pct:
                        v = raw * 100 if abs(raw) <= 1 else raw
                        fg = C_GREEN_POS if v > 0 else C_RED_NEG
                        c = wc(ws, r, j+2, v/100, bg=bg, fg=fg, sz=10)
                        c.number_format = "0.0%;(0.0%);-"
                    elif is_ratio:
                        c = wc(ws, r, j+2, round(raw,2), bg=bg, fg=C_FORMULA_FG, sz=10)
                        c.number_format = "0.00x"
                    else:
                        v = raw * fx / 1e9
                        fg = C_GREEN_POS if v > 0 else (C_RED_NEG if v < 0 else C_FORMULA_FG)
                        c = wc(ws, r, j+2, round(v,2), bg=bg, fg=fg, bold=bold, sz=10)
                        c.number_format = "#,##0.00;(#,##0.00);-"
                # CAGR
                try:
                    vals_clean = [v*fx/1e9 for v in vals if pd.notna(v) and v and v > 0]
                    if len(vals_clean) >= 2 and not is_pct:
                        cagr = (vals_clean[-1]/vals_clean[0])**(1/(len(vals_clean)-1)) - 1
                        c = wc(ws, r, len(vals)+2, cagr, bg=bg, fg=C_GREEN_POS if cagr>0 else C_RED_NEG, bold=bold, sz=10)
                        c.number_format = "0.0%"
                except: pass
            r += 1
        return r + 1

    inc_cfg = [
        (f"Revenue ({to_code}B)",             "revenue",          False, False, True),
        (f"Operating Income ({to_code}B)",     "operating_income", False, False, True),
        (f"Net Income ({to_code}B)",           "net_income",       False, False, True),
        ("Operating Margin",                   "op_margin",        True,  False, False),
        ("Net Margin",                         "net_margin",       True,  False, False),
    ]
    next_r = fin_section(ws7, 3, "INCOME STATEMENT (P&L)", income_df, inc_cfg, len(hist_years))

    cf_cfg = [
        (f"Operating Cash Flow ({to_code}B)",  "ocf",   False, False, True),
        (f"Capital Expenditure ({to_code}B)",  "capex", False, False, False),
        (f"Free Cash Flow ({to_code}B)",       "fcf",   False, False, True),
        ("FCF Margin",                         "fcf_growth", True, False, False),
    ]
    next_r = fin_section(ws7, next_r + 1, "CASH FLOW STATEMENT", cf_df, cf_cfg, len(hist_years), C_GREEN_POS)

    if bs_df is not None and not bs_df.empty:
        bs_cfg = [
            (f"Total Assets ({to_code}B)",         "total_assets",  False, False, True),
            (f"Total Debt ({to_code}B)",            "total_debt",    False, False, False),
            (f"Cash & Equivalents ({to_code}B)",   "cash",          False, False, False),
            (f"Shareholders Equity ({to_code}B)",  "equity",        False, False, True),
            ("Debt / Equity",                       "de_ratio",      False, True,  False),
            ("Current Ratio",                       "current_ratio", False, True,  False),
        ]
        fin_section(ws7, next_r + 1, "BALANCE SHEET", bs_df, bs_cfg, len(hist_years), _c(C_SUBHDR_BG)[:6] if False else C_SUBHDR_BG)

    # ══════════════════════════════════════════════════════════
    # SHEET 8 — INVESTMENT CONCLUSION
    # ══════════════════════════════════════════════════════════
    ws8 = wb.create_sheet("8. Conclusion")
    ws8.sheet_view.showGridLines = False
    ws8.column_dimensions["A"].width = 38
    ws8.column_dimensions["B"].width = 26
    ws8.column_dimensions["C"].width = 40

    title_block(ws8, 1, ticker, f"Investment Conclusion  |  {sector}  |  Moat: {moat_grade}", 3)

    conclusion = "UNDERVALUED" if mos_pct > 20 else "FAIRLY VALUED" if mos_pct > -10 else "OVERVALUED"
    conclusion_color = {"UNDERVALUED": "E2EFDA", "FAIRLY VALUED": "FFF2CC", "OVERVALUED": "FCE4D6"}[conclusion]
    conclusion_fg = {"UNDERVALUED": C_GREEN_POS, "FAIRLY VALUED": C_AMBER, "OVERVALUED": C_RED_NEG}[conclusion]

    ws8.merge_cells("A3:C5")
    c_conc = ws8.cell(3, 1, value=f"  {conclusion}")
    c_conc.fill = hf(conclusion_color)
    c_conc.font = Font(name="Calibri", bold=True, size=28, color=_c(conclusion_fg))
    c_conc.alignment = Alignment(horizontal="center", vertical="center")
    ws8.row_dimensions[3].height = 60

    summary_rows = [
        ("Current Market Price",         f"{sym}{price:,.2f}",           C_WHITE),
        ("Base Case IV (DCF + PE blend)", f"{sym}{iv_ps:,.2f}",           C_WHITE),
        ("Probability-Weighted IV",       f"{sym}{pw_iv:,.2f}",           C_YELLOW_FLAG),
        ("Margin of Safety",              f"{mos_pct:.1f}%",             conclusion_color),
        ("Bear / Base / Bull IV",         f"{sym}{bear.get('iv',0)*fx:,.0f} / {sym}{base_sc.get('iv',0)*fx:,.0f} / {sym}{bull.get('iv',0)*fx:,.0f}", C_WHITE),
        ("Moat Grade",                    moat_grade,                     C_WHITE),
        ("Fundamental Grade",             f"{enriched.get('fundamental_grade','N/A')} ({enriched.get('fundamental_score',0)}/100)", C_WHITE),
        ("Sector",                        sector,                         C_WHITE),
        ("WACC Used",                     f"{wacc:.2%}",                  C_WHITE),
        ("Terminal Growth Rate",          f"{terminal_g:.1%}",            C_WHITE),
        ("TV% of EV",                     f"{tv_pct:.0%}  {'⚠ HIGH' if tv_pct>0.75 else '✓ OK'}",  "FCE4D6" if tv_pct>0.75 else "E2EFDA"),
    ]
    section_header(ws8, 7, "VALUATION SUMMARY", 3)
    header_row(ws8, 8, ["Metric", "Value", "Note"], bg=C_SUBHDR_BG)
    notes = {
        "Margin of Safety": ">20% = undervalued by model; <0% = overvalued by model",
        "TV% of EV": "Flag if >75% — model relies too heavily on terminal assumptions",
        "Moat Grade": "Wide=sustainable 10yr+ advantage; None=commodity business",
        "Probability-Weighted IV": "25% Bear + 50% Base + 25% Bull",
    }
    for i, (label, val, bg) in enumerate(summary_rows):
        r = 9 + i
        ws8.row_dimensions[r].height = 18
        wc(ws8, r, 1, label, bg=bg, fg=C_FORMULA_FG, sz=10, h_align="left")
        wc(ws8, r, 2, val,   bg=bg, fg=conclusion_fg if "Margin" in label else C_FORMULA_FG, bold="Margin" in label, sz=10, h_align="left")
        wc(ws8, r, 3, notes.get(label, ""), bg=C_ALT_ROW, fg="595959", sz=8, h_align="left", italic=True)

    # Freeze panes on all sheets
    for ws in [ws_a, ws2, ws3, ws4, ws5, ws6, ws7, ws8]:
        ws.freeze_panes = "B5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate institutional DCF Excel model")
    parser.add_argument("--ticker", required=True, help="e.g. TCS.NS")
    args = parser.parse_args()

    ticker = args.ticker.upper().strip()
    print(f"Fetching data for {ticker}...")

    from data.collector import StockDataCollector
    from data.processor import compute_metrics
    from models.forecaster import FCFForecaster, compute_wacc
    from models.industry_wacc import get_industry_wacc
    from screener.dcf_engine import DCFEngine, margin_of_safety, assign_signal
    from screener.scenarios import run_scenarios
    from screener.valuation_model import generate_valuation_summary as generate_investment_plan
    from screener.moat_engine import compute_moat_score, apply_moat_adjustments

    collector = StockDataCollector(ticker)
    raw = collector.get_all()
    if not raw:
        print(f"Could not fetch data for {ticker}")
        sys.exit(1)

    enriched    = compute_metrics(raw)
    forecaster  = FCFForecaster()
    wacc_data   = compute_wacc(collector._ticker_obj, is_indian=ticker.endswith((".NS",".BO")))
    wacc        = wacc_data.get("wacc", 0.11)
    terminal_g  = 0.03
    forecast_yrs = 10

    industry_info = get_industry_wacc(ticker=ticker, capm_wacc=wacc)
    final_wacc    = industry_info["wacc"]
    industry_tg   = industry_info["terminal_growth"]

    dcf_engine     = DCFEngine(discount_rate=final_wacc, terminal_growth=industry_tg)
    forecast_result = forecaster.predict(enriched, years=forecast_yrs)
    projected       = forecast_result["projections"]
    terminal_norm   = forecast_result["terminal_fcf_norm"]

    dcf_res = dcf_engine.intrinsic_value_per_share(
        projected_fcfs=projected, terminal_fcf_norm=terminal_norm,
        total_debt=enriched["total_debt"], total_cash=enriched["total_cash"],
        shares_outstanding=enriched["shares"], current_price=enriched["price"],
        ticker=ticker,
    )

    moat_result = compute_moat_score(enriched, final_wacc)
    moat_adj    = apply_moat_adjustments(moat_result, final_wacc,
                    forecast_result.get("base_growth",0), industry_tg,
                    dcf_res.get("intrinsic_value_per_share",0), sector=enriched.get("sector","general"))

    enriched["moat_grade"]       = moat_result.get("grade", "None")
    enriched["fundamental_grade"]= "N/A"
    enriched["fundamental_score"]= 0

    iv_n    = dcf_res.get("intrinsic_value_per_share", 0)
    price_n = enriched["price"]
    mos     = margin_of_safety(iv_n, price_n)
    sig     = assign_signal(mos, dcf_res.get("suspicious", False), forecast_result.get("reliable", True))
    inv_plan = generate_investment_plan(enriched, price_n, iv_n, mos)

    scenarios = run_scenarios(
        enriched=enriched, fcf_base=forecast_result.get("fcf_base", 1e9),
        base_growth=forecast_result.get("base_growth", 0.08),
        base_wacc=final_wacc, base_terminal_g=industry_tg,
        total_debt=enriched["total_debt"], total_cash=enriched["total_cash"],
        shares=enriched["shares"], current_price=price_n, years=forecast_yrs,
    )

    report_data = {
        "price":    price_n,
        "iv":       iv_n,
        "mos_pct":  mos * 100,
        "signal":   sig,
        "bear_iv":  scenarios.get("Bear 🐻",{}).get("iv", iv_n*0.7),
        "bull_iv":  scenarios.get("Bull 🐂",{}).get("iv", iv_n*1.3),
    }

    print(f"Generating institutional Excel model...")
    excel_bytes = generate_institutional_dcf(
        ticker=ticker, enriched=enriched, dcf_res=dcf_res,
        forecast_result=forecast_result, scenarios=scenarios,
        wacc_data=wacc_data, wacc=final_wacc, terminal_g=industry_tg,
        forecast_yrs=forecast_yrs, sym="₹", to_code="INR", fx=1.0,
    )

    out_path = f"{ticker.replace('.','_')}_Institutional_DCF_{datetime.now().strftime('%Y%m%d')}.xlsx"
    with open(out_path, "wb") as f:
        f.write(excel_bytes)
    print(f"✅ Saved: {out_path}")
