"""Smoke tests for backend/services/excel_export_service.

Verifies the workbook builder:
  * Returns non-empty bytes that openpyxl can re-open.
  * Has the four required sheets (Inputs, DCF, Scenarios, Source Data).
  * Defines the workbook-scoped names the DCF sheet formulas depend on.
  * Survives a minimal fixture (P/BV financials path with no DCF metadata)
    without raising.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

# Repo-root on sys.path so `backend.*` imports work whether pytest is
# invoked from repo root or tests/.
_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from openpyxl import load_workbook

from backend.services.excel_export_service import build_workbook


def _full_dcf_fixture() -> dict:
    """Mimics a healthy AnalysisResponse for a DCF-valued large-cap."""
    return {
        "ticker": "TCS.NS",
        "company": {
            "name": "Tata Consultancy Services",
            "sector": "Information Technology",
            "industry": "IT Services",
            "currency": "INR",
            "market_cap": 1_200_000_000_000.0,
            "shares_outstanding": 3_660_000_000.0,
            "market_cap_source": "yfinance",
            "market_cap_as_of": "2026-05-03T00:00:00Z",
            "shares_outstanding_source": "yfinance",
        },
        "valuation": {
            "fair_value": 4200.0,
            "current_price": 3850.0,
            "margin_of_safety": 9.1,
            "verdict": "fairly_valued",
            "bear_case": 3500.0,
            "base_case": 4200.0,
            "bull_case": 4900.0,
            "wacc": 0.115,
            "terminal_growth": 0.04,
            "fcf_growth_rate": 0.09,
            "confidence_score": 78,
            "wacc_industry_min": 0.10,
            "wacc_industry_max": 0.13,
            "tv_pct_of_ev": 0.62,
            "pv_fcfs": 8_500_000_000_000.0,
            "pv_terminal": 5_500_000_000_000.0,
            "enterprise_value": 14_000_000_000_000.0,
            "equity_value": 13_500_000_000_000.0,
            "valuation_model": "dcf",
            "fair_value_computed_at": "2026-05-03T08:00:00Z",
            "current_price_source": "nse_parquet",
            "current_price_as_of": "2026-05-03T08:00:00Z",
            "valuation_engine_used": "dcf",
        },
        "scenarios": {
            "bear": {"iv": 3500.0, "mos_pct": -9.1, "growth": 0.05, "wacc": 0.13},
            "base": {"iv": 4200.0, "mos_pct": 9.1,  "growth": 0.09, "wacc": 0.115},
            "bull": {"iv": 4900.0, "mos_pct": 27.3, "growth": 0.13, "wacc": 0.10},
        },
        "quality": {
            "score": 82,
            "grade": "A",
            "piotroski_score": 7,
            "moat_grade": "Wide",
            "latest_filing_period_end": "2026-03-31",
            "revenue_cagr_window": "5y",
        },
    }


def _financials_fixture() -> dict:
    """Mimics a P/BV financials path (DCF metadata absent)."""
    return {
        "ticker": "HDFCBANK.NS",
        "company": {
            "name": "HDFC Bank",
            "sector": "Financial Services",
            "industry": "Banks",
            "currency": "INR",
            "market_cap": 12_000_000_000_000.0,
            "shares_outstanding": 7_590_000_000.0,
        },
        "valuation": {
            "fair_value": 1850.0,
            "current_price": 1620.0,
            "margin_of_safety": 14.2,
            "verdict": "undervalued",
            "bear_case": 1300.0,
            "base_case": 1850.0,
            "bull_case": 2400.0,
            "wacc": 0.0,
            "terminal_growth": 0.0,
            "fcf_growth_rate": 0.0,
            "valuation_model": "pb_ratio",
            "valuation_engine_used": "pb_residual_income",
            "confidence_score": 65,
        },
        "scenarios": {
            "bear": {"iv": 1300.0},
            "base": {"iv": 1850.0},
            "bull": {"iv": 2400.0},
        },
        "quality": {"score": 75, "grade": "B+"},
    }


def test_build_workbook_returns_bytes():
    out = build_workbook(_full_dcf_fixture())
    assert isinstance(out, bytes)
    assert len(out) > 4_000  # a real .xlsx is far bigger than this


def test_workbook_has_required_sheets():
    out = build_workbook(_full_dcf_fixture())
    wb = load_workbook(io.BytesIO(out))
    assert wb.sheetnames == ["Inputs", "DCF", "Scenarios", "Source Data"]


def test_inputs_sheet_defines_workbook_names():
    """The DCF sheet formulas reference defined names — they must exist."""
    out = build_workbook(_full_dcf_fixture())
    wb = load_workbook(io.BytesIO(out))
    expected = {"WACC", "Term_g", "Yrs", "FCF0", "G_FCF",
                "Shares", "NetDebt", "CurrPrice"}
    actual = set(wb.defined_names)
    missing = expected - actual
    assert not missing, f"missing defined names: {missing}"


def test_dcf_sheet_uses_formulas_not_constants():
    out = build_workbook(_full_dcf_fixture())
    wb = load_workbook(io.BytesIO(out))
    ws = wb["DCF"]
    # Row 6 = projected FCF; row 8 = PV of FCF; both must be formulas.
    fcf_year_1 = ws.cell(row=6, column=2).value
    pv_year_1 = ws.cell(row=8, column=2).value
    assert isinstance(fcf_year_1, str) and fcf_year_1.startswith("=")
    assert "FCF0" in fcf_year_1 and "G_FCF" in fcf_year_1
    assert isinstance(pv_year_1, str) and pv_year_1.startswith("=")
    # Intrinsic value / share row.
    iv_share = ws.cell(row=23, column=2).value
    assert isinstance(iv_share, str) and iv_share.startswith("=")


def test_scenarios_sheet_has_three_rows():
    out = build_workbook(_full_dcf_fixture())
    wb = load_workbook(io.BytesIO(out))
    ws = wb["Scenarios"]
    labels = [ws.cell(row=r, column=1).value for r in (6, 7, 8)]
    assert labels == ["Bear", "Base", "Bull"]
    # Intrinsic-value column populated for each.
    ivs = [ws.cell(row=r, column=2).value for r in (6, 7, 8)]
    assert all(isinstance(v, (int, float)) and v > 0 for v in ivs)


def test_financials_path_does_not_raise():
    """P/BV financials have no DCF metadata — builder must still succeed."""
    out = build_workbook(_financials_fixture())
    wb = load_workbook(io.BytesIO(out))
    # Should still produce all four sheets.
    assert "DCF" in wb.sheetnames
    # Source Data sheet should carry the non-DCF warning.
    src = wb["Source Data"]
    found_warning = False
    for row in src.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and "non-DCF" in cell:
                found_warning = True
                break
        if found_warning:
            break
    assert found_warning, "expected non-DCF warning in Source Data sheet"
