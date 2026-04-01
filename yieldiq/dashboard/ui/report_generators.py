"""dashboard/ui/report_generators.py
DCF text report and Excel model generators.
These are pure functions — no st.* calls, no session state.
"""
from __future__ import annotations
import io
from datetime import datetime

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ══════════════════════════════════════════════════════════════
# DCF REPORT GENERATOR
# ══════════════════════════════════════════════════════════════
def generate_dcf_report(ticker, result_data: dict, scenarios: dict, sym: str) -> bytes:
    """Generate a downloadable text-based DCF report."""
    r = result_data
    lines = [
        "=" * 65,
        f"  YieldIQ Valuation Report — {ticker}",
        f"  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        "=" * 65,
        "",
        "VALUATION SUMMARY",
        "-" * 40,
        f"  Current Price      : {sym}{r.get('price', 0):,.2f}",
        f"  Intrinsic Value    : {sym}{r.get('iv', 0):,.2f}",
        f"  Margin of Safety   : {r.get('mos_pct', 0):.1f}%",
        f"  Signal             : {r.get('signal', '')}",
        f"  WACC Used          : {r.get('wacc', 0):.1%}",
        f"  Terminal Growth    : {r.get('term_g', 0):.1%}",
        "",
        "FUNDAMENTALS",
        "-" * 40,
        f"  Revenue Growth     : {r.get('rev_growth', 0):.1%} p.a.",
        f"  FCF Growth         : {r.get('fcf_growth', 0):.1%} p.a.",
        f"  Operating Margin   : {r.get('op_margin', 0):.1%}",
        f"  Fundamental Grade  : {r.get('fund_grade', 'N/A')} ({r.get('fund_score', 0)}/100)",
        "",
        "INVESTMENT PLAN",
        "-" * 40,
        f"  Entry Signal       : {r.get('entry_signal', '')}",
        f"  Buy Price          : {sym}{r.get('buy_price', 0):,.2f}",
        f"  Target Price       : {sym}{r.get('target_price', 0):,.2f}",
        f"  Stop Loss          : {sym}{r.get('stop_loss', 0):,.2f}  (-{r.get('sl_pct', 0):.1f}%)",
        f"  Risk/Reward        : {r.get('rr_ratio', 0):.1f}x",
        f"  Holding Period     : {r.get('holding_period', 'N/A')}",
        "",
        "THREE SCENARIO ANALYSIS",
        "-" * 40,
    ]

    for sname, sdata in scenarios.items():
        lines += [
            f"  {sname}",
            f"    Growth: {sdata['growth']:.1%}  WACC: {sdata['wacc']:.1%}  Terminal g: {sdata['term_g']:.1%}",
            f"    Intrinsic Value: {sym}{sdata['iv']:,.2f}  |  MoS: {sdata['mos_pct']:.1f}%",
            "",
        ]

    lines += [
        "DCF WATERFALL  (raw DCF — before PE blend)",
        "-" * 40,
        f"  PV of FCFs         : {sym}{r.get('sum_pv_fcfs', 0):,.0f}",
        f"  PV Terminal Value  : {sym}{r.get('pv_tv', 0):,.0f}",
        f"  Enterprise Value   : {sym}{r.get('ev', 0):,.0f}",
        f"  Less: Total Debt   : {sym}{r.get('debt', 0):,.0f}",
        f"  Plus: Cash         : {sym}{r.get('cash', 0):,.0f}",
        f"  Equity Value       : {sym}{r.get('equity', 0):,.0f}",
        f"  Shares Outstanding : {r.get('shares', 0)/1e9:.3f}B",
        f"  DCF IV/share       : {sym}{r.get('dcf_only_iv', r.get('iv', 0)):,.2f}",
        f"  PE-blended IV/sh   : {sym}{r.get('iv', 0):,.2f}  ← headline number",
        "",
        "=" * 65,
        "DISCLAIMER: Educational purposes only. Not financial advice.",
        "Always conduct independent research before investing.",
        "=" * 65,
    ]
    return "\n".join(lines).encode("utf-8")


def generate_excel_dcf_model(
    ticker: str,
    enriched: dict,
    dcf_res: dict,
    forecast_result: dict,
    scenarios: dict,
    inv_plan: dict,
    report_data: dict,
    sensitivity_df: pd.DataFrame,
    sym: str,
    to_code: str,
    fx: float,
    wacc: float,
    terminal_g: float,
    forecast_yrs: int,
) -> bytes:
    """
    Generate a professional multi-sheet Excel DCF model.
    Sheets: Summary | DCF Model | FCFF Build | ROIC Analysis |
            P&L | Cash Flow | Balance Sheet | Scenarios |
            Sensitivity | Quality Checks | Historical Charts | Assumptions
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.chart.series import DataPoint
    except ImportError:
        return None

    wb = openpyxl.Workbook()

    # ── Palette ────────────────────────────────────────────────
    DARK_BG  = "0D1424"; BLUE_HDR = "1D4ED8"; GREEN_HDR = "065F46"
    RED_HDR  = "7F1D1D"; AMBER_HDR= "78350F"; CYAN_HDR  = "164E63"
    PURP_HDR = "3B0764"; MID_BG   = "0A1020"; BORDER_C  = "1E2D45"
    TEXT_MAIN= "F1F5F9"; TEXT_DIM = "94A3B8"; TEXT_NUM  = "E2E8F0"
    GREEN_POS= "10B981"; RED_NEG  = "EF4444"; AMBER_VAL = "F59E0B"
    WARN_BG  = "451A03"; OK_BG    = "052E16"

    def _c(s):
        """Convert any hex color to valid 8-char aRGB for openpyxl.
        Handles 6-char, 8-char, 10-char (from +"44" appends), # prefix."""
        if not s: return "FFF1F5F9"
        s = str(s).lstrip("#").upper()
        if len(s) == 8 and s[:2] == "FF": s = s[2:]  # strip FF prefix -> 6-char
        if len(s) == 6:  return "FF" + s    # fully opaque
        if len(s) == 8:  return s            # already aRGB
        if len(s) >= 10: return s[:8]        # e.g. "1D4ED844" extra chars
        return "FFF1F5F9"

    def hf(c): return PatternFill("solid", fgColor=_c(c))
    def tb():
        s = Side(style='thin', color=_c(BORDER_C))
        return Border(left=s, right=s, top=s, bottom=s)
    def hdr_font(bold=True, sz=11, color=TEXT_MAIN):
        return Font(name="Calibri", bold=bold, size=sz, color=_c(color))
    def val_font(bold=False, sz=11, color=TEXT_NUM, mono=False):
        return Font(name="Courier New" if mono else "Calibri", bold=bold, size=sz, color=_c(color))
    def center(): return Alignment(horizontal="center", vertical="center")
    def right():  return Alignment(horizontal="right",  vertical="center")
    def left():   return Alignment(horizontal="left",   vertical="center")
    def wrap_left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def wc(ws, row, col, value, fill=None, font=None, align=None, nf=None, h=22):
        c = ws.cell(row=row, column=col, value=value)
        if fill:  c.fill = hf(fill)
        if font:  c.font = font
        if align: c.alignment = align
        if nf:    c.number_format = nf
        c.border = tb()
        ws.row_dimensions[row].height = h
        return c

    def sec(ws, row, text, cols, fill=BLUE_HDR, sz=12):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        c = ws.cell(row=row, column=1, value=f"  {text}")
        c.fill = hf(fill); c.font = hdr_font(True, sz); c.alignment = left()
        c.border = tb(); ws.row_dimensions[row].height = 24

    def title_row(ws, text, cols, fill=BLUE_HDR, sz=15):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
        c = ws["A1"]; c.value = text; c.fill = hf(fill)
        c.font = Font(name="Calibri", bold=True, size=sz, color="FFFFFFFF")
        c.alignment = center(); ws.row_dimensions[1].height = 38

    def subtitle(ws, text, cols, fill=MID_BG):
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
        c = ws["A2"]; c.value = text; c.fill = hf(fill)
        c.font = Font(name="Calibri", size=10, color=_c(TEXT_DIM))
        c.alignment = center(); ws.row_dimensions[2].height = 18

    # ── Shared data ────────────────────────────────────────────
    projected  = forecast_result["projections"]
    pv_fcfs    = dcf_res.get("pv_fcfs", [])
    gs         = forecast_result.get("growth_schedule", [])
    yr_labels  = [f"Year {i+1}" for i in range(forecast_yrs)]
    income_df  = enriched.get("income_df", pd.DataFrame())
    cf_df      = enriched.get("cf_df",     pd.DataFrame())
    bs_df      = enriched.get("bs_df",     pd.DataFrame())
    pt         = inv_plan["price_targets"]
    hp         = inv_plan["holding_period"]
    fs         = inv_plan["fundamental"]
    price_d    = report_data["price"]
    iv_d       = report_data["iv"]
    mos_pct    = report_data["mos_pct"]
    pv_tv_d    = dcf_res.get("pv_tv", 0) * fx
    ev_d       = dcf_res.get("enterprise_value", 0) * fx
    tv_pct     = dcf_res.get("tv_pct_of_ev", 0)

    # Get historical years from income_df
    hist_years = []
    if not income_df.empty and "year" in income_df.columns:
        hist_years = [str(int(y)) for y in income_df["year"].tolist()]

    # ═══════════════════════════════════════════════════════════
    # SHEET 1 — SUMMARY DASHBOARD
    # ═══════════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = "📊 Summary"
    ws1.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [34, 22, 22, 22, 22]):
        ws1.column_dimensions[col].width = w

    title_row(ws1, f"YieldIQ Valuation Report — {ticker}", 5)
    subtitle(ws1, f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Currency: {to_code}  |  WACC: {wacc:.2%}  |  Terminal g: {terminal_g:.2%}  |  Horizon: {forecast_yrs} yrs", 5)

    sec(ws1, 4, "VALUATION SUMMARY", 5, BLUE_HDR)
    tv_flag = "⚠️  EXCEEDS 70% — HIGH SENSITIVITY" if tv_pct > 0.70 else "✅  Within safe limit"
    val_rows = [
        ("Current Market Price",          f"{sym}{price_d:,.2f}",                  DARK_BG, MID_BG),
        ("Intrinsic Value (Base DCF)",    f"{sym}{iv_d:,.2f}",                     DARK_BG, MID_BG),
        ("Discount to Fair Value",         f"{mos_pct:.1f}%",                       DARK_BG, MID_BG),
        ("Signal",                        report_data["signal"],                   DARK_BG, MID_BG),
        ("Bear Case IV",                  f"{sym}{report_data.get('bear_iv',0):,.2f}", DARK_BG, RED_HDR),
        ("Base Case IV",                  f"{sym}{iv_d:,.2f}",                     DARK_BG, AMBER_HDR),
        ("Bull Case IV",                  f"{sym}{report_data.get('bull_iv',0):,.2f}", DARK_BG, GREEN_HDR),
        ("Terminal Value % of EV",        f"{tv_pct:.1%}  {tv_flag}",             DARK_BG, WARN_BG if tv_pct > 0.70 else OK_BG),
    ]
    for i, (label, value, lbg, vbg) in enumerate(val_rows):
        r = 5 + i
        wc(ws1, r, 1, label, fill=lbg, font=hdr_font(False, 11, TEXT_DIM),  align=left())
        wc(ws1, r, 2, value, fill=vbg, font=val_font(True,  12, TEXT_MAIN, True), align=right())

    sec(ws1, 14, "INVESTMENT ACTION PLAN", 5, GREEN_HDR)
    inv_rows = [
        ("Entry Signal",      pt.get("entry_signal", "")),
        ("Buy Zone Price",    f"{sym}{(pt.get('buy_price') or 0)*fx:,.2f}"),
        ("Target Price",      f"{sym}{(pt.get('target_price') or 0)*fx:,.2f}"),
        ("Stop Loss",         f"{sym}{(pt.get('stop_loss') or 0)*fx:,.2f}  (−{pt.get('sl_pct',0):.1f}%)"),
        ("Risk / Reward",     f"{pt.get('rr_ratio',0):.2f}x"),
        ("Suggested Holding", hp.get("label", "N/A")),
        ("Rationale",         hp.get("rationale","")[:80]),
    ]
    for i, (label, value) in enumerate(inv_rows):
        r = 15 + i
        wc(ws1, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
        wc(ws1, r, 2, value, fill=MID_BG,  font=val_font(True, 11, TEXT_MAIN), align=left())

    sec(ws1, 23, "FUNDAMENTAL STRENGTH", 5, CYAN_HDR)
    fund_rows = [
        ("Overall Grade",     f"{fs.get('grade','N/A')} — {fs.get('score',0)}/100"),
        ("Revenue Growth",    f"{enriched.get('revenue_growth',0)*100:.1f}% p.a."),
        ("Operating Margin",  f"{enriched.get('op_margin',0)*100:.1f}%"),
        ("FCF Growth",        f"{enriched.get('fcf_growth',0)*100:.1f}% p.a."),
        ("FCF Positive",      "Yes" if enriched.get("latest_fcf",0) > 0 else "No"),
    ]
    for i, (label, value) in enumerate(fund_rows):
        r = 24 + i
        wc(ws1, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
        wc(ws1, r, 2, value, fill=MID_BG,  font=val_font(True, 11, TEXT_MAIN), align=left())

    # ═══════════════════════════════════════════════════════════
    # SHEET 2 — DCF MODEL (CORE)
    # ═══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("🧮 DCF Model")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 36
    for i in range(1, forecast_yrs + 3):
        ws2.column_dimensions[get_column_letter(i+1)].width = 15

    title_row(ws2, f"DCF MODEL — {ticker}", forecast_yrs + 2, BLUE_HDR)
    subtitle(ws2, f"WACC: {wacc:.2%}  |  Terminal g: {terminal_g:.2%}  |  Forecast Horizon: {forecast_yrs} years", forecast_yrs + 2)

    # Year headers
    sec(ws2, 4, "FREE CASH FLOW PROJECTIONS", forecast_yrs + 2, BLUE_HDR)
    wc(ws2, 5, 1, "Metric", fill=BLUE_HDR, font=hdr_font(sz=11), align=left())
    for j, lbl in enumerate(yr_labels):
        wc(ws2, 5, j+2, lbl, fill=BLUE_HDR, font=hdr_font(sz=11), align=center())
    wc(ws2, 5, forecast_yrs+2, "Terminal", fill=AMBER_HDR, font=hdr_font(sz=11), align=center())

    term_fcf = forecast_result.get("terminal_fcf_norm", 0) * fx / 1e9
    pv_tv_bn = dcf_res.get("pv_tv", 0) * fx / 1e9

    proj_rows = [
        (f"Projected FCF ({to_code}B)",  [v*fx/1e9 for v in projected],  term_fcf,  "#,##0.00", TEXT_MAIN, True),
        ("YoY Growth Rate (%)",           [g*100 for g in gs],           "",         "0.0",      AMBER_VAL, False),
        (f"PV of FCF ({to_code}B)",       [v*fx/1e9 for v in pv_fcfs],   pv_tv_bn,  "#,##0.00", TEXT_NUM,  False),
    ]
    for k, (label, vals, term_val, nf, clr, bold) in enumerate(proj_rows):
        r = 6 + k
        wc(ws2, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
        for j, v in enumerate(vals):
            wc(ws2, r, j+2, round(float(v), 3) if v != "" else v,
               fill=MID_BG, font=val_font(bold, 11, clr, True), align=right(), nf=nf)
        if term_val != "":
            wc(ws2, r, forecast_yrs+2, round(float(term_val), 3),
               fill=MID_BG, font=val_font(bold, 11, clr, True), align=right(), nf=nf)

    # DCF Waterfall
    sec(ws2, 10, "DCF WATERFALL — BRIDGE TO EQUITY VALUE", forecast_yrs + 2, BLUE_HDR)
    waterfall = [
        (f"Σ PV of FCFs ({to_code}B)",           dcf_res.get("sum_pv_fcfs",0)*fx/1e9,  TEXT_NUM),
        (f"PV of Terminal Value ({to_code}B)",    pv_tv_bn,                              TEXT_NUM),
        (f"Enterprise Value ({to_code}B)",         ev_d/1e9,                             TEXT_MAIN),
        (f"Less: Total Debt ({to_code}B)",         enriched.get("total_debt",0)*fx/1e9, RED_NEG),
        (f"Plus: Cash & Equivalents ({to_code}B)", enriched.get("total_cash",0)*fx/1e9, GREEN_POS),
        (f"Equity Value ({to_code}B)",             dcf_res.get("equity_value",0)*fx/1e9, TEXT_MAIN),
        ("Shares Outstanding (Billions)",          enriched.get("shares",0)/1e9,          TEXT_NUM),
        (f"Intrinsic Value Per Share ({sym})",     iv_d,                                  GREEN_POS if mos_pct > 0 else RED_NEG),
        (f"Current Market Price ({sym})",          price_d,                               AMBER_VAL),
        ("Discount to fair value (%)",                   mos_pct,                               GREEN_POS if mos_pct > 0 else RED_NEG),
    ]
    for i, (label, value, clr) in enumerate(waterfall):
        r = 11 + i
        wc(ws2, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
        wc(ws2, r, 2, round(float(value), 2), fill=MID_BG,
           font=val_font(True, 12, clr, True), align=right(), nf="#,##0.00")

    # ═══════════════════════════════════════════════════════════
    # SHEET 3 — FCFF REINVESTMENT MODEL (NEW)
    # ═══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("🔬 FCFF Build")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 38
    for i in range(1, len(hist_years) + forecast_yrs + 3):
        ws3.column_dimensions[get_column_letter(i+1)].width = 16

    all_labels  = (["Historical: " + y for y in hist_years] +
                   [f"Forecast Y{i+1}" for i in range(forecast_yrs)])
    total_cols  = len(all_labels)

    title_row(ws3, f"FCFF REINVESTMENT BUILD — {ticker}", total_cols + 2, PURP_HDR)
    subtitle(ws3, "FCFF = NOPAT + D&A − CapEx − ΔWorking Capital  |  Professional bottom-up FCF construction", total_cols + 2, MID_BG)

    # Column headers
    wc(ws3, 4, 1, "Line Item", fill=PURP_HDR, font=hdr_font(sz=11), align=left(), h=24)
    for j, lbl in enumerate(all_labels):
        is_hist = j < len(hist_years)
        wc(ws3, 4, j+2, lbl, fill=BLUE_HDR if is_hist else PURP_HDR,
           font=hdr_font(sz=10), align=center(), h=24)

    # Pull historical data from income_df and cf_df
    def get_col(df, col, default=None):
        if df is not None and not df.empty and col in df.columns:
            return df[col].tolist()
        return [default] * (len(df) if df is not None and not df.empty else 0)

    hist_rev   = get_col(income_df, "revenue",        0)
    hist_opinc = get_col(income_df, "operating_income",0)
    hist_ni    = get_col(income_df, "net_income",     0)
    hist_cfo   = get_col(cf_df,    "cfo",             0)
    hist_capex = get_col(cf_df,    "capex",           0)
    hist_fcf   = get_col(cf_df,    "fcf",             0)

    # Estimate NOPAT, reinvestment, ROIC from historical data
    latest_rev    = enriched.get("latest_revenue", 0)
    latest_opinc  = enriched.get("latest_opinc", hist_opinc[-1] if hist_opinc else 0)
    latest_fcf_v  = enriched.get("latest_fcf", 0)
    op_margin     = enriched.get("op_margin", 0)
    rev_growth    = enriched.get("revenue_growth", 0)
    tax_rate      = 0.25  # typical effective tax rate assumption

    # Build NOPAT and reinvestment rate for historical + forecast
    nopat_hist = [v * (1 - tax_rate) * fx / 1e9 for v in hist_opinc]

    # Forecast reinvestment model: project Revenue → EBIT → NOPAT → FCFF
    base_rev = latest_rev * fx / 1e9 if latest_rev else 1.0
    fc_rev, fc_ebit, fc_nopat, fc_da, fc_capex_f, fc_dwc, fc_reinv, fc_fcff = [], [], [], [], [], [], [], []
    growth_fade = [max(rev_growth * np.exp(-0.3 * i), terminal_g) for i in range(1, forecast_yrs+1)]

    running_rev = base_rev
    for i, g in enumerate(growth_fade):
        running_rev = running_rev * (1 + g)
        ebit   = running_rev * op_margin
        nopat  = ebit * (1 - tax_rate)
        da     = running_rev * 0.035          # D&A ~3.5% of revenue
        capex_ = running_rev * 0.045          # CapEx ~4.5% of revenue
        dwc    = running_rev * g * 0.08       # ΔWC ~8% of incremental revenue
        reinv  = capex_ - da + dwc            # net reinvestment
        fcff   = nopat - reinv
        fc_rev.append(round(running_rev, 2)); fc_ebit.append(round(ebit, 3))
        fc_nopat.append(round(nopat, 3));     fc_da.append(round(da, 3))
        fc_capex_f.append(round(capex_, 3));  fc_dwc.append(round(dwc, 3))
        fc_reinv.append(round(reinv, 3));     fc_fcff.append(round(fcff, 3))

    def fcff_row(ws, row, label, hist_vals, fc_vals, fill_h, fill_f, clr, bold=False, nf="#,##0.00"):
        wc(ws, row, 1, label, fill=fill_h if hist_vals else DARK_BG,
           font=hdr_font(bold, 11, TEXT_DIM if not bold else TEXT_MAIN), align=left())
        for j, v in enumerate(hist_vals):
            val = v * fx / 1e9 if v else 0
            wc(ws, row, j+2, round(val, 3), fill=fill_h,
               font=val_font(bold, 11, clr, True), align=right(), nf=nf)
        for j, v in enumerate(fc_vals):
            wc(ws, row, len(hist_vals)+j+2, v, fill=fill_f,
               font=val_font(bold, 11, clr, True), align=right(), nf=nf)

    sec(ws3, 5,  "INCOME BRIDGE", total_cols + 2, PURP_HDR)
    fcff_row(ws3, 6,  f"Revenue ({to_code}B)",           hist_rev,   fc_rev,    BLUE_HDR+"44", PURP_HDR+"44", TEXT_MAIN, True)
    fcff_row(ws3, 7,  "Operating Margin (%)",            [v/r if r else 0 for v, r in zip(hist_opinc, hist_rev)], [op_margin]*forecast_yrs, DARK_BG, DARK_BG, AMBER_VAL, nf="0.0%")
    fcff_row(ws3, 8,  f"Operating Income / EBIT ({to_code}B)", hist_opinc, fc_ebit, MID_BG, MID_BG, TEXT_NUM)
    fcff_row(ws3, 9,  f"NOPAT (after tax @ {tax_rate:.0%}) ({to_code}B)", nopat_hist, fc_nopat, MID_BG, MID_BG, TEXT_NUM)

    sec(ws3, 11, "REINVESTMENT COMPONENTS", total_cols + 2, GREEN_HDR)
    fcff_row(ws3, 12, f"D&A (est. ~3.5% rev) ({to_code}B)",     [0]*len(hist_years), fc_da,     DARK_BG, MID_BG, GREEN_POS)
    fcff_row(ws3, 13, f"Capital Expenditure ({to_code}B)",       hist_capex, fc_capex_f, MID_BG, MID_BG, RED_NEG)
    fcff_row(ws3, 14, f"ΔWorking Capital ({to_code}B)",          [0]*len(hist_years), fc_dwc,    DARK_BG, MID_BG, RED_NEG)
    fcff_row(ws3, 15, f"Net Reinvestment ({to_code}B)",          [0]*len(hist_years), fc_reinv,  DARK_BG, MID_BG, AMBER_VAL)

    sec(ws3, 17, "FCFF DERIVATION  —  NOPAT − Net Reinvestment", total_cols + 2, CYAN_HDR)
    fcff_row(ws3, 18, f"Historical FCF Actual ({to_code}B)",     hist_fcf,   [],        CYAN_HDR+"44", DARK_BG, TEXT_MAIN, True)
    fcff_row(ws3, 19, f"FCFF (Reinvestment Model) ({to_code}B)", [],         fc_fcff,   DARK_BG, CYAN_HDR+"44", GREEN_POS, True)

    sec(ws3, 21, "REINVESTMENT RATE CHECK", total_cols + 2, AMBER_HDR)
    reinv_rates = [round(r/n, 3) if n else 0 for r, n in zip(fc_reinv, fc_nopat)]
    fcff_row(ws3, 22, "Reinvestment Rate (Reinv / NOPAT)",       [], reinv_rates, DARK_BG, MID_BG, AMBER_VAL, nf="0.0%")

    # ═══════════════════════════════════════════════════════════
    # SHEET 4 — ROIC vs WACC (NEW)
    # ═══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("📐 ROIC vs WACC")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 36
    for col in ["B","C","D","E","F","G"]:
        ws4.column_dimensions[col].width = 20

    title_row(ws4, f"ROIC vs WACC — VALUE CREATION ANALYSIS — {ticker}", 7, CYAN_HDR)
    subtitle(ws4, "ROIC > WACC = Value Creation  |  ROIC < WACC = Value Destruction  |  Spread × Invested Capital = Economic Profit", 7)

    # Compute ROIC from historical data
    # ROIC = NOPAT / Invested Capital  (where IC = Total Assets - Non-interest-bearing Current Liabilities)
    # Proxy: IC = Total Equity + Total Debt (book value of capital)
    total_debt_v  = enriched.get("total_debt",  0) * fx / 1e9
    total_cash_v  = enriched.get("total_cash",  0) * fx / 1e9
    shares_v      = enriched.get("shares",      0)
    latest_fcf_bn = enriched.get("latest_fcf",  0) * fx / 1e9

    # Estimate invested capital and ROIC
    latest_rev_bn = latest_rev * fx / 1e9 if latest_rev else 1.0
    nopat_latest  = latest_rev_bn * op_margin * (1 - tax_rate)

    # IC proxy: use EV / Revenue multiple as sanity check
    ev_bn = ev_d / 1e9
    ic_estimate = max(ev_bn * 0.6, total_debt_v + latest_rev_bn * 0.3)  # rough IC proxy
    roic_estimate = nopat_latest / ic_estimate if ic_estimate > 0 else 0
    spread = roic_estimate - wacc
    ep_estimate = spread * ic_estimate  # Economic Profit = Spread × IC

    sec(ws4, 4, "ROIC CALCULATION", 7, CYAN_HDR)
    roic_rows = [
        ("Latest Revenue",                      f"{to_code}B",    f"{latest_rev_bn:,.2f}"),
        ("Operating Margin",                    "%",              f"{op_margin*100:.1f}%"),
        ("EBIT",                                f"{to_code}B",    f"{latest_rev_bn*op_margin:,.2f}"),
        ("Effective Tax Rate (assumed)",        "%",              f"{tax_rate*100:.0f}%"),
        ("NOPAT (Net Operating Profit After Tax)", f"{to_code}B", f"{nopat_latest:,.2f}"),
        ("Invested Capital (estimated proxy)",  f"{to_code}B",   f"{ic_estimate:,.2f}"),
        ("ROIC",                                "%",              f"{roic_estimate*100:.1f}%"),
    ]
    wc(ws4, 5, 1, "Metric",       fill=CYAN_HDR, font=hdr_font(sz=11), align=left())
    wc(ws4, 5, 2, "Unit",         fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 5, 3, "Value",        fill=CYAN_HDR, font=hdr_font(sz=11), align=right())
    for i, (label, unit, value) in enumerate(roic_rows):
        r = 6 + i
        wc(ws4, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
        wc(ws4, r, 2, unit,  fill=DARK_BG, font=val_font(False, 11, TEXT_DIM), align=center())
        wc(ws4, r, 3, value, fill=MID_BG,  font=val_font(True, 12, TEXT_MAIN, True), align=right())

    sec(ws4, 14, "ROIC vs WACC SPREAD — VALUE CREATION TEST", 7, CYAN_HDR)
    spread_ok = roic_estimate > wacc
    spread_rows = [
        ("ROIC",                  f"{roic_estimate*100:.2f}%",  GREEN_POS if spread_ok else RED_NEG),
        ("WACC",                  f"{wacc*100:.2f}%",           AMBER_VAL),
        ("Value Creation Spread (ROIC − WACC)",
                                  f"{spread*100:+.2f}%",        GREEN_POS if spread_ok else RED_NEG),
        ("Economic Profit (Spread × IC)",
                                  f"{to_code}B {ep_estimate:,.2f}", GREEN_POS if spread_ok else RED_NEG),
        ("Verdict",
                                  "✅ ROIC > WACC — Growth creates value" if spread_ok
                                  else "⚠️ ROIC < WACC — Growth destroys value",
                                  GREEN_POS if spread_ok else RED_NEG),
    ]
    wc(ws4, 15, 1, "Metric",      fill=CYAN_HDR, font=hdr_font(sz=11), align=left())
    wc(ws4, 15, 2, "Value",       fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 15, 3, "Interpretation", fill=CYAN_HDR, font=hdr_font(sz=11), align=left())
    interp = {
        "ROIC":          "Return earned on every unit of capital deployed",
        "WACC":          "Minimum return required by capital providers",
        "Value Creation Spread (ROIC − WACC)": "Positive = value creation; Negative = destruction",
        "Economic Profit (Spread × IC)": "Total economic value added (or destroyed) per year",
        "Verdict":       "Key quality test — positive spread supports DCF growth assumptions",
    }
    for i, (label, value, clr) in enumerate(spread_rows):
        r = 16 + i
        bg = OK_BG if spread_ok and i < 4 else (WARN_BG if not spread_ok and i < 4 else DARK_BG)
        wc(ws4, r, 1, label,  fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
        wc(ws4, r, 2, value,  fill=bg,      font=val_font(True, 12, clr, True),  align=center())
        wc(ws4, r, 3, interp.get(label,""), fill=DARK_BG, font=Font(name="Calibri", italic=True, size=10, color=TEXT_DIM), align=wrap_left())

    # Forecast ROIC trend
    sec(ws4, 23, "FORECAST ROIC TREND", 7, CYAN_HDR)
    wc(ws4, 24, 1, "Year",           fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 24, 2, f"Revenue ({to_code}B)", fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 24, 3, f"NOPAT ({to_code}B)",  fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 24, 4, "ROIC (%)",       fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 24, 5, "WACC (%)",       fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 24, 6, "Spread (%)",     fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    for i, (rev, nopat, fcff_v) in enumerate(zip(fc_rev, fc_nopat, fc_fcff)):
        r = 25 + i
        fc_roic   = nopat / ic_estimate if ic_estimate > 0 else 0
        fc_spread = fc_roic - wacc
        wc(ws4, r, 1, yr_labels[i], fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=center())
        wc(ws4, r, 2, rev,          fill=MID_BG,  font=val_font(False, 11, TEXT_NUM, True), align=right(), nf="#,##0.00")
        wc(ws4, r, 3, nopat,        fill=MID_BG,  font=val_font(False, 11, TEXT_NUM, True), align=right(), nf="#,##0.00")
        wc(ws4, r, 4, f"{fc_roic*100:.1f}%",   fill=MID_BG, font=val_font(True, 11, GREEN_POS if fc_roic > wacc else RED_NEG, True), align=right())
        wc(ws4, r, 5, f"{wacc*100:.1f}%",       fill=MID_BG, font=val_font(False, 11, AMBER_VAL, True), align=right())
        wc(ws4, r, 6, f"{fc_spread*100:+.1f}%", fill=OK_BG if fc_spread > 0 else WARN_BG,
           font=val_font(True, 11, GREEN_POS if fc_spread > 0 else RED_NEG, True), align=right())

    # ═══════════════════════════════════════════════════════════
    # SHEET 5 — INCOME STATEMENT
    # ═══════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("📋 Income Statement")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 36
    for j, yr in enumerate(hist_years):
        ws5.column_dimensions[get_column_letter(j+2)].width = 18

    def fin_sheet_build(ws, title_text, df, rows_cfg, accent):
        title_row(ws, title_text, max(len(hist_years)+1, 5), accent)
        if df is None or df.empty:
            ws.cell(row=3, column=1).value = "No data available"
            return
        wc(ws, 4, 1, "Line Item", fill=accent, font=hdr_font(sz=11), align=left(), h=22)
        for j, yr in enumerate(hist_years):
            wc(ws, 4, j+2, yr, fill=accent, font=hdr_font(sz=11), align=center(), h=22)

        for k, (label, col, is_pct, bold, is_sec) in enumerate(rows_cfg):
            r = 5 + k
            if is_sec:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(hist_years)+1)
                c = ws.cell(row=r, column=1, value=f"  {label}")
                c.fill = hf(accent); c.font = hdr_font(True, 11); c.alignment = left()
                c.border = tb(); ws.row_dimensions[r].height = 22
            else:
                wc(ws, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM if not bold else TEXT_MAIN), align=left())
                if col and col in df.columns:
                    for j, raw in enumerate(df[col].tolist()):
                        if pd.isna(raw) or raw is None:
                            wc(ws, r, j+2, "—", fill=MID_BG, font=val_font(bold, 11, TEXT_DIM, True), align=right())
                        elif is_pct:
                            v = raw * 100 if abs(raw) <= 1 else raw
                            clr = GREEN_POS if v > 0 else RED_NEG
                            wc(ws, r, j+2, f"{v:.1f}%", fill=MID_BG, font=val_font(bold, 11, clr, True), align=right())
                        else:
                            v = raw * fx / 1e9
                            clr = GREEN_POS if v > 0 else (RED_NEG if v < 0 else TEXT_NUM)
                            wc(ws, r, j+2, round(v, 2), fill=MID_BG, font=val_font(bold, 11, clr, True), align=right(), nf="#,##0.00")

    inc_cfg = [
        ("REVENUE & PROFITABILITY", None,              False, True,  True),
        (f"Revenue ({to_code}B)",   "revenue",         False, True,  False),
        (f"Gross Profit ({to_code}B)","gross_profit",  False, False, False),
        (f"Operating Income ({to_code}B)","operating_income",False,True,False),
        (f"Net Income ({to_code}B)","net_income",      False, True,  False),
        ("MARGINS",                 None,              False, True,  True),
        ("Gross Margin",            "gross_margin",    True,  False, False),
        ("Operating Margin",        "op_margin",       True,  True,  False),
        ("Net Margin",              "net_margin",      True,  False, False),
    ]
    fin_sheet_build(ws5, f"INCOME STATEMENT — {ticker}  ({to_code} Billions)", income_df, inc_cfg, BLUE_HDR)

    # ═══════════════════════════════════════════════════════════
    # SHEET 6 — CASH FLOW STATEMENT
    # ═══════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("💰 Cash Flow")
    ws6.sheet_view.showGridLines = False
    ws6.column_dimensions["A"].width = 36
    for j in range(len(hist_years)):
        ws6.column_dimensions[get_column_letter(j+2)].width = 18

    cf_cfg = [
        ("OPERATING ACTIVITIES",       None,      False, True,  True),
        (f"Operating Cash Flow ({to_code}B)","cfo",False,True,  False),
        (f"Capital Expenditure ({to_code}B)","capex",False,False,False),
        ("FREE CASH FLOW",             None,      False, True,  True),
        (f"Free Cash Flow ({to_code}B)","fcf",    False, True,  False),
        ("GROWTH",                     None,      False, True,  True),
        ("FCF YoY Growth",             "fcf_growth",True,False, False),
    ]
    fin_sheet_build(ws6, f"CASH FLOW STATEMENT — {ticker}  ({to_code} Billions)", cf_df, cf_cfg, GREEN_HDR)

    # ═══════════════════════════════════════════════════════════
    # SHEET 7 — BALANCE SHEET
    # ═══════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("🏦 Balance Sheet")
    ws7.sheet_view.showGridLines = False
    ws7.column_dimensions["A"].width = 36
    for j in range(len(hist_years)):
        ws7.column_dimensions[get_column_letter(j+2)].width = 18

    if bs_df is not None and not bs_df.empty:
        bs_cfg = [
            ("ASSETS",                              None,           False, True,  True),
            (f"Total Assets ({to_code}B)",          "total_assets", False, True,  False),
            (f"Cash & Equivalents ({to_code}B)",    "cash",         False, False, False),
            (f"Current Assets ({to_code}B)",        "current_assets",False,False, False),
            ("LIABILITIES",                         None,           False, True,  True),
            (f"Total Debt ({to_code}B)",            "total_debt",   False, True,  False),
            (f"Current Liabilities ({to_code}B)",   "current_liab", False, False, False),
            ("EQUITY",                              None,           False, True,  True),
            (f"Shareholders' Equity ({to_code}B)",  "equity",       False, True,  False),
            ("Debt / Equity Ratio",                 "de_ratio",     False, False, False),
            ("Current Ratio",                       "current_ratio",False, False, False),
        ]
        fin_sheet_build(ws7, f"BALANCE SHEET — {ticker}  ({to_code} Billions)", bs_df, bs_cfg, CYAN_HDR)
    else:
        title_row(ws7, f"BALANCE SHEET SNAPSHOT — {ticker}", 4, CYAN_HDR)
        for i, (label, val) in enumerate([
            (f"Cash ({to_code}B)", enriched.get("total_cash",0)*fx/1e9),
            (f"Total Debt ({to_code}B)", enriched.get("total_debt",0)*fx/1e9),
        ]):
            wc(ws7, 4+i, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
            wc(ws7, 4+i, 2, round(val,2), fill=MID_BG, font=val_font(True,11,TEXT_MAIN,True), align=right(), nf="#,##0.00")

    # ═══════════════════════════════════════════════════════════
    # SHEET 8 — SCENARIO ANALYSIS
    # ═══════════════════════════════════════════════════════════
    ws8 = wb.create_sheet("🎭 Scenarios")
    ws8.sheet_view.showGridLines = False
    ws8.column_dimensions["A"].width = 30
    for col, w in zip(["B","C","D"], [22,22,22]):
        ws8.column_dimensions[col].width = w

    title_row(ws8, f"BEAR / BASE / BULL SCENARIO ANALYSIS — {ticker}", 4, AMBER_HDR)
    subtitle(ws8, "Each scenario adjusts FCF growth, WACC, and terminal growth to stress-test the valuation", 4)

    bear = scenarios.get("Bear 🐻", {}); base_sc = scenarios.get("Base 📊", {}); bull = scenarios.get("Bull 🐂", {})
    for j, (hdr, clr) in enumerate(zip(["Metric", "Bear 🐻", "Base 📊", "Bull 🐂"], [BLUE_HDR, RED_HDR, AMBER_HDR, GREEN_HDR])):
        wc(ws8, 4, j+1, hdr, fill=clr, font=hdr_font(sz=12), align=center(), h=24)

    sc_data = [
        ("ASSUMPTIONS", True, None, None, None),
        ("FCF Growth Rate",     False, f"{bear.get('growth',0):.1%}", f"{base_sc.get('growth',0):.1%}", f"{bull.get('growth',0):.1%}"),
        ("WACC",                False, f"{bear.get('wacc',0):.1%}",   f"{base_sc.get('wacc',0):.1%}",   f"{bull.get('wacc',0):.1%}"),
        ("Terminal Growth",     False, f"{bear.get('term_g',0):.1%}", f"{base_sc.get('term_g',0):.1%}", f"{bull.get('term_g',0):.1%}"),
        ("OUTPUTS", True, None, None, None),
        (f"Intrinsic Value ({sym})", False, f"{sym}{bear.get('iv',0)*fx:,.2f}", f"{sym}{base_sc.get('iv',0)*fx:,.2f}", f"{sym}{bull.get('iv',0)*fx:,.2f}"),
        ("Discount to fair value", False, f"{bear.get('mos_pct',0):+.1f}%", f"{base_sc.get('mos_pct',0):+.1f}%", f"{bull.get('mos_pct',0):+.1f}%"),
        (f"Current Price ({sym})", False, f"{sym}{price_d:,.2f}", f"{sym}{price_d:,.2f}", f"{sym}{price_d:,.2f}"),
    ]
    for i, row_d in enumerate(sc_data):
        r = 5 + i
        label, is_sec, *vals = row_d
        if is_sec:
            ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            c = ws8.cell(row=r, column=1, value=f"  {label}")
            c.fill = hf(AMBER_HDR); c.font = hdr_font(True, 11); c.alignment = left()
            c.border = tb(); ws8.row_dimensions[r].height = 22
        else:
            wc(ws8, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
            for j, (v, clr) in enumerate(zip(vals, [RED_NEG, TEXT_MAIN, GREEN_POS])):
                wc(ws8, r, j+2, v, fill=MID_BG, font=val_font(True, 11, clr, True), align=center())

    # ═══════════════════════════════════════════════════════════
    # SHEET 9 — SENSITIVITY HEATMAP
    # ═══════════════════════════════════════════════════════════
    ws9 = wb.create_sheet("🔥 Sensitivity")
    ws9.sheet_view.showGridLines = False
    ws9.column_dimensions["A"].width = 14
    for j in range(len(sensitivity_df.columns)):
        ws9.column_dimensions[get_column_letter(j+2)].width = 14

    title_row(ws9, f"SENSITIVITY — {ticker}  |  IV/share ({sym})  |  WACC × Terminal Growth", len(sensitivity_df.columns)+2, CYAN_HDR)
    ws9.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(sensitivity_df.columns)+2)
    c = ws9["A2"]; c.value = f"★  Current Market Price: {sym}{price_d:,.2f}  |  Green = Undervalued vs Price  |  Red = Overvalued"
    c.fill = hf(AMBER_HDR); c.font = Font(name="Calibri", size=11, color="FFFFFF"); c.alignment = center()
    ws9.row_dimensions[2].height = 22

    wc(ws9, 4, 1, "WACC \\ g →", fill=BLUE_HDR, font=hdr_font(sz=11), align=center(), h=24)
    for j, col_name in enumerate(sensitivity_df.columns):
        wc(ws9, 4, j+2, col_name, fill=BLUE_HDR, font=hdr_font(sz=11), align=center(), h=24)
    for i, (idx, row) in enumerate(sensitivity_df.iterrows()):
        r = 5 + i
        wc(ws9, r, 1, str(idx), fill=BLUE_HDR, font=hdr_font(sz=11), align=center())
        for j, val in enumerate(row):
            v = val * fx if pd.notna(val) else 0
            is_green = v > price_d
            wc(ws9, r, j+2, round(v, 2), fill="0D4A2A" if is_green else "4A0D0D",
               font=val_font(True, 11, GREEN_POS if is_green else RED_NEG, True),
               align=right(), nf=f'"{sym}"#,##0.00')

    # ═══════════════════════════════════════════════════════════
    # SHEET 10 — QUALITY CHECKS (TV WEIGHT + SANITY MULTIPLES)
    # ═══════════════════════════════════════════════════════════
    ws10 = wb.create_sheet("✅ Quality Checks")
    ws10.sheet_view.showGridLines = False
    ws10.column_dimensions["A"].width = 38
    ws10.column_dimensions["B"].width = 22
    ws10.column_dimensions["C"].width = 22
    ws10.column_dimensions["D"].width = 36

    title_row(ws10, f"DCF QUALITY CHECKS — {ticker}", 4, RED_HDR)
    subtitle(ws10, "Sanity tests every analyst should run before trusting a DCF output", 4)

    # 1. Terminal Value Weight
    sec(ws10, 4, "CHECK 1 — TERMINAL VALUE WEIGHT", 4, BLUE_HDR)
    tv_safe = tv_pct <= 0.70
    tv_checks = [
        ("PV of Forecast FCFs",             f"{sym}{dcf_res.get('sum_pv_fcfs',0)*fx/1e9:,.2f}B",  "—"),
        ("PV of Terminal Value",             f"{sym}{pv_tv_bn:,.2f}B",                             "—"),
        ("Enterprise Value",                 f"{sym}{ev_d/1e9:,.2f}B",                            "—"),
        ("Terminal Value as % of EV",        f"{tv_pct:.1%}",
         "✅ < 70% — Healthy" if tv_safe else "⚠️ > 70% — Over-reliant on terminal assumptions"),
        ("Threshold (Professional Standard)", "70%",                                               "Flag if exceeded"),
    ]
    wc(ws10, 5, 1, "Metric",      fill=BLUE_HDR, font=hdr_font(sz=11), align=left())
    wc(ws10, 5, 2, "Value",       fill=BLUE_HDR, font=hdr_font(sz=11), align=center())
    wc(ws10, 5, 3, "Status / Note", fill=BLUE_HDR, font=hdr_font(sz=11), align=left())
    for i, (label, value, note) in enumerate(tv_checks):
        r = 6 + i
        flag_bg = (OK_BG if tv_safe else WARN_BG) if "Terminal Value as %" in label else DARK_BG
        wc(ws10, r, 1, label, fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=left())
        wc(ws10, r, 2, value, fill=MID_BG,  font=val_font(True,11,TEXT_MAIN,True), align=right())
        wc(ws10, r, 3, note,  fill=flag_bg, font=Font(name="Calibri",bold=True,size=11,
           color=GREEN_POS if tv_safe else RED_NEG), align=left())

    # 2. Implied Multiples Sanity Check
    latest_ni  = (hist_ni[-1]  if hist_ni  else 0) * fx / 1e9
    latest_rev_check = (hist_rev[-1] if hist_rev else 0) * fx / 1e9
    latest_ebitda = latest_rev_check * (op_margin + 0.035)  # add back D&A est.
    latest_opinc_check = latest_rev_check * op_margin

    equity_val_d = dcf_res.get("equity_value", 0) * fx / 1e9
    mkt_cap_d    = price_d * shares_v / 1e9

    implied_pe        = (mkt_cap_d / latest_ni)     if latest_ni     > 0 else 0
    dcf_implied_pe    = (equity_val_d / latest_ni)  if latest_ni     > 0 else 0
    implied_ev_ebitda = (ev_d/1e9 / latest_ebitda)  if latest_ebitda > 0 else 0
    implied_ev_rev    = (ev_d/1e9 / latest_rev_check) if latest_rev_check > 0 else 0

    def mult_flag(label, val):
        if "P/E" in label:
            if val <= 0:   return "N/A (negative earnings)", TEXT_DIM,  DARK_BG
            if val > 60:   return f"⚠️ {val:.1f}x — Very expensive; check model",  RED_NEG,  WARN_BG
            if val > 35:   return f"⚡ {val:.1f}x — High; growth stock territory",  AMBER_VAL, DARK_BG
            return         f"✅ {val:.1f}x — Reasonable",                           GREEN_POS, OK_BG
        if "EV/EBITDA" in label:
            if val <= 0:   return "N/A",                                TEXT_DIM,   DARK_BG
            if val > 30:   return f"⚠️ {val:.1f}x — Very high",        RED_NEG,    WARN_BG
            if val > 15:   return f"⚡ {val:.1f}x — Premium",           AMBER_VAL,  DARK_BG
            return         f"✅ {val:.1f}x — Reasonable",               GREEN_POS,  OK_BG
        if "EV/Rev" in label:
            if val > 10:   return f"⚠️ {val:.1f}x — Elevated",         RED_NEG,    WARN_BG
            return         f"✅ {val:.1f}x — OK",                       GREEN_POS,  OK_BG
        return f"{val:.1f}x", TEXT_MAIN, DARK_BG

    sec(ws10, 13, "CHECK 2 — DCF SANITY MULTIPLES  (DCF Implied vs Market)", 4, RED_HDR)
    wc(ws10, 14, 1, "Multiple",        fill=RED_HDR, font=hdr_font(sz=11), align=left())
    wc(ws10, 14, 2, "Market (Current)",fill=RED_HDR, font=hdr_font(sz=11), align=center())
    wc(ws10, 14, 3, "DCF Implied",     fill=RED_HDR, font=hdr_font(sz=11), align=center())
    wc(ws10, 14, 4, "Sanity Flag",     fill=RED_HDR, font=hdr_font(sz=11), align=left())

    mult_rows = [
        ("Price / Earnings (P/E)",      mkt_cap_d / latest_ni   if latest_ni > 0 else 0,  dcf_implied_pe),
        ("EV / EBITDA",                 ev_d/1e9 / latest_ebitda if latest_ebitda > 0 else 0, implied_ev_ebitda),
        ("EV / Revenue",                ev_d/1e9 / latest_rev_check if latest_rev_check > 0 else 0, implied_ev_rev),
        ("Price / Book (proxy)",        mkt_cap_d / max(equity_val_d * 0.6, 0.01), equity_val_d / max(equity_val_d * 0.6, 0.01)),
    ]
    for i, (mult_name, mkt_val, dcf_val) in enumerate(mult_rows):
        r = 15 + i
        flag_text, flag_color, flag_bg = mult_flag(mult_name, dcf_val)
        wc(ws10, r, 1, mult_name,                fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=left())
        wc(ws10, r, 2, f"{mkt_val:.1f}x" if mkt_val else "N/A", fill=MID_BG, font=val_font(True,11,TEXT_NUM,True), align=right())
        wc(ws10, r, 3, f"{dcf_val:.1f}x" if dcf_val else "N/A", fill=MID_BG, font=val_font(True,12,TEXT_MAIN,True), align=right())
        wc(ws10, r, 4, flag_text, fill=flag_bg, font=Font(name="Calibri",bold=True,size=11,color=flag_color), align=left())

    # 3. MoS Reliability check
    sec(ws10, 21, "CHECK 3 — MODEL RELIABILITY SCORECARD", 4, CYAN_HDR)
    reliable = forecast_result.get("reliable", True)
    suspicious = dcf_res.get("suspicious", False)
    quality_checks = [
        ("DCF Reliable Flag",         "✅ Reliable" if reliable else "⚠️ Unreliable",    reliable),
        ("IV Hard Cap Triggered",     "⚠️ Yes — IV was capped at 5× price" if suspicious else "✅ No — IV within bounds", not suspicious),
        ("Terminal Value < 70% EV",   "✅ Pass" if tv_pct <= 0.70 else f"⚠️ Fail — {tv_pct:.0%}", tv_pct <= 0.70),
        ("ROIC > WACC",               "✅ Value Creation" if roic_estimate > wacc else "⚠️ Value Destruction", roic_estimate > wacc),
        ("FCF Positive",              "✅ Yes" if enriched.get("latest_fcf",0) > 0 else "⚠️ Negative FCF", enriched.get("latest_fcf",0) > 0),
        ("Op Margin > 8%",            "✅ Pass" if enriched.get("op_margin",0) > 0.08 else "⚠️ Below threshold", enriched.get("op_margin",0) > 0.08),
        ("DCF P/E Sanity",            "✅ Reasonable" if 0 < dcf_implied_pe < 60 else ("⚠️ Suspicious" if dcf_implied_pe >= 60 else "N/A"), 0 < dcf_implied_pe < 60),
    ]
    wc(ws10, 22, 1, "Check",   fill=CYAN_HDR, font=hdr_font(sz=11), align=left())
    wc(ws10, 22, 2, "Result",  fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    for i, (check, result, passed) in enumerate(quality_checks):
        r = 23 + i
        wc(ws10, r, 1, check,  fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=left())
        wc(ws10, r, 2, result, fill=OK_BG if passed else WARN_BG,
           font=Font(name="Calibri",bold=True,size=11,color=GREEN_POS if passed else RED_NEG), align=left())

    # ═══════════════════════════════════════════════════════════
    # SHEET 11 — HISTORICAL CHARTS
    # ═══════════════════════════════════════════════════════════
    ws11 = wb.create_sheet("📈 Historical Charts")
    ws11.sheet_view.showGridLines = False

    title_row(ws11, f"HISTORICAL TRENDS — {ticker}", 10, PURP_HDR)
    subtitle(ws11, "Revenue Growth  |  Operating Margin  |  Free Cash Flow Trend", 10)

    # Write data tables for charts
    # Table 1: Revenue
    sec(ws11, 4, "REVENUE TREND", 10, BLUE_HDR)
    wc(ws11, 5, 1, "Year",              fill=BLUE_HDR, font=hdr_font(sz=11), align=center())
    wc(ws11, 5, 2, f"Revenue ({to_code}B)", fill=BLUE_HDR, font=hdr_font(sz=11), align=center())
    wc(ws11, 5, 3, "YoY Growth %",     fill=BLUE_HDR, font=hdr_font(sz=11), align=center())

    rev_list = []
    if not income_df.empty and "revenue" in income_df.columns:
        rev_list = [(str(int(y)), v*fx/1e9) for y, v in zip(income_df["year"], income_df["revenue"]) if pd.notna(v)]

    for i, (yr, v) in enumerate(rev_list):
        r = 6 + i
        prev_v = rev_list[i-1][1] if i > 0 else v
        growth = (v - prev_v) / prev_v * 100 if prev_v > 0 and i > 0 else 0
        wc(ws11, r, 1, yr,         fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=center())
        wc(ws11, r, 2, round(v,2), fill=MID_BG,  font=val_font(True,11,TEXT_MAIN,True), align=right(), nf="#,##0.00")
        wc(ws11, r, 3, f"{growth:.1f}%", fill=MID_BG,
           font=val_font(True,11,GREEN_POS if growth > 0 else RED_NEG,True), align=right())

    # Table 2: Operating Margin
    sec(ws11, 6 + len(rev_list) + 1, "OPERATING MARGIN TREND", 10, GREEN_HDR)
    margin_start = 6 + len(rev_list) + 2
    wc(ws11, margin_start, 1, "Year",             fill=GREEN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws11, margin_start, 2, "Operating Margin", fill=GREEN_HDR, font=hdr_font(sz=11), align=center())

    margin_list = []
    if not income_df.empty and "op_margin" in income_df.columns:
        margin_list = [(str(int(y)), v*100) for y, v in zip(income_df["year"], income_df["op_margin"]) if pd.notna(v)]
    elif not income_df.empty and "operating_income" in income_df.columns and "revenue" in income_df.columns:
        margin_list = [(str(int(y)), (oi/rev*100 if rev else 0))
                       for y, oi, rev in zip(income_df["year"], income_df["operating_income"], income_df["revenue"])
                       if pd.notna(oi) and pd.notna(rev)]

    for i, (yr, v) in enumerate(margin_list):
        r = margin_start + 1 + i
        wc(ws11, r, 1, yr,             fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=center())
        wc(ws11, r, 2, f"{v:.1f}%",   fill=MID_BG,  font=val_font(True,11,GREEN_POS if v > 0 else RED_NEG,True), align=right())

    # Table 3: FCF Trend
    fcf_sec_row = margin_start + len(margin_list) + 2
    sec(ws11, fcf_sec_row, "FREE CASH FLOW TREND", 10, CYAN_HDR)
    fcf_hdr_row = fcf_sec_row + 1
    wc(ws11, fcf_hdr_row, 1, "Year",              fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws11, fcf_hdr_row, 2, f"FCF ({to_code}B)", fill=CYAN_HDR, font=hdr_font(sz=11), align=center())

    fcf_list = []
    if not cf_df.empty and "fcf" in cf_df.columns:
        fcf_list = [(str(int(y)), v*fx/1e9) for y, v in zip(cf_df["year"], cf_df["fcf"]) if pd.notna(v) and abs(v) > 1e6]

    for i, (yr, v) in enumerate(fcf_list):
        r = fcf_hdr_row + 1 + i
        wc(ws11, r, 1, yr,         fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=center())
        wc(ws11, r, 2, round(v,2), fill=MID_BG,  font=val_font(True,11,GREEN_POS if v > 0 else RED_NEG,True), align=right(), nf="#,##0.00")

    # Add bar charts if data available
    chart_col = 5  # Charts go in column E

    if len(rev_list) >= 2:
        chart1 = BarChart()
        chart1.type = "col"; chart1.grouping = "clustered"
        chart1.title = f"Revenue Trend — {ticker}"
        chart1.y_axis.title = f"{to_code}B"; chart1.x_axis.title = "Year"
        chart1.style = 10; chart1.width = 18; chart1.height = 12
        data_ref  = Reference(ws11, min_col=2, min_row=5, max_row=5+len(rev_list))
        cat_ref   = Reference(ws11, min_col=1, min_row=6, max_row=5+len(rev_list))
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cat_ref)
        chart1.series[0].graphicalProperties.solidFill = "1D4ED8"
        ws11.add_chart(chart1, f"E4")

    if len(margin_list) >= 2:
        chart2 = LineChart()
        chart2.title = f"Operating Margin — {ticker}"
        chart2.y_axis.title = "%"; chart2.x_axis.title = "Year"
        chart2.style = 10; chart2.width = 18; chart2.height = 12
        data_ref2 = Reference(ws11, min_col=2, min_row=margin_start, max_row=margin_start+len(margin_list))
        cat_ref2  = Reference(ws11, min_col=1, min_row=margin_start+1, max_row=margin_start+len(margin_list))
        chart2.add_data(data_ref2, titles_from_data=True)
        chart2.set_categories(cat_ref2)
        chart2.series[0].graphicalProperties.line.solidFill = "10B981"
        chart2.series[0].graphicalProperties.line.width = 25000
        ws11.add_chart(chart2, f"E24")

    if len(fcf_list) >= 2:
        chart3 = BarChart()
        chart3.type = "col"; chart3.grouping = "clustered"
        chart3.title = f"Free Cash Flow — {ticker}"
        chart3.y_axis.title = f"{to_code}B"; chart3.x_axis.title = "Year"
        chart3.style = 10; chart3.width = 18; chart3.height = 12
        data_ref3 = Reference(ws11, min_col=2, min_row=fcf_hdr_row, max_row=fcf_hdr_row+len(fcf_list))
        cat_ref3  = Reference(ws11, min_col=1, min_row=fcf_hdr_row+1, max_row=fcf_hdr_row+len(fcf_list))
        chart3.add_data(data_ref3, titles_from_data=True)
        chart3.set_categories(cat_ref3)
        chart3.series[0].graphicalProperties.solidFill = "06B6D4"
        ws11.add_chart(chart3, f"E44")

    # ═══════════════════════════════════════════════════════════
    # SHEET 12 — KEY ASSUMPTIONS
    # ═══════════════════════════════════════════════════════════
    ws12 = wb.create_sheet("⚙️ Assumptions")
    ws12.sheet_view.showGridLines = False
    ws12.column_dimensions["A"].width = 36
    ws12.column_dimensions["B"].width = 28
    ws12.column_dimensions["C"].width = 50

    title_row(ws12, f"KEY MODEL ASSUMPTIONS — {ticker}", 3, BLUE_HDR)
    wc(ws12, 4, 1, "Parameter",   fill=BLUE_HDR, font=hdr_font(sz=12), align=left())
    wc(ws12, 4, 2, "Value Used",  fill=BLUE_HDR, font=hdr_font(sz=12), align=center())
    wc(ws12, 4, 3, "Explanation", fill=BLUE_HDR, font=hdr_font(sz=12), align=left())

    assump_rows = [
        (True,  "DISCOUNT RATE",         None,   None),
        (False, "WACC",                  f"{wacc:.2%}", "Weighted Average Cost of Capital — auto CAPM or manual"),
        (False, "Cost of Equity",        "Auto CAPM", "Re = Rf + β×(Rm−Rf)"),
        (False, "Risk-Free Rate",        "~7.0% (IN) / ~4.5% (US)", "10-yr government bond yield"),
        (True,  "GROWTH ASSUMPTIONS",    None,   None),
        (False, "Base FCF Growth",       f"{forecast_result.get('base_growth',0):.2%}", "Historical CAGR with exponential fade to terminal g"),
        (False, "Long-run growth rate",  f"{terminal_g:.2%}", "Long-run GDP growth assumption"),
        (False, "Forecast Horizon",      f"{forecast_yrs} years", "Explicit modelling period"),
        (False, "FCF Base Method",       forecast_result.get("fcf_base_method","Auto"), "How the starting FCF was selected"),
        (True,  "REINVESTMENT MODEL",    None,   None),
        (False, "D&A (% of Revenue)",    "3.5%", "Estimated depreciation & amortisation"),
        (False, "CapEx (% of Revenue)",  "4.5%", "Estimated capital expenditure"),
        (False, "ΔWC (% of ΔRevenue)",  "8.0%", "Working capital investment on incremental revenue"),
        (False, "Tax Rate",              f"{tax_rate:.0%}", "Effective corporate tax rate assumption"),
        (True,  "VALUATION GUARDRAILS",  None,   None),
        (False, "IV Hard Cap",           "5× current price", "Prevents outlier results from bad data"),
        (False, "TV Warning Threshold",  "70% of EV", "Flags terminal-value-heavy models"),
        (False, "MoS Formula",          "(IV − Price) / Price", "Standard margin of safety relative to price"),
        (True,  "QUALITY FILTERS",       None,   None),
        (False, "Min Operating Margin",  "8%",   "Below this → DCF flagged unreliable"),
        (False, "Max FCF Margin",        "30%",  "Suspiciously high FCF → flagged for review"),
        (False, "Microcap Filter",       "₹2000 Cr / $200M", "Small companies excluded from screener"),
        (True,  "DISCLAIMER",            None,   None),
        (False, "Purpose",               "Educational / Research Only", "Not financial advice"),
        (False, "Data Source",           "Yahoo Finance (yfinance)", "Prices and financials via yfinance API"),
    ]
    for i, (is_sec, label, value, expl) in enumerate(assump_rows):
        r = 5 + i
        if is_sec:
            ws12.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            c = ws12.cell(row=r, column=1, value=f"  {label}")
            c.fill = hf(BLUE_HDR); c.font = hdr_font(True, 12); c.alignment = left()
            c.border = tb(); ws12.row_dimensions[r].height = 24
        else:
            wc(ws12, r, 1, label,        fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=left())
            wc(ws12, r, 2, value or "",  fill=MID_BG,  font=val_font(True,11,TEXT_MAIN,True), align=center())
            c3 = ws12.cell(row=r, column=3, value=expl or "")
            c3.fill = hf(DARK_BG); c3.font = Font(name="Calibri",size=10,color=TEXT_DIM,italic=True)
            c3.alignment = wrap_left(); c3.border = tb()

    # ── Freeze panes & tab ordering ────────────────────────────
    for ws in [ws1,ws2,ws3,ws4,ws5,ws6,ws7,ws8,ws9,ws10,ws11,ws12]:
        ws.freeze_panes = "B5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



# ══════════════════════════════════════════════════════════════
# SIDEBAR
