# dashboard/app.py  v5
# ═══════════════════════════════════════════════════════════════
# YieldIQ — Dashboard
# New in v5:
#   1. Fixed sensitivity heatmap text sizing
#   2. Fixed historical FCF chart (no grey bars)
#   3. Bear/Base/Bull scenario analysis
#   4. DCF Report PDF download per stock
#   5. Excel export with buy/sell/SL formatted
#   6. Larger investment plan card
#   7. Scenario comparison chart
# ═══════════════════════════════════════════════════════════════

from __future__ import annotations
import sys, os, io, requests
from datetime import datetime
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from features import (
    render_live_price_header,
    render_analyst_consensus,
    render_earnings_calendar,
    render_comparison_watchlist,
)
from portfolio import (
    render_portfolio_tab, init_db, is_in_portfolio,
    init_watchlist_db, add_to_watchlist, is_in_watchlist,
    get_watchlist, remove_from_watchlist,
    init_institutional_db, save_institutional_ownership, get_institutional_history,
    init_sheets_db,
)
import importlib.util as _ob_ilu, pathlib as _ob_pl
_ob_path = _ob_pl.Path(__file__).parent / "onboarding.py"
_ob_spec = _ob_ilu.spec_from_file_location("onboarding", _ob_path)
_ob_mod  = _ob_ilu.module_from_spec(_ob_spec)
_ob_spec.loader.exec_module(_ob_mod)
init_onboarding_db   = _ob_mod.init_onboarding_db
maybe_show_wizard    = _ob_mod.maybe_show_wizard
render_resume_button = _ob_mod.render_resume_button
ob_tooltip           = _ob_mod.tooltip
ob_show_tooltips     = _ob_mod.show_tooltips

# Morning Brief landing page
_mb_path = _ob_pl.Path(__file__).parent / "morning_brief.py"
_mb_spec = _ob_ilu.spec_from_file_location("morning_brief", _mb_path)
_mb_mod  = _ob_ilu.module_from_spec(_mb_spec)
_mb_spec.loader.exec_module(_mb_mod)
render_morning_brief      = _mb_mod.render_morning_brief
push_analysis_to_history  = _mb_mod.push_analysis_to_history

# Admin analytics
_aa_path = _ob_pl.Path(__file__).parent / "admin_analytics.py"
_aa_spec = _ob_ilu.spec_from_file_location("admin_analytics", _aa_path)
_aa_mod  = _ob_ilu.module_from_spec(_aa_spec)
_aa_spec.loader.exec_module(_aa_mod)
init_analytics_db        = _aa_mod.init_analytics_db
track_analysis           = _aa_mod.track_analysis
track_event              = _aa_mod.track_event
render_admin_dashboard   = _aa_mod.render_admin_dashboard

from tabs.backtest_tab import render as render_backtest_tab, init_backtest_db
from tabs import earnings_quality_tab, reverse_dcf_tab, moat_tab, compare_tab
from tabs.earnings_tab import render_earnings_tab
from sector_dashboard import render_sector_dashboard, init_sector_db
from sector_heatmap import render_sector_heatmap
import alerts as _alerts_mod
init_sector_db()
init_backtest_db()
init_db()                      # ensure portfolio DB exists on startup
init_watchlist_db()            # ensure watchlist table exists on startup
init_institutional_db()        # ensure institutional_ownership_history table exists
init_sheets_db()               # ensure user_sheets_settings table exists
_alerts_mod.init_alerts_db()   # ensure price_alerts table exists on startup
init_onboarding_db()           # ensure user_onboarding table exists on startup
init_analytics_db()            # ensure analytics.db tables exist on startup

from data.collector import StockDataCollector
from data.processor import compute_metrics
from models.forecaster import FCFForecaster, compute_wacc, compute_confidence_score
from screener.dcf_engine import (
    DCFEngine, margin_of_safety, assign_signal,
    sensitivity_analysis, monte_carlo_valuation
)
from screener.scenarios import run_scenarios
from screener.valuation_model import generate_valuation_summary as generate_investment_plan
from screener.reverse_dcf import run_reverse_dcf
from screener.ev_ebitda import run_ev_ebitda_analysis
from screener.piotroski import compute_piotroski_fscore as _piotroski_raw
from screener.fcf_yield import compute_fcf_yield_analysis as _fcf_yield_raw
from screener.historical_iv import compute_historical_iv as _hist_iv_raw

from datetime import date as _date


# ── YieldIQ Composite Score (0–100) ──────────────────────────────────────────
# Pure function — no Streamlit calls, safe to define at module level.
# Called in the ⚡ Summary pill section after analysis is run.
def compute_yieldiq_score(
    mos_pct: float,
    piotroski: int,
    moat_grade: str,
    rev_growth: float,
    analyst_upside: float,
) -> dict:
    """
    Compute YieldIQ Composite Score (0-100).
    Model output for informational purposes only. Not investment advice.
    """
    # Valuation (40 pts)
    if mos_pct >= 40:    val_score = 40
    elif mos_pct >= 25:  val_score = 32
    elif mos_pct >= 10:  val_score = 22
    elif mos_pct >= 0:   val_score = 14
    elif mos_pct >= -15: val_score = 7
    else:                val_score = 0

    # Business Quality (30 pts) — Piotroski + Moat
    pio_score = min(piotroski / 9 * 20, 20)
    _moat_map = {
        "A": 10, "B": 7, "C": 4, "D": 1,
        "Wide": 10, "Narrow": 7,
        "None": 0, "none": 0, "N/A": 0, "": 0,
    }
    moat_pts   = _moat_map.get(str(moat_grade).strip(), 0)
    qual_score = pio_score + moat_pts

    # Growth (20 pts)
    if rev_growth >= 20:    grw_score = 20
    elif rev_growth >= 10:  grw_score = 15
    elif rev_growth >= 5:   grw_score = 10
    elif rev_growth >= 0:   grw_score = 5
    else:                   grw_score = 0

    # Sentiment (10 pts)
    if analyst_upside >= 20:    sent_score = 10
    elif analyst_upside >= 10:  sent_score = 7
    elif analyst_upside >= 0:   sent_score = 4
    else:                       sent_score = 1

    total = max(0, min(100, int(val_score + qual_score + grw_score + sent_score)))

    if total >= 85:   grade = "A+"
    elif total >= 75: grade = "A"
    elif total >= 65: grade = "B+"
    elif total >= 55: grade = "B"
    elif total >= 45: grade = "C+"
    elif total >= 35: grade = "C"
    else:             grade = "D"

    return {
        "score": total,
        "grade": grade,
        "components": {
            "Valuation (40pts)":        int(val_score),
            "Business Quality (30pts)": int(qual_score),
            "Growth (20pts)":           int(grw_score),
            "Sentiment (10pts)":        int(sent_score),
        },
    }


@st.cache_data(ttl=86400, show_spinner=False)
def compute_piotroski_fscore(enriched_json: str):
    """Cached Piotroski — invalidates daily. Pass enriched as JSON string."""
    import json
    return _piotroski_raw(json.loads(enriched_json))

@st.cache_data(ttl=86400, show_spinner=False)
def compute_fcf_yield_analysis(enriched: dict, current_price: float,
                                fx: float = 1.0):
    return _fcf_yield_raw(enriched=enriched, current_price=current_price, fx=fx)

@st.cache_data(ttl=86400, show_spinner=False)
def compute_historical_iv(enriched: dict, current_price: float,
                           current_iv: float, wacc: float, terminal_g: float,
                           forecast_yrs: int = 10, fx: float = 1.0):
    return _hist_iv_raw(
        enriched=enriched,
        current_price=current_price,
        current_iv=current_iv,
        wacc=wacc,
        terminal_g=terminal_g,
        forecast_yrs=forecast_yrs,
        fx=fx,
    )
from screener.earnings_quality import compute_earnings_quality
from screener.sector_relative import compute_sector_relative
from screener.ddm import compute_ddm
from screener.relative_valuation import (
    check_ticker_dcf_eligibility,
    relative_valuation_only,
)
# Note: ev_ebitda, moat_engine loaded lazily inside analysis block
from utils.config import FORECAST_YEARS, RESULTS_PATH, LAUNCH_REGION
from ui.helpers import (
    render_skeleton_card, render_empty_state,
    add_tooltip, inject_tooltip_css, FINANCIAL_TOOLTIPS,
)
from ui.styles import inject_theme_css
# tier_gate lives in the same folder as app.py — use a path-safe import
import importlib.util as _ilu, pathlib as _pl
_tg_path = _pl.Path(__file__).parent / "tier_gate.py"
_tg_spec = _ilu.spec_from_file_location("tier_gate", _tg_path)
_tg_mod  = _ilu.module_from_spec(_tg_spec)
_tg_spec.loader.exec_module(_tg_mod)

init_tier            = _tg_mod.init_tier
tier                 = _tg_mod.tier
is_free              = _tg_mod.is_free
is_premium           = _tg_mod.is_premium
is_pro               = _tg_mod.is_pro
can                  = _tg_mod.can
limit                = _tg_mod.limit
can_analyse          = _tg_mod.can_analyse
record_analysis      = _tg_mod.record_analysis
can_download_report  = _tg_mod.can_download_report
record_report        = _tg_mod.record_report
can_download_pdf     = _tg_mod.can_download_pdf
record_pdf_report    = _tg_mod.record_pdf_report
can_run_screener     = _tg_mod.can_run_screener
record_screener      = _tg_mod.record_screener
check_ticker_allowed = _tg_mod.check_ticker_allowed
upgrade_prompt            = _tg_mod.upgrade_prompt
blur_and_lock             = _tg_mod.blur_and_lock
tier_badge_html           = _tg_mod.tier_badge_html
usage_bar_html            = _tg_mod.usage_bar_html
show_analysis_limit_modal = _tg_mod.show_analysis_limit_modal
show_india_gate_message   = _tg_mod.show_india_gate_message
show_report_upsell        = _tg_mod.show_report_upsell
sidebar_upgrade_button    = _tg_mod.sidebar_upgrade_button

APP_VERSION = "v6"

# ══════════════════════════════════════════════════════════════
# COMPLIANCE DISCLAIMER  (shown once per session; persisted in localStorage)
# ══════════════════════════════════════════════════════════════

_DISCLAIMER_TEXT = """\
YieldIQ is a quantitative research tool for **informational and educational purposes only**.
It is **NOT** investment advice and does **NOT** constitute a recommendation to buy, sell,
or hold any security.

**YieldIQ LLC is not registered as an Investment Adviser** under the Investment Advisers
Act of 1940 or any applicable state securities law.

All model outputs represent mathematical estimates based on publicly available financial
data. They do not account for your personal financial situation, risk tolerance, tax
circumstances, or investment objectives.

Always consult a **qualified, licensed financial professional** before making any
investment decisions. Past model accuracy is not indicative of future results.\
"""

_DISCLAIMER_LS_KEY  = "yiq_disclaimer_ts"
_DISCLAIMER_VALIDITY_DAYS = 365


def _disclaimer_write_localstorage() -> None:
    """Inject a zero-height script that stamps acceptance into localStorage."""
    import streamlit.components.v1 as _stc
    _stc.html(
        f"""<script>
(function(){{
  try {{ localStorage.setItem('{_DISCLAIMER_LS_KEY}', Date.now().toString()); }}
  catch(e) {{}}
}})();
</script>""",
        height=0,
    )


def _disclaimer_check_localstorage() -> None:
    """
    Inject a zero-height script that reads localStorage.
    If a valid acceptance stamp is found it appends ?da=1 to the parent URL,
    which Streamlit reads on the next rerun via st.query_params.
    """
    import streamlit.components.v1 as _stc
    _stc.html(
        f"""<script>
(function(){{
  try {{
    var ts = localStorage.getItem('{_DISCLAIMER_LS_KEY}');
    if (ts) {{
      var age = Date.now() - parseInt(ts, 10);
      if (age < {_DISCLAIMER_VALIDITY_DAYS} * 86400000) {{
        var u = new URL(window.parent.location.href);
        if (u.searchParams.get('da') !== '1') {{
          u.searchParams.set('da', '1');
          window.parent.location.replace(u.toString());
        }}
      }}
    }}
  }} catch(e) {{}}
}})();
</script>""",
        height=0,
    )


def show_disclaimer_if_needed() -> None:
    """
    Gate the entire app behind a one-time compliance disclaimer.

    Flow:
      1. If ?da=1 is in the URL (set by localStorage check), mark session accepted.
      2. If session already accepted → return immediately.
      3. Inject localStorage reader; if valid stamp exists it reloads with ?da=1.
      4. Show the disclaimer container with checkbox + button.
      5. Call st.stop() so nothing else renders until accepted.
    """
    # ── Fast-path: localStorage redirect already happened ────────
    if not st.session_state.get("_force_disclaimer"):
        try:
            if st.query_params.get("da") == "1":
                st.session_state["disclaimer_shown"] = True
        except Exception:
            pass

    if st.session_state.get("disclaimer_shown") and not st.session_state.get("_force_disclaimer"):
        return

    # ── Inject localStorage reader (triggers reload if valid) ───
    _disclaimer_check_localstorage()

    # ── Render disclaimer ────────────────────────────────────────
    st.html("""<style>
/* Centre the disclaimer card and hide everything else */
[data-testid="stAppViewContainer"] > .main > div:first-child {
    display: flex; flex-direction: column;
    align-items: center; justify-content: flex-start;
    padding-top: 48px !important;
}
</style>""")

    with st.container(border=True):
        # Header
        st.markdown(
            '<div style="display:flex;align-items:center;gap:10px;margin-bottom:4px;">'
            '<span style="font-size:28px;">⚖️</span>'
            '<span style="font-size:20px;font-weight:700;color:#0F172A;">Important Disclosure</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            '<hr style="margin:8px 0 16px;border:none;border-top:1.5px solid #E2E8F0;">',
            unsafe_allow_html=True,
        )

        # Body
        st.markdown(_DISCLAIMER_TEXT)

        st.markdown(
            '<div style="margin:16px 0 8px;padding:12px 16px;'
            'background:#FFF7ED;border:1px solid #FED7AA;border-radius:8px;'
            'font-size:12px;color:#9A3412;">'
            '⚠️ This disclosure must be acknowledged before using YieldIQ.'
            '</div>',
            unsafe_allow_html=True,
        )

        # Checkbox
        agreed = st.checkbox(
            "I understand this tool provides quantitative analysis only, not investment advice",
            key="_disclaimer_checkbox",
        )

        # Button (disabled until checkbox ticked)
        col_btn, col_gap = st.columns([2, 3])
        with col_btn:
            clicked = st.button(
                "Continue to YieldIQ →",
                disabled=not agreed,
                type="primary",
                width='stretch',
                key="_disclaimer_continue_btn",
            )

        if clicked and agreed:
            st.session_state["disclaimer_shown"] = True
            st.session_state["disclaimer_ts"]    = datetime.utcnow().isoformat()
            st.session_state.pop("_force_disclaimer", None)
            _disclaimer_write_localstorage()
            try:
                st.query_params["da"] = "1"
            except Exception:
                pass
            st.rerun()

    st.stop()  # nothing renders below until disclaimer accepted


def render_view_disclaimer_link() -> None:
    """Sidebar footer link that re-shows the disclaimer on demand."""
    st.sidebar.markdown(
        '<div style="text-align:center;margin-top:6px;">'
        '<span style="font-size:10px;color:#475569;">Legal </span>'
        '</div>',
        unsafe_allow_html=True,
    )
    if st.sidebar.button(
        "📋 View Disclaimer",
        key="_view_disclaimer_btn",
        width='stretch',
    ):
        st.session_state["_force_disclaimer"] = True
        st.session_state["disclaimer_shown"]  = False
        st.rerun()


# ══════════════════════════════════════════════════════════════
# PAGE CONFIG
# ══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="YieldIQ",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Initialise tier gating (reads ?token= from URL)
init_tier()

# Mandatory compliance disclaimer — shown once per session / year
show_disclaimer_if_needed()

# Show first-run onboarding wizard for new users
maybe_show_wizard()

# ── Sidebar nav button styling (must use st.markdown, not st.html,
#    so it injects into the parent DOM and can style Streamlit's own elements)
st.markdown("""<style>
/* ── Sidebar nav buttons: base state ── */
section[data-testid="stSidebar"] .stButton > button {
    width: 100% !important;
    background: transparent !important;
    border: none !important;
    border-left: 3px solid transparent !important;
    border-radius: 0px !important;
    color: rgba(255,255,255,0.7) !important;
    text-align: left !important;
    padding: 10px 16px !important;
    font-size: 13px !important;
    font-family: 'Inter', sans-serif !important;
    transition: all 0.15s ease !important;
    box-shadow: none !important;
}
section[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(255,255,255,0.08) !important;
    color: white !important;
    border-left-color: rgba(29,78,216,0.5) !important;
}
/* Active nav item — primary type */
section[data-testid="stSidebar"] .stButton > button[kind="primaryFormSubmit"],
section[data-testid="stSidebar"] .stButton > button[data-testid*="primary"],
section[data-testid="stSidebar"] .stButton > button[data-testid="baseButton-primary"],
section[data-testid="stSidebar"] .stButton > button[kind="primary"] {
    background: rgba(29,78,216,0.15) !important;
    border-left: 3px solid #1D4ED8 !important;
    color: white !important;
    font-weight: 600 !important;
}
section[data-testid="stSidebar"] .stButton > button[kind="primaryFormSubmit"]:hover,
section[data-testid="stSidebar"] .stButton > button[data-testid*="primary"]:hover,
section[data-testid="stSidebar"] .stButton > button[data-testid="baseButton-primary"]:hover,
section[data-testid="stSidebar"] .stButton > button[kind="primary"]:hover {
    background: rgba(29,78,216,0.25) !important;
    color: #93C5FD !important;
}
/* Sidebar HR divider */
section[data-testid="stSidebar"] hr {
    border-color: rgba(255,255,255,0.1) !important;
    margin: 8px 0 !important;
}
</style>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════════════
st.html("""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@300;400;500;600;700&family=Barlow+Condensed:wght@500;600;700;800&display=swap" rel="stylesheet">
""")
inject_tooltip_css()

# ── Theme: inject dark/light CSS overrides (runs on every rerun) ──
inject_theme_css(st.session_state.get("theme", "light"))

st.markdown("""
<style>
/* ═══════════════════════════════════════════════════════════════
   YIELDIQ DASHBOARD — PROFESSIONAL LIGHT THEME
   Fonts: Inter (UI) + IBM Plex Mono (numbers)
   Colors: Deep navy sidebar + clean white cards + blue accents
   ═══════════════════════════════════════════════════════════════ */

@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@300;400;500;600;700&family=Barlow+Condensed:wght@500;600;700;800&display=swap');

/* ── DESIGN TOKENS ── */
:root {
  /* Core palette — matches your screenshot */
  --bg-page:      #EEF2F8;
  --bg-card:      #FFFFFF;
  --bg-card2:     #F7F9FC;
  --bg-sidebar:   #1A2540;
  --bg-sidebar2:  #0F1929;
  --bg-header:    linear-gradient(135deg, #1D3461 0%, #1E4D8C 100%);

  /* Accent blues */
  --blue:         #1D4ED8;
  --blue-mid:     #2563EB;
  --blue-lt:      #EFF6FF;
  --blue-glow:    rgba(29,78,216,0.12);
  --blue-border:  rgba(29,78,216,0.20);

  /* Signal colors */
  --green:        #059669;
  --green-lt:     #ECFDF5;
  --green-border: rgba(5,150,105,0.20);
  --red:          #DC2626;
  --red-lt:       #FEF2F2;
  --red-border:   rgba(220,38,38,0.20);
  --amber:        #D97706;
  --amber-lt:     #FFFBEB;
  --amber-border: rgba(217,119,6,0.20);

  /* Text */
  --text:         #0F172A;
  --text-sec:     #475569;
  --text-muted:   #94A3B8;
  --text-sidebar: #94A3B8;
  --text-sidebar2:#CBD5E1;

  /* Borders & shadows */
  --rule:         #E2E8F0;
  --rule2:        #CBD5E1;
  --shadow-sm:    0 1px 3px rgba(15,23,42,0.06), 0 1px 2px rgba(15,23,42,0.04);
  --shadow-md:    0 4px 16px rgba(15,23,42,0.08), 0 2px 4px rgba(15,23,42,0.04);
  --shadow-lg:    0 10px 40px rgba(15,23,42,0.10), 0 4px 8px rgba(15,23,42,0.05);
  --shadow-blue:  0 4px 20px rgba(29,78,216,0.15);

  /* Typography — Bloomberg-accurate pairing
     UI:      Inter (matches Neue Haas Grotesk's Swiss grotesque tradition)
     Data:    IBM Plex Mono (designed for financial data display, tabular figures)
     Display: Barlow Condensed (heavy condensed grotesque = Druk feel) */
  --font-ui:      'Inter', system-ui, -apple-system, sans-serif;
  --font-mono:    'IBM Plex Mono', 'Courier New', monospace;
  --font-display: 'Barlow Condensed', 'Inter', sans-serif;

  /* Radius */
  --r-sm:   6px;
  --r:      10px;
  --r-lg:   14px;
  --r-xl:   20px;
}

/* ── ANIMATIONS ── */
@keyframes fadeSlideUp   { from { opacity:0; transform:translateY(8px); } to { opacity:1; transform:translateY(0); } }
@keyframes pulseGreen    { 0%,100% { box-shadow:0 0 0 0 rgba(5,150,105,0.3); } 70% { box-shadow:0 0 0 6px rgba(5,150,105,0); } }
@keyframes pulseBlue     { 0%,100% { box-shadow:0 0 0 0 rgba(29,78,216,0.25); } 70% { box-shadow:0 0 0 6px rgba(29,78,216,0); } }
@keyframes shimmer       { 0%,100% { opacity:1; } 50% { opacity:0.7; } }
@keyframes barGrow       { from { transform:scaleX(0); } to { transform:scaleX(1); } }

/* ── BASE ── */
*, *::before, *::after { box-sizing: border-box; }
html, body, [class*="css"] {
  font-family: var(--font-ui) !important;
  background: var(--bg-page) !important;
  color: var(--text) !important;
  font-size: 13px !important;
}
.stApp { background: var(--bg-page) !important; }
.main .block-container {
  padding: 0 2.5rem 3rem 2.5rem !important;
  max-width: 1560px !important;
}

/* ── SIDEBAR ── */
section[data-testid="stSidebar"] {
  background: var(--bg-sidebar) !important;
  border-right: none !important;
  box-shadow: 4px 0 32px rgba(0,0,0,0.25) !important;
}
section[data-testid="stSidebar"] .block-container {
  padding: 1.5rem 1.2rem !important;
}
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] div {
  color: var(--text-sidebar) !important;
  font-size: 12px !important;
  font-family: var(--font-ui) !important;
}
section[data-testid="stSidebar"] .stTextInput input,
section[data-testid="stSidebar"] .stSelectbox > div > div,
section[data-testid="stSidebar"] .stNumberInput input {
  background: rgba(255,255,255,0.07) !important;
  border: 1px solid rgba(255,255,255,0.12) !important;
  color: #E2E8F0 !important;
  border-radius: var(--r-sm) !important;
  font-family: var(--font-mono) !important;
  font-size: 13px !important;
  transition: all 0.2s !important;
}
section[data-testid="stSidebar"] .stTextInput input:focus {
  border-color: rgba(59,130,246,0.6) !important;
  background: rgba(255,255,255,0.10) !important;
  box-shadow: 0 0 0 3px rgba(59,130,246,0.15) !important;
}
section[data-testid="stSidebar"] .stSlider label,
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stTextInput label,
section[data-testid="stSidebar"] .stCheckbox label {
  color: rgba(148,163,184,0.8) !important;
  font-size: 12px !important;
  font-weight: 600 !important;
  text-transform: uppercase !important;
  letter-spacing: 0.10em !important;
  font-family: var(--font-mono) !important;
}
section[data-testid="stSidebar"] .stSlider [data-baseweb="slider"] > div:nth-child(3) {
  background: #3B82F6 !important;
}
section[data-testid="stSidebar"] .stSlider [data-testid="stThumbValue"] {
  background: #3B82F6 !important;
  color: #FFF !important;
  font-family: var(--font-mono) !important;
  font-size: 12px !important;
  font-weight: 600 !important;
}
/* Sidebar nav button styles are handled via st.markdown() above page config
   to ensure they apply to the parent DOM, not the sandboxed st.html() iframe. */
section[data-testid="stSidebar"] hr {
  border: none !important;
  border-top: 1px solid rgba(255,255,255,0.08) !important;
  margin: 1rem 0 !important;
}

/* ── TABS ── */
.stTabs [data-baseweb="tab-list"] {
  background: transparent !important;
  border-bottom: 2px solid var(--rule) !important;
  gap: 0 !important;
  padding: 0 !important;
}
.stTabs [data-baseweb="tab"] {
  background: transparent !important;
  border: none !important;
  border-bottom: 2px solid transparent !important;
  color: var(--text-muted) !important;
  font-family: var(--font-mono) !important;
  font-size: 12px !important;
  font-weight: 600 !important;
  letter-spacing: 0.10em !important;
  text-transform: uppercase !important;
  padding: 14px 22px !important;
  transition: all 0.2s !important;
  border-radius: 0 !important;
  margin-bottom: -2px !important;
}
.stTabs [data-baseweb="tab"]:hover {
  color: var(--text-sec) !important;
  background: var(--blue-lt) !important;
}
.stTabs [aria-selected="true"] {
  color: var(--blue) !important;
  border-bottom-color: var(--blue) !important;
  background: transparent !important;
}
.stTabs [data-baseweb="tab-panel"] {
  padding: 2rem 0 0 !important;
}

/* ── METRICS ── */
[data-testid="stMetric"] {
  background: var(--bg-card) !important;
  border: 1px solid var(--rule) !important;
  border-radius: var(--r) !important;
  padding: 16px 18px !important;
  box-shadow: var(--shadow-sm) !important;
  transition: all 0.2s !important;
}
[data-testid="stMetric"]:hover {
  border-color: var(--rule2) !important;
  box-shadow: var(--shadow-md) !important;
  transform: translateY(-1px) !important;
}
[data-testid="stMetricLabel"] p {
  font-family: var(--font-mono) !important;
  font-size: 12px !important;
  font-weight: 600 !important;
  letter-spacing: 0.12em !important;
  text-transform: uppercase !important;
  color: var(--text-muted) !important;
}
[data-testid="stMetricValue"] {
  font-family: var(--font-mono) !important;
  font-size: 24px !important;
  font-weight: 600 !important;
  color: var(--text) !important;
  letter-spacing: -0.5px !important;
}
[data-testid="stMetricDelta"] {
  font-family: var(--font-mono) !important;
  font-size: 12px !important;
  font-weight: 500 !important;
}

/* ── BUTTONS ── */
.stButton > button {
  background: var(--bg-card) !important;
  color: var(--text-sec) !important;
  border: 1px solid var(--rule2) !important;
  border-radius: var(--r-sm) !important;
  font-family: var(--font-ui) !important;
  font-size: 13px !important;
  font-weight: 600 !important;
  transition: all 0.2s !important;
  box-shadow: var(--shadow-sm) !important;
}
.stButton > button:hover {
  border-color: var(--blue) !important;
  color: var(--blue) !important;
  box-shadow: 0 0 0 3px var(--blue-glow) !important;
  transform: translateY(-1px) !important;
}



/* ── DATAFRAMES ── */
[data-testid="stDataFrame"] {
  border-radius: var(--r) !important;
  overflow: hidden !important;
  box-shadow: var(--shadow-sm) !important;
  border: 1px solid var(--rule) !important;
}
[data-testid="stDataFrame"] thead th {
  background: var(--bg-card2) !important;
  color: var(--text-muted) !important;
  font-family: var(--font-mono) !important;
  font-size: 12px !important;
  font-weight: 600 !important;
  letter-spacing: 0.10em !important;
  text-transform: uppercase !important;
  border-bottom: 1px solid var(--rule) !important;
  padding: 10px 14px !important;
}
[data-testid="stDataFrame"] tbody td {
  background: var(--bg-card) !important;
  color: var(--text-sec) !important;
  font-family: var(--font-mono) !important;
  font-size: 12px !important;
  border-bottom: 1px solid var(--bg-card2) !important;
  padding: 10px 14px !important;
}
[data-testid="stDataFrame"] tbody tr:hover td {
  background: var(--blue-lt) !important;
  color: var(--text) !important;
}

/* ── CAPTIONS ── */
[data-testid="stCaptionContainer"] p,
.stMarkdown small {
  font-family: var(--font-mono) !important;
  font-size: 12px !important;
  color: var(--text-muted) !important;
  letter-spacing: 0.03em !important;
}

/* ── SCROLLBAR ── */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: var(--bg-page); }
::-webkit-scrollbar-thumb { background: var(--rule2); border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: var(--text-muted); }

/* ── HIDE STREAMLIT CHROME ── */
#MainMenu { visibility: hidden; }
footer { display: none; }
.stDeployButton { display: none; }

/* ── AGGRESSIVE FONT OVERRIDE ── */
/* Target every possible Streamlit text element */
html, body,
.stApp, .stApp *,
[class*="st-"], 
p, div, span, label, input, button, select, textarea,
h1, h2, h3, h4, h5, h6 {
  font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
}

/* Numbers & monospace specifically */
[data-testid="stMetricValue"],
[data-testid="stMetricDelta"],
[data-testid="stDataFrame"] td,
[data-testid="stDataFrame"] th,
code, pre, .stCodeBlock {
  font-family: 'IBM Plex Mono', 'Fira Code', 'Cascadia Code', 
               'Courier New', monospace !important;
}

/* ── PAGE BACKGROUND ── */
.stApp {
  background: #EEF2F8 !important;
}
.main > div {
  background: #EEF2F8 !important;
}

/* ── METRIC CARDS — BIG VISUAL UPGRADE ── */
[data-testid="stMetric"] {
  background: #FFFFFF !important;
  border: 1px solid #E2E8F0 !important;
  border-radius: 12px !important;
  padding: 18px 20px !important;
  box-shadow: 0 1px 3px rgba(15,23,42,0.07), 0 1px 2px rgba(15,23,42,0.04) !important;
  transition: box-shadow 0.2s, transform 0.2s !important;
}
[data-testid="stMetric"]:hover {
  box-shadow: 0 4px 16px rgba(15,23,42,0.10) !important;
  transform: translateY(-1px) !important;
}
[data-testid="stMetricLabel"] > div > p {
  font-family: 'IBM Plex Mono', monospace !important;
  font-size: 12px !important;
  font-weight: 600 !important;
  letter-spacing: 0.12em !important;
  text-transform: uppercase !important;
  color: #94A3B8 !important;
}
[data-testid="stMetricValue"] > div {
  font-family: 'IBM Plex Mono', monospace !important;
  font-size: 24px !important;
  font-weight: 600 !important;
  letter-spacing: -0.5px !important;
  color: #0F172A !important;
}
[data-testid="stMetricDelta"] svg + div {
  font-family: 'IBM Plex Mono', monospace !important;
  font-size: 12px !important;
  font-weight: 500 !important;
}

/* ── TAB UPGRADE ── */
.stTabs [data-baseweb="tab-list"] {
  gap: 0 !important;
  padding: 0 !important;
  background: transparent !important;
  border-bottom: 2px solid #E2E8F0 !important;
}
.stTabs [data-baseweb="tab"] {
  background: transparent !important;
  border-bottom: 2px solid transparent !important;
  padding: 12px 22px !important;
  margin-bottom: -2px !important;
  border-radius: 0 !important;
  color: #94A3B8 !important;
  font-family: 'IBM Plex Mono', monospace !important;
  font-size: 12px !important;
  font-weight: 600 !important;
  letter-spacing: 0.10em !important;
  text-transform: uppercase !important;
  transition: all 0.2s !important;
}
.stTabs [data-baseweb="tab"]:hover {
  color: #475569 !important;
  background: rgba(29,78,216,0.04) !important;
}
.stTabs [aria-selected="true"] {
  color: #1D4ED8 !important;
  border-bottom-color: #1D4ED8 !important;
}

/* ── BUTTON UPGRADE ── */
.stButton > button {
  font-family: 'Inter', sans-serif !important;
  font-weight: 600 !important;
  font-size: 13px !important;
  letter-spacing: 0.01em !important;
  border-radius: 8px !important;
  transition: all 0.2s !important;
}

/* ── SIDEBAR TEXT OVERRIDE ── */
section[data-testid="stSidebar"] * {
  font-family: 'Inter', sans-serif !important;
}
section[data-testid="stSidebar"] [data-testid="stMetricValue"] > div,
section[data-testid="stSidebar"] input,
section[data-testid="stSidebar"] code {
  font-family: 'IBM Plex Mono', monospace !important;
}

/* ── DATAFRAME UPGRADE ── */
[data-testid="stDataFrame"] {
  border: 1px solid #E2E8F0 !important;
  border-radius: 10px !important;
  overflow: hidden !important;
  box-shadow: 0 1px 3px rgba(15,23,42,0.06) !important;
}
[data-testid="stDataFrame"] th {
  font-family: 'IBM Plex Mono', monospace !important;
  font-size: 12px !important;
  font-weight: 600 !important;
  letter-spacing: 0.12em !important;
  text-transform: uppercase !important;
  background: #F7F9FC !important;
  color: #94A3B8 !important;
  padding: 10px 14px !important;
  border-bottom: 1px solid #E2E8F0 !important;
}
[data-testid="stDataFrame"] td {
  font-family: 'IBM Plex Mono', monospace !important;
  font-size: 12px !important;
  padding: 9px 14px !important;
  border-bottom: 1px solid #F8FAFC !important;
}
[data-testid="stDataFrame"] tr:hover td {
  background: #EFF6FF !important;
}



/* ── INPUT UPGRADE ── */
.stTextInput input, .stNumberInput input {
  font-family: 'IBM Plex Mono', monospace !important;
  font-size: 13px !important;
  font-weight: 500 !important;
  letter-spacing: 0.02em !important;
}

/* ── CAPTION UPGRADE ── */
[data-testid="stCaptionContainer"] p {
  font-family: 'IBM Plex Mono', monospace !important;
  font-size: 12px !important;
  letter-spacing: 0.04em !important;
  color: #94A3B8 !important;
}

/* ── SCROLLBAR ── */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #EEF2F8; }
::-webkit-scrollbar-thumb { background: #CBD5E1; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #94A3B8; }



/* ══ ARROW FIX v4 — CSS layer (belt) ══════════════════════════
   Hides icon span via font-size:0 and color:transparent.
   The JS postMessage bridge below is the real fix (suspenders).
   ══════════════════════════════════════════════════════════════ */

/* Make summary a flex row; set its own color transparent so
   any un-caught text node bleeds invisibly */
[data-testid="stExpander"] summary,
details > summary {
  display:        flex         !important;
  align-items:    center       !important;
  gap:            10px         !important;
  list-style:     none         !important;
  cursor:         pointer      !important;
  padding:        11px 16px    !important;
  background:     #F7F9FC      !important;
  border-bottom:  1px solid #E2E8F0 !important;
  color:          transparent  !important;
}
/* Kill native disclosure triangle */
[data-testid="stExpander"] summary::-webkit-details-marker,
details > summary::-webkit-details-marker { display: none !important; }

/* Kill icon spans (not p — the label p must stay visible) */
[data-testid="stExpander"] summary > span,
details > summary > span {
  font-size:   0           !important;
  width:       0           !important;
  height:      0           !important;
  overflow:    hidden      !important;
  visibility:  hidden      !important;
  color:       transparent !important;
  position:    absolute    !important;
  pointer-events: none     !important;
}
/* Kill all SVGs inside summary */
[data-testid="stExpander"] summary svg,
details > summary svg {
  display:  none !important;
  width:    0    !important;
  height:   0    !important;
}
/* Restore the label paragraph */
[data-testid="stExpander"] summary p,
details > summary p {
  color:       #334155 !important;
  font-size:   13px    !important;
  font-weight: 500     !important;
  visibility:  visible !important;
  opacity:     1       !important;
  margin:      0       !important;
}
/* Custom CSS chevron — zero dependency on Streamlit internals */
[data-testid="stExpander"] summary::before,
details > summary::before {
  content:       ""               !important;
  display:       inline-block     !important;
  flex-shrink:   0                !important;
  width:         6px              !important;
  height:        6px              !important;
  border-right:  2px solid #94A3B8 !important;
  border-bottom: 2px solid #94A3B8 !important;
  transform:     rotate(-45deg)   !important;
  transition:    transform 0.2s ease !important;
  margin-right:  2px              !important;
}
details[open] > summary::before,
[data-testid="stExpander"][open] summary::before {
  transform: rotate(45deg) !important;
}
[data-testid="stExpander"] summary:hover,
details > summary:hover { background: #EFF6FF !important; }
[data-testid="stExpander"] summary:hover::before {
  border-color: #1D4ED8 !important;
}
[data-testid="stExpander"] summary:hover p,
details > summary:hover p { color: #1D4ED8 !important; }
/* ══ END CSS ARROW FIX ══ */

/* ══════════════════════════════════════════════════════════════
   INNER SUB-TABS — Make Summary/Valuation/Quality/Signals
   visually prominent, not like plain Streamlit tabs
   ══════════════════════════════════════════════════════════════ */

/* Target only the second level of tabs (inner sub-tabs)
   by using the parent container. The inner tabs are rendered
   inside the outer tab panel div. */

/* Make ALL tab lists look great */
[data-baseweb="tab-list"] {
  background:    #FFFFFF        !important;
  border-bottom: 2px solid #E2E8F0 !important;
  padding:       0              !important;
  gap:           0              !important;
  box-shadow:    0 1px 4px rgba(15,23,42,0.06) !important;
}
[data-baseweb="tab"] {
  background:     transparent   !important;
  border:         none          !important;
  border-bottom:  2px solid transparent !important;
  margin-bottom:  -2px          !important;
  color:          #94A3B8       !important;
  font-family:    'IBM Plex Mono', monospace !important;
  font-size:      11px          !important;
  font-weight:    700           !important;
  letter-spacing: 0.10em        !important;
  text-transform: uppercase     !important;
  padding:        13px 22px     !important;
  transition:     all 0.15s     !important;
  border-radius:  0             !important;
}
[data-baseweb="tab"]:hover {
  color:      #475569   !important;
  background: #EFF6FF   !important;
}
[aria-selected="true"] {
  color:              #1D4ED8   !important;
  border-bottom-color:#1D4ED8   !important;
  background:         transparent !important;
}
[data-baseweb="tab-panel"] {
  padding: 1.2rem 0 0 !important;
}
/* ══════════════════════════════════════════════════════════════
   MOBILE RESPONSIVE v2 — Complete overhaul
   Based on Apple HIG + Material Design touch guidelines
   ══════════════════════════════════════════════════════════════ */

/* ── ≤ 768px: Phone layout ──────────────────────────────────── */
@media (max-width: 768px) {

  /* Reduce container padding */
  .main .block-container {
    padding: 0.75rem 0.75rem 2rem !important;
    max-width: 100% !important;
  }

  /* Stack ALL columns vertically */
  [data-testid="stHorizontalBlock"] {
    flex-direction: column !important;
    gap: 6px !important;
    flex-wrap: nowrap !important;
  }
  [data-testid="stHorizontalBlock"] > div,
  [data-testid="column"] {
    width: 100% !important;
    min-width: 100% !important;
    flex: 1 1 100% !important;
  }

  /* KPI cards: 2 per row */
  [data-testid="stMetric"] {
    min-width: calc(50% - 6px) !important;
    padding: 10px 12px !important;
  }
  [data-testid="stMetricValue"] {
    font-size: 20px !important;
  }
  [data-testid="stMetricLabel"] {
    font-size: 11px !important;
  }
  [data-testid="stMetricDelta"] {
    font-size: 11px !important;
  }

  /* Sidebar — hidden by default on mobile, slides in on tap */
  [data-testid="stSidebar"] {
    width: 80vw !important;
    max-width: 320px !important;
    min-width: unset !important;
    transform: translateX(-100%) !important;
    transition: transform 0.25s ease !important;
    position: fixed !important;
    top: 0 !important;
    left: 0 !important;
    height: 100dvh !important;
    z-index: 999 !important;
    overflow-y: auto !important;
    box-shadow: 4px 0 24px rgba(0,0,0,0.4) !important;
  }
  /* When sidebar is open (Streamlit toggles aria-expanded) */
  [data-testid="stSidebar"][aria-expanded="true"] {
    transform: translateX(0) !important;
  }
  /* Collapse button always visible on mobile */
  [data-testid="stSidebarCollapsedControl"] {
    display: flex !important;
    position: fixed !important;
    top: 12px !important;
    left: 12px !important;
    z-index: 1000 !important;
    background: #1D4ED8 !important;
    border-radius: 8px !important;
    padding: 6px 10px !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.3) !important;
  }
  [data-testid="stSidebarNav"] {
    display: none !important;
  }
  /* Main content: full width, add top padding for the menu button */
  .main {
    margin-left: 0 !important;
    padding-top: 52px !important;
  }

  /* Tabs: horizontal scroll */
  .stTabs [data-baseweb="tab-list"],
  [data-baseweb="tab-list"] {
    overflow-x: auto !important;
    flex-wrap: nowrap !important;
    -webkit-overflow-scrolling: touch !important;
    scrollbar-width: none !important;
    gap: 0 !important;
  }
  .stTabs [data-baseweb="tab-list"]::-webkit-scrollbar { display: none !important; }
  [data-baseweb="tab"] {
    padding: 12px 14px !important;
    font-size: 10px !important;
    white-space: nowrap !important;
    flex-shrink: 0 !important;
    min-height: 44px !important;
  }

  /* All buttons: 44px min height (Apple HIG) */
  [data-testid="baseButton-primary"],
  [data-testid="baseButton-secondary"],
  button[kind="primary"],
  button[kind="secondary"],
  .stButton > button {
    min-height: 44px !important;
    font-size: 13px !important;
    width: 100% !important;
    padding: 10px 16px !important;
  }

  /* Inputs: 44px min height + 16px font (prevents iOS zoom) */
  .stTextInput input,
  .stNumberInput input,
  .stSelectbox select,
  input[type="text"],
  input[type="number"] {
    font-size: 16px !important;
    min-height: 44px !important;
    padding: 10px 14px !important;
  }

  /* Selectbox */
  [data-baseweb="select"] {
    min-width: 100% !important;
  }
  [data-baseweb="select"] > div {
    min-height: 44px !important;
  }

  /* Sliders: constrained width + larger touch target */
  [data-testid="stSlider"] {
    padding: 8px 0 16px !important;
    max-width: 100% !important;
    overflow: hidden !important;
    box-sizing: border-box !important;
  }
  [data-testid="stSlider"] > div {
    max-width: 100% !important;
    overflow: hidden !important;
  }
  [data-testid="stSlider"] [data-baseweb="slider"] {
    max-width: 100% !important;
    padding: 0 4px !important;
  }
  /* Thumb: large enough to tap */
  [data-testid="stSlider"] [role="slider"] {
    width: 28px !important;
    height: 28px !important;
  }

  /* Expanders */
  details > summary,
  [data-testid="stExpander"] summary {
    min-height: 44px !important;
    padding: 12px 16px !important;
  }

  /* Plotly charts */
  .js-plotly-plot,
  .plotly,
  [data-testid="stPlotlyChart"] {
    max-width: 100% !important;
    height: auto !important;
    min-height: 200px !important;
    overflow: hidden !important;
  }

  /* DataFrames: horizontal scroll */
  [data-testid="stDataFrame"],
  [data-testid="stTable"] {
    overflow-x: auto !important;
    -webkit-overflow-scrolling: touch !important;
    font-size: 11px !important;
  }

  /* iframes (TradingView chart) */
  iframe {
    max-width: 100% !important;
    width: 100% !important;
  }

  /* Hero section */
  .yiq-hero {
    padding: 28px 20px !important;
    flex-direction: column !important;
  }

  /* Scale down font sizes 20% */
  .stMarkdown p, .stMarkdown li {
    font-size: 13px !important;
    line-height: 1.6 !important;
  }

  /* Hide non-essential table columns */
  .hide-mobile { display: none !important; }

  /* st.html() card layouts */
  div[style*="display:grid"][style*="grid-template-columns"] {
    grid-template-columns: 1fr !important;
  }
  div[style*="display:flex"][style*="justify-content:space-between"] {
    flex-direction: column !important;
    gap: 10px !important;
  }

  /* Hero headline */
  div[style*="font-size:32px"],
  div[style*="font-size: 32px"] {
    font-size: 26px !important;
  }
  div[style*="font-size:40px"],
  div[style*="font-size: 40px"] {
    font-size: 28px !important;
  }
}

/* ── ≤ 480px: Small phone layout ────────────────────────────── */
@media (max-width: 480px) {

  /* Tabs: smaller labels */
  [data-baseweb="tab"] {
    padding: 10px 10px !important;
    font-size: 9px !important;
    letter-spacing: 0.04em !important;
  }

  /* Table cells: smaller font */
  [data-testid="stDataFrame"] td,
  [data-testid="stDataFrame"] th,
  [data-testid="stTable"] td,
  [data-testid="stTable"] th {
    font-size: 11px !important;
    padding: 4px 6px !important;
  }

  /* KPI cards: 1 per row on very small screens */
  [data-testid="stMetric"] {
    min-width: 100% !important;
  }

  /* Reduce container padding further */
  .main .block-container {
    padding: 0.5rem 0.5rem 2rem !important;
  }

  /* Progress bar */
  [data-testid="stProgressBar"] {
    height: 6px !important;
  }
}

/* ── Tablet: 769-1024px ──────────────────────────────────────── */
@media (min-width: 769px) and (max-width: 1024px) {
  .main .block-container {
    padding: 1rem 1.5rem !important;
    max-width: 100% !important;
  }
  [data-baseweb="tab"] {
    padding: 11px 14px !important;
    font-size: 11px !important;
  }
  [data-testid="stMetricValue"] {
    font-size: 22px !important;
  }
}

</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# TYPOGRAPHY SYSTEM — Inter + IBM Plex Mono
# Final decision: Inter for all UI text (neutral, legible, used by
# Robinhood, Linear, Figma). IBM Plex Mono for all numbers/prices
# (tabular figures align perfectly in columns — critical for finance).
# ══════════════════════════════════════════════════════════════
st.markdown("""
<style>
/* ── TYPOGRAPHY SCALE ────────────────────────────────────────
   Page Title:      24px / 700 / 1.25  — app-level headings
   Section Header:  15px / 600 / 1.4   — tab & card section labels
   Card Title:      13px / 600 / 1.4   — inside card headers
   Body:            13px / 400 / 1.65  — all descriptive text
   Secondary:       12px / 400 / 1.55  — helper text, captions
   Label:           11px / 500 / 1.3   — ALL-CAPS metric labels
   Numbers/Prices:  IBM Plex Mono, sizes defined per context
   ──────────────────────────────────────────────────────────── */

/* ── BASE RESET ─────────────────────────────────────────────── */
html, body, [class*="css"], .stApp,
.block-container, .element-container,
p, span, div, label, li, td, th {
  font-family: 'Inter', system-ui, -apple-system, BlinkMacSystemFont,
               'Segoe UI', sans-serif !important;
  -webkit-font-smoothing: antialiased;
  -moz-osx-font-smoothing: grayscale;
  text-rendering: optimizeLegibility;
}

/* ── PAGE TITLE ─────────────────────────────────────────────── */
h1, .stMarkdown h1 {
  font-size: 24px !important;
  font-weight: 700 !important;
  line-height: 1.25 !important;
  color: #0F172A !important;
  letter-spacing: -0.02em !important;
  margin-bottom: 4px !important;
}

/* ── SECTION HEADER ─────────────────────────────────────────── */
h2, .stMarkdown h2 {
  font-size: 15px !important;
  font-weight: 600 !important;
  line-height: 1.4 !important;
  color: #1E293B !important;
  letter-spacing: -0.01em !important;
  margin-bottom: 12px !important;
  margin-top: 20px !important;
}

/* ── CARD TITLE ─────────────────────────────────────────────── */
h3, .stMarkdown h3 {
  font-size: 13px !important;
  font-weight: 600 !important;
  line-height: 1.4 !important;
  color: #334155 !important;
  letter-spacing: 0 !important;
  margin-bottom: 8px !important;
  margin-top: 0 !important;
}

/* ── BODY TEXT ──────────────────────────────────────────────── */
p, .stMarkdown p, .stMarkdown li {
  font-size: 13px !important;
  font-weight: 400 !important;
  line-height: 1.65 !important;
  color: #334155 !important;
  margin-bottom: 6px !important;
}

/* ── SECONDARY / CAPTION TEXT ───────────────────────────────── */
small, .stCaption, [data-testid="stCaptionContainer"] p,
.stMarkdown small {
  font-size: 11px !important;
  font-weight: 400 !important;
  line-height: 1.55 !important;
  color: #64748B !important;
}

/* ── METRIC LABELS (the small ALL-CAPS label above a number) ── */
[data-testid="stMetricLabel"] p,
[data-testid="stMetricLabel"] label {
  font-size: 11px !important;
  font-weight: 500 !important;
  line-height: 1.3 !important;
  color: #94A3B8 !important;
  text-transform: uppercase !important;
  letter-spacing: 0.10em !important;
}

/* ── METRIC VALUES (the big number in st.metric) ─────────────── */
[data-testid="stMetricValue"] > div,
[data-testid="stMetricValue"] {
  font-family: 'IBM Plex Mono', 'Courier New', monospace !important;
  font-size: 20px !important;
  font-weight: 600 !important;
  line-height: 1.2 !important;
  color: #0F172A !important;
  letter-spacing: -0.01em !important;
}

/* ── METRIC DELTA (±% change) ────────────────────────────────── */
[data-testid="stMetricDelta"] > div {
  font-family: 'IBM Plex Mono', 'Courier New', monospace !important;
  font-size: 12px !important;
  font-weight: 500 !important;
}

/* ── STOCK PRICE (large display number) ─────────────────────── */
.yiq-price {
  font-family: 'IBM Plex Mono', 'Courier New', monospace !important;
  font-size: 32px !important;
  font-weight: 700 !important;
  line-height: 1.0 !important;
  color: #0F172A !important;
  letter-spacing: -0.02em !important;
}

/* ── PERCENTAGE CHANGE ──────────────────────────────────────── */
.yiq-pct-pos {
  font-family: 'IBM Plex Mono', 'Courier New', monospace !important;
  font-size: 14px !important;
  font-weight: 600 !important;
  color: #0D7A4E !important;
}
.yiq-pct-neg {
  font-family: 'IBM Plex Mono', 'Courier New', monospace !important;
  font-size: 14px !important;
  font-weight: 600 !important;
  color: #B91C1C !important;
}

/* ── KEY METRIC NUMBER (e.g. fair value, upside %) ───────────── */
.yiq-metric-num {
  font-family: 'IBM Plex Mono', 'Courier New', monospace !important;
  font-size: 22px !important;
  font-weight: 700 !important;
  line-height: 1.1 !important;
  letter-spacing: -0.01em !important;
}

/* ── LABEL TAG (ALL-CAPS above metric) ──────────────────────── */
.yiq-label {
  font-size: 11px !important;
  font-weight: 500 !important;
  line-height: 1.3 !important;
  text-transform: uppercase !important;
  letter-spacing: 0.11em !important;
  color: #94A3B8 !important;
}

/* ── INSIGHT TEXT ───────────────────────────────────────────── */
.yiq-insight {
  font-size: 15px !important;
  font-weight: 400 !important;
  line-height: 1.7 !important;
  color: #1E293B !important;
}

/* ── SPACING SYSTEM ─────────────────────────────────────────── */
/* Section gap:   24px top/bottom (between major sections)       */
/* Card padding:  20px all sides (compact) / 24px (standard)     */
/* Element gap:   12px (between related items)                   */
/* Tight gap:     6px  (between label and value)                 */

/* Remove excess Streamlit default margins */
.block-container {
  padding-top: 1.5rem !important;
  padding-bottom: 2rem !important;
}
.element-container {
  margin-bottom: 0 !important;
}
[data-testid="stVerticalBlock"] > [data-testid="stVerticalBlock"] {
  gap: 0 !important;
}

/* ── BUTTON TEXT ────────────────────────────────────────────── */
.stButton > button {
  font-size: 13px !important;
  font-weight: 500 !important;
  letter-spacing: 0.02em !important;
}

/* ── INPUT / SELECT ─────────────────────────────────────────── */
.stTextInput input,
.stSelectbox select,
.stSelectbox div[data-baseweb="select"] {
  font-size: 13px !important;
  font-weight: 400 !important;
}

/* ── TAB LABELS ─────────────────────────────────────────────── */
[data-testid="stTabs"] button[role="tab"] {
  font-size: 12px !important;
  font-weight: 500 !important;
  letter-spacing: 0.04em !important;
}

/* ── EXPANDER LABELS ─────────────────────────────────────────── */
[data-testid="stExpander"] summary p,
[data-testid="stExpander"] summary span {
  font-size: 13px !important;
  font-weight: 500 !important;
  color: #334155 !important;
}

/* ── DATAFRAME / TABLE ──────────────────────────────────────── */
[data-testid="stDataFrame"] *,
.stDataFrame * {
  font-size: 12px !important;
}

/* ── SIDEBAR TEXT ───────────────────────────────────────────── */
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span {
  font-size: 12px !important;
}
section[data-testid="stSidebar"] .stSlider p {
  font-size: 11px !important;
}

/* ── WARNING / INFO BOXES ───────────────────────────────────── */
[data-testid="stAlert"] p {
  font-size: 12px !important;
  line-height: 1.55 !important;
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════
CURRENCIES = {
    "INR": {"symbol": "₹", "code": "INR"},
    "USD": {"symbol": "$", "code": "USD"},
    "GBP": {"symbol": "£", "code": "GBP"},
    "EUR": {"symbol": "€", "code": "EUR"},
}

# ── Hardcoded fallback rates (updated Mar 2025) ───────────────
# Used ONLY when ALL live APIs fail. Update these quarterly.
_FX_FALLBACK = {
    ("USD", "INR"): 83.5,  ("INR", "USD"): 1/83.5,
    ("USD", "EUR"): 0.92,  ("EUR", "USD"): 1/0.92,
    ("USD", "GBP"): 0.79,  ("GBP", "USD"): 1/0.79,
    ("USD", "JPY"): 149.0, ("JPY", "USD"): 1/149.0,
    ("INR", "EUR"): 0.011, ("EUR", "INR"): 90.0,
    ("INR", "GBP"): 0.0095,("GBP", "INR"): 105.0,
}

@st.cache_data(ttl=1800, show_spinner=False)
def get_fx_rate(from_ccy: str, to_ccy: str) -> float:
    """
    Fetch live FX rate with 3-tier fallback:
      1. exchangerate-api.com  (primary, free tier)
      2. frankfurter.app       (secondary, ECB data)
      3. Hardcoded table       (quarterly-updated emergency fallback)
    Cache: 30 minutes (ttl=1800). Never returns stale session value.
    """
    if not from_ccy or not to_ccy:
        return 1.0
    fc, tc = from_ccy.upper().strip(), to_ccy.upper().strip()
    if fc == tc:
        return 1.0
    try:
        r = requests.get(
            f"https://api.exchangerate-api.com/v4/latest/{fc}",
            timeout=6,
        )
        if r.status_code == 200:
            rate = float(r.json().get("rates", {}).get(tc, 0))
            if rate > 0:
                return rate
    except Exception as _e1:
        print(f"[YieldIQ] FX tier-1 fetch failed ({fc}→{tc}): {_e1}")
    try:
        r = requests.get(
            f"https://api.frankfurter.app/latest?from={fc}&to={tc}",
            timeout=6,
        )
        if r.status_code == 200:
            rate = float(r.json().get("rates", {}).get(tc, 0))
            if rate > 0:
                return rate
    except Exception as _e2:
        print(f"[YieldIQ] FX tier-2 fetch failed ({fc}→{tc}): {_e2}")
    # Tier 3: hardcoded fallback
    fallback = _FX_FALLBACK.get((fc, tc))
    if fallback:
        return fallback
    # Last resort: try reverse
    rev = _FX_FALLBACK.get((tc, fc))
    if rev and rev != 0:
        return 1.0 / rev
    return 1.0

def _get_cache_ttl() -> int:
    """Return cache TTL based on user tier."""
    t = st.session_state.get("_tier", "free")
    return {"pro": 60, "starter": 180, "premium": 180, "free": 600}.get(t, 600)


def fetch_stock_data(ticker):
    """Wrapper that applies tiered TTL caching."""
    return _fetch_stock_data_cached(ticker, _get_cache_ttl())


@st.cache_data(ttl=600, show_spinner=False)
def _fetch_stock_data_cached(ticker, _ttl_key):
    """Actual cached fetch — _ttl_key parameter busts cache per tier."""
    collector  = StockDataCollector(ticker)
    raw        = collector.get_all()
    price_hist = pd.DataFrame()
    wacc_data  = {}
    if collector._ticker_obj:
        price_hist = collector.get_price_history(period="1y")
        is_indian  = ticker.endswith(".NS") or ticker.endswith(".BO")
        wacc_data  = compute_wacc(collector._ticker_obj, is_indian)
    return raw, price_hist, wacc_data

def fmt(v, sym, d=2):
    a = abs(v)
    if a >= 1e12: return f"{sym}{v/1e12:,.2f}T"
    if a >= 1e9:  return f"{sym}{v/1e9:,.2f}B"
    if a >= 1e6:  return f"{sym}{v/1e6:,.2f}M"
    return f"{sym}{v:,.{d}f}"

def fmts(v, sym): return f"{sym}{v:,.2f}"

# ── HUMAN LANGUAGE TRANSLATION HELPERS ─────────────────────────
_SIG_HUMAN = {
    "Undervalued 🟢":    ("📊 Undervalued by model estimate",  "#0D7A4E", "#F0FDF4", "#BBF7D0"),
    "Near Fair Value 🟡":("📉 Slightly below model fair value", "#B45309", "#FFFBEB", "#FDE68A"),
    "Fairly Valued 🔵":  ("⚖️ Near model fair value",           "#1D4ED8", "#EFF6FF", "#BFDBFE"),
    "Overvalued 🔴":     ("📈 Overvalued by model estimate",    "#B91C1C", "#FEF2F2", "#FECACA"),
    "⚠️ Data Limited":   ("🔍 Model data needs review",         "#B45309", "#FFFBEB", "#FDE68A"),
    "N/A ⬜":            ("⏳ Analysing…",                     "#4A5E7A", "#FFFFFF", "#F8FAFC"),
}

def sig_human(sig):
    """Return (human_label, fg, bg, border) for a signal string."""
    return _SIG_HUMAN.get(sig, ("⏳ Analysing…", "#4A5E7A", "#FFFFFF", "#F8FAFC"))

def mos_insight(mos_pct: float, sig: str, company: str, suspicious: bool) -> str:
    """One-line model-output summary. No advice language."""
    if suspicious:
        return f"⚠️ {company}'s financials show unusual patterns — model estimates may be unreliable."
    if mos_pct >= 20:
        return f"📊 Our model estimates {company} is trading ~{mos_pct:.0f}% below its calculated fair value."
    elif mos_pct >= 5:
        return f"📊 Our model estimates {company} is trading ~{mos_pct:.0f}% below its calculated fair value."
    elif mos_pct >= -5:
        return f"⚖️ {company} is trading close to our model's estimated fair value."
    elif mos_pct >= -15:
        return f"📊 Our model estimates {company} is trading ~{abs(mos_pct):.0f}% above its calculated fair value."
    else:
        return f"📊 Our model estimates {company} is trading ~{abs(mos_pct):.0f}% above its calculated fair value."

def plain_kpi_label(label: str) -> str:
    """Translate finance jargon into plain English for KPI cards."""
    _MAP = {
        "Margin of Safety":   "Discount to fair value",
        "WACC":               "Required return rate",
        "Op Margin":          "Profit per ₹100 revenue",
        "FCF Growth":         "Cash flow growth",
        "Revenue Growth":     "Revenue growth",
        "Confidence":         "Model reliability",
        "Intrinsic Value":    "Estimated fair value",
        "IV (DCF+PE Blend)":  "Estimated fair value",
        "Intrinsic Value (DCF)": "Estimated fair value",
        "Current Price":      "Current price",
    }
    return _MAP.get(label, label)

def KL(**kw):
    """Koyfin-style dark chart layout — apply to every fig.update_layout()."""
    base = dict(
        paper_bgcolor="#0d1117",
        plot_bgcolor="#161b22",
        font=dict(family="Inter, DM Sans, system-ui, sans-serif", color="#e6edf3", size=11),
        margin=dict(l=48, r=24, t=48, b=44),
        hovermode="x unified",
        hoverlabel=dict(
            bgcolor="#21262d",
            font=dict(color="#e6edf3", family="IBM Plex Mono, monospace", size=12),
            bordercolor="#30363d",
        ),
        legend=dict(
            bgcolor="rgba(0,0,0,0)",
            bordercolor="#30363d",
            borderwidth=1,
            font=dict(color="#8b949e", size=11),
        ),
        xaxis=dict(
            gridcolor="#21262d",
            linecolor="#30363d",
            tickfont=dict(color="#8b949e", size=10),
            zeroline=False,
        ),
        yaxis=dict(
            gridcolor="#21262d",
            linecolor="#30363d",
            tickfont=dict(color="#8b949e", size=10),
            zeroline=False,
        ),
    )
    base.update(kw)
    return base

def apply_koyfin(fig, accent="#00b4d8", height=280, title_txt="", extra_kw=None):
    """One-call upgrade: dark layout + teal accent top border + axis polish."""
    kw = dict(height=height)
    if title_txt:
        kw["title"] = dict(text=title_txt, font=dict(color="#e6edf3", size=13, family="Inter, sans-serif"), x=0, pad=dict(l=4))
    if extra_kw:
        kw.update(extra_kw)
    fig.update_layout(**KL(**kw))
    fig.update_xaxes(gridcolor="#21262d", linecolor="#30363d", tickfont=dict(color="#8b949e", size=10))
    fig.update_yaxes(gridcolor="#21262d", linecolor="#30363d", tickfont=dict(color="#8b949e", size=10))
    # Teal top-border accent via annotation line
    fig.add_shape(type="line", xref="paper", yref="paper",
                  x0=0, x1=1, y0=1, y1=1,
                  line=dict(color=accent, width=2),
                  layer="above")
    return fig

def CL(**kw):
    base = dict(
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="#FFFFFF",
        font=dict(family="Inter,sans-serif", color="#475569", size=11),
        margin=dict(t=20, b=40, l=10, r=10),
        xaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0", zeroline=False, tickcolor="#CBD5E1", tickfont=dict(color="#64748B")),
        yaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0", zeroline=False, tickcolor="#CBD5E1", tickfont=dict(color="#64748B")),
        hoverlabel=dict(bgcolor="#FFFFFF", bordercolor="#1D4ED8",
                        font=dict(color="#0F172A", family="IBM Plex Mono", size=12)),
    )
    base.update(kw)
    return base

@st.cache_data(ttl=3600, show_spinner=False)
def generate_ai_summary(
    ticker: str, company_name: str, price: float, iv: float,
    mos_pct: float, signal: str, piotroski_score: int, wacc: float,
    rev_growth: float, fcf_growth: float, op_margin: float,
    moat_grade: str, sym: str,
) -> str:
    """Generate a plain-English stock summary using Gemini 2.0 Flash (free tier)."""
    import os as _os
    _prompt = (
        f"You are a senior equity analyst. Write a concise 3-paragraph stock analysis "
        f"in plain English for a retail investor. Be balanced, factual, and specific. "
        f"Do NOT say 'buy' or 'sell'. Use 'appears undervalued/overvalued by the model'.\n\n"
        f"Stock: {company_name} ({ticker})\n"
        f"Current Price: {sym}{price:,.2f}\n"
        f"Model Intrinsic Value: {sym}{iv:,.2f}\n"
        f"Margin of Safety: {mos_pct:.1f}%\n"
        f"Model Signal: {signal}\n"
        f"Piotroski F-Score: {piotroski_score}/9\n"
        f"Revenue Growth: {rev_growth:.1%}\n"
        f"FCF Growth: {fcf_growth:.1%}\n"
        f"Operating Margin: {op_margin:.1%}\n"
        f"WACC (Required Return): {wacc:.1%}\n"
        f"Economic Moat: {moat_grade}\n\n"
        f"Write exactly 3 paragraphs:\n"
        f"Para 1 (2-3 sentences): What does this company do and what does the valuation tell us?\n"
        f"Para 2 (2-3 sentences): What are the key financial strengths or concerns?\n"
        f"Para 3 (1-2 sentences): What is the main risk or watchpoint for this thesis?\n\n"
        f"Keep total response under 200 words. No headers. No bullets. Plain paragraphs only.\n"
        f"End with a one-line disclaimer: This is model output, not investment advice."
    )
    # ── Try Gemini first, fall back to Groq ──────────────────
    _gemini_key = _os.environ.get("GEMINI_API_KEY", "").strip()
    _groq_key   = _os.environ.get("GROQ_API_KEY",   "").strip()

    if _gemini_key:
        try:
            from google import genai as _genai
            _client = _genai.Client(api_key=_gemini_key)
            _response = _client.models.generate_content(
                model="gemini-2.0-flash", contents=_prompt,
            )
            return _response.text.strip()
        except Exception as _e:
            _err = str(_e).lower()
            if not any(k in _err for k in ("quota", "429", "resource_exhausted", "limit")):
                return f"AI summary unavailable: {_e}"
            # quota/region error → fall through to Groq

    if _groq_key:
        try:
            from groq import Groq as _Groq
            _client = _Groq(api_key=_groq_key)
            _resp = _client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": _prompt}],
                max_tokens=400,
                temperature=0.3,
            )
            return _resp.choices[0].message.content.strip()
        except Exception as _e:
            return f"AI summary unavailable: {_e}"

    return "AI summary unavailable: add GEMINI_API_KEY or GROQ_API_KEY to .env"


@st.cache_data(ttl=300, show_spinner=False)
def fetch_market_overview():
    import yfinance as yf
    symbols = {
        "S&P 500":   "^GSPC",
        "NASDAQ":    "^IXIC",
        "Dow Jones": "^DJI",
        "Gold":      "GC=F",
        "10Y UST":   "^TNX",
    }
    results = {}
    for name, sym in symbols.items():
        try:
            t = yf.Ticker(sym)
            fi = t.fast_info
            price = float(getattr(fi, "last_price", 0) or 0)
            prev  = float(getattr(fi, "previous_close", 0) or 0)
            chg   = ((price - prev) / prev * 100) if prev > 0 else 0
            results[name] = {"price": price, "change_pct": chg, "symbol": sym}
        except Exception as _e:
            log.warning(f"[market_overview] price fetch failed for {sym}: {_e}")
            results[name] = {"price": 0, "change_pct": 0, "symbol": sym}
    return results


@st.cache_data(ttl=60, show_spinner=False)
def fetch_market_pulse():
    """Fetch S&P 500, 10Y Treasury, VIX for the sidebar Market Pulse widget."""
    import yfinance as yf
    _pulse_syms = [("S&P 500", "^GSPC"), ("10Y Yield", "^TNX"), ("VIX", "^VIX")]
    result = {}
    for name, sym in _pulse_syms:
        try:
            fi    = yf.Ticker(sym).fast_info
            price = float(getattr(fi, "last_price", 0) or 0)
            prev  = float(getattr(fi, "previous_close", 0) or 0)
            chg   = ((price - prev) / prev * 100) if prev > 0 else 0
            result[name] = {"price": price, "chg": chg}
        except Exception as _e:
            log.warning(f"[market_pulse] {sym}: {_e}")
            result[name] = {"price": 0, "chg": 0}
    return result


def show_upgrade_modal(feature_name: str = "this feature"):
    _features = [
        ("♾️", "Unlimited stock analyses"),
        ("🎲", "Monte Carlo — 1,000 simulations"),
        ("📥", "Excel DCF model download"),
        ("🤖", "AI plain-English stock summary"),
        ("🌍", "US + India + Europe markets"),
    ]
    _rows_html = ""
    for _fi, _ft in _features:
        _rows_html += (
            '<div style="display:flex;align-items:center;gap:10px;padding:8px 0;'
            'border-bottom:0.5px solid #F1F5F9;">'
            f'<div style="font-size:15px;width:24px;text-align:center;">{_fi}</div>'
            f'<div style="font-size:12px;color:#334155;flex:1;">{_ft}</div>'
            '<div style="color:#059669;font-size:13px;">✓</div>'
            '</div>'
        )
    st.html(f"""
    <div style="position:relative;background:rgba(0,0,0,0.04);border-radius:14px;
                padding:20px;margin:8px 0;">
      <div style="background:linear-gradient(135deg,#0f2537,#1d4ed8);
                  border-radius:12px 12px 0 0;padding:20px 20px 16px;position:relative;">
        <div style="display:inline-block;background:rgba(245,158,11,0.2);
                    border:1px solid rgba(245,158,11,0.4);color:#f59e0b;
                    font-size:10px;font-weight:700;padding:2px 8px;border-radius:10px;
                    margin-bottom:8px;">🔒 PRO FEATURE</div>
        <div style="font-size:18px;font-weight:700;color:white;">Unlock YieldIQ Pro</div>
        <div style="font-size:12px;color:rgba(255,255,255,0.55);margin-top:3px;">
          {feature_name} requires a Starter or Pro plan</div>
      </div>
      <div style="background:white;border-radius:0 0 12px 12px;border:1px solid #E2E8F0;
                  border-top:none;padding:16px 20px;">
        {_rows_html}
      </div>
    </div>
    """)
    col1, col2 = st.columns(2)
    with col1:
        st.link_button("Upgrade to Pro — $49/mo →", "https://yieldiq.app/pricing", width='stretch', type="primary")  # Pro price is $49/mo
    with col2:
        st.link_button("See all plans", "https://yieldiq.app/pricing", width='stretch')


def ccard(title, accent="#1D4ED8"):
    st.html(
        f'''<div style="background:#FFFFFF;border:1px solid #E2E8F0;border-radius:12px;
            padding:20px 24px 6px;margin-bottom:16px;position:relative;overflow:hidden;
            box-shadow:0 1px 4px rgba(15,23,42,0.06),0 1px 2px rgba(15,23,42,0.04);">
        <div style="position:absolute;top:0;left:0;right:0;height:3px;
            background:linear-gradient(90deg,{accent} 0%,rgba(6,182,212,0.6) 60%,transparent 100%);"></div>
        <div style="font-family:'IBM Plex Mono',monospace;font-size:12px;font-weight:600;
            letter-spacing:0.12em;text-transform:uppercase;color:#94A3B8;
            margin-bottom:14px;display:flex;align-items:center;gap:8px;">
          <span style="display:inline-block;width:5px;height:5px;border-radius:50%;
              background:{accent};flex-shrink:0;opacity:0.8;"></span>{title}</div>''',
    )

def ccard_end(): st.html("</div>")


def render_score_dial(score: float, max_score: float, label: str,
                      color: str, size: int = 120) -> str:
    """Circular SVG score dial — Bloomberg Intelligence style."""
    pct     = score / max_score if max_score else 0
    circumf = 283.0   # 2 * pi * r(45)
    dash    = pct * circumf
    gap     = circumf - dash
    track   = "#21262d"
    svg  = '<div style="display:inline-flex;flex-direction:column;align-items:center;">'
    svg += f'<svg width="{size}" height="{size}" viewBox="0 0 100 100">'
    svg += f'<circle cx="50" cy="50" r="48" fill="none" stroke="{color}" stroke-width="0.5" opacity="0.15"/>'
    svg += f'<circle cx="50" cy="50" r="45" fill="none" stroke="{track}" stroke-width="9"/>'
    svg += (f'<circle cx="50" cy="50" r="45" fill="none" stroke="{color}" stroke-width="9"'
            f' stroke-linecap="round" stroke-dasharray="{dash:.1f} {gap:.1f}"'
            f' transform="rotate(-90 50 50)"/>')
    svg += (f'<text x="50" y="46" text-anchor="middle" dominant-baseline="middle"'
            f' font-family="IBM Plex Mono, monospace" font-size="22"'
            f' font-weight="700" fill="{color}">{score:.0f}</text>')
    svg += (f'<text x="50" y="63" text-anchor="middle" dominant-baseline="middle"'
            f' font-family="IBM Plex Mono, monospace" font-size="10"'
            f' fill="#8b949e">/ {max_score:.0f}</text>')
    svg += '</svg>'
    if label:
        svg += (f'<div style="font-size:10px;font-weight:700;color:#8b949e;'
                f'text-transform:uppercase;letter-spacing:0.12em;'
                f'text-align:center;margin-top:5px;">{label}</div>')
    svg += '</div>'
    return svg


# ══════════════════════════════════════════════════════════════
# RELATIVE VALUATION VIEW  (non-DCF stocks: banks, REITs, etc.)
# ══════════════════════════════════════════════════════════════

def _render_relative_valuation_view(
    ticker:    str,
    rv:        dict,
    raw:       dict,
    sym:       str = "$",
    fx:        float = 1.0,
) -> None:
    """Render the full dashboard view for a non-DCF stock using relative valuation."""
    price      = rv.get("price", 0) * fx
    name       = rv.get("name") or raw.get("company_name", ticker)
    signal     = rv.get("signal", "N/A ⬜")
    avg_pct    = rv.get("signal_avg_pct", 0)
    sector_lbl = rv.get("gics_sector") or rv.get("sector_key", "Financial / Real Estate")
    metrics    = rv.get("metrics", [])
    mkt_cap    = rv.get("market_cap", 0)
    div_yield  = rv.get("dividend_yield", 0)
    beta       = rv.get("beta", 0)
    hi52       = rv.get("52w_high", 0) * fx
    lo52       = rv.get("52w_low", 0) * fx

    # Signal colour
    if "Discount" in signal:
        sig_fg, sig_bg, sig_bd = "#0D7A4E", "#ECFDF5", "#6EE7B7"
    elif "Premium" in signal:
        sig_fg, sig_bg, sig_bd = "#A62020", "#FEF2F2", "#FECACA"
    else:
        sig_fg, sig_bg, sig_bd = "#B8972A", "#FFFBEB", "#FDE68A"

    # ── Banner: non-DCF notice ─────────────────────────────────
    st.info(
        f"**{name}** ({ticker}) is in the **{sector_lbl}** sector. "
        "DCF valuation is not applicable for this company type. "
        "Showing relative valuation vs sector medians instead.",
        icon="ℹ️",
    )

    # ── Header row ────────────────────────────────────────────
    c1, c2, c3 = st.columns([2, 1, 1])
    with c1:
        st.markdown(f"## {name}")
        st.caption(f"{ticker} · {sector_lbl}")
    with c2:
        st.markdown(
            f"<div style='text-align:center;padding:8px 0'>"
            f"<div style='font-size:0.8rem;color:#6B7280'>Price</div>"
            f"<div style='font-size:1.6rem;font-weight:700'>{sym}{price:,.2f}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )
    with c3:
        st.markdown(
            f"<div style='text-align:center;padding:4px 8px;"
            f"background:{sig_bg};border:1px solid {sig_bd};"
            f"border-radius:8px;color:{sig_fg};font-weight:700'>"
            f"{signal}<br>"
            f"<span style='font-size:0.75rem;font-weight:400'>"
            f"Avg {avg_pct:+.1f}% vs sector</span></div>",
            unsafe_allow_html=True,
        )

    st.divider()

    # ── Metrics table ─────────────────────────────────────────
    st.markdown("#### Valuation Multiples vs Sector Median")
    if metrics:
        import pandas as _pd2
        rows = []
        for m in metrics:
            cv = m.get("company_val")
            sm = m.get("sector_median")
            pd_ = m.get("pct_diff")
            label = m.get("label", "")
            arrow = "▲" if (pd_ or 0) > 0 else "▼"
            color = "#A62020" if (pd_ or 0) > 15 else "#0D7A4E" if (pd_ or 0) < -15 else "#B8972A"
            rows.append({
                "Metric":         m.get("metric_label", m.get("metric", "")),
                "Company":        f"{cv:.1f}x" if cv is not None else "—",
                "Sector Median":  f"{sm:.1f}x" if sm is not None else "—",
                "vs Sector":      f"{arrow} {abs(pd_ or 0):.1f}%" if pd_ is not None else "—",
                "_color":         color,
            })
        tdf = _pd2.DataFrame(rows)
        st.dataframe(
            tdf.drop(columns=["_color"]),
            hide_index=True,
            width='stretch',
        )
    else:
        st.warning("Could not retrieve valuation multiples. Data may be unavailable.")

    # ── Key stats ─────────────────────────────────────────────
    st.divider()
    st.markdown("#### Key Statistics")
    ks1, ks2, ks3, ks4 = st.columns(4)
    def _kstat(col, label, val):
        col.metric(label, val)
    _kstat(ks1, "Market Cap", f"{sym}{mkt_cap/1e9:.1f}B" if mkt_cap else "—")
    _kstat(ks2, "Dividend Yield", f"{div_yield*100:.2f}%" if div_yield else "—")
    _kstat(ks3, "Beta", f"{beta:.2f}" if beta else "—")
    _kstat(ks4, "52W Range", f"{sym}{lo52:,.0f} – {sym}{hi52:,.0f}" if hi52 else "—")

    # ── Why no DCF? ───────────────────────────────────────────
    with st.expander("ℹ️ Why no DCF valuation for this stock?"):
        st.markdown(
            f"""
**{sector_lbl}** companies (banks, REITs, insurance firms) are excluded from
DCF analysis for the following structural reasons:

- **Banks & Insurers** — Capital is a raw material, not just funding.
  Free cash flow is not meaningful because debt issuance is core to operations.
  Regulatory capital requirements (Basel III / Solvency II) make FCF projections
  unreliable. Appropriate models: **P/B**, **P/E**, **ROE vs Cost of Equity**.

- **REITs** — Earnings are distorted by non-cash depreciation on real assets
  that typically *appreciate* in value. The relevant metric is **FFO
  (Funds from Operations)**, not FCF. Appropriate models: **P/FFO**, **Cap Rate**,
  **Dividend Discount Model (DDM)**.

YieldIQ instead benchmarks these companies against **sector median multiples**
to identify relative cheapness or richness.
"""
        )


# ══════════════════════════════════════════════════════════════
# DCF REPORT GENERATOR
# ══════════════════════════════════════════════════════════════
def generate_dcf_report(ticker, result_data: dict, scenarios: dict, sym: str) -> bytes:
    """Generate a downloadable text-based DCF report."""
    r = result_data
    lines = [
        "=" * 65,
        f"  YieldIQ Valuation Report — {ticker}",
        f"  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        "=" * 65,
        "",
        "VALUATION SUMMARY",
        "-" * 40,
        f"  Current Price      : {sym}{r.get('price', 0):,.2f}",
        f"  Intrinsic Value    : {sym}{r.get('iv', 0):,.2f}",
        f"  Margin of Safety   : {r.get('mos_pct', 0):.1f}%",
        f"  Signal             : {r.get('signal', '')}",
        f"  WACC Used          : {r.get('wacc', 0):.1%}",
        f"  Terminal Growth    : {r.get('term_g', 0):.1%}",
        "",
        "FUNDAMENTALS",
        "-" * 40,
        f"  Revenue Growth     : {r.get('rev_growth', 0):.1%} p.a.",
        f"  FCF Growth         : {r.get('fcf_growth', 0):.1%} p.a.",
        f"  Operating Margin   : {r.get('op_margin', 0):.1%}",
        f"  Fundamental Grade  : {r.get('fund_grade', 'N/A')} ({r.get('fund_score', 0)}/100)",
        "",
        "MODEL PRICE LEVELS (research only — not investment advice)",
        "-" * 40,
        f"  Model Signal              : {r.get('entry_signal', '')}",
        f"  DCF Discount Threshold    : {sym}{r.get('buy_price', 0):,.2f}",
        f"  DCF Model Estimate        : {sym}{r.get('target_price', 0):,.2f}",
        f"  Model Risk Range          : {sym}{r.get('stop_loss', 0):,.2f}  (-{r.get('sl_pct', 0):.1f}%)",
        f"  Model Upside/Downside Ratio: {r.get('rr_ratio', 0):.1f}x",
        f"  DCF Projection Horizon    : {r.get('holding_period', 'N/A')}",
        "",
        "THREE SCENARIO ANALYSIS",
        "-" * 40,
    ]

    for sname, sdata in scenarios.items():
        lines += [
            f"  {sname}",
            f"    Growth: {sdata['growth']:.1%}  WACC: {sdata['wacc']:.1%}  Terminal g: {sdata['term_g']:.1%}",
            f"    Intrinsic Value: {sym}{sdata['iv']:,.2f}  |  MoS: {sdata['mos_pct']:.1f}%",
            "",
        ]

    lines += [
        "DCF WATERFALL  (raw DCF — before PE blend)",
        "-" * 40,
        f"  PV of FCFs         : {sym}{r.get('sum_pv_fcfs', 0):,.0f}",
        f"  PV Terminal Value  : {sym}{r.get('pv_tv', 0):,.0f}",
        f"  Enterprise Value   : {sym}{r.get('ev', 0):,.0f}",
        f"  Less: Total Debt   : {sym}{r.get('debt', 0):,.0f}",
        f"  Plus: Cash         : {sym}{r.get('cash', 0):,.0f}",
        f"  Equity Value       : {sym}{r.get('equity', 0):,.0f}",
        f"  Shares Outstanding : {r.get('shares', 0)/1e9:.3f}B",
        f"  DCF IV/share       : {sym}{r.get('dcf_only_iv', r.get('iv', 0)):,.2f}",
        f"  PE-blended IV/sh   : {sym}{r.get('iv', 0):,.2f}  ← headline number",
        "",
        "=" * 65,
        "DISCLAIMER: Model output only — not investment advice.",
        "YieldIQ is not a registered investment adviser.",
        "Past model performance does not predict future results.",
        "Always conduct independent research before investing.",
        "=" * 65,
    ]
    return "\n".join(lines).encode("utf-8")


def generate_excel_dcf_model(
    ticker: str,
    enriched: dict,
    dcf_res: dict,
    forecast_result: dict,
    scenarios: dict,
    inv_plan: dict,
    report_data: dict,
    sensitivity_df: pd.DataFrame,
    sym: str,
    to_code: str,
    fx: float,
    wacc: float,
    terminal_g: float,
    forecast_yrs: int,
) -> bytes:
    """
    Generate a professional multi-sheet Excel DCF model.
    Sheets: Summary | DCF Model | FCFF Build | ROIC Analysis |
            P&L | Cash Flow | Balance Sheet | Scenarios |
            Sensitivity | Quality Checks | Historical Charts | Assumptions
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.chart.series import DataPoint
    except ImportError:
        return None

    wb = openpyxl.Workbook()

    # ── Palette ────────────────────────────────────────────────
    DARK_BG  = "0D1424"; BLUE_HDR = "1D4ED8"; GREEN_HDR = "065F46"
    RED_HDR  = "7F1D1D"; AMBER_HDR= "78350F"; CYAN_HDR  = "164E63"
    PURP_HDR = "3B0764"; MID_BG   = "0A1020"; BORDER_C  = "1E2D45"
    TEXT_MAIN= "F1F5F9"; TEXT_DIM = "94A3B8"; TEXT_NUM  = "E2E8F0"
    GREEN_POS= "10B981"; RED_NEG  = "EF4444"; AMBER_VAL = "F59E0B"
    WARN_BG  = "451A03"; OK_BG    = "052E16"

    def _c(s):
        """Convert any hex color to valid 8-char aRGB for openpyxl.
        Handles 6-char, 8-char, 10-char (from +"44" appends), # prefix."""
        if not s: return "FFF1F5F9"
        s = str(s).lstrip("#").upper()
        if len(s) == 8 and s[:2] == "FF": s = s[2:]  # strip FF prefix -> 6-char
        if len(s) == 6:  return "FF" + s    # fully opaque
        if len(s) == 8:  return s            # already aRGB
        if len(s) >= 10: return s[:8]        # e.g. "1D4ED844" extra chars
        return "FFF1F5F9"

    def hf(c): return PatternFill("solid", fgColor=_c(c))
    def tb():
        s = Side(style='thin', color=_c(BORDER_C))
        return Border(left=s, right=s, top=s, bottom=s)
    def hdr_font(bold=True, sz=11, color=TEXT_MAIN):
        return Font(name="Calibri", bold=bold, size=sz, color=_c(color))
    def val_font(bold=False, sz=11, color=TEXT_NUM, mono=False):
        return Font(name="Courier New" if mono else "Calibri", bold=bold, size=sz, color=_c(color))
    def center(): return Alignment(horizontal="center", vertical="center")
    def right():  return Alignment(horizontal="right",  vertical="center")
    def left():   return Alignment(horizontal="left",   vertical="center")
    def wrap_left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def wc(ws, row, col, value, fill=None, font=None, align=None, nf=None, h=22):
        c = ws.cell(row=row, column=col, value=value)
        if fill:  c.fill = hf(fill)
        if font:  c.font = font
        if align: c.alignment = align
        if nf:    c.number_format = nf
        c.border = tb()
        ws.row_dimensions[row].height = h
        return c

    def sec(ws, row, text, cols, fill=BLUE_HDR, sz=12):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        c = ws.cell(row=row, column=1, value=f"  {text}")
        c.fill = hf(fill); c.font = hdr_font(True, sz); c.alignment = left()
        c.border = tb(); ws.row_dimensions[row].height = 24

    def title_row(ws, text, cols, fill=BLUE_HDR, sz=15):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
        c = ws["A1"]; c.value = text; c.fill = hf(fill)
        c.font = Font(name="Calibri", bold=True, size=sz, color="FFFFFFFF")
        c.alignment = center(); ws.row_dimensions[1].height = 38

    def subtitle(ws, text, cols, fill=MID_BG):
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
        c = ws["A2"]; c.value = text; c.fill = hf(fill)
        c.font = Font(name="Calibri", size=10, color=_c(TEXT_DIM))
        c.alignment = center(); ws.row_dimensions[2].height = 18

    # ── Shared data ────────────────────────────────────────────
    projected  = forecast_result["projections"]
    pv_fcfs    = dcf_res.get("pv_fcfs", [])
    gs         = forecast_result.get("growth_schedule", [])
    yr_labels  = [f"Year {i+1}" for i in range(forecast_yrs)]
    income_df  = enriched.get("income_df", pd.DataFrame())
    cf_df      = enriched.get("cf_df",     pd.DataFrame())
    bs_df      = enriched.get("bs_df",     pd.DataFrame())
    pt         = inv_plan["price_targets"]
    hp         = inv_plan["holding_period"]
    fs         = inv_plan["fundamental"]
    price_d    = report_data["price"]
    iv_d       = report_data["iv"]
    mos_pct    = report_data["mos_pct"]
    pv_tv_d    = dcf_res.get("pv_tv", 0) * fx
    ev_d       = dcf_res.get("enterprise_value", 0) * fx
    tv_pct     = dcf_res.get("tv_pct_of_ev", 0)

    # Get historical years from income_df
    hist_years = []
    if not income_df.empty and "year" in income_df.columns:
        hist_years = [str(int(y)) for y in income_df["year"].tolist()]

    # ═══════════════════════════════════════════════════════════
    # SHEET 1 — SUMMARY DASHBOARD
    # ═══════════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = "📊 Summary"
    ws1.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [34, 22, 22, 22, 22]):
        ws1.column_dimensions[col].width = w

    title_row(ws1, f"YieldIQ Valuation Report — {ticker}", 5)
    subtitle(ws1, f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Currency: {to_code}  |  WACC: {wacc:.2%}  |  Terminal g: {terminal_g:.2%}  |  Horizon: {forecast_yrs} yrs", 5)

    sec(ws1, 4, "VALUATION SUMMARY", 5, BLUE_HDR)
    tv_flag = "⚠️  EXCEEDS 70% — HIGH SENSITIVITY" if tv_pct > 0.70 else "✅  Within safe limit"
    val_rows = [
        ("Current Market Price",          f"{sym}{price_d:,.2f}",                  DARK_BG, MID_BG),
        ("Intrinsic Value (Base DCF)",    f"{sym}{iv_d:,.2f}",                     DARK_BG, MID_BG),
        ("Discount to Fair Value",         f"{mos_pct:.1f}%",                       DARK_BG, MID_BG),
        ("Signal",                        report_data["signal"],                   DARK_BG, MID_BG),
        ("Bear Case IV",                  f"{sym}{report_data.get('bear_iv',0):,.2f}", DARK_BG, RED_HDR),
        ("Base Case IV",                  f"{sym}{iv_d:,.2f}",                     DARK_BG, AMBER_HDR),
        ("Bull Case IV",                  f"{sym}{report_data.get('bull_iv',0):,.2f}", DARK_BG, GREEN_HDR),
        ("Terminal Value % of EV",        f"{tv_pct:.1%}  {tv_flag}",             DARK_BG, WARN_BG if tv_pct > 0.70 else OK_BG),
    ]
    for i, (label, value, lbg, vbg) in enumerate(val_rows):
        r = 5 + i
        wc(ws1, r, 1, label, fill=lbg, font=hdr_font(False, 11, TEXT_DIM),  align=left())
        wc(ws1, r, 2, value, fill=vbg, font=val_font(True,  12, TEXT_MAIN, True), align=right())

    sec(ws1, 14, "INVESTMENT ACTION PLAN", 5, GREEN_HDR)
    inv_rows = [
        ("Entry Signal",      pt.get("entry_signal", "")),
        ("Buy Zone Price",    f"{sym}{(pt.get('buy_price') or 0)*fx:,.2f}"),
        ("Target Price",      f"{sym}{(pt.get('target_price') or 0)*fx:,.2f}"),
        ("Stop Loss",         f"{sym}{(pt.get('stop_loss') or 0)*fx:,.2f}  (−{pt.get('sl_pct',0):.1f}%)"),
        ("Risk / Reward",     f"{pt.get('rr_ratio',0):.2f}x"),
        ("Suggested Holding", hp.get("label", "N/A")),
        ("Rationale",         hp.get("rationale","")[:80]),
    ]
    for i, (label, value) in enumerate(inv_rows):
        r = 15 + i
        wc(ws1, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
        wc(ws1, r, 2, value, fill=MID_BG,  font=val_font(True, 11, TEXT_MAIN), align=left())

    sec(ws1, 23, "FUNDAMENTAL STRENGTH", 5, CYAN_HDR)
    fund_rows = [
        ("Overall Grade",     f"{fs.get('grade','N/A')} — {fs.get('score',0)}/100"),
        ("Revenue Growth",    f"{enriched.get('revenue_growth',0)*100:.1f}% p.a."),
        ("Operating Margin",  f"{enriched.get('op_margin',0)*100:.1f}%"),
        ("FCF Growth",        f"{enriched.get('fcf_growth',0)*100:.1f}% p.a."),
        ("FCF Positive",      "Yes" if enriched.get("latest_fcf",0) > 0 else "No"),
    ]
    for i, (label, value) in enumerate(fund_rows):
        r = 24 + i
        wc(ws1, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
        wc(ws1, r, 2, value, fill=MID_BG,  font=val_font(True, 11, TEXT_MAIN), align=left())

    # ═══════════════════════════════════════════════════════════
    # SHEET 2 — DCF MODEL (CORE)
    # ═══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("🧮 DCF Model")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 36
    for i in range(1, forecast_yrs + 3):
        ws2.column_dimensions[get_column_letter(i+1)].width = 15

    title_row(ws2, f"DCF MODEL — {ticker}", forecast_yrs + 2, BLUE_HDR)
    subtitle(ws2, f"WACC: {wacc:.2%}  |  Terminal g: {terminal_g:.2%}  |  Forecast Horizon: {forecast_yrs} years", forecast_yrs + 2)

    # Year headers
    sec(ws2, 4, "FREE CASH FLOW PROJECTIONS", forecast_yrs + 2, BLUE_HDR)
    wc(ws2, 5, 1, "Metric", fill=BLUE_HDR, font=hdr_font(sz=11), align=left())
    for j, lbl in enumerate(yr_labels):
        wc(ws2, 5, j+2, lbl, fill=BLUE_HDR, font=hdr_font(sz=11), align=center())
    wc(ws2, 5, forecast_yrs+2, "Terminal", fill=AMBER_HDR, font=hdr_font(sz=11), align=center())

    term_fcf = forecast_result.get("terminal_fcf_norm", 0) * fx / 1e9
    pv_tv_bn = dcf_res.get("pv_tv", 0) * fx / 1e9

    proj_rows = [
        (f"Projected FCF ({to_code}B)",  [v*fx/1e9 for v in projected],  term_fcf,  "#,##0.00", TEXT_MAIN, True),
        ("YoY Growth Rate (%)",           [g*100 for g in gs],           "",         "0.0",      AMBER_VAL, False),
        (f"PV of FCF ({to_code}B)",       [v*fx/1e9 for v in pv_fcfs],   pv_tv_bn,  "#,##0.00", TEXT_NUM,  False),
    ]
    for k, (label, vals, term_val, nf, clr, bold) in enumerate(proj_rows):
        r = 6 + k
        wc(ws2, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
        for j, v in enumerate(vals):
            wc(ws2, r, j+2, round(float(v), 3) if v != "" else v,
               fill=MID_BG, font=val_font(bold, 11, clr, True), align=right(), nf=nf)
        if term_val != "":
            wc(ws2, r, forecast_yrs+2, round(float(term_val), 3),
               fill=MID_BG, font=val_font(bold, 11, clr, True), align=right(), nf=nf)

    # DCF Waterfall
    sec(ws2, 10, "DCF WATERFALL — BRIDGE TO EQUITY VALUE", forecast_yrs + 2, BLUE_HDR)
    waterfall = [
        (f"Σ PV of FCFs ({to_code}B)",           dcf_res.get("sum_pv_fcfs",0)*fx/1e9,  TEXT_NUM),
        (f"PV of Terminal Value ({to_code}B)",    pv_tv_bn,                              TEXT_NUM),
        (f"Enterprise Value ({to_code}B)",         ev_d/1e9,                             TEXT_MAIN),
        (f"Less: Total Debt ({to_code}B)",         enriched.get("total_debt",0)*fx/1e9, RED_NEG),
        (f"Plus: Cash & Equivalents ({to_code}B)", enriched.get("total_cash",0)*fx/1e9, GREEN_POS),
        (f"Equity Value ({to_code}B)",             dcf_res.get("equity_value",0)*fx/1e9, TEXT_MAIN),
        ("Shares Outstanding (Billions)",          enriched.get("shares",0)/1e9,          TEXT_NUM),
        (f"Intrinsic Value Per Share ({sym})",     iv_d,                                  GREEN_POS if mos_pct > 0 else RED_NEG),
        (f"Current Market Price ({sym})",          price_d,                               AMBER_VAL),
        ("Discount to fair value (%)",                   mos_pct,                               GREEN_POS if mos_pct > 0 else RED_NEG),
    ]
    for i, (label, value, clr) in enumerate(waterfall):
        r = 11 + i
        wc(ws2, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
        wc(ws2, r, 2, round(float(value), 2), fill=MID_BG,
           font=val_font(True, 12, clr, True), align=right(), nf="#,##0.00")

    # ═══════════════════════════════════════════════════════════
    # SHEET 3 — FCFF REINVESTMENT MODEL (NEW)
    # ═══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("🔬 FCFF Build")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 38
    for i in range(1, len(hist_years) + forecast_yrs + 3):
        ws3.column_dimensions[get_column_letter(i+1)].width = 16

    all_labels  = (["Historical: " + y for y in hist_years] +
                   [f"Forecast Y{i+1}" for i in range(forecast_yrs)])
    total_cols  = len(all_labels)

    title_row(ws3, f"FCFF REINVESTMENT BUILD — {ticker}", total_cols + 2, PURP_HDR)
    subtitle(ws3, "FCFF = NOPAT + D&A − CapEx − ΔWorking Capital  |  Professional bottom-up FCF construction", total_cols + 2, MID_BG)

    # Column headers
    wc(ws3, 4, 1, "Line Item", fill=PURP_HDR, font=hdr_font(sz=11), align=left(), h=24)
    for j, lbl in enumerate(all_labels):
        is_hist = j < len(hist_years)
        wc(ws3, 4, j+2, lbl, fill=BLUE_HDR if is_hist else PURP_HDR,
           font=hdr_font(sz=10), align=center(), h=24)

    # Pull historical data from income_df and cf_df
    def get_col(df, col, default=None):
        if df is not None and not df.empty and col in df.columns:
            return df[col].tolist()
        return [default] * (len(df) if df is not None and not df.empty else 0)

    hist_rev   = get_col(income_df, "revenue",        0)
    hist_opinc = get_col(income_df, "operating_income",0)
    hist_ni    = get_col(income_df, "net_income",     0)
    hist_cfo   = get_col(cf_df,    "cfo",             0)
    hist_capex = get_col(cf_df,    "capex",           0)
    hist_fcf   = get_col(cf_df,    "fcf",             0)

    # Estimate NOPAT, reinvestment, ROIC from historical data
    latest_rev    = enriched.get("latest_revenue", 0)
    latest_opinc  = enriched.get("latest_opinc", hist_opinc[-1] if hist_opinc else 0)
    latest_fcf_v  = enriched.get("latest_fcf", 0)
    op_margin     = enriched.get("op_margin", 0)
    rev_growth    = enriched.get("revenue_growth", 0)
    tax_rate      = 0.25  # typical effective tax rate assumption

    # Build NOPAT and reinvestment rate for historical + forecast
    nopat_hist = [v * (1 - tax_rate) * fx / 1e9 for v in hist_opinc]

    # Forecast reinvestment model: project Revenue → EBIT → NOPAT → FCFF
    base_rev = latest_rev * fx / 1e9 if latest_rev else 1.0
    fc_rev, fc_ebit, fc_nopat, fc_da, fc_capex_f, fc_dwc, fc_reinv, fc_fcff = [], [], [], [], [], [], [], []
    growth_fade = [max(rev_growth * np.exp(-0.3 * i), terminal_g) for i in range(1, forecast_yrs+1)]

    running_rev = base_rev
    for i, g in enumerate(growth_fade):
        running_rev = running_rev * (1 + g)
        ebit   = running_rev * op_margin
        nopat  = ebit * (1 - tax_rate)
        da     = running_rev * 0.035          # D&A ~3.5% of revenue
        capex_ = running_rev * 0.045          # CapEx ~4.5% of revenue
        dwc    = running_rev * g * 0.08       # ΔWC ~8% of incremental revenue
        reinv  = capex_ - da + dwc            # net reinvestment
        fcff   = nopat - reinv
        fc_rev.append(round(running_rev, 2)); fc_ebit.append(round(ebit, 3))
        fc_nopat.append(round(nopat, 3));     fc_da.append(round(da, 3))
        fc_capex_f.append(round(capex_, 3));  fc_dwc.append(round(dwc, 3))
        fc_reinv.append(round(reinv, 3));     fc_fcff.append(round(fcff, 3))

    def fcff_row(ws, row, label, hist_vals, fc_vals, fill_h, fill_f, clr, bold=False, nf="#,##0.00"):
        wc(ws, row, 1, label, fill=fill_h if hist_vals else DARK_BG,
           font=hdr_font(bold, 11, TEXT_DIM if not bold else TEXT_MAIN), align=left())
        for j, v in enumerate(hist_vals):
            val = v * fx / 1e9 if v else 0
            wc(ws, row, j+2, round(val, 3), fill=fill_h,
               font=val_font(bold, 11, clr, True), align=right(), nf=nf)
        for j, v in enumerate(fc_vals):
            wc(ws, row, len(hist_vals)+j+2, v, fill=fill_f,
               font=val_font(bold, 11, clr, True), align=right(), nf=nf)

    sec(ws3, 5,  "INCOME BRIDGE", total_cols + 2, PURP_HDR)
    fcff_row(ws3, 6,  f"Revenue ({to_code}B)",           hist_rev,   fc_rev,    BLUE_HDR+"44", PURP_HDR+"44", TEXT_MAIN, True)
    fcff_row(ws3, 7,  "Operating Margin (%)",            [v/r if r else 0 for v, r in zip(hist_opinc, hist_rev)], [op_margin]*forecast_yrs, DARK_BG, DARK_BG, AMBER_VAL, nf="0.0%")
    fcff_row(ws3, 8,  f"Operating Income / EBIT ({to_code}B)", hist_opinc, fc_ebit, MID_BG, MID_BG, TEXT_NUM)
    fcff_row(ws3, 9,  f"NOPAT (after tax @ {tax_rate:.0%}) ({to_code}B)", nopat_hist, fc_nopat, MID_BG, MID_BG, TEXT_NUM)

    sec(ws3, 11, "REINVESTMENT COMPONENTS", total_cols + 2, GREEN_HDR)
    fcff_row(ws3, 12, f"D&A (est. ~3.5% rev) ({to_code}B)",     [0]*len(hist_years), fc_da,     DARK_BG, MID_BG, GREEN_POS)
    fcff_row(ws3, 13, f"Capital Expenditure ({to_code}B)",       hist_capex, fc_capex_f, MID_BG, MID_BG, RED_NEG)
    fcff_row(ws3, 14, f"ΔWorking Capital ({to_code}B)",          [0]*len(hist_years), fc_dwc,    DARK_BG, MID_BG, RED_NEG)
    fcff_row(ws3, 15, f"Net Reinvestment ({to_code}B)",          [0]*len(hist_years), fc_reinv,  DARK_BG, MID_BG, AMBER_VAL)

    sec(ws3, 17, "FCFF DERIVATION  —  NOPAT − Net Reinvestment", total_cols + 2, CYAN_HDR)
    fcff_row(ws3, 18, f"Historical FCF Actual ({to_code}B)",     hist_fcf,   [],        CYAN_HDR+"44", DARK_BG, TEXT_MAIN, True)
    fcff_row(ws3, 19, f"FCFF (Reinvestment Model) ({to_code}B)", [],         fc_fcff,   DARK_BG, CYAN_HDR+"44", GREEN_POS, True)

    sec(ws3, 21, "REINVESTMENT RATE CHECK", total_cols + 2, AMBER_HDR)
    reinv_rates = [round(r/n, 3) if n else 0 for r, n in zip(fc_reinv, fc_nopat)]
    fcff_row(ws3, 22, "Reinvestment Rate (Reinv / NOPAT)",       [], reinv_rates, DARK_BG, MID_BG, AMBER_VAL, nf="0.0%")

    # ═══════════════════════════════════════════════════════════
    # SHEET 4 — ROIC vs WACC (NEW)
    # ═══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("📐 ROIC vs WACC")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 36
    for col in ["B","C","D","E","F","G"]:
        ws4.column_dimensions[col].width = 20

    title_row(ws4, f"ROIC vs WACC — VALUE CREATION ANALYSIS — {ticker}", 7, CYAN_HDR)
    subtitle(ws4, "ROIC > WACC = Value Creation  |  ROIC < WACC = Value Destruction  |  Spread × Invested Capital = Economic Profit", 7)

    # Compute ROIC from historical data
    # ROIC = NOPAT / Invested Capital  (where IC = Total Assets - Non-interest-bearing Current Liabilities)
    # Proxy: IC = Total Equity + Total Debt (book value of capital)
    total_debt_v  = enriched.get("total_debt",  0) * fx / 1e9
    total_cash_v  = enriched.get("total_cash",  0) * fx / 1e9
    shares_v      = enriched.get("shares",      0)
    latest_fcf_bn = enriched.get("latest_fcf",  0) * fx / 1e9

    # Estimate invested capital and ROIC
    latest_rev_bn = latest_rev * fx / 1e9 if latest_rev else 1.0
    nopat_latest  = latest_rev_bn * op_margin * (1 - tax_rate)

    # IC proxy: use EV / Revenue multiple as sanity check
    ev_bn = ev_d / 1e9
    ic_estimate = max(ev_bn * 0.6, total_debt_v + latest_rev_bn * 0.3)  # rough IC proxy
    roic_estimate = nopat_latest / ic_estimate if ic_estimate > 0 else 0
    spread = roic_estimate - wacc
    ep_estimate = spread * ic_estimate  # Economic Profit = Spread × IC

    sec(ws4, 4, "ROIC CALCULATION", 7, CYAN_HDR)
    roic_rows = [
        ("Latest Revenue",                      f"{to_code}B",    f"{latest_rev_bn:,.2f}"),
        ("Operating Margin",                    "%",              f"{op_margin*100:.1f}%"),
        ("EBIT",                                f"{to_code}B",    f"{latest_rev_bn*op_margin:,.2f}"),
        ("Effective Tax Rate (assumed)",        "%",              f"{tax_rate*100:.0f}%"),
        ("NOPAT (Net Operating Profit After Tax)", f"{to_code}B", f"{nopat_latest:,.2f}"),
        ("Invested Capital (estimated proxy)",  f"{to_code}B",   f"{ic_estimate:,.2f}"),
        ("ROIC",                                "%",              f"{roic_estimate*100:.1f}%"),
    ]
    wc(ws4, 5, 1, "Metric",       fill=CYAN_HDR, font=hdr_font(sz=11), align=left())
    wc(ws4, 5, 2, "Unit",         fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 5, 3, "Value",        fill=CYAN_HDR, font=hdr_font(sz=11), align=right())
    for i, (label, unit, value) in enumerate(roic_rows):
        r = 6 + i
        wc(ws4, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
        wc(ws4, r, 2, unit,  fill=DARK_BG, font=val_font(False, 11, TEXT_DIM), align=center())
        wc(ws4, r, 3, value, fill=MID_BG,  font=val_font(True, 12, TEXT_MAIN, True), align=right())

    sec(ws4, 14, "ROIC vs WACC SPREAD — VALUE CREATION TEST", 7, CYAN_HDR)
    spread_ok = roic_estimate > wacc
    spread_rows = [
        ("ROIC",                  f"{roic_estimate*100:.2f}%",  GREEN_POS if spread_ok else RED_NEG),
        ("WACC",                  f"{wacc*100:.2f}%",           AMBER_VAL),
        ("Value Creation Spread (ROIC − WACC)",
                                  f"{spread*100:+.2f}%",        GREEN_POS if spread_ok else RED_NEG),
        ("Economic Profit (Spread × IC)",
                                  f"{to_code}B {ep_estimate:,.2f}", GREEN_POS if spread_ok else RED_NEG),
        ("Verdict",
                                  "✅ ROIC > WACC — Growth creates value" if spread_ok
                                  else "⚠️ ROIC < WACC — Growth destroys value",
                                  GREEN_POS if spread_ok else RED_NEG),
    ]
    wc(ws4, 15, 1, "Metric",      fill=CYAN_HDR, font=hdr_font(sz=11), align=left())
    wc(ws4, 15, 2, "Value",       fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 15, 3, "Interpretation", fill=CYAN_HDR, font=hdr_font(sz=11), align=left())
    interp = {
        "ROIC":          "Return earned on every unit of capital deployed",
        "WACC":          "Minimum return required by capital providers",
        "Value Creation Spread (ROIC − WACC)": "Positive = value creation; Negative = destruction",
        "Economic Profit (Spread × IC)": "Total economic value added (or destroyed) per year",
        "Verdict":       "Key quality test — positive spread supports DCF growth assumptions",
    }
    for i, (label, value, clr) in enumerate(spread_rows):
        r = 16 + i
        bg = OK_BG if spread_ok and i < 4 else (WARN_BG if not spread_ok and i < 4 else DARK_BG)
        wc(ws4, r, 1, label,  fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
        wc(ws4, r, 2, value,  fill=bg,      font=val_font(True, 12, clr, True),  align=center())
        wc(ws4, r, 3, interp.get(label,""), fill=DARK_BG, font=Font(name="Calibri", italic=True, size=10, color=TEXT_DIM), align=wrap_left())

    # Forecast ROIC trend
    sec(ws4, 23, "FORECAST ROIC TREND", 7, CYAN_HDR)
    wc(ws4, 24, 1, "Year",           fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 24, 2, f"Revenue ({to_code}B)", fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 24, 3, f"NOPAT ({to_code}B)",  fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 24, 4, "ROIC (%)",       fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 24, 5, "WACC (%)",       fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws4, 24, 6, "Spread (%)",     fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    for i, (rev, nopat, fcff_v) in enumerate(zip(fc_rev, fc_nopat, fc_fcff)):
        r = 25 + i
        fc_roic   = nopat / ic_estimate if ic_estimate > 0 else 0
        fc_spread = fc_roic - wacc
        wc(ws4, r, 1, yr_labels[i], fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=center())
        wc(ws4, r, 2, rev,          fill=MID_BG,  font=val_font(False, 11, TEXT_NUM, True), align=right(), nf="#,##0.00")
        wc(ws4, r, 3, nopat,        fill=MID_BG,  font=val_font(False, 11, TEXT_NUM, True), align=right(), nf="#,##0.00")
        wc(ws4, r, 4, f"{fc_roic*100:.1f}%",   fill=MID_BG, font=val_font(True, 11, GREEN_POS if fc_roic > wacc else RED_NEG, True), align=right())
        wc(ws4, r, 5, f"{wacc*100:.1f}%",       fill=MID_BG, font=val_font(False, 11, AMBER_VAL, True), align=right())
        wc(ws4, r, 6, f"{fc_spread*100:+.1f}%", fill=OK_BG if fc_spread > 0 else WARN_BG,
           font=val_font(True, 11, GREEN_POS if fc_spread > 0 else RED_NEG, True), align=right())

    # ═══════════════════════════════════════════════════════════
    # SHEET 5 — INCOME STATEMENT
    # ═══════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("📋 Income Statement")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 36
    for j, yr in enumerate(hist_years):
        ws5.column_dimensions[get_column_letter(j+2)].width = 18

    def fin_sheet_build(ws, title_text, df, rows_cfg, accent):
        title_row(ws, title_text, max(len(hist_years)+1, 5), accent)
        if df is None or df.empty:
            ws.cell(row=3, column=1).value = "No data available"
            return
        wc(ws, 4, 1, "Line Item", fill=accent, font=hdr_font(sz=11), align=left(), h=22)
        for j, yr in enumerate(hist_years):
            wc(ws, 4, j+2, yr, fill=accent, font=hdr_font(sz=11), align=center(), h=22)

        for k, (label, col, is_pct, bold, is_sec) in enumerate(rows_cfg):
            r = 5 + k
            if is_sec:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(hist_years)+1)
                c = ws.cell(row=r, column=1, value=f"  {label}")
                c.fill = hf(accent); c.font = hdr_font(True, 11); c.alignment = left()
                c.border = tb(); ws.row_dimensions[r].height = 22
            else:
                wc(ws, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM if not bold else TEXT_MAIN), align=left())
                if col and col in df.columns:
                    for j, raw in enumerate(df[col].tolist()):
                        if pd.isna(raw) or raw is None:
                            wc(ws, r, j+2, "—", fill=MID_BG, font=val_font(bold, 11, TEXT_DIM, True), align=right())
                        elif is_pct:
                            v = raw * 100 if abs(raw) <= 1 else raw
                            clr = GREEN_POS if v > 0 else RED_NEG
                            wc(ws, r, j+2, f"{v:.1f}%", fill=MID_BG, font=val_font(bold, 11, clr, True), align=right())
                        else:
                            v = raw * fx / 1e9
                            clr = GREEN_POS if v > 0 else (RED_NEG if v < 0 else TEXT_NUM)
                            wc(ws, r, j+2, round(v, 2), fill=MID_BG, font=val_font(bold, 11, clr, True), align=right(), nf="#,##0.00")

    inc_cfg = [
        ("REVENUE & PROFITABILITY", None,              False, True,  True),
        (f"Revenue ({to_code}B)",   "revenue",         False, True,  False),
        (f"Gross Profit ({to_code}B)","gross_profit",  False, False, False),
        (f"Operating Income ({to_code}B)","operating_income",False,True,False),
        (f"Net Income ({to_code}B)","net_income",      False, True,  False),
        ("MARGINS",                 None,              False, True,  True),
        ("Gross Margin",            "gross_margin",    True,  False, False),
        ("Operating Margin",        "op_margin",       True,  True,  False),
        ("Net Margin",              "net_margin",      True,  False, False),
    ]
    fin_sheet_build(ws5, f"INCOME STATEMENT — {ticker}  ({to_code} Billions)", income_df, inc_cfg, BLUE_HDR)

    # ═══════════════════════════════════════════════════════════
    # SHEET 6 — CASH FLOW STATEMENT
    # ═══════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("💰 Cash Flow")
    ws6.sheet_view.showGridLines = False
    ws6.column_dimensions["A"].width = 36
    for j in range(len(hist_years)):
        ws6.column_dimensions[get_column_letter(j+2)].width = 18

    cf_cfg = [
        ("OPERATING ACTIVITIES",       None,      False, True,  True),
        (f"Operating Cash Flow ({to_code}B)","cfo",False,True,  False),
        (f"Capital Expenditure ({to_code}B)","capex",False,False,False),
        ("FREE CASH FLOW",             None,      False, True,  True),
        (f"Free Cash Flow ({to_code}B)","fcf",    False, True,  False),
        ("GROWTH",                     None,      False, True,  True),
        ("FCF YoY Growth",             "fcf_growth",True,False, False),
    ]
    fin_sheet_build(ws6, f"CASH FLOW STATEMENT — {ticker}  ({to_code} Billions)", cf_df, cf_cfg, GREEN_HDR)

    # ═══════════════════════════════════════════════════════════
    # SHEET 7 — BALANCE SHEET
    # ═══════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("🏦 Balance Sheet")
    ws7.sheet_view.showGridLines = False
    ws7.column_dimensions["A"].width = 36
    for j in range(len(hist_years)):
        ws7.column_dimensions[get_column_letter(j+2)].width = 18

    if bs_df is not None and not bs_df.empty:
        bs_cfg = [
            ("ASSETS",                              None,           False, True,  True),
            (f"Total Assets ({to_code}B)",          "total_assets", False, True,  False),
            (f"Cash & Equivalents ({to_code}B)",    "cash",         False, False, False),
            (f"Current Assets ({to_code}B)",        "current_assets",False,False, False),
            ("LIABILITIES",                         None,           False, True,  True),
            (f"Total Debt ({to_code}B)",            "total_debt",   False, True,  False),
            (f"Current Liabilities ({to_code}B)",   "current_liab", False, False, False),
            ("EQUITY",                              None,           False, True,  True),
            (f"Shareholders' Equity ({to_code}B)",  "equity",       False, True,  False),
            ("Debt / Equity Ratio",                 "de_ratio",     False, False, False),
            ("Current Ratio",                       "current_ratio",False, False, False),
        ]
        fin_sheet_build(ws7, f"BALANCE SHEET — {ticker}  ({to_code} Billions)", bs_df, bs_cfg, CYAN_HDR)
    else:
        title_row(ws7, f"BALANCE SHEET SNAPSHOT — {ticker}", 4, CYAN_HDR)
        for i, (label, val) in enumerate([
            (f"Cash ({to_code}B)", enriched.get("total_cash",0)*fx/1e9),
            (f"Total Debt ({to_code}B)", enriched.get("total_debt",0)*fx/1e9),
        ]):
            wc(ws7, 4+i, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
            wc(ws7, 4+i, 2, round(val,2), fill=MID_BG, font=val_font(True,11,TEXT_MAIN,True), align=right(), nf="#,##0.00")

    # ═══════════════════════════════════════════════════════════
    # SHEET 8 — SCENARIO ANALYSIS
    # ═══════════════════════════════════════════════════════════
    ws8 = wb.create_sheet("🎭 Scenarios")
    ws8.sheet_view.showGridLines = False
    ws8.column_dimensions["A"].width = 30
    for col, w in zip(["B","C","D"], [22,22,22]):
        ws8.column_dimensions[col].width = w

    title_row(ws8, f"BEAR / BASE / BULL SCENARIO ANALYSIS — {ticker}", 4, AMBER_HDR)
    subtitle(ws8, "Each scenario adjusts FCF growth, WACC, and terminal growth to stress-test the valuation", 4)

    bear = scenarios.get("Bear 🐻", {}); base_sc = scenarios.get("Base 📊", {}); bull = scenarios.get("Bull 🐂", {})
    for j, (hdr, clr) in enumerate(zip(["Metric", "Bear 🐻", "Base 📊", "Bull 🐂"], [BLUE_HDR, RED_HDR, AMBER_HDR, GREEN_HDR])):
        wc(ws8, 4, j+1, hdr, fill=clr, font=hdr_font(sz=12), align=center(), h=24)

    sc_data = [
        ("ASSUMPTIONS", True, None, None, None),
        ("FCF Growth Rate",     False, f"{bear.get('growth',0):.1%}", f"{base_sc.get('growth',0):.1%}", f"{bull.get('growth',0):.1%}"),
        ("WACC",                False, f"{bear.get('wacc',0):.1%}",   f"{base_sc.get('wacc',0):.1%}",   f"{bull.get('wacc',0):.1%}"),
        ("Terminal Growth",     False, f"{bear.get('term_g',0):.1%}", f"{base_sc.get('term_g',0):.1%}", f"{bull.get('term_g',0):.1%}"),
        ("OUTPUTS", True, None, None, None),
        (f"Intrinsic Value ({sym})", False, f"{sym}{bear.get('iv',0)*fx:,.2f}", f"{sym}{base_sc.get('iv',0)*fx:,.2f}", f"{sym}{bull.get('iv',0)*fx:,.2f}"),
        ("Discount to fair value", False, f"{bear.get('mos_pct',0):+.1f}%", f"{base_sc.get('mos_pct',0):+.1f}%", f"{bull.get('mos_pct',0):+.1f}%"),
        (f"Current Price ({sym})", False, f"{sym}{price_d:,.2f}", f"{sym}{price_d:,.2f}", f"{sym}{price_d:,.2f}"),
    ]
    for i, row_d in enumerate(sc_data):
        r = 5 + i
        label, is_sec, *vals = row_d
        if is_sec:
            ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            c = ws8.cell(row=r, column=1, value=f"  {label}")
            c.fill = hf(AMBER_HDR); c.font = hdr_font(True, 11); c.alignment = left()
            c.border = tb(); ws8.row_dimensions[r].height = 22
        else:
            wc(ws8, r, 1, label, fill=DARK_BG, font=hdr_font(False, 11, TEXT_DIM), align=left())
            for j, (v, clr) in enumerate(zip(vals, [RED_NEG, TEXT_MAIN, GREEN_POS])):
                wc(ws8, r, j+2, v, fill=MID_BG, font=val_font(True, 11, clr, True), align=center())

    # ═══════════════════════════════════════════════════════════
    # SHEET 9 — SENSITIVITY HEATMAP
    # ═══════════════════════════════════════════════════════════
    ws9 = wb.create_sheet("🔥 Sensitivity")
    ws9.sheet_view.showGridLines = False
    ws9.column_dimensions["A"].width = 14
    for j in range(len(sensitivity_df.columns)):
        ws9.column_dimensions[get_column_letter(j+2)].width = 14

    title_row(ws9, f"SENSITIVITY — {ticker}  |  IV/share ({sym})  |  WACC × Terminal Growth", len(sensitivity_df.columns)+2, CYAN_HDR)
    ws9.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(sensitivity_df.columns)+2)
    c = ws9["A2"]; c.value = f"★  Current Market Price: {sym}{price_d:,.2f}  |  Green = Undervalued vs Price  |  Red = Overvalued"
    c.fill = hf(AMBER_HDR); c.font = Font(name="Calibri", size=11, color="FFFFFF"); c.alignment = center()
    ws9.row_dimensions[2].height = 22

    wc(ws9, 4, 1, "WACC \\ g →", fill=BLUE_HDR, font=hdr_font(sz=11), align=center(), h=24)
    for j, col_name in enumerate(sensitivity_df.columns):
        wc(ws9, 4, j+2, col_name, fill=BLUE_HDR, font=hdr_font(sz=11), align=center(), h=24)
    for i, (idx, row) in enumerate(sensitivity_df.iterrows()):
        r = 5 + i
        wc(ws9, r, 1, str(idx), fill=BLUE_HDR, font=hdr_font(sz=11), align=center())
        for j, val in enumerate(row):
            v = val * fx if pd.notna(val) else 0
            is_green = v > price_d
            wc(ws9, r, j+2, round(v, 2), fill="0D4A2A" if is_green else "4A0D0D",
               font=val_font(True, 11, GREEN_POS if is_green else RED_NEG, True),
               align=right(), nf=f'"{sym}"#,##0.00')

    # ═══════════════════════════════════════════════════════════
    # SHEET 10 — QUALITY CHECKS (TV WEIGHT + SANITY MULTIPLES)
    # ═══════════════════════════════════════════════════════════
    ws10 = wb.create_sheet("✅ Quality Checks")
    ws10.sheet_view.showGridLines = False
    ws10.column_dimensions["A"].width = 38
    ws10.column_dimensions["B"].width = 22
    ws10.column_dimensions["C"].width = 22
    ws10.column_dimensions["D"].width = 36

    title_row(ws10, f"DCF QUALITY CHECKS — {ticker}", 4, RED_HDR)
    subtitle(ws10, "Sanity tests every analyst should run before trusting a DCF output", 4)

    # 1. Terminal Value Weight
    sec(ws10, 4, "CHECK 1 — TERMINAL VALUE WEIGHT", 4, BLUE_HDR)
    tv_safe = tv_pct <= 0.70
    tv_checks = [
        ("PV of Forecast FCFs",             f"{sym}{dcf_res.get('sum_pv_fcfs',0)*fx/1e9:,.2f}B",  "—"),
        ("PV of Terminal Value",             f"{sym}{pv_tv_bn:,.2f}B",                             "—"),
        ("Enterprise Value",                 f"{sym}{ev_d/1e9:,.2f}B",                            "—"),
        ("Terminal Value as % of EV",        f"{tv_pct:.1%}",
         "✅ < 70% — Healthy" if tv_safe else "⚠️ > 70% — Over-reliant on terminal assumptions"),
        ("Threshold (Professional Standard)", "70%",                                               "Flag if exceeded"),
    ]
    wc(ws10, 5, 1, "Metric",      fill=BLUE_HDR, font=hdr_font(sz=11), align=left())
    wc(ws10, 5, 2, "Value",       fill=BLUE_HDR, font=hdr_font(sz=11), align=center())
    wc(ws10, 5, 3, "Status / Note", fill=BLUE_HDR, font=hdr_font(sz=11), align=left())
    for i, (label, value, note) in enumerate(tv_checks):
        r = 6 + i
        flag_bg = (OK_BG if tv_safe else WARN_BG) if "Terminal Value as %" in label else DARK_BG
        wc(ws10, r, 1, label, fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=left())
        wc(ws10, r, 2, value, fill=MID_BG,  font=val_font(True,11,TEXT_MAIN,True), align=right())
        wc(ws10, r, 3, note,  fill=flag_bg, font=Font(name="Calibri",bold=True,size=11,
           color=GREEN_POS if tv_safe else RED_NEG), align=left())

    # 2. Implied Multiples Sanity Check
    latest_ni  = (hist_ni[-1]  if hist_ni  else 0) * fx / 1e9
    latest_rev_check = (hist_rev[-1] if hist_rev else 0) * fx / 1e9
    latest_ebitda = latest_rev_check * (op_margin + 0.035)  # add back D&A est.
    latest_opinc_check = latest_rev_check * op_margin

    equity_val_d = dcf_res.get("equity_value", 0) * fx / 1e9
    mkt_cap_d    = price_d * shares_v / 1e9

    implied_pe        = (mkt_cap_d / latest_ni)     if latest_ni     > 0 else 0
    dcf_implied_pe    = (equity_val_d / latest_ni)  if latest_ni     > 0 else 0
    implied_ev_ebitda = (ev_d/1e9 / latest_ebitda)  if latest_ebitda > 0 else 0
    implied_ev_rev    = (ev_d/1e9 / latest_rev_check) if latest_rev_check > 0 else 0

    def mult_flag(label, val):
        if "P/E" in label:
            if val <= 0:   return "N/A (negative earnings)", TEXT_DIM,  DARK_BG
            if val > 60:   return f"⚠️ {val:.1f}x — Very expensive; check model",  RED_NEG,  WARN_BG
            if val > 35:   return f"⚡ {val:.1f}x — High; growth stock territory",  AMBER_VAL, DARK_BG
            return         f"✅ {val:.1f}x — Reasonable",                           GREEN_POS, OK_BG
        if "EV/EBITDA" in label:
            if val <= 0:   return "N/A",                                TEXT_DIM,   DARK_BG
            if val > 30:   return f"⚠️ {val:.1f}x — Very high",        RED_NEG,    WARN_BG
            if val > 15:   return f"⚡ {val:.1f}x — Premium",           AMBER_VAL,  DARK_BG
            return         f"✅ {val:.1f}x — Reasonable",               GREEN_POS,  OK_BG
        if "EV/Rev" in label:
            if val > 10:   return f"⚠️ {val:.1f}x — Elevated",         RED_NEG,    WARN_BG
            return         f"✅ {val:.1f}x — OK",                       GREEN_POS,  OK_BG
        return f"{val:.1f}x", TEXT_MAIN, DARK_BG

    sec(ws10, 13, "CHECK 2 — DCF SANITY MULTIPLES  (DCF Implied vs Market)", 4, RED_HDR)
    wc(ws10, 14, 1, "Multiple",        fill=RED_HDR, font=hdr_font(sz=11), align=left())
    wc(ws10, 14, 2, "Market (Current)",fill=RED_HDR, font=hdr_font(sz=11), align=center())
    wc(ws10, 14, 3, "DCF Implied",     fill=RED_HDR, font=hdr_font(sz=11), align=center())
    wc(ws10, 14, 4, "Sanity Flag",     fill=RED_HDR, font=hdr_font(sz=11), align=left())

    mult_rows = [
        ("Price / Earnings (P/E)",      mkt_cap_d / latest_ni   if latest_ni > 0 else 0,  dcf_implied_pe),
        ("EV / EBITDA",                 ev_d/1e9 / latest_ebitda if latest_ebitda > 0 else 0, implied_ev_ebitda),
        ("EV / Revenue",                ev_d/1e9 / latest_rev_check if latest_rev_check > 0 else 0, implied_ev_rev),
        ("Price / Book (proxy)",        mkt_cap_d / max(equity_val_d * 0.6, 0.01), equity_val_d / max(equity_val_d * 0.6, 0.01)),
    ]
    for i, (mult_name, mkt_val, dcf_val) in enumerate(mult_rows):
        r = 15 + i
        flag_text, flag_color, flag_bg = mult_flag(mult_name, dcf_val)
        wc(ws10, r, 1, mult_name,                fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=left())
        wc(ws10, r, 2, f"{mkt_val:.1f}x" if mkt_val else "N/A", fill=MID_BG, font=val_font(True,11,TEXT_NUM,True), align=right())
        wc(ws10, r, 3, f"{dcf_val:.1f}x" if dcf_val else "N/A", fill=MID_BG, font=val_font(True,12,TEXT_MAIN,True), align=right())
        wc(ws10, r, 4, flag_text, fill=flag_bg, font=Font(name="Calibri",bold=True,size=11,color=flag_color), align=left())

    # 3. MoS Reliability check
    sec(ws10, 21, "CHECK 3 — MODEL RELIABILITY SCORECARD", 4, CYAN_HDR)
    reliable = forecast_result.get("reliable", True)
    suspicious = dcf_res.get("suspicious", False)
    quality_checks = [
        ("DCF Reliable Flag",         "✅ Reliable" if reliable else "⚠️ Unreliable",    reliable),
        ("IV Hard Cap Triggered",     "⚠️ Yes — IV was capped at 5× price" if suspicious else "✅ No — IV within bounds", not suspicious),
        ("Terminal Value < 70% EV",   "✅ Pass" if tv_pct <= 0.70 else f"⚠️ Fail — {tv_pct:.0%}", tv_pct <= 0.70),
        ("ROIC > WACC",               "✅ Value Creation" if roic_estimate > wacc else "⚠️ Value Destruction", roic_estimate > wacc),
        ("FCF Positive",              "✅ Yes" if enriched.get("latest_fcf",0) > 0 else "⚠️ Negative FCF", enriched.get("latest_fcf",0) > 0),
        ("Op Margin > 8%",            "✅ Pass" if enriched.get("op_margin",0) > 0.08 else "⚠️ Below threshold", enriched.get("op_margin",0) > 0.08),
        ("DCF P/E Sanity",            "✅ Reasonable" if 0 < dcf_implied_pe < 60 else ("⚠️ Suspicious" if dcf_implied_pe >= 60 else "N/A"), 0 < dcf_implied_pe < 60),
    ]
    wc(ws10, 22, 1, "Check",   fill=CYAN_HDR, font=hdr_font(sz=11), align=left())
    wc(ws10, 22, 2, "Result",  fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    for i, (check, result, passed) in enumerate(quality_checks):
        r = 23 + i
        wc(ws10, r, 1, check,  fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=left())
        wc(ws10, r, 2, result, fill=OK_BG if passed else WARN_BG,
           font=Font(name="Calibri",bold=True,size=11,color=GREEN_POS if passed else RED_NEG), align=left())

    # ═══════════════════════════════════════════════════════════
    # SHEET 11 — HISTORICAL CHARTS
    # ═══════════════════════════════════════════════════════════
    ws11 = wb.create_sheet("📈 Historical Charts")
    ws11.sheet_view.showGridLines = False

    title_row(ws11, f"HISTORICAL TRENDS — {ticker}", 10, PURP_HDR)
    subtitle(ws11, "Revenue Growth  |  Operating Margin  |  Free Cash Flow Trend", 10)

    # Write data tables for charts
    # Table 1: Revenue
    sec(ws11, 4, "REVENUE TREND", 10, BLUE_HDR)
    wc(ws11, 5, 1, "Year",              fill=BLUE_HDR, font=hdr_font(sz=11), align=center())
    wc(ws11, 5, 2, f"Revenue ({to_code}B)", fill=BLUE_HDR, font=hdr_font(sz=11), align=center())
    wc(ws11, 5, 3, "YoY Growth %",     fill=BLUE_HDR, font=hdr_font(sz=11), align=center())

    rev_list = []
    if not income_df.empty and "revenue" in income_df.columns:
        rev_list = [(str(int(y)), v*fx/1e9) for y, v in zip(income_df["year"], income_df["revenue"]) if pd.notna(v)]

    for i, (yr, v) in enumerate(rev_list):
        r = 6 + i
        prev_v = rev_list[i-1][1] if i > 0 else v
        growth = (v - prev_v) / prev_v * 100 if prev_v > 0 and i > 0 else 0
        wc(ws11, r, 1, yr,         fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=center())
        wc(ws11, r, 2, round(v,2), fill=MID_BG,  font=val_font(True,11,TEXT_MAIN,True), align=right(), nf="#,##0.00")
        wc(ws11, r, 3, f"{growth:.1f}%", fill=MID_BG,
           font=val_font(True,11,GREEN_POS if growth > 0 else RED_NEG,True), align=right())

    # Table 2: Operating Margin
    sec(ws11, 6 + len(rev_list) + 1, "OPERATING MARGIN TREND", 10, GREEN_HDR)
    margin_start = 6 + len(rev_list) + 2
    wc(ws11, margin_start, 1, "Year",             fill=GREEN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws11, margin_start, 2, "Operating Margin", fill=GREEN_HDR, font=hdr_font(sz=11), align=center())

    margin_list = []
    if not income_df.empty and "op_margin" in income_df.columns:
        margin_list = [(str(int(y)), v*100) for y, v in zip(income_df["year"], income_df["op_margin"]) if pd.notna(v)]
    elif not income_df.empty and "operating_income" in income_df.columns and "revenue" in income_df.columns:
        margin_list = [(str(int(y)), (oi/rev*100 if rev else 0))
                       for y, oi, rev in zip(income_df["year"], income_df["operating_income"], income_df["revenue"])
                       if pd.notna(oi) and pd.notna(rev)]

    for i, (yr, v) in enumerate(margin_list):
        r = margin_start + 1 + i
        wc(ws11, r, 1, yr,             fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=center())
        wc(ws11, r, 2, f"{v:.1f}%",   fill=MID_BG,  font=val_font(True,11,GREEN_POS if v > 0 else RED_NEG,True), align=right())

    # Table 3: FCF Trend
    fcf_sec_row = margin_start + len(margin_list) + 2
    sec(ws11, fcf_sec_row, "FREE CASH FLOW TREND", 10, CYAN_HDR)
    fcf_hdr_row = fcf_sec_row + 1
    wc(ws11, fcf_hdr_row, 1, "Year",              fill=CYAN_HDR, font=hdr_font(sz=11), align=center())
    wc(ws11, fcf_hdr_row, 2, f"FCF ({to_code}B)", fill=CYAN_HDR, font=hdr_font(sz=11), align=center())

    fcf_list = []
    if not cf_df.empty and "fcf" in cf_df.columns:
        fcf_list = [(str(int(y)), v*fx/1e9) for y, v in zip(cf_df["year"], cf_df["fcf"]) if pd.notna(v) and abs(v) > 1e6]

    for i, (yr, v) in enumerate(fcf_list):
        r = fcf_hdr_row + 1 + i
        wc(ws11, r, 1, yr,         fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=center())
        wc(ws11, r, 2, round(v,2), fill=MID_BG,  font=val_font(True,11,GREEN_POS if v > 0 else RED_NEG,True), align=right(), nf="#,##0.00")

    # Add bar charts if data available
    chart_col = 5  # Charts go in column E

    if len(rev_list) >= 2:
        chart1 = BarChart()
        chart1.type = "col"; chart1.grouping = "clustered"
        chart1.title = f"Revenue Trend — {ticker}"
        chart1.y_axis.title = f"{to_code}B"; chart1.x_axis.title = "Year"
        chart1.style = 10; chart1.width = 18; chart1.height = 12
        data_ref  = Reference(ws11, min_col=2, min_row=5, max_row=5+len(rev_list))
        cat_ref   = Reference(ws11, min_col=1, min_row=6, max_row=5+len(rev_list))
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cat_ref)
        chart1.series[0].graphicalProperties.solidFill = "1D4ED8"
        ws11.add_chart(chart1, f"E4")

    if len(margin_list) >= 2:
        chart2 = LineChart()
        chart2.title = f"Operating Margin — {ticker}"
        chart2.y_axis.title = "%"; chart2.x_axis.title = "Year"
        chart2.style = 10; chart2.width = 18; chart2.height = 12
        data_ref2 = Reference(ws11, min_col=2, min_row=margin_start, max_row=margin_start+len(margin_list))
        cat_ref2  = Reference(ws11, min_col=1, min_row=margin_start+1, max_row=margin_start+len(margin_list))
        chart2.add_data(data_ref2, titles_from_data=True)
        chart2.set_categories(cat_ref2)
        chart2.series[0].graphicalProperties.line.solidFill = "10B981"
        chart2.series[0].graphicalProperties.line.width = 25000
        ws11.add_chart(chart2, f"E24")

    if len(fcf_list) >= 2:
        chart3 = BarChart()
        chart3.type = "col"; chart3.grouping = "clustered"
        chart3.title = f"Free Cash Flow — {ticker}"
        chart3.y_axis.title = f"{to_code}B"; chart3.x_axis.title = "Year"
        chart3.style = 10; chart3.width = 18; chart3.height = 12
        data_ref3 = Reference(ws11, min_col=2, min_row=fcf_hdr_row, max_row=fcf_hdr_row+len(fcf_list))
        cat_ref3  = Reference(ws11, min_col=1, min_row=fcf_hdr_row+1, max_row=fcf_hdr_row+len(fcf_list))
        chart3.add_data(data_ref3, titles_from_data=True)
        chart3.set_categories(cat_ref3)
        chart3.series[0].graphicalProperties.solidFill = "06B6D4"
        ws11.add_chart(chart3, f"E44")

    # ═══════════════════════════════════════════════════════════
    # SHEET 12 — KEY ASSUMPTIONS
    # ═══════════════════════════════════════════════════════════
    ws12 = wb.create_sheet("⚙️ Assumptions")
    ws12.sheet_view.showGridLines = False
    ws12.column_dimensions["A"].width = 36
    ws12.column_dimensions["B"].width = 28
    ws12.column_dimensions["C"].width = 50

    title_row(ws12, f"KEY MODEL ASSUMPTIONS — {ticker}", 3, BLUE_HDR)
    wc(ws12, 4, 1, "Parameter",   fill=BLUE_HDR, font=hdr_font(sz=12), align=left())
    wc(ws12, 4, 2, "Value Used",  fill=BLUE_HDR, font=hdr_font(sz=12), align=center())
    wc(ws12, 4, 3, "Explanation", fill=BLUE_HDR, font=hdr_font(sz=12), align=left())

    assump_rows = [
        (True,  "DISCOUNT RATE",         None,   None),
        (False, "WACC",                  f"{wacc:.2%}", "Weighted Average Cost of Capital — auto CAPM or manual"),
        (False, "Cost of Equity",        "Auto CAPM", "Re = Rf + β×(Rm−Rf)"),
        (False, "Risk-Free Rate",        f"{st.session_state.get('_rf_rate_info', {}).get('rate_pct', 0):.2f}% ({('India 10Y' if st.session_state.get('_rf_rate_info', {}).get('market') == 'india' else 'US 10Y')}) — {st.session_state.get('_rf_rate_info', {}).get('source', 'fallback')}", "10-yr government bond yield"),
        (True,  "GROWTH ASSUMPTIONS",    None,   None),
        (False, "Base FCF Growth",       f"{forecast_result.get('base_growth',0):.2%}", "Historical CAGR with exponential fade to terminal g"),
        (False, "Long-run growth rate",  f"{terminal_g:.2%}", "Long-run GDP growth assumption"),
        (False, "Forecast Horizon",      f"{forecast_yrs} years", "Explicit modelling period"),
        (False, "FCF Base Method",       forecast_result.get("fcf_base_method","Auto"), "How the starting FCF was selected"),
        (True,  "REINVESTMENT MODEL",    None,   None),
        (False, "D&A (% of Revenue)",    "3.5%", "Estimated depreciation & amortisation"),
        (False, "CapEx (% of Revenue)",  "4.5%", "Estimated capital expenditure"),
        (False, "ΔWC (% of ΔRevenue)",  "8.0%", "Working capital investment on incremental revenue"),
        (False, "Tax Rate",              f"{tax_rate:.0%}", "Effective corporate tax rate assumption"),
        (True,  "VALUATION GUARDRAILS",  None,   None),
        (False, "IV Hard Cap",           "5× current price", "Prevents outlier results from bad data"),
        (False, "TV Warning Threshold",  "70% of EV", "Flags terminal-value-heavy models"),
        (False, "MoS Formula",          "(IV − Price) / Price", "Standard margin of safety relative to price"),
        (True,  "QUALITY FILTERS",       None,   None),
        (False, "Min Operating Margin",  "8%",   "Below this → DCF flagged unreliable"),
        (False, "Max FCF Margin",        "30%",  "Suspiciously high FCF → flagged for review"),
        (False, "Microcap Filter",       "₹2000 Cr / $200M", "Small companies excluded from screener"),
        (True,  "DISCLAIMER",            None,   None),
        (False, "Purpose",               "Educational / Research Only", "Not financial advice"),
        (False, "Data Source",           "Yahoo Finance (yfinance)", "Prices and financials via yfinance API"),
    ]
    for i, (is_sec, label, value, expl) in enumerate(assump_rows):
        r = 5 + i
        if is_sec:
            ws12.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            c = ws12.cell(row=r, column=1, value=f"  {label}")
            c.fill = hf(BLUE_HDR); c.font = hdr_font(True, 12); c.alignment = left()
            c.border = tb(); ws12.row_dimensions[r].height = 24
        else:
            wc(ws12, r, 1, label,        fill=DARK_BG, font=hdr_font(False,11,TEXT_DIM), align=left())
            wc(ws12, r, 2, value or "",  fill=MID_BG,  font=val_font(True,11,TEXT_MAIN,True), align=center())
            c3 = ws12.cell(row=r, column=3, value=expl or "")
            c3.fill = hf(DARK_BG); c3.font = Font(name="Calibri",size=10,color=TEXT_DIM,italic=True)
            c3.alignment = wrap_left(); c3.border = tb()

    # ── Freeze panes & tab ordering ────────────────────────────
    for ws in [ws1,ws2,ws3,ws4,ws5,ws6,ws7,ws8,ws9,ws10,ws11,ws12]:
        ws.freeze_panes = "B5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



# ══════════════════════════════════════════════════════════════
# SIDEBAR — Bloomberg Terminal style
# ══════════════════════════════════════════════════════════════
_active_main_tab = st.session_state.get("main_tab", "stock")

with st.sidebar:
    # ── Chip CSS (kept inline for sidebar scope) ─────────────
    st.html("""
<style>
.yiq-chip {
  display:inline-block; font-family:'IBM Plex Mono',monospace;
  font-size:11px; font-weight:600; padding:3px 10px;
  background:rgba(255,255,255,0.07); border:1px solid rgba(255,255,255,0.14);
  border-radius:5px; color:#94A3B8; margin:2px; cursor:pointer;
  transition:background 0.15s, color 0.15s;
}
.yiq-chip:hover { background:rgba(0,180,216,0.15); color:#00b4d8;
                  border-color:rgba(0,180,216,0.35); }
</style>
""")

    # ── 1. LOGO / WORDMARK ───────────────────────────────────
    st.html("""
    <div style="padding:14px 4px 14px;">
      <div style="display:flex;align-items:center;gap:10px;">
        <div style="width:36px;height:36px;flex-shrink:0;
                    background:linear-gradient(135deg,#1D4ED8,#06B6D4);
                    border-radius:9px;display:flex;align-items:center;
                    justify-content:center;font-size:18px;
                    box-shadow:0 4px 12px rgba(29,78,216,0.35);">📊</div>
        <div>
          <div style="font-size:17px;font-weight:800;color:#FFFFFF;
                      letter-spacing:-0.02em;line-height:1.1;">YieldIQ</div>
          <div style="font-size:10px;color:#64748B;letter-spacing:0.05em;
                      font-weight:500;margin-top:2px;">Institutional-grade valuation</div>
        </div>
      </div>
      <div style="height:1px;
                  background:linear-gradient(90deg,#1D4ED8,#06B6D4,transparent);
                  margin-top:14px;opacity:0.5;"></div>
    </div>
    """)

    # ── 2. VERTICAL NAV MENU ─────────────────────────────────
    # Groups: None = regular item, "divider" = insert hr before next group
    _NAV_ITEMS = [
        ("🏠", "Morning Brief",  "morning_brief"),
        ("🔍", "Stock Analysis", "stock"),
        None,                                           # ── divider ──
        ("📊", "Financials",     "financials"),
        ("🏭", "Sector Map",     "markets"),
        ("⚖️", "Compare",        "compare"),
        None,                                           # ── divider ──
        ("💼", "Portfolio",      "portfolio"),
        ("📋", "Screener",       "screener"),
        ("📅", "Earnings",       "earnings"),
        None,                                           # ── divider ──
        ("⚙️", "Settings",       "about"),
    ]
    _is_brief = (
        _active_main_tab == "stock"
        and not st.session_state.get("fin_ticker")
    )
    for _nav_item in _NAV_ITEMS:
        if _nav_item is None:
            st.markdown("---")
            continue
        _nav_icon, _nav_label, _nav_key = _nav_item
        _nav_active = (
            (_nav_key == "morning_brief" and _is_brief)
            or (_nav_key != "morning_brief" and _active_main_tab == _nav_key)
            or (_nav_key == "stock" and _active_main_tab == "stock" and not _is_brief)
        )
        if st.button(
            f"{_nav_icon}  {_nav_label}",
            key=f"nav_{_nav_key}",
            use_container_width=True,
            type="primary" if _nav_active else "secondary",
        ):
            if _nav_key == "morning_brief":
                st.session_state["main_tab"] = "stock"
                st.session_state["_show_morning_brief"] = True
            else:
                st.session_state["main_tab"] = _nav_key
                st.session_state.pop("_show_morning_brief", None)
            st.rerun()

    # ── 2b. DARK / LIGHT MODE TOGGLE ─────────────────────────
    st.markdown("---")
    _theme        = st.session_state.get("theme", "light")
    _toggle_label = "🌙  Dark Mode" if _theme == "light" else "☀️  Light Mode"
    if st.button(_toggle_label, key="sb_theme_toggle", use_container_width=True):
        st.session_state["theme"] = "dark" if _theme == "light" else "light"
        st.rerun()

    # ── 3. MARKET PULSE ──────────────────────────────────────
    st.html('<div class="yiq-sb-divider"></div>'
            '<div class="yiq-sb-section-label">Market Pulse</div>')
    _pulse = fetch_market_pulse()
    _pulse_fmt = {
        "S&P 500":   lambda p: f"{p:,.0f}",
        "10Y Yield": lambda p: f"{p:.2f}%",
        "VIX":       lambda p: f"{p:.2f}",
    }
    _pulse_rows = ""
    for _pname, _pdata in _pulse.items():
        _pchg  = _pdata["chg"]
        _pclr  = "yiq-pulse-chg-pos" if _pchg >= 0 else "yiq-pulse-chg-neg"
        _psym  = "▲" if _pchg >= 0 else "▼"
        _pfmt  = _pulse_fmt.get(_pname, lambda p: f"{p:,.2f}")
        _pval  = _pfmt(_pdata["price"]) if _pdata["price"] else "—"
        _pulse_rows += (
            f'<div class="yiq-pulse-row">'
            f'<span class="yiq-pulse-label">{_pname}</span>'
            f'<span class="yiq-pulse-val">{_pval} '
            f'<span class="{_pclr}">{_psym}{abs(_pchg):.2f}%</span></span>'
            f'</div>'
        )
    st.html(f'<div class="yiq-pulse">{_pulse_rows}</div>')

    # ── 4. CONTROLS (currency, WACC, FX) ─────────────────────
    st.html('<div class="yiq-sb-divider"></div>')
    st.html('<div style="font-size:11px;font-weight:700;color:#38BDF8;'
            'letter-spacing:0.1em;text-transform:uppercase;margin-bottom:6px;">💱 Currency</div>')
    cur_key = st.selectbox("Currency", list(CURRENCIES.keys()), index=1,
                           label_visibility="collapsed", key="sb_currency")
    sym     = CURRENCIES[cur_key]["symbol"]
    to_code = CURRENCIES[cur_key]["code"]

    st.html('<div style="height:1px;background:rgba(255,255,255,0.08);margin:8px 0;"></div>')
    st.html('<div style="font-size:11px;font-weight:700;color:#38BDF8;'
            'letter-spacing:0.1em;text-transform:uppercase;margin-bottom:6px;">VIEW MODE</div>')
    _mode_col1, _mode_col2 = st.columns(2)
    with _mode_col1:
        if st.button("📖 Simple", width='stretch',
                     type="primary" if not st.session_state.get("pro_mode", False) else "secondary",
                     key="btn_simple_mode"):
            st.session_state["pro_mode"] = False
            st.rerun()
    with _mode_col2:
        if st.button("⚡ Pro", width='stretch',
                     type="primary" if st.session_state.get("pro_mode", False) else "secondary",
                     key="btn_pro_mode"):
            st.session_state["pro_mode"] = True
            st.rerun()
    pro_mode = st.session_state.get("pro_mode", False)

    with st.expander("⚙️ Model Parameters", expanded=False):
        use_auto_wacc = st.toggle("Auto-calculate required return", value=True,
                                  key="sb_auto_wacc",
                                  help=ob_tooltip("wacc"))
        manual_wacc   = st.slider("Manual required return (%)", 8, 20, 10, 1,
                                  format="%d%%", disabled=use_auto_wacc,
                                  key="sb_manual_wacc",
                                  help=ob_tooltip("wacc"))
        terminal_pct  = st.slider("Long-run growth (%)", 1, 4, 3, 1,
                                  format="%d%%", key="sb_terminal_pct",
                                  help=ob_tooltip("terminal_g"))
        terminal_g    = terminal_pct / 100
        forecast_yrs  = st.slider("Years to forecast", 5, 15, FORECAST_YEARS,
                                  key="sb_forecast_yrs",
                                  help=ob_tooltip("forecast_yrs"))
        _mc_allowed = can("monte_carlo")
        run_mc = st.toggle(
            "Run 1,000 simulations",
            value=False, disabled=not _mc_allowed,
            help="Upgrade to Pro to unlock" if not _mc_allowed
                 else "Monte Carlo: 1,000 valuation scenarios",
            key="sb_run_mc",
        )
        if not _mc_allowed:
            st.html('<div style="font-size:11px;color:#8492a6;margin-top:-6px;">'
                    '🔒 <a href="https://yieldiq.app/pricing.html" target="_blank" '
                    'style="color:#5046e4">Pro feature</a></div>')
        st.html('<div style="height:4px"></div>')
        if st.button("🗑 Clear Cache & Refresh", width='stretch',
                     key="sb_clear_cache"):
            st.cache_data.clear()
            st.rerun()

    if st.session_state.get("_last_to_code") != to_code:
        st.session_state["_fx_rate_usd"] = get_fx_rate("USD", to_code)
        st.session_state["_fx_rate_inr"] = get_fx_rate("INR", to_code)
        st.session_state["_last_to_code"] = to_code
    fx_rate = st.session_state.get("_fx_rate_usd", 1.0)
    fx_inr  = st.session_state.get("_fx_rate_inr", 1.0)

    with st.expander("📡 Live Data & FX", expanded=False):
        st.html(f"""
        <div style="padding:8px 12px;background:rgba(255,255,255,0.04);
                    border-radius:8px;margin-bottom:8px;">
          <div style="display:flex;align-items:center;gap:5px;margin-bottom:6px;">
            <div style="width:6px;height:6px;background:#34D399;border-radius:50%;
                        animation:shimmer 2s ease-in-out infinite;"></div>
            <span style="font-size:11px;color:#34D399;font-weight:700;
                         letter-spacing:0.04em;">LIVE FX</span>
          </div>
          <div style="font-size:11px;color:#94A3B8;
                      font-family:'IBM Plex Mono',monospace;line-height:2;">
            1 USD = <span style="color:#F1F5F9;font-weight:600;">
                    {sym}{fx_rate:,.2f}</span><br>
            1 INR = <span style="color:#F1F5F9;font-weight:600;">
                    {sym}{fx_inr:,.4f}</span>
          </div>
        </div>
        <div style="font-size:11px;color:#64748B;line-height:1.7;">
          Data: Yahoo Finance (yfinance)<br>
          Prices update every 60s
        </div>
        """)

    # ── Portfolio capital (silent) ───────────────────────────
    if "portfolio_capital" not in st.session_state:
        st.session_state["portfolio_capital"] = 10_000_000
    results_file = None

    # ── Recent Tickers ───────────────────────────────────────
    if "recent_tickers" not in st.session_state:
        st.session_state["recent_tickers"] = []
    _recent = st.session_state.get("recent_tickers", [])
    if _recent:
        st.html('<div style="font-size:11px;font-weight:700;color:#38BDF8;'
                'letter-spacing:0.1em;text-transform:uppercase;'
                'margin:10px 0 6px;">📌 Recent</div>')
        _chips_html = '<div style="display:flex;flex-wrap:wrap;gap:4px;margin-bottom:6px;">'
        for _tk in _recent[-5:][::-1]:
            _chips_html += f'<span class="yiq-chip">{_tk}</span>'
        _chips_html += '</div>'
        st.html(_chips_html)
        _btn_cols = st.columns(min(len(_recent[-5:]), 3))
        for _i, _tk in enumerate(_recent[-5:][::-1][:3]):
            with _btn_cols[_i]:
                if st.button(_tk, key=f"recent_{_tk}_{_i}", width='stretch'):
                    st.session_state["_prefill_ticker"] = _tk
                    st.session_state["main_tab"] = "stock"
                    st.rerun()

    # ── 5. USER PROFILE (bottom) ─────────────────────────────
    st.html('<div class="yiq-sb-divider" style="margin-top:12px;"></div>')
    _up_email    = st.session_state.get("auth_email", "guest")
    _up_tier     = tier()
    _up_tname    = {"free": "FREE", "starter": "STARTER",
                    "premium": "STARTER", "pro": "PRO"}.get(_up_tier, "FREE")
    _up_tclr_cls = {"free": "yiq-tier-free", "starter": "yiq-tier-starter",
                    "premium": "yiq-tier-starter", "pro": "yiq-tier-pro"
                    }.get(_up_tier, "yiq-tier-free")
    _up_email_disp = (
        (_up_email[:22] + "…") if len(_up_email) > 24 else _up_email
    )
    _up_email_html = (
        f'<div class="yiq-profile-email">{_up_email_disp}</div>'
        if _up_email and _up_email != "guest"
        else '<div class="yiq-profile-email" style="color:#475569;">Not signed in</div>'
    )
    _usage_html = usage_bar_html()
    st.html(
        f'<div class="yiq-profile">'
        f'{_up_email_html}'
        f'<span class="yiq-tier-badge {_up_tclr_cls}">{_up_tname}</span>'
        f'<div style="margin-top:8px;">{_usage_html}</div>'
        f'</div>'
    )
    sidebar_upgrade_button()
    render_resume_button()

    # ── Disclaimer strip ─────────────────────────────────────
    st.html("""
    <div style="margin-top:10px;padding:8px 10px;
                background:rgba(251,191,36,0.06);
                border:1px solid rgba(251,191,36,0.15);border-radius:7px;">
      <div style="font-size:10px;color:#92400E;line-height:1.7;">
        ⚠ Model output only — not investment advice<br>
        <span style="color:#475569;">YieldIQ is not a registered RIA</span>
      </div>
    </div>
    """)

# ── "View Disclaimer" sidebar link ────────────────────────────
render_view_disclaimer_link()

# ── PRO MODE CSS INJECTION ────────────────────────────────────
if pro_mode:
    st.html("""
    <style>
    /* PRO MODE — dark dense layout */
    .main .block-container { background: #0d1117 !important; }
    .stApp { background: #0d1117 !important; }
    [data-testid="stMetric"] {
      background: #161b22 !important;
      border: 1px solid #21262d !important;
      border-radius: 8px !important;
      padding: 12px !important;
    }
    [data-testid="stMetricLabel"] { color: #8b949e !important; font-size: 11px !important; }
    [data-testid="stMetricValue"] { color: #e6edf3 !important; font-family: 'IBM Plex Mono', monospace !important; }
    [data-testid="stMetricDelta"] { font-family: 'IBM Plex Mono', monospace !important; }
    [data-testid="stExpander"] summary { border-bottom: 1px solid #E2E8F0 !important; }
    [data-testid="stExpander"] summary p { color: #1E293B !important; font-weight: 600 !important; }
    [data-testid="stExpander"] > div { background: #F8FAFC !important; }
    .stTabs [data-baseweb="tab-list"] { background: #161b22 !important; border-bottom: 1px solid #21262d !important; }
    .stTabs [data-baseweb="tab"] { color: #8b949e !important; }
    .stTabs [aria-selected="true"] { color: #00b4d8 !important; border-bottom-color: #00b4d8 !important; }
    p, div, span, label { color: #e6edf3 !important; }
    </style>
    """)

# ══════════════════════════════════════════════════════════════
# TOP NAV
# ══════════════════════════════════════════════════════════════
st.html(f"""
<div style="padding:16px 0 8px;">
  <div style="display:flex;align-items:center;justify-content:space-between;
              padding:14px 24px;
              background:linear-gradient(135deg,#1E2A3A 0%,#1D4ED8 60%,#0891B2 100%);
              border-radius:14px;margin-bottom:16px;
              box-shadow:0 4px 24px rgba(29,78,216,0.25);">
    <div style="display:flex;align-items:center;gap:12px;">
      <div style="width:36px;height:36px;background:rgba(255,255,255,0.15);
                  border-radius:9px;display:flex;align-items:center;justify-content:center;
                  font-size:20px;backdrop-filter:blur(4px);">📊</div>
      <div>
        <div style="font-size:20px;font-weight:800;color:#FFFFFF;letter-spacing:-0.02em;">YieldIQ</div>
        <div style="font-size:12px;color:rgba(255,255,255,0.6);letter-spacing:0.08em;text-transform:uppercase;">Institutional DCF Platform</div>
      </div>
      <div style="margin-left:8px;padding:4px 10px;background:rgba(255,255,255,0.15);
                  border-radius:20px;font-size:12px;color:#FFFFFF;font-weight:700;letter-spacing:0.5px;">{APP_VERSION}</div>
    </div>
    <div style="display:flex;align-items:center;gap:20px;">
      <div style="text-align:center;">
        <div style="font-size:12px;color:rgba(255,255,255,0.5);text-transform:uppercase;letter-spacing:0.04em;">WACC</div>
        <div style="font-size:13px;font-weight:700;color:#FFFFFF;font-family:'IBM Plex Mono',monospace;">{"Auto" if use_auto_wacc else f"{manual_wacc}%"}</div>
      </div>
      <div style="width:1px;height:28px;background:rgba(255,255,255,0.2);"></div>
      <div style="text-align:center;">
        <div style="font-size:12px;color:rgba(255,255,255,0.5);text-transform:uppercase;letter-spacing:0.04em;">Term. g</div>
        <div style="font-size:13px;font-weight:700;color:#FFFFFF;font-family:'IBM Plex Mono',monospace;">{terminal_pct}%</div>
      </div>
      <div style="width:1px;height:28px;background:rgba(255,255,255,0.2);"></div>
      <div style="text-align:center;">
        <div style="font-size:12px;color:rgba(255,255,255,0.5);text-transform:uppercase;letter-spacing:0.04em;">FX Rate</div>
        <div style="font-size:13px;font-weight:700;color:#34D399;font-family:'IBM Plex Mono',monospace;">1 USD={sym}{fx_rate:,.2f}</div>
      </div>
      <div style="display:flex;align-items:center;gap:6px;padding:6px 12px;
                  background:rgba(52,211,153,0.2);border-radius:20px;
                  border:1px solid rgba(52,211,153,0.3);">
        <div style="width:7px;height:7px;background:#34D399;border-radius:50%;
                    animation:shimmer 2s ease-in-out infinite;"></div>
        <span style="font-size:12px;color:#34D399;font-weight:700;">LIVE</span>
      </div>
    </div>
  </div>
</div>
""")


mkt = fetch_market_overview()

# ── Scrolling ticker bar ──────────────────────────────────────
ticker_items = ""
for name, data in mkt.items():
    chg = data["change_pct"]
    chg_color = "#34d399" if chg >= 0 else "#ef4444"
    chg_sym = "▲" if chg >= 0 else "▼"
    price_fmt = f"{data['price']:,.2f}"
    ticker_items += f"""
    <div style="display:inline-flex;align-items:center;gap:8px;padding:4px 18px;
                border-right:1px solid rgba(255,255,255,0.08);white-space:nowrap;">
      <span style="font-size:11px;font-weight:700;color:#e2e8f0;font-family:'IBM Plex Mono',monospace;">{name}</span>
      <span style="font-size:11px;color:#e2e8f0;font-family:monospace;">{price_fmt}</span>
      <span style="font-size:10px;font-weight:600;color:{chg_color};">{chg_sym} {abs(chg):.2f}%</span>
    </div>"""
st.html(f"""
<div style="background:#0f2537;border-radius:8px;padding:8px 0;overflow:hidden;margin-bottom:12px;">
  <div style="display:flex;overflow:hidden;">{ticker_items}</div>
</div>
""")

# ── 2x2 market cards grid ────────────────────────────────────
mkt_list = list(mkt.items())[:4]
cols = st.columns(4)
for col, (name, data) in zip(cols, mkt_list):
    chg = data["change_pct"]
    chg_color = "#059669" if chg >= 0 else "#dc2626"
    chg_sym = "▲" if chg >= 0 else "▼"
    bar_w = min(abs(chg) * 15 + 40, 95)
    bar_col = "#059669" if chg >= 0 else "#dc2626"
    col.html(f"""
    <div style="background:#FFFFFF;border:1px solid #E2E8F0;border-radius:10px;
                padding:12px 14px;margin-bottom:8px;">
      <div style="font-size:10px;font-weight:600;color:#94A3B8;text-transform:uppercase;
                  letter-spacing:0.08em;margin-bottom:4px;">{name}</div>
      <div style="font-size:16px;font-weight:700;color:#0F172A;
                  font-family:'IBM Plex Mono',monospace;">{data['price']:,.2f}</div>
      <div style="font-size:11px;font-weight:600;color:{chg_color};margin-top:2px;">
        {chg_sym} {abs(chg):.2f}%</div>
      <div style="height:3px;background:#F1F5F9;border-radius:2px;margin-top:8px;">
        <div style="height:100%;width:{bar_w:.0f}%;background:{bar_col};border-radius:2px;"></div>
      </div>
    </div>""")

# ── _arrow_right KILLER — postMessage bridge (definitive fix) ──────────
# PROBLEM: st.html() runs in a sandboxed iframe. The iframe's `sandbox`
# attribute may block `window.parent` access. The ONLY guaranteed way to
# reach the parent document from a sandboxed iframe in Streamlit is via
# window.parent.postMessage, with a listener injected by a SECOND
# st.components.v1.html() call that runs at the TOP LEVEL (no sandbox).
#
# Strategy:
#   Script A (st.html): sends postMessage "CLEAN_ARROWS" every 500ms
#   Script B (st.components.v1.html height=0): listens, runs cleanAll()
#   Script B also runs its own MutationObserver on document.body directly
# ─────────────────────────────────────────────────────────────────────────
import streamlit.components.v1 as _stc

# Script B — runs in the MAIN document (no sandbox), does the actual cleaning
_stc.html("""
<script>
(function() {
  'use strict';
  var ICON_RE = /^_[a-z][a-z_]+$/;
  function isIcon(t) {
    if (!t) return false;
    t = t.trim();
    return t === '_arrow_right' || t === '_arrow_drop_down' ||
           t === '_expand_more' || t.startsWith('_arrow') ||
           t.startsWith('_expand') || t.startsWith('_chevron') ||
           ICON_RE.test(t);
  }
  function cleanNode(node) {
    var toRemove = [];
    node.childNodes.forEach(function(n) {
      if (n.nodeType === 3 && isIcon(n.textContent)) {
        n.textContent = ''; return;
      }
      if (n.nodeType === 1) {
        var tag = n.tagName.toLowerCase();
        var txt = (n.textContent || '').trim();
        if (tag === 'svg') { toRemove.push(n); return; }
        if (['span','i','em','b'].indexOf(tag) !== -1 && isIcon(txt)) {
          toRemove.push(n); return;
        }
        if (isIcon(txt) && !n.querySelector('p,div,button,a')) {
          toRemove.push(n); return;
        }
      }
    });
    toRemove.forEach(function(n) {
      try { n.parentNode && n.parentNode.removeChild(n); } catch(e) {}
    });
  }
  function cleanAll() {
    var sels = [
      '[data-testid="stExpander"] summary',
      'details > summary',
      '.streamlit-expanderHeader',
    ];
    sels.forEach(function(s) {
      document.querySelectorAll(s).forEach(cleanNode);
    });
  }
  // Run immediately and at staggered delays
  cleanAll();
  [100, 300, 600, 1200, 2500, 5000].forEach(function(ms) {
    setTimeout(cleanAll, ms);
  });
  // MutationObserver — watches for new expanders added by React
  var obs = new MutationObserver(function(muts) {
    var needs = false;
    for (var i = 0; i < muts.length; i++) {
      if (muts[i].addedNodes.length > 0 || muts[i].type === 'characterData') {
        needs = true; break;
      }
    }
    if (needs) cleanAll();
  });
  var cfg = { childList: true, subtree: true, characterData: true };
  obs.observe(document.body, cfg);
  var root = document.getElementById('root');
  if (root) obs.observe(root, cfg);
  // Re-observe after full React remounts (Streamlit reruns)
  [3000, 6000, 10000].forEach(function(ms) {
    setTimeout(function() {
      var r = document.getElementById('root');
      if (r) { try { obs.observe(r, cfg); } catch(e) {} }
      cleanAll();
    }, ms);
  });
  // Also listen for postMessage from sandboxed st.html() iframes
  window.addEventListener('message', function(e) {
    if (e.data === 'YIELDIQ_CLEAN_ARROWS') cleanAll();
  });
})();
</script>
""", height=0, scrolling=False)

# ══════════════════════════════════════════════════════════════
# TRIGGERED-ALERT BANNER
# Check at most once every 5 minutes per browser session.
# Newly-fired alerts are accumulated in session_state so the banner
# persists across reruns until the user clears it in the Alerts tab.
# ══════════════════════════════════════════════════════════════
import time as _time
_al_email = st.session_state.get("auth_email", "")
_al_uid   = _alerts_mod._get_user_id(_al_email)
if _al_uid:
    _al_last = st.session_state.get("_al_last_check_ts", 0)
    if _time.time() - _al_last > 300:   # 5-minute TTL
        _al_newly = _alerts_mod.check_alerts(_al_uid)
        st.session_state["_al_last_check_ts"] = _time.time()
        st.session_state["_al_fired"] = (
            st.session_state.get("_al_fired", []) + _al_newly
        )

_al_fired = st.session_state.get("_al_fired", [])
if _al_fired:
    _al_lines = []
    for _f in _al_fired:
        _al_lines.append(
            f"**{_f['ticker']}** — {_f['label']} "
            f"${_f['target_price']:,.2f} "
            f"(now ${_f['current_price']:,.2f})"
        )
    st.warning(
        "🔔 **Price alert" + ("s" if len(_al_fired) > 1 else "") + " triggered:** "
        + "  ·  ".join(_al_lines)
        + "  —  [Clear in the Alerts tab]"
    )

# ══════════════════════════════════════════════════════════════
# MAIN CONTENT — session-state tab routing
# ══════════════════════════════════════════════════════════════
_ADMIN_MODE = os.environ.get("YIELDIQ_ADMIN", "0") == "1"
# _active_main_tab is already set above (before the sidebar block)


# ══════════════════════════════════════════════════════════════
# TAB — COMPARE STOCKS
# ══════════════════════════════════════════════════════════════
if _active_main_tab == "compare":
    compare_tab.render(st.container())


# ══════════════════════════════════════════════════════════════
# TAB — MARKETS (Sector Heatmap)
# ══════════════════════════════════════════════════════════════
if _active_main_tab == "markets":
    render_sector_heatmap()


# ══════════════════════════════════════════════════════════════
# TAB 1 — SINGLE STOCK ANALYSIS
# ══════════════════════════════════════════════════════════════
if _active_main_tab == "stock":
    sc1, sc2, sc3 = st.columns([2, 1, 3])
    with sc1:
        # ── Pre-fill from recent ticker click or popular pill ────
        _pf = st.session_state.pop("_prefill_ticker", None) or st.session_state.pop("ticker_input", None)
        if LAUNCH_REGION == "US":
            _default_ticker = _pf if _pf else "AAPL"
            _ticker_placeholder = "AAPL · MSFT · GOOGL · NVDA · JPM"
        else:
            _default_ticker = _pf if _pf else "TCS.NS"
            _ticker_placeholder = "TCS.NS · RELIANCE.NS · AAPL"
        ticker_input = st.text_input(
            "Ticker", value=_default_ticker,
            placeholder=_ticker_placeholder,
            label_visibility="collapsed"
        ).upper().strip()
    with sc2:
        analyse_btn = (
            st.button("🔍 Analyse this stock", width='stretch')
            or st.session_state.pop("_auto_analyse", False)
        )

    # Must be defined here — used by hero check below AND by redisplay logic later
    _has_results = (
        st.session_state.get("fin_ticker") and
        st.session_state.get("fin_enriched") is not None
    )

    with sc3:
        # 🏠 Home button when analysis results exist — lets users return to Morning Brief
        if _has_results:
            if st.button("🏠 Home", width='stretch', key="btn_go_home",
                         help="Return to the Morning Brief dashboard"):
                st.session_state["_show_morning_brief"] = True
                st.rerun()
        else:
            if LAUNCH_REGION == "US":
                st.html("""
                <div style="padding:10px 16px;background:#FFFFFF;border:1px solid #E2E8F0;
                            border-radius:2px;font-size:12px;color:#64748B;letter-spacing:0.05em;">
                  US: <span style="color:#475569;">AAPL · MSFT · GOOGL · NVDA · JPM · JNJ · TSLA · AMZN</span>
                </div>
                """)
            else:
                st.html("""
                <div style="padding:10px 16px;background:#FFFFFF;border:1px solid #E2E8F0;
                            border-radius:2px;font-size:12px;color:#64748B;letter-spacing:0.05em;">
                  NSE: <span style="color:#475569;">TCS.NS · INFY.NS · RELIANCE.NS · HDFCBANK.NS</span>
                  &nbsp;·&nbsp; US: <span style="color:#475569;">AAPL · MSFT · GOOGL · NVDA</span>
                </div>
                """)

    # ── Popular ticker pills ──────────────────────────────────────
    st.caption("Popular stocks:")
    _popular = ["AAPL", "MSFT", "NVDA", "GOOGL", "TSLA", "JPM", "AMZN", "META"]
    _pop_cols = st.columns(len(_popular))
    for _i, _t in enumerate(_popular):
        with _pop_cols[_i]:
            if st.button(_t, key=f"pop_{_t}", width="stretch"):
                st.session_state.pop("fin_ticker",   None)   # clear stale analysis
                st.session_state.pop("fin_enriched", None)
                st.session_state["_prefill_ticker"]      = _t
                st.session_state["_auto_analyse"]        = True
                st.session_state["_show_morning_brief"]  = False
                st.rerun()

    # Show Morning Brief when:
    #   (a) no analysis triggered AND no cached results, OR
    #   (b) user explicitly navigated home via session state flag
    _show_brief = st.session_state.get("_show_morning_brief", True)
    if analyse_btn:
        st.session_state["_show_morning_brief"] = False
        _show_brief = False

    if not analyse_btn and not _has_results:
        # ── First visit / no analysis yet — empty state ─────────
        render_empty_state(sym)
        # Functional ticker chip buttons (styled as chips via CSS)
        _es_tickers = ["AAPL", "MSFT", "GOOGL", "AMZN", "TSLA", "NVDA"]
        _es_cols = st.columns(len(_es_tickers))
        for _ei, _et in enumerate(_es_tickers):
            with _es_cols[_ei]:
                if st.button(_et, key=f"es_{_et}", width="stretch"):
                    st.session_state["_prefill_ticker"] = _et
                    st.session_state["_auto_analyse"]   = True
                    st.session_state["_show_morning_brief"] = False
                    st.rerun()
    elif not analyse_btn and _show_brief:
        # ── Returning user navigated home — Morning Brief ────────
        render_morning_brief(
            watchlist_rows=get_watchlist(),
            sym=sym,
            has_prior_results=_has_results,
            theme=st.session_state.get("theme", "light"),
        )

    if False:  # dead code — original hero kept here for reference only
        st.html("""
<style>
/* ── Hero animations ─────────────────────────────────────── */
@keyframes yiq-hero-shift {
  0%   { background-position: 0%   50%; }
  50%  { background-position: 100% 50%; }
  100% { background-position: 0%   50%; }
}
@keyframes yiq-fade-up {
  from { opacity: 0; transform: translateY(16px); }
  to   { opacity: 1; transform: translateY(0);    }
}
@keyframes yiq-marquee {
  0%   { transform: translateX(0); }
  100% { transform: translateX(-50%); }
}
@keyframes yiq-blink {
  0%,49%  { opacity: 1; }
  50%,100%{ opacity: 0; }
}
/* Typewriter cycle: show each hint for 2s, total 8s loop */
@keyframes tw-show { 0%,25%{opacity:1;width:auto} 26%,100%{opacity:0;width:0} }

.yiq-hero {
  margin-top: 24px;
  border-radius: 16px;
  overflow: hidden;
  background: linear-gradient(135deg, #020817 0%, #0d1f35 30%, #0a1628 60%, #061220 100%);
  background-size: 300% 300%;
  animation: yiq-hero-shift 10s ease infinite;
  padding: 52px 48px 40px;
  text-align: center;
  position: relative;
}
.yiq-hero::before {
  content: "";
  position: absolute; inset: 0;
  background-image:
    linear-gradient(rgba(0,180,216,0.05) 1px, transparent 1px),
    linear-gradient(90deg, rgba(0,180,216,0.05) 1px, transparent 1px);
  background-size: 40px 40px;
  pointer-events: none;
}
.yiq-hero::after {
  content: "";
  position: absolute;
  top: -60px; left: 50%; transform: translateX(-50%);
  width: 400px; height: 280px;
  background: radial-gradient(ellipse, rgba(0,180,216,0.15) 0%, transparent 70%);
  pointer-events: none;
}
.yiq-hero-inner {
  position: relative; z-index: 2;
  animation: yiq-fade-up 0.6s ease both;
}
.yiq-hero-eyebrow {
  display: inline-block;
  font-size: 11px; font-weight: 600;
  letter-spacing: 0.18em; text-transform: uppercase;
  color: #00b4d8;
  background: rgba(0,180,216,0.12);
  border: 1px solid rgba(0,180,216,0.3);
  border-radius: 20px;
  padding: 4px 14px;
  margin-bottom: 12px;
}

/* ── Scrolling stats ticker ──────────────────────────────── */
.yiq-ticker-wrap {
  overflow: hidden;
  width: 100%;
  margin: 0 auto 20px;
  max-width: 760px;
  mask-image: linear-gradient(90deg, transparent 0%, black 8%, black 92%, transparent 100%);
  -webkit-mask-image: linear-gradient(90deg, transparent 0%, black 8%, black 92%, transparent 100%);
}
.yiq-ticker-track {
  display: inline-flex;
  gap: 8px;
  white-space: nowrap;
  animation: yiq-marquee 28s linear infinite;
}
.yiq-ticker-chip {
  display: inline-flex;
  align-items: center;
  gap: 5px;
  padding: 4px 12px;
  background: rgba(255,255,255,0.05);
  border: 1px solid rgba(255,255,255,0.08);
  border-radius: 20px;
  font-size: 11px;
  color: rgba(255,255,255,0.55);
  white-space: nowrap;
  flex-shrink: 0;
}

.yiq-hero-headline {
  font-family: 'Barlow Condensed', 'Inter', sans-serif;
  font-size: 42px; font-weight: 700; line-height: 1.15;
  color: #FFFFFF;
  letter-spacing: -0.01em;
  margin-bottom: 16px;
  max-width: 680px; margin-left: auto; margin-right: auto;
}
.yiq-hero-headline span {
  background: linear-gradient(90deg, #00b4d8, #38e8ff);
  -webkit-background-clip: text; -webkit-text-fill-color: transparent;
  background-clip: text;
}
.yiq-hero-sub {
  font-size: 15px; font-weight: 400; line-height: 1.7;
  color: rgba(255,255,255,0.6);
  max-width: 520px; margin: 0 auto 32px;
}

/* ── Value prop cards ─────────────────────────────────────── */
.yiq-cards {
  display: grid;
  grid-template-columns: 1fr 1fr 1fr;
  gap: 12px;
  max-width: 760px;
  margin: 0 auto 28px;
}
.yiq-card {
  background: rgba(255,255,255,0.05);
  border: 1px solid rgba(255,255,255,0.09);
  border-radius: 10px;
  padding: 16px;
  text-align: left;
  transition: background 0.2s, border-color 0.2s;
}
.yiq-card:hover {
  background: rgba(0,180,216,0.09);
  border-color: rgba(0,180,216,0.3);
}
.yiq-card-icon { font-size: 20px; margin-bottom: 6px; display: block; }
.yiq-card-title { font-size: 13px; font-weight: 600; color: #FFFFFF; margin-bottom: 3px; }
.yiq-card-desc  { font-size: 11px; color: rgba(255,255,255,0.45); line-height: 1.5; }

/* ── Social proof stat blocks ────────────────────────────── */
.yiq-stats {
  display: flex; align-items: center; justify-content: center;
  gap: 0; max-width: 680px; margin: 0 auto 24px;
  border: 1px solid rgba(255,255,255,0.08);
  border-radius: 10px; overflow: hidden;
}
.yiq-stat {
  flex: 1; padding: 14px 16px; text-align: center;
  border-right: 1px solid rgba(255,255,255,0.08);
}
.yiq-stat:last-child { border-right: none; }
.yiq-stat-num  { font-size: 18px; font-weight: 700; color: #00b4d8; line-height: 1.1; }
.yiq-stat-sub  { font-size: 10px; color: rgba(255,255,255,0.4); margin-top: 2px; line-height: 1.4; }

/* ── Trust bar ───────────────────────────────────────────── */
.yiq-trust {
  font-size: 11px; color: rgba(255,255,255,0.3);
  letter-spacing: 0.06em;
  margin-bottom: 20px;
}
.yiq-trust strong { color: rgba(255,255,255,0.5); font-weight: 500; }

/* ── Ticker chips ────────────────────────────────────────── */
.yiq-tickers {
  display: flex; align-items: center; justify-content: center;
  flex-wrap: wrap; gap: 6px;
  margin-bottom: 16px;
}
.yiq-ticker-pill {
  font-family: 'IBM Plex Mono', monospace;
  font-size: 11px; font-weight: 500;
  color: rgba(255,255,255,0.5);
  background: rgba(255,255,255,0.06);
  border: 1px solid rgba(255,255,255,0.1);
  border-radius: 4px;
  padding: 3px 9px;
}
.yiq-ticker-sep { color: rgba(255,255,255,0.2); font-size: 11px; }

/* ── Scrolling stock news ticker ─────────────────────────── */
@keyframes yiq-stock-scroll {
  0%   { transform: translateX(0); }
  100% { transform: translateX(-50%); }
}
.yiq-stock-ticker-wrap {
  width: 100%;
  overflow: hidden;
  margin: 6px 0 14px;
  mask-image: linear-gradient(90deg, transparent 0%, black 8%, black 92%, transparent 100%);
  -webkit-mask-image: linear-gradient(90deg, transparent 0%, black 8%, black 92%, transparent 100%);
}
.yiq-stock-ticker-track {
  display: inline-flex;
  align-items: center;
  gap: 0;
  white-space: nowrap;
  animation: yiq-stock-scroll 35s linear infinite;
}
.yiq-stock-ticker-track:hover {
  animation-play-state: paused;
}
.yiq-stock-item {
  display: inline-flex;
  align-items: center;
  gap: 6px;
  padding: 4px 14px;
  cursor: pointer;
  transition: opacity 0.2s;
}
.yiq-stock-item:hover { opacity: 0.75; }
.yiq-avatar {
  width: 18px;
  height: 18px;
  border-radius: 4px;
  display: inline-flex;
  align-items: center;
  justify-content: center;
  font-size: 10px;
  font-weight: 800;
  color: #fff;
  flex-shrink: 0;
  font-family: 'Inter', sans-serif;
}
.yiq-stock-name {
  font-family: 'Inter', sans-serif;
  font-size: 11px;
  font-weight: 500;
  color: rgba(255,255,255,0.55);
}
.yiq-stock-sym {
  font-family: 'IBM Plex Mono', monospace;
  font-size: 11px;
  font-weight: 700;
  color: #60A5FA;
  letter-spacing: 0.03em;
}
.yiq-stock-sep {
  color: rgba(255,255,255,0.12);
  font-size: 14px;
  flex-shrink: 0;
}

/* ── Typewriter search hints ─────────────────────────────── */
.yiq-tw-wrap {
  font-size: 12px; color: rgba(255,255,255,0.3);
  letter-spacing: 0.04em; height: 18px;
  overflow: hidden; margin-bottom: 4px;
}
.yiq-tw-item {
  display: inline-block;
  overflow: hidden; white-space: nowrap;
  opacity: 0; width: 0;
  animation: tw-show 8s infinite;
}
.yiq-tw-item:nth-child(1) { animation-delay: 0s;   }
.yiq-tw-item:nth-child(2) { animation-delay: 2s;   }
.yiq-tw-item:nth-child(3) { animation-delay: 4s;   }
.yiq-tw-item:nth-child(4) { animation-delay: 6s;   }
.yiq-cursor {
  display: inline-block; width: 2px; height: 12px;
  background: #00b4d8; margin-left: 2px; vertical-align: middle;
  animation: yiq-blink 0.8s step-end infinite;
}
</style>

<div class="yiq-hero">
  <div class="yiq-hero-inner">

    <div class="yiq-hero-eyebrow">Institutional-Grade Stock Analysis</div>

    <!-- Scrolling stats ticker -->
    <div class="yiq-ticker-wrap">
      <div class="yiq-ticker-track">
        <span class="yiq-ticker-chip">📊 47,293 stocks analysed</span>
        <span class="yiq-ticker-chip">🎯 89.2% DCF accuracy (backtest)</span>
        <span class="yiq-ticker-chip">⚡ Real-time price data</span>
        <span class="yiq-ticker-chip">🌍 US + India markets</span>
        <span class="yiq-ticker-chip">🔒 Model-grade analysis</span>
        <span class="yiq-ticker-chip">📈 10-yr FCF forecasting</span>
        <span class="yiq-ticker-chip">🧮 Monte Carlo simulation</span>
        <span class="yiq-ticker-chip">⚖️ Economic moat scoring</span>
        <!-- Duplicate for seamless loop -->
        <span class="yiq-ticker-chip">📊 47,293 stocks analysed</span>
        <span class="yiq-ticker-chip">🎯 89.2% DCF accuracy (backtest)</span>
        <span class="yiq-ticker-chip">⚡ Real-time price data</span>
        <span class="yiq-ticker-chip">🌍 US + India markets</span>
        <span class="yiq-ticker-chip">🔒 Model-grade analysis</span>
        <span class="yiq-ticker-chip">📈 10-yr FCF forecasting</span>
        <span class="yiq-ticker-chip">🧮 Monte Carlo simulation</span>
        <span class="yiq-ticker-chip">⚖️ Economic moat scoring</span>
      </div>
    </div>

    <div class="yiq-hero-headline">
      Know What a Stock Is<br><span>Really Worth</span> — Before You Invest
    </div>

    <div class="yiq-hero-sub">
      DCF-powered intrinsic value, margin of safety, scenario analysis,
      and quality scoring — in plain English.
    </div>

    <div class="yiq-cards">
      <div class="yiq-card">
        <span class="yiq-card-icon">📊</span>
        <div class="yiq-card-title">Intrinsic Value</div>
        <div class="yiq-card-desc">DCF-based fair value with margin of safety and 3-scenario analysis</div>
      </div>
      <div class="yiq-card">
        <span class="yiq-card-icon">🏢</span>
        <div class="yiq-card-title">Company Quality</div>
        <div class="yiq-card-desc">Revenue trends, profit margins, debt levels and earnings consistency</div>
      </div>
      <div class="yiq-card">
        <span class="yiq-card-icon">⚡</span>
        <div class="yiq-card-title">Live Data</div>
        <div class="yiq-card-desc">Real-time price vs estimated value — updated every minute</div>
      </div>
    </div>

    <!-- Social proof stat blocks -->
    <div class="yiq-stats">
      <div class="yiq-stat">
        <div class="yiq-stat-num">10+</div>
        <div class="yiq-stat-sub">Valuation Models<br>DCF, DDM, EV/EBITDA, Moat Score &amp; more</div>
      </div>
      <div class="yiq-stat">
        <div class="yiq-stat-num">US + India</div>
        <div class="yiq-stat-sub">Markets Supported<br>NYSE, NASDAQ, NSE, BSE</div>
      </div>
      <div class="yiq-stat">
        <div class="yiq-stat-num">1,000×</div>
        <div class="yiq-stat-sub">Monte Carlo Ready<br>Simulation scenarios per stock</div>
      </div>
    </div>

    <div class="yiq-trust">
      <strong>Trusted analytical framework</strong> · DCF · Economic Moat · Monte Carlo · Piotroski Score
    </div>

    <!-- Typewriter search hints -->
    <div class="yiq-tw-wrap">
      <span class="yiq-tw-item">Try AAPL...</span>
      <span class="yiq-tw-item">Try NVDA...</span>
      <span class="yiq-tw-item">Try MSFT...</span>
      <span class="yiq-tw-item">Try GOOGL...</span>
      <span class="yiq-cursor"></span>
    </div>

    <!-- Scrolling stock news ticker (CSS letter avatars, no external deps) -->
    <div class="yiq-stock-ticker-wrap">
      <div class="yiq-stock-ticker-track">
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#555;">🍎</span><span class="yiq-stock-name">Apple</span><span class="yiq-stock-sym">AAPL</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#0078d4;">M</span><span class="yiq-stock-name">Microsoft</span><span class="yiq-stock-sym">MSFT</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#76b900;">N</span><span class="yiq-stock-name">NVIDIA</span><span class="yiq-stock-sym">NVDA</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#4285f4;">G</span><span class="yiq-stock-name">Alphabet</span><span class="yiq-stock-sym">GOOGL</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#ff9900;">A</span><span class="yiq-stock-name">Amazon</span><span class="yiq-stock-sym">AMZN</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#cc0000;">T</span><span class="yiq-stock-name">Tesla</span><span class="yiq-stock-sym">TSLA</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#0668e1;">f</span><span class="yiq-stock-name">Meta</span><span class="yiq-stock-sym">META</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#1a5276;">J</span><span class="yiq-stock-name">JPMorgan</span><span class="yiq-stock-sym">JPM</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#1b4f72;">B</span><span class="yiq-stock-name">Berkshire</span><span class="yiq-stock-sym">BRK-B</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#1a6b3c;">U</span><span class="yiq-stock-name">UnitedHealth</span><span class="yiq-stock-sym">UNH</span></span>
        <!-- Duplicate set for seamless loop -->
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#555;">🍎</span><span class="yiq-stock-name">Apple</span><span class="yiq-stock-sym">AAPL</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#0078d4;">M</span><span class="yiq-stock-name">Microsoft</span><span class="yiq-stock-sym">MSFT</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#76b900;">N</span><span class="yiq-stock-name">NVIDIA</span><span class="yiq-stock-sym">NVDA</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#4285f4;">G</span><span class="yiq-stock-name">Alphabet</span><span class="yiq-stock-sym">GOOGL</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#ff9900;">A</span><span class="yiq-stock-name">Amazon</span><span class="yiq-stock-sym">AMZN</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#cc0000;">T</span><span class="yiq-stock-name">Tesla</span><span class="yiq-stock-sym">TSLA</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#0668e1;">f</span><span class="yiq-stock-name">Meta</span><span class="yiq-stock-sym">META</span></span>
        <span class="yiq-stock-sep">·</span>
        <span class="yiq-stock-item"><span class="yiq-avatar" style="background:#1a5276;">J</span><span class="yiq-stock-name">JPMorgan</span><span class="yiq-stock-sym">JPM</span></span>
      </div>
    </div>

  </div>
</div>
        """)

    # Re-render from session state if we already have results
    # This prevents page reset on any rerun (form submit, button click etc.)
    # NOTE: _has_results is defined earlier (before hero check) — do not redefine
    _should_analyse = analyse_btn and ticker_input
    _should_redisplay = (
        _has_results and
        not analyse_btn and
        not _show_brief
    )

    # Re-display from session state (after form submit / button click rerun)
    if _should_redisplay:
        ticker_input = st.session_state.get("fin_ticker", "")
        _should_analyse = True

    if _should_analyse and ticker_input:
        # ── Skip fetch/compute if redisplaying from session state ──
        _from_cache = (
            _should_redisplay and
            st.session_state.get("fin_ticker") == ticker_input and
            st.session_state.get("fin_enriched") is not None and
            st.session_state.get("_dcf_res") is not None
        )

        if _from_cache:
            # ── Short-circuit for non-DCF tickers (from cache) ──────
            if not st.session_state.get("_dcf_eligible", True):
                _rel_val = st.session_state.get("_rel_val", {})
                raw      = st.session_state.get("fin_raw", {})
                _render_relative_valuation_view(
                    ticker=ticker_input,
                    rv=_rel_val or {},
                    raw=raw,
                    sym=st.session_state.get("fin_sym", "$"),
                    fx=st.session_state.get("fin_fx", 1.0),
                )
                st.stop()

            # Restore all computed variables from session state
            import pandas as _pd
            enriched        = st.session_state.get("fin_enriched", {})
            raw             = st.session_state.get("fin_raw", {})
            fx              = st.session_state.get("fin_fx", 1.0)
            to_code         = st.session_state.get("fin_to_code", "USD")
            sym             = st.session_state.get("fin_sym", "$")
            iv_d            = st.session_state.get("fin_iv_d", 0)
            mos_pct         = st.session_state.get("fin_mos_pct", 0)
            sig             = st.session_state.get("fin_signal", "N/A ⬜")
            dcf_res         = st.session_state.get("_dcf_res", {})
            forecast_result = st.session_state.get("_forecast_result", {})
            scenarios       = st.session_state.get("_scenarios", {})
            inv_plan        = st.session_state.get("_inv_plan", {
                "price_targets":{}, "holding_period":{},
                "fundamental":{"grade":"N/A","score":0}
            })
            confidence      = st.session_state.get("_confidence", {"grade":"N/A","score":0})
            price_hist      = st.session_state.get("_price_hist", _pd.DataFrame())
            wacc_data       = st.session_state.get("_wacc_data", {})
            use_auto_wacc   = st.session_state.get("_use_auto_wacc", True)
            forecast_yrs    = st.session_state.get("_forecast_yrs", 10)
            terminal_pct    = st.session_state.get("_terminal_pct", 3)
            _rf_rate_info   = st.session_state.get("_rf_rate_info", {})
            _insider_adj    = st.session_state.get("_insider_adj", 0.0)
            _insider_data   = raw.get("finnhub_insider", {})
            _insider_sent   = _insider_data.get("sentiment", "NEUTRAL")
            # Recompute derived display vars
            price_n         = enriched.get("price", 0)
            price_d         = price_n * fx
            iv_n            = iv_d / fx if fx else iv_d
            mos             = mos_pct / 100
            terminal_g      = terminal_pct / 100
            native_ccy      = raw.get("native_ccy", "USD")
            company_name    = raw.get("company_name", ticker_input)
            wacc            = enriched.get("wacc_used", 0.10)
            wacc_source     = enriched.get("wacc_source", "Auto CAPM")
            projected       = forecast_result.get("projections", [])
            terminal_norm   = forecast_result.get("terminal_fcf_norm", 0)
            pv_tv_d         = dcf_res.get("pv_tv", 0) * fx
            proj_d          = [v * fx for v in projected]
            pv_fcfs_d       = [v * fx for v in dcf_res.get("pv_fcfs", [])]
            fx_rate         = st.session_state.get("_fx_rate", 1.0)
            pt              = inv_plan.get("price_targets", {})
            hp              = inv_plan.get("holding_period", {})
            fs              = inv_plan.get("fundamental", {"grade":"N/A","score":0})
            suspicious      = dcf_res.get("suspicious", False)
            _show_plan      = can("action_plan")
            _show_quality   = can("quality_score")
            _show_scenarios = can("scenarios")
            _show_sensitive = can("sensitivity")
            _show_mc        = can("monte_carlo")
            mos_color = "#0D7A4E" if mos_pct>20 else "#B8972A" if mos_pct>0 else "#A62020"
            mos_w     = min(max(abs(mos_pct), 2), 100)
            _fs_color_map = {"STRONG":"#0D7A4E","GOOD":"#2563EB","AVERAGE":"#B8972A","WEAK":"#A62020"}
            _fs_color = _fs_color_map.get(fs.get("grade",""), "#4A5E7A")
            _fs_label = fs.get("grade", "N/A")
            # Use human language signal helper for colors
            _h_label_c, sig_fg, sig_bg, sig_bd = sig_human(sig)
            years_labels    = [f"Y{i+1}" for i in range(forecast_yrs)]
            # Derive vars that come from forecast_result
            growth_schedule = forecast_result.get("growth_schedule", [])
            base_growth     = forecast_result.get("base_growth", 0)
            fcf_base        = forecast_result.get("fcf_base", 0)
            if len(projected) > 0 and len(growth_schedule) > 0 and growth_schedule[0] > -1:
                fcf_base_for_scenarios = projected[0] / (1 + growth_schedule[0])
            else:
                fcf_base_for_scenarios = fcf_base if fcf_base > 0 else enriched.get("latest_fcf", 1e6)
            # Derive moat display vars from enriched
            moat_grade      = enriched.get("moat_grade", "None")
            moat_score      = enriched.get("moat_score", 0)
            _moat_result    = st.session_state.get("fin_moat", {})
            _moat_adj       = st.session_state.get("fin_moat_adj", {})
            native_ccy      = raw.get("native_ccy", "USD")
            company_name    = raw.get("company_name", ticker_input)
            # Restore wacc/terminal from enriched
            wacc            = enriched.get("wacc_used", enriched.get("wacc", 0.10))
            terminal_g      = enriched.get("terminal_g_used", terminal_g)
            wacc_source     = enriched.get("wacc_source", "Auto CAPM")
            # inv_plan sub-vars
            buy_d    = (pt.get("buy_price")    or 0) * fx
            tgt_d    = (pt.get("target_price") or 0) * fx
            sl_d     = (pt.get("stop_loss")    or 0) * fx
            upside   = ((tgt_d - price_d) / price_d * 100) if price_d > 0 else 0
            downside = pt.get("sl_pct", 15)
            rr       = pt.get("rr_ratio", 0)
            # Enriched sub-vars used in display
            shares             = enriched.get("shares", 0)
            shares_outstanding = enriched.get("shares", 0)
            total_debt         = enriched.get("total_debt", 0)
            total_cash         = enriched.get("total_cash", 0)
            current_price      = price_n
            _sector            = enriched.get("sector", "general")
            sector             = _sector
            _pe_iv             = enriched.get("pe_iv", 0)
            projected_fcfs     = projected
            moat_adj           = _moat_adj
            moat_summary       = enriched.get("moat_summary", "")
            moat_types         = enriched.get("moat_types", [])
            iv_delta_pct       = _moat_adj.get("iv_delta_pct", 0)
            mc_result          = st.session_state.get("_mc_result", {})
            _conf_warnings     = confidence.get("warnings", [])
            years              = forecast_yrs

        else:
            # ── TIER CHECK: daily limit ─────────────────────────
            if not can_analyse():
                show_analysis_limit_modal()
                st.stop()

            # ── TIER CHECK / REGION CHECK: market access ────────
            allowed, reason = check_ticker_allowed(ticker_input)
            if not allowed:
                if reason == "india_region":
                    # US launch mode — Indian tickers not yet available
                    st.info(
                        "🇮🇳 **Indian market stocks are coming soon.** "
                        "YieldIQ is currently serving US markets only. "
                        "Stay tuned for NSE/BSE coverage!",
                        icon="🚧",
                    )
                elif reason == "europe_region":
                    st.info(
                        "🌍 **European market stocks are coming soon.** "
                        "YieldIQ is currently serving US markets only.",
                        icon="🚧",
                    )
                elif reason == "india":
                    show_india_gate_message()
                else:
                    upgrade_prompt(reason)
                st.stop()

            # ── Multi-step progress card ──────────────────────────
            _prog_ph = st.empty()
            _PROG_STEPS = [
                "Fetching live price data",
                "Loading financial statements",
                "Running AI growth forecast",
                "Computing DCF valuation",
                "Running scenario analysis",
            ]

            def _update_progress(step: int, detail: str = "") -> None:
                """Re-render the progress card at the given step index (0-based)."""
                _pct = min((step + 1) * 20, 100)
                _rows = ""
                for _si, _sl in enumerate(_PROG_STEPS):
                    if _si < step:
                        _ic, _col, _fw, _dots = "✅", "#059669", "500", ""
                    elif _si == step:
                        _ic, _col, _fw, _dots = "⏳", "#1D4ED8", "600", "…"
                    else:
                        _ic, _col, _fw, _dots = "◻", "#CBD5E1", "400", ""
                    _rows += (
                        f'<div style="display:flex;align-items:center;gap:10px;'
                        f'padding:5px 0;font-size:12px;color:{_col};font-weight:{_fw};">'
                        f'<span style="width:16px;text-align:center;">{_ic}</span>'
                        f'<span>{_sl}{_dots}</span></div>'
                    )
                _det = (
                    f'<div style="margin-top:10px;padding:8px 10px;background:#F8FAFC;'
                    f'border-radius:6px;font-size:11px;color:#475569;'
                    f'font-family:\'IBM Plex Mono\',monospace;">{detail}</div>'
                ) if detail else ""
                _prog_ph.markdown(f"""
<div style="background:#FFFFFF;border:1px solid #E2E8F0;border-radius:14px;
            padding:22px 26px;box-shadow:0 4px 20px rgba(15,23,42,0.07);
            margin:16px 0;max-width:540px;">
  <div style="display:flex;align-items:center;gap:12px;margin-bottom:14px;">
    <div style="width:36px;height:36px;background:linear-gradient(135deg,#1D4ED8,#06B6D4);
                border-radius:9px;display:flex;align-items:center;justify-content:center;
                font-size:16px;flex-shrink:0;">🔍</div>
    <div>
      <div style="font-size:13px;font-weight:700;color:#0F172A;">
        Analyzing <code style="background:#EFF6FF;color:#1D4ED8;
        padding:1px 6px;border-radius:4px;font-size:12px;">{ticker_input}</code>
      </div>
      <div style="font-size:11px;color:#94A3B8;margin-top:1px;">
        Model-based analysis · not investment advice
      </div>
    </div>
  </div>
  <div style="height:4px;background:#E2E8F0;border-radius:2px;margin-bottom:16px;">
    <div style="height:100%;width:{_pct}%;
                background:linear-gradient(90deg,#1D4ED8,#06B6D4);
                border-radius:2px;transition:width 0.35s ease;"></div>
  </div>
  {_rows}
  {_det}
</div>
""", unsafe_allow_html=True)

            _update_progress(0)
            raw, price_hist, wacc_data = fetch_stock_data(ticker_input)

        if not _from_cache:
            if raw is None:
                # ── Retry: try appending .NS for Indian stocks entered without suffix ──
                if not ticker_input.endswith((".NS", ".BO")):
                    try:
                        _r2, _ph2, _wd2 = fetch_stock_data(ticker_input + ".NS")
                        if _r2 is not None:
                            raw, price_hist, wacc_data = _r2, _ph2, _wd2
                            ticker_input = ticker_input + ".NS"
                    except Exception:
                        pass

            if raw is None:
                if "_prog_ph" in dir():
                    try: _prog_ph.empty()
                    except Exception: pass
                # ── Common ticker alias suggestions ───────────────────────
                _ticker_suggestions: dict[str, str] = {
                    "GOOG":    "Try <b>GOOGL</b> — Alphabet Inc. Class A shares.",
                    "BRK":     "Try <b>BRK-B</b> — Berkshire Hathaway Class B.",
                    "BRKB":    "Try <b>BRK-B</b>.",
                    "FB":      "Meta Platforms now trades as <b>META</b>.",
                    "TWITTER": "Twitter is private. Try <b>META</b> or <b>SNAP</b>.",
                    "GOOGLE":  "Try <b>GOOGL</b> or <b>GOOG</b>.",
                }
                _sugg = _ticker_suggestions.get(ticker_input.upper(), "")
                _err_obj = st.session_state.get("_last_fetch_error", "")
                if "429" in str(_err_obj) or "rate" in str(_err_obj).lower():
                    _reason = "Data provider rate limit reached. Please wait 60 seconds."
                    _action = "→ Wait a moment, then try again"
                elif "connection" in str(_err_obj).lower() or "timeout" in str(_err_obj).lower():
                    _reason = "Could not reach the data provider. Check your internet connection."
                    _action = "→ Check your connection and try again"
                elif _sugg:
                    _reason = f"Ticker not found — possible alias detected."
                    _action = _sugg
                else:
                    _reason = (
                        "No data found for this ticker. "
                        "It may be delisted, private, or entered incorrectly. "
                        "Indian stocks need the suffix: <b>RELIANCE.NS</b>"
                    )
                    _action = "→ Double-check the symbol and try again"
                st.html(f"""
<div style="border-left:4px solid #DC2626;background:#FEF2F2;
            border-radius:0 12px 12px 0;padding:18px 22px;margin:16px 0;
            box-shadow:0 2px 8px rgba(220,38,38,0.07);">
  <div style="display:flex;align-items:flex-start;gap:14px;">
    <span style="font-size:22px;flex-shrink:0;">⚠️</span>
    <div>
      <div style="font-size:14px;font-weight:700;color:#991B1B;
                  margin-bottom:6px;font-family:'Inter',sans-serif;">
        Could not load data for
        <code style="background:rgba(220,38,38,0.1);color:#991B1B;
               padding:1px 6px;border-radius:4px;font-size:13px;">{ticker_input}</code>
      </div>
      <div style="font-size:13px;color:#7F1D1D;line-height:1.65;
                  font-family:'Inter',sans-serif;">{_reason}</div>
      <div style="margin-top:10px;font-size:12px;color:#1D4ED8;
                  font-weight:600;font-family:'Inter',sans-serif;">{_action}</div>
    </div>
  </div>
</div>
""")
                st.stop()

            # Record this analysis
            record_analysis()
            st.session_state["last_fetch_time"] = _time.time()
            # ── Track recent tickers ──────────────────────────────────
            _rt = st.session_state.get("recent_tickers", [])
            if ticker_input not in _rt:
                _rt.append(ticker_input)
            st.session_state["recent_tickers"] = _rt[-10:]  # keep last 10

        if not _from_cache:
            _price_n_tmp = (raw or {}).get("price", 0)
            _price_str   = f"{sym}{_price_n_tmp:,.2f}" if _price_n_tmp else ""
            _update_progress(1, f"{ticker_input}: {_price_str}" if _price_str else "")
            enriched   = compute_metrics(raw)
            _update_progress(2)
            forecaster = FCFForecaster()

            wacc = wacc_data.get("wacc", manual_wacc/100) if use_auto_wacc and wacc_data.get("auto_computed") else manual_wacc/100
            wacc_source = f"Auto CAPM ({wacc:.1%})" if use_auto_wacc and wacc_data.get("auto_computed") else f"Manual ({manual_wacc}%)"

            # ── Sector detection + industry WACC adjustment ────────
            # This is the critical step that was missing from the dashboard.
            # Detects sector from ticker, applies industry-appropriate WACC,
            # and stores sector in enriched for PE blending to use correctly.
            try:
                from models.industry_wacc import get_industry_wacc, detect_sector
                _yf_sector   = raw.get("sector_name", "") if raw else ""
                _sector_key  = detect_sector(ticker_input, _yf_sector)
                # Pass yf_sector so data/analytics cos (SPGI,MCO) don't fall into us_banks
                _ind_info    = get_industry_wacc(ticker=ticker_input, yf_sector=_yf_sector, capm_wacc=wacc)
                enriched["sector"]      = _sector_key
                enriched["sector_name"] = _ind_info.get("sector_name", _sector_key)
                # Only use industry WACC if auto mode is on
                if use_auto_wacc:
                    wacc       = _ind_info["wacc"]
                    terminal_g = _ind_info["terminal_growth"]
                    wacc_source = f"Industry-Adjusted ({wacc:.1%}, {_sector_key})"
            except Exception as _se:
                _sector_key = "general"
                enriched.setdefault("sector", "general")

            # ── Non-DCF eligibility check ──────────────────────────
            # Banks, REITs, and insurance companies skip DCF entirely.
            # Route them to relative valuation and stop here.
            _gics_sector = raw.get("sector_name", "") or enriched.get("sector_name", "")
            _dcf_eligible = check_ticker_dcf_eligibility(
                ticker_input,
                sector_key=_sector_key,
                gics_sector=_gics_sector,
            )
            if not _dcf_eligible:
                _update_progress(3, "Running relative valuation (non-DCF sector)…")
                _rel_val = relative_valuation_only(
                    ticker=ticker_input,
                    sector_key=_sector_key,
                    raw=raw,
                    gics_sector=_gics_sector,
                )
                _prog_ph.empty()
                # Compute FX for non-DCF view
                _native_ccy = raw.get("native_ccy", "USD")
                _fx_rv = get_fx_rate(_native_ccy, to_code)
                # Persist to session state so cache path works on rerun
                st.session_state["fin_ticker"]     = ticker_input
                st.session_state["fin_raw"]        = raw
                st.session_state["_dcf_eligible"]  = False
                st.session_state["_rel_val"]       = _rel_val
                st.session_state["fin_enriched"]   = enriched
                st.session_state["_dcf_res"]       = {"_non_dcf": True}
                st.session_state["fin_sym"]        = sym
                st.session_state["fin_fx"]         = _fx_rv
                st.session_state["fin_to_code"]    = to_code
                _render_relative_valuation_view(
                    ticker=ticker_input,
                    rv=_rel_val or {},
                    raw=raw,
                    sym=sym,
                    fx=_fx_rv,
                )
                st.stop()

            _update_progress(3)
            dcf_engine      = DCFEngine(discount_rate=wacc, terminal_growth=terminal_g)
            forecast_result = forecaster.predict(enriched, years=forecast_yrs)
            projected       = forecast_result["projections"]
            terminal_norm   = forecast_result["terminal_fcf_norm"]
            base_growth     = forecast_result["base_growth"]
            fcf_base        = forecast_result["fcf_base"]
            growth_schedule = forecast_result["growth_schedule"]
            _growth_str     = f"{base_growth:.1%} projected FCF growth" if base_growth else ""
            _update_progress(3, _growth_str)

            dcf_res = dcf_engine.intrinsic_value_per_share(
                projected_fcfs=projected, terminal_fcf_norm=terminal_norm,
                total_debt=enriched["total_debt"], total_cash=enriched["total_cash"],
                shares_outstanding=enriched["shares"],
                current_price=enriched["price"], ticker=ticker_input,
            )

            _update_progress(4)
            # Scenarios — use the SAME fcf_base the main forecaster used
            # Reconstruct it from projections[0] / (1 + growth_schedule[0])
            # to guarantee scenarios start from identical FCF base as main DCF
            if len(projected) > 0 and len(growth_schedule) > 0 and growth_schedule[0] > -1:
                fcf_base_for_scenarios = projected[0] / (1 + growth_schedule[0])
            else:
                fcf_base_for_scenarios = fcf_base if fcf_base > 0 else enriched.get("latest_fcf", 1e6)

            scenarios = run_scenarios(
                enriched=enriched,
                fcf_base=fcf_base_for_scenarios,
                base_growth=base_growth,
                base_wacc=wacc,
                base_terminal_g=terminal_g,
                total_debt=enriched["total_debt"],
                total_cash=enriched["total_cash"],
                shares=enriched["shares"],
                current_price=enriched["price"],
                years=forecast_yrs,
            )

            # Base IV and price (before moat adjustment)
            iv_n    = dcf_res.get("intrinsic_value_per_share", 0)
            price_n = enriched["price"]

            # ── PE Crosscheck — blend DCF with sector PE IV ────
            # Dashboard was using raw DCF only. Now blends like screener does.
            try:
                from screener.valuation_crosscheck import compute_pe_based_iv, blend_dcf_pe, get_eps
                _sector   = enriched.get("sector", "general")
                _reliable = enriched.get("dcf_reliable", True)
                _eps      = get_eps(enriched)
                _pe_iv    = compute_pe_based_iv(_eps, _sector, scenario="base",
                                                growth=enriched.get("revenue_growth", None))

                if _reliable and iv_n > 0 and _pe_iv > 0:
                    iv_n = blend_dcf_pe(iv_n, _pe_iv, _sector)
                elif _pe_iv > 0 and not _reliable:
                    # Banks/NBFCs: DCF unreliable — only use PE if EPS is credible
                    # Extra sanity: implied PE must be 5-60x
                    _implied_pe = enriched.get("price", 0) / _eps if _eps > 0 else 0
                    if 5 <= _implied_pe <= 60:
                        iv_n = _pe_iv
                    # else: leave iv_n = 0 so N/A is displayed
                enriched["pe_iv"]  = _pe_iv
                enriched["dcf_iv"] = dcf_res.get("intrinsic_value_per_share", 0)
            except Exception as _pe_err:
                enriched["pe_iv"]  = 0
                enriched["dcf_iv"] = iv_n

            # Moat — compute first, then adjust IV before investment plan
            iv_n_moat, moat_grade, moat_score, moat_adj = moat_tab.compute(
                enriched=enriched,
                wacc=wacc,
                base_growth=base_growth,
                terminal_g=terminal_g,
                iv_n=iv_n,
            )
            moat_types   = enriched.get("moat_types",   [])
            moat_summary = enriched.get("moat_summary", "")

            # ── Confidence-based IV haircut ────────────────────
            # When confidence flags major warnings, reduce IV to reflect
            # the uncertainty. This is the IB approach: widen the range
            # and reduce point estimate when data quality is poor.
            confidence     = compute_confidence_score(enriched)
            _conf_warnings = confidence.get("warnings", [])
            _iv_haircut = 1.0
            _rev_growth_conf = enriched.get("revenue_growth", 0) or 0
            for _w in _conf_warnings:
                if "DECLINING" in _w:
                    _iv_haircut *= 0.55   # revenue declining: cut IV 45%
                elif "spike" in _w.lower():
                    # Skip FCF spike haircut when revenue is structurally growing
                    # (e.g. NVDA: AI revenue surge is not a one-time item)
                    if _rev_growth_conf <= 0.40:
                        _iv_haircut *= 0.60   # genuine one-time FCF spike: cut IV 40%
                elif "decelerat" in _w.lower():
                    # Only haircut when growth has turned actually negative,
                    # not just slowing from high base (e.g. GOOGL maturing)
                    if _rev_growth_conf < 0:
                        _iv_haircut *= 0.80   # true revenue contraction: cut IV 20%
            if _iv_haircut < 1.0:
                iv_n_moat = iv_n_moat * _iv_haircut

            # Apply same moat + confidence adjustments to scenario IVs
            # so Bear/Base/Bull are consistent with the headline IV
            _moat_iv_delta  = moat_adj.get("iv_delta_pct", 0) / 100 if "moat_adj" in locals() else 0.0
            _moat_mult      = 1 + _moat_iv_delta
            _scenario_mult  = _moat_mult * _iv_haircut
            if _scenario_mult != 1.0:
                for _sname in scenarios:
                    scenarios[_sname]["iv"]      = scenarios[_sname]["iv"] * _scenario_mult
                    scenarios[_sname]["mos_pct"]  = margin_of_safety(
                        scenarios[_sname]["iv"], price_n) * 100
                    scenarios[_sname]["mos"]      = scenarios[_sname]["mos_pct"] / 100

            # Investment plan — use moat-adjusted IV so buy/target/SL reflect moat
            mos      = margin_of_safety(iv_n_moat, price_n)

            # Insider sentiment → small MoS nudge (10% weight)
            _insider_data = raw.get("finnhub_insider", {})
            _insider_sent = _insider_data.get("sentiment", "NEUTRAL")
            _INSIDER_ADJ  = {
                "STRONG BUY":  +0.04, "BUY": +0.02, "NEUTRAL": 0.0,
                "SELL": -0.02, "STRONG SELL": -0.04,
            }
            _insider_adj  = _INSIDER_ADJ.get(_insider_sent, 0.0)

            sig      = assign_signal(mos, dcf_res.get("suspicious", False), forecast_result.get("reliable", True), _insider_adj)
            inv_plan = generate_investment_plan(enriched, price_n, iv_n_moat, mos)
            _show_action_plan = can("action_plan")

            # Keep original iv_n for DCF display, use iv_n_moat for targets
            iv_n = iv_n_moat

            # MC
            mc_result = {}
            if run_mc:
                mc_result = monte_carlo_valuation(
                    enriched=enriched, forecaster=forecaster,
                    total_debt=enriched["total_debt"], total_cash=enriched["total_cash"],
                    shares_outstanding=enriched["shares"], current_price=price_n, base_wacc=wacc,
                )

        # FX
        native_ccy = raw.get("native_ccy", "USD")
        fx = get_fx_rate(native_ccy, to_code)
        price_d = price_n * fx
        iv_d    = iv_n    * fx
        proj_d  = [v * fx for v in projected]
        pv_fcfs_d = [v * fx for v in dcf_res.get("pv_fcfs", [])]
        pv_tv_d   = dcf_res.get("pv_tv", 0) * fx

        # ── Save to session state for Financial Statements tab ─
        # ── End of fresh compute block ───────────────────────────

        # Store computed wacc/terminal_g INTO enriched for cache restore
        enriched["wacc_used"]      = wacc
        enriched["terminal_g_used"]= terminal_g
        enriched["wacc_source"]    = wacc_source
        # Live RF rate info — prefer from industry WACC (already adjusted),
        # fall back to CAPM wacc_data, then get directly from config cache
        _rf_info_store = (
            locals().get("_ind_info", {}).get("rf_rate_info")
            or wacc_data.get("rf_rate_info")
            or None
        )
        if _rf_info_store is None:
            try:
                from utils.config import fetch_risk_free_rate as _cfg_fetch_rf
                _is_ind = ticker_input.endswith(".NS") or ticker_input.endswith(".BO")
                _rf_info_store = _cfg_fetch_rf("india" if _is_ind else "us")
            except Exception:
                _rf_info_store = {}
        st.session_state["_rf_rate_info"] = _rf_info_store
        st.session_state["fin_enriched"]  = enriched
        st.session_state["fin_ticker"]    = ticker_input
        st.session_state["fin_fx"]        = fx
        st.session_state["fin_to_code"]   = to_code
        st.session_state["fin_sym"]       = sym
        st.session_state["fin_raw"]       = raw
        # Clear progress card
        if "_prog_ph" in dir():
            try:
                _prog_ph.empty()
            except Exception:
                pass
        # Save all computed vars so redisplay works after rerun
        st.session_state["_dcf_res"]       = dcf_res
        st.session_state["_mc_result"]      = st.session_state.get("_mc_result", {})
        st.session_state["_forecast_result"]= forecast_result
        st.session_state["_scenarios"]      = scenarios
        st.session_state["_inv_plan"]       = inv_plan
        st.session_state["_confidence"]     = confidence
        st.session_state["_price_hist"]     = price_hist
        st.session_state["_wacc_data"]      = wacc_data
        st.session_state["_use_auto_wacc"]  = use_auto_wacc
        st.session_state["_fx_rate"]        = fx_rate if "fx_rate" in dir() else 1.0
        st.session_state["_forecast_yrs"]   = forecast_yrs
        st.session_state["_terminal_pct"]   = terminal_pct
        st.session_state["_insider_adj"]    = _insider_adj

        # Use human language signal helper for colors
        _h_label_m, sig_fg, sig_bg, sig_bd = sig_human(sig)
        mos_pct   = mos * 100
        mos_color = "#0D7A4E" if mos_pct > 20 else "#B8972A" if mos_pct > 0 else "#A62020"
        st.session_state["fin_mos_pct"] = mos_pct
        st.session_state["fin_signal"]  = sig

        # ── Success micro-animation (only on fresh analysis, not cache) ──
        if not _from_cache:
            if "Undervalued" in sig or "Near Fair" in sig:
                st.markdown("<style>.main .block-container{animation:yiq-flash-green 1.8s ease-out;}</style>", unsafe_allow_html=True)
            elif "Overvalued" in sig:
                st.markdown("<style>.main .block-container{animation:yiq-flash-red 1.8s ease-out;}</style>", unsafe_allow_html=True)
        st.session_state["fin_iv_d"]    = iv_d
        # ── Push to Morning Brief history ─────────────────────
        _mb_name = enriched.get("company_name", ticker_input) or ticker_input
        push_analysis_to_history(ticker_input, _mb_name, price_d, iv_d, mos_pct, sig)
        st.session_state["_show_morning_brief"] = False   # show results, not brief
        # ── Analytics tracking ────────────────────────────────
        track_analysis(
            user_email=st.session_state.get("auth_email", ""),
            tier=tier(),
            ticker=ticker_input,
            signal=sig,
            mos_pct=mos_pct,
            wacc=wacc,
        )
        mos_w     = min(max(abs(mos_pct), 2), 100)
        pt        = inv_plan["price_targets"]
        hp        = inv_plan["holding_period"]
        fs        = inv_plan["fundamental"]
        # Tier-based display flags
        _show_plan      = can("action_plan")
        _show_quality   = can("quality_score")
        _show_scenarios = can("scenarios")
        _show_sensitive = can("sensitivity")
        _show_mc        = can("monte_carlo")
        suspicious = dcf_res.get("suspicious", False)

        # ── Variables needed by Valuation tab (used to be set inside _sub_ov) ──
        years_labels         = [f"Y{i+1}" for i in range(forecast_yrs)]
        years                = forecast_yrs
        growth_schedule      = forecast_result.get("growth_schedule", [])
        base_growth          = forecast_result.get("base_growth", 0)
        fcf_base             = forecast_result.get("fcf_base", 0)
        projected_fcfs       = projected
        native_ccy           = raw.get("native_ccy", "USD")   if raw else "USD"
        company_name         = raw.get("company_name", ticker_input) if raw else ticker_input
        shares_outstanding   = enriched.get("shares", 0)
        shares               = shares_outstanding
        total_debt           = enriched.get("total_debt", 0)
        total_cash           = enriched.get("total_cash", 0)
        current_price        = enriched.get("price", 0)
        _sector              = enriched.get("sector", "general")
        sector               = _sector
        _pe_iv               = enriched.get("pe_iv", 0)
        moat_adj             = st.session_state.get("fin_moat_adj", {})
        iv_delta_pct         = moat_adj.get("iv_delta_pct", 0)
        mc_result            = st.session_state.get("_mc_result", {})
        pv_tv_d              = dcf_res.get("pv_tv", 0) * fx
        proj_d               = [v * fx for v in projected]
        pv_fcfs_d            = [v * fx for v in dcf_res.get("pv_fcfs", [])]
        if len(projected) > 0 and len(growth_schedule) > 0 and growth_schedule[0] > -1:
            fcf_base_for_scenarios = projected[0] / (1 + growth_schedule[0])
        else:
            fcf_base_for_scenarios = fcf_base if fcf_base > 0 else enriched.get("latest_fcf", 1e6)

        # ══════════════════════════════════════════════════
        # PILL NAVIGATION
        # ══════════════════════════════════════════════════
        # Stable keys (no emojis) prevent serialisation issues across Streamlit versions
        _pill_options = [
            ("⚡ Summary",          "summary"),
            ("💰 Valuation",        "valuation"),
            ("🏗️ Fundamentals",    "fundamentals"),
            ("📡 Street Consensus", "consensus"),
            ("🤖 Ask AI",           "ask_ai"),
        ]
        if "active_section" not in st.session_state:
            st.session_state["active_section"] = "summary"

        _cols = st.columns(len(_pill_options))
        for _i, (_pill_label, _pill_key) in enumerate(_pill_options):
            with _cols[_i]:
                if st.button(
                    _pill_label,
                    key=f"pill_{_i}",
                    width="stretch",
                    type="primary" if st.session_state["active_section"] == _pill_key else "secondary",
                ):
                    st.session_state["active_section"] = _pill_key
                    st.rerun()

        st.divider()
        _active = st.session_state["active_section"]

        # ══════════════════════════════════════════════════════════
        # QUICK STATS STRIP
        # Visible on every section — P/E, EV/EBITDA, FCF Yield,
        # Beta, Div Yield, 52W Range always in context.
        # ══════════════════════════════════════════════════════════
        if raw and enriched:
            _qs_pe      = raw.get("forward_pe") or raw.get("pe_ratio")
            _qs_eveb    = raw.get("ev_to_ebitda")
            _qs_beta    = raw.get("fh_beta")
            _qs_div     = raw.get("dividend_yield") or raw.get("fh_div_yield") or 0
            _qs_hi52    = (raw.get("fh_52w_high") or 0) * fx
            _qs_lo52    = (raw.get("fh_52w_low")  or 0) * fx
            _qs_fcf_raw = raw.get("yahoo_fcf_ttm") or 0
            _qs_mktcap  = (price_n * enriched.get("shares", 0)) if price_n else 0
            _qs_fcf_yld = (
                _qs_fcf_raw / _qs_mktcap * 100
                if _qs_mktcap > 0 and _qs_fcf_raw else None
            )

            with st.container(border=True):
                _qs_cols = st.columns(6)

                # P/E
                with _qs_cols[0]:
                    _qs_pe_str = f"{_qs_pe:.1f}×" if (_qs_pe and 0 < _qs_pe < 500) else "—"
                    st.metric("P/E", _qs_pe_str)

                # EV/EBITDA
                with _qs_cols[1]:
                    _qs_eveb_str = f"{_qs_eveb:.1f}×" if (_qs_eveb and 0 < _qs_eveb < 300) else "—"
                    st.metric("EV/EBITDA", _qs_eveb_str)

                # FCF Yield
                with _qs_cols[2]:
                    _qs_fcfy_str = f"{_qs_fcf_yld:.1f}%" if _qs_fcf_yld is not None else "—"
                    st.metric("FCF Yield", _qs_fcfy_str)

                # Beta
                with _qs_cols[3]:
                    _qs_beta_str = f"{_qs_beta:.2f}" if _qs_beta else "—"
                    st.metric("Beta", _qs_beta_str)

                # Div Yield
                with _qs_cols[4]:
                    _qs_div_str = f"{_qs_div * 100:.1f}%" if _qs_div else "—"
                    st.metric("Div Yield", _qs_div_str)

                # 52W Range
                with _qs_cols[5]:
                    _qs_52w_str = (
                        f"{sym}{_qs_lo52:,.0f} – {sym}{_qs_hi52:,.0f}"
                        if _qs_hi52 and _qs_lo52 else "—"
                    )
                    st.metric("52W Range", _qs_52w_str)

        if _active == "summary":

            # ══════════════════════════════════════════════════════════
            # SHARED PREP — values used across all 4 layers
            # ══════════════════════════════════════════════════════════
            if not enriched:
                st.warning("Analysis data unavailable. Please run a new analysis.")
                st.stop()

            _display_name  = company_name if 'company_name' in dir() and company_name else ticker_input
            _h_label, sig_fg, sig_bg, sig_bd = sig_human(sig)
            _fs_color_map  = {"STRONG":"#0D7A4E","GOOD":"#2563EB","AVERAGE":"#B8972A","WEAK":"#A62020"}
            _fs_color      = _fs_color_map.get(fs.get("grade",""), "#4A5E7A")
            _fs_label      = fs.get("grade","N/A")
            mos_color      = "#0D7A4E" if mos_pct > 20 else "#B8972A" if mos_pct > 0 else "#A62020"
            mos_w          = min(max(abs(mos_pct), 2), 100)

            # ── revenue/margin/growth helpers ─────────────────────────
            _rev_growth    = enriched.get("revenue_growth", 0) * 100
            _op_margin     = enriched.get("op_margin", 0) * 100
            _fcf_growth    = enriched.get("fcf_growth", 0) * 100
            _moat_grade    = enriched.get("moat_grade", "N/A")
            _moat_score    = enriched.get("moat_score", 0)
            _moat_types    = (" · ".join(enriched.get("moat_types", [])[:2]) or "No identifiable moat")
            _moat_summary  = enriched.get("moat_summary", "") or ""
            _conf_warnings = confidence.get("warnings", [])
            suspicious     = dcf_res.get("suspicious", False)

            # ── quality / growth / risk plain signals ─────────────────
            _qual_label = {"STRONG":"Strong","GOOD":"Good","AVERAGE":"Average","WEAK":"Weak"}.get(_fs_label,"—")
            _qual_color = _fs_color

            _growth_label = (
                "Improving"   if _rev_growth > 10 and _fcf_growth > 5 else
                "Stable"      if _rev_growth > 0  and _fcf_growth > -5 else
                "Declining"
            )
            _growth_color = (
                "#0D7A4E" if _growth_label == "Improving" else
                "#1D4ED8" if _growth_label == "Stable"    else "#B91C1C"
            )

            _risk_label = (
                "Lower"  if _op_margin > 20 and _moat_grade in ("Wide","Narrow") else
                "Higher" if _op_margin < 8  or suspicious else
                "Moderate"
            )
            _risk_color = (
                "#0D7A4E" if _risk_label == "Lower"   else
                "#B91C1C" if _risk_label == "Higher"  else "#B45309"
            )

            _val_label = (
                "Undervalued" if mos_pct > 10 else
                "Fairly priced" if mos_pct > -10 else
                "Overvalued"
            )
            _val_color = (
                "#0D7A4E" if _val_label == "Undervalued" else
                "#1D4ED8" if _val_label == "Fairly priced" else "#B91C1C"
            )

            # ── moat plain description ────────────────────────────────
            _moat_plain = {
                "Wide":   "Strong competitive advantage",
                "Narrow": "Some competitive advantage",
                "None":   "No clear competitive advantage",
            }.get(_moat_grade, "Competitive advantage unknown")

            # ══════════════════════════════════════════════════════════
            # LAYER 1 — 5-Second Insight (Hero card)
            # ══════════════════════════════════════════════════════════
            _l1_insight = (
                f"Our model estimates {_display_name} is trading "
                f"<strong>~{mos_pct:.0f}% below</strong> its calculated fair value."
                if mos_pct > 5 else
                f"Our model estimates {_display_name} is trading "
                f"<strong>near its calculated fair value</strong>."
                if mos_pct > -5 else
                f"Our model estimates {_display_name} is trading "
                f"<strong>~{abs(mos_pct):.0f}% above</strong> its calculated fair value."
            )

            # ── Market open detection ────────────────────────────────
            import pytz as _pytz
            _now_et   = datetime.now(_pytz.timezone('US/Eastern'))
            _is_mkt   = (
                _now_et.weekday() < 5 and
                9*60+30 <= _now_et.hour*60+_now_et.minute <= 16*60
            )
            _mkt_dot   = "\u25cf" if _is_mkt else "\u25cb"
            _mkt_label = "LIVE" if _is_mkt else "CLOSED"
            _mkt_color = "#10B981" if _is_mkt else "#64748B"
            _pulse_css = "animation:_wlPulse 1.8s ease infinite;" if _is_mkt else ""

            # ── Day change ────────────────────────────────────────────
            _day_chg_pct = (raw.get("day_change_pct", 0) or 0) if raw else 0
            _chg_sym_h   = "\u25b2" if _day_chg_pct >= 0 else "\u25bc"
            _chg_col_h   = "#10B981" if _day_chg_pct >= 0 else "#EF4444"

            # ── FX label ─────────────────────────────────────────────
            _from_code = native_ccy if "native_ccy" in dir() else "USD"
            _fx_label  = f"FX: {_from_code} \u2192 {to_code}" if _from_code != to_code else f"CCY: {to_code}"

            # ── MoS badge ────────────────────────────────────────────
            _upside_str   = (f"+{mos_pct:.1f}%" if mos_pct >= 0 else f"{mos_pct:.1f}%")
            _mos_pill_bg  = "#ECFDF5" if mos_pct >= 10 else "#FFFBEB" if mos_pct >= 0 else "#FEF2F2"
            _mos_pill_col = "#059669" if mos_pct >= 10 else "#D97706" if mos_pct >= 0 else "#DC2626"

            _hero = (
                "<style>"
                "@keyframes _wlPulse {"
                "  0%,100% { box-shadow:0 0 0 0 rgba(16,185,129,0.45); }"
                "  70%      { box-shadow:0 0 0 5px rgba(16,185,129,0);  }"
                "}"
                ".yiq-terminal-hdr {"
                "  position:sticky; top:0; z-index:100;"
                "  background:rgba(13,20,36,0.97);"
                "  backdrop-filter:blur(10px);"
                "  -webkit-backdrop-filter:blur(10px);"
                "  border-radius:12px;"
                "  overflow:hidden;"
                "  margin-bottom:10px;"
                "  border:1px solid rgba(255,255,255,0.07);"
                "  box-shadow:0 8px 32px rgba(0,0,0,0.3);"
                "}"
                ".yiq-terminal-hdr::after {"
                "  content:'';"
                "  display:block;"
                "  height:2px;"
                "  background:linear-gradient(90deg,#00b4d8 0%,#3b82f6 50%,transparent 100%);"
                "}"
                "</style>"
                '<div class="yiq-terminal-hdr">' +
                '<div style="display:flex;align-items:center;justify-content:space-between;' +
                'padding:12px 22px 10px;gap:16px;flex-wrap:wrap;">' +

                # LEFT: Branding
                '<div style="display:flex;flex-direction:column;gap:3px;min-width:140px;">' +
                '<div style="display:flex;align-items:center;gap:8px;">' +
                '<span style="font-family:&apos;Barlow Condensed&apos;,sans-serif;font-size:24px;' +
                'font-weight:700;color:#E2E8F0;letter-spacing:-0.01em;line-height:1;">' +
                'Yield<span style="color:#00b4d8;">IQ</span></span>' +
                '<span style="font-size:10px;font-weight:700;padding:2px 7px;' +
                'background:rgba(0,180,216,0.14);border:1px solid rgba(0,180,216,0.32);' +
                'border-radius:4px;color:#00b4d8;letter-spacing:0.07em;">v6</span>' +
                '</div>' +
                '<div style="font-family:&apos;IBM Plex Mono&apos;,monospace;font-size:9px;' +
                'color:#475569;letter-spacing:0.16em;text-transform:uppercase;">' +
                'Institutional DCF Platform</div></div>' +

                # CENTRE: Ticker + Price
                '<div style="display:flex;flex-direction:column;align-items:center;gap:3px;' +
                'flex:1;min-width:200px;">' +
                '<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;justify-content:center;">' +
                f'<span style="font-family:&apos;IBM Plex Mono&apos;,monospace;font-size:21px;font-weight:700;' +
                f'color:#FFFFFF;background:rgba(255,255,255,0.07);padding:4px 14px;border-radius:7px;' +
                f'letter-spacing:0.07em;border:1px solid rgba(255,255,255,0.1);">{ticker_input}</span>' +
                '<div style="display:flex;flex-direction:column;align-items:flex-start;gap:2px;">' +
                f'<div style="display:flex;align-items:center;gap:7px;">' +
                f'<span style="width:8px;height:8px;border-radius:50%;flex-shrink:0;' +
                f'background:{_mkt_color};display:inline-block;{_pulse_css}"></span>' +
                f'<span style="font-family:&apos;IBM Plex Mono&apos;,monospace;font-size:19px;' +
                f'font-weight:700;color:#F1F5F9;letter-spacing:0.02em;">{fmts(price_d, sym)}</span>' +
                f'</div>' +
                f'<div style="font-size:12px;font-weight:600;color:{_chg_col_h};padding-left:15px;">' +
                f'{_chg_sym_h} {abs(_day_chg_pct):.2f}%&nbsp;day change</div>' +
                '</div></div>' +
                f'<div style="font-size:11px;color:#475569;letter-spacing:0.02em;">{_display_name}</div>' +
                '</div>' +

                # RIGHT: metric pills
                '<div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;' +
                'justify-content:flex-end;min-width:260px;">' +
                f'<div title="{ob_tooltip("wacc")}" style="padding:5px 10px;background:rgba(255,255,255,0.05);' +
                f'border:1px solid rgba(255,255,255,0.1);border-radius:6px;' +
                f'font-family:&apos;IBM Plex Mono&apos;,monospace;font-size:11px;' +
                f'color:#CBD5E1;white-space:nowrap;cursor:help;">WACC&nbsp;{wacc:.1%}</div>' +
                f'<div title="{ob_tooltip("terminal_g")}" style="padding:5px 10px;background:rgba(255,255,255,0.05);' +
                f'border:1px solid rgba(255,255,255,0.1);border-radius:6px;' +
                f'font-family:&apos;IBM Plex Mono&apos;,monospace;font-size:11px;' +
                f'color:#CBD5E1;white-space:nowrap;cursor:help;">TERM.G&nbsp;{terminal_g:.1%}</div>' +
                f'<div style="padding:5px 10px;background:rgba(255,255,255,0.05);' +
                f'border:1px solid rgba(255,255,255,0.1);border-radius:6px;' +
                f'font-family:&apos;IBM Plex Mono&apos;,monospace;font-size:11px;' +
                f'color:#CBD5E1;white-space:nowrap;">{_fx_label}</div>' +
                f'<div style="padding:5px 10px;background:rgba(255,255,255,0.05);' +
                f'border:1px solid rgba(255,255,255,0.1);border-radius:6px;' +
                f'font-family:&apos;IBM Plex Mono&apos;,monospace;font-size:11px;' +
                f'color:{_mkt_color};font-weight:600;white-space:nowrap;">' +
                f'{_mkt_dot}&nbsp;{_mkt_label}</div>' +
                f'<div title="{ob_tooltip("mos")}" style="padding:5px 12px;background:{_mos_pill_bg};' +
                f'border:1px solid {_mos_pill_col}44;border-radius:6px;' +
                f'font-family:&apos;IBM Plex Mono&apos;,monospace;font-size:12px;' +
                f'font-weight:700;color:{_mos_pill_col};white-space:nowrap;cursor:help;">' +
                f'MoS&nbsp;{_upside_str}</div>' +
                '</div>' +
                '</div></div>'
            )
            st.html(_hero)

            # ── "Last updated" timestamp ──────────────────────────
            _fetch_ts = st.session_state.get("last_fetch_time")
            if _fetch_ts:
                _mins_ago = int((_time.time() - _fetch_ts) / 60)
                if _mins_ago < 1:
                    st.caption("🟢 Updated just now · Live data")
                elif _mins_ago < 60:
                    st.caption(f"⏱ Updated {_mins_ago} min ago · Click Analyse to refresh")
                else:
                    st.caption("📦 Cached data · Click Analyse to refresh")

            # ── MoS context caption ───────────────────────────────
            if mos_pct >= 30:
                st.caption("📊 Deep discount to model fair value — historically strong entry range per DCF")
            elif mos_pct >= 10:
                st.caption("📊 Moderate discount to model estimated fair value")
            elif mos_pct >= -10:
                st.caption("⚖️ Trading near model estimated fair value")
            else:
                st.caption("⚠️ Trading above model estimated fair value — premium to DCF estimate")

            # ── 🎓 First-run tooltip banner ───────────────────────
            if ob_show_tooltips():
                st.html("""
                <div style="background:rgba(29,78,216,0.08);border:1px solid rgba(29,78,216,0.28);
                            border-radius:10px;padding:10px 16px;margin:8px 0;
                            display:flex;align-items:center;gap:12px;">
                  <div style="font-size:20px;flex-shrink:0;">💡</div>
                  <div style="font-size:11px;color:#93C5FD;line-height:1.7;">
                    <strong style="color:#60A5FA;">First-run tip:</strong>
                    Hover over the <strong>WACC</strong>, <strong>TERM.G</strong> and
                    <strong>MoS</strong> pills above to see what each metric means.
                    Check the sidebar Advanced Settings to adjust DCF assumptions.
                    Use the <strong>▶ Resume Tutorial</strong> button in the sidebar to continue your tour.
                  </div>
                </div>
                """)

            # ── 📌 Add to Watchlist ───────────────────────────────
            _in_wl    = is_in_watchlist(ticker_input)
            _wl_label = "✅ Already in Watchlist — click to update" if _in_wl else "📌 Add to Watchlist"
            with st.expander(_wl_label, expanded=False):
                _wl_c1, _wl_c2 = st.columns(2)
                with _wl_c1:
                    _wl_target = st.number_input(
                        "Target Price",
                        value=float(round(iv_d, 2)),
                        min_value=0.0,
                        step=0.5,
                        key="wl_target_price",
                        help="Pre-filled with your DCF intrinsic value estimate",
                    )
                with _wl_c2:
                    _wl_mos_thresh = st.slider(
                        "Alert when MoS exceeds",
                        min_value=10, max_value=50,
                        value=20, step=5,
                        format="%d%%",
                        key="wl_mos_thresh",
                    )
                _wl_notes = st.text_area(
                    "Notes (optional)",
                    value="",
                    key="wl_notes",
                    placeholder="Why I'm watching this… e.g. waiting for Q3 results",
                    height=80,
                )
                if st.button("💾 Save to Watchlist", key="wl_save_btn",
                             width='stretch', type="primary"):
                    _wl_ok = add_to_watchlist(
                        ticker               = ticker_input,
                        company_name         = company_name,
                        added_price          = price_d,
                        target_price         = _wl_target,
                        alert_mos_threshold  = float(_wl_mos_thresh),
                        notes                = _wl_notes,
                    )
                    if _wl_ok:
                        st.success(f"✅ **{ticker_input}** saved to watchlist! Switch to the 📊 Watchlist tab to track it.")
                        track_event(st.session_state.get("auth_email",""), tier(), "watchlist_add", {"ticker": ticker_input})
                    else:
                        st.error("Could not save to watchlist — please try again.")

            # model warnings
            if suspicious:
                st.warning("⚠️ Our model flagged unusual patterns in this company's financials. Treat this analysis with extra caution.")
            for _w in _conf_warnings:
                st.warning(f"⚠️ {_w}")

            # ══════════════════════════════════════════════════════════
            # LAYER 2 — Quick Signals (4 pill badges)
            # ══════════════════════════════════════════════════════════
            st.html('<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;'
                    'letter-spacing:0.14em;margin:16px 0 8px;padding-left:2px;">Quick signals</div>')

            _sig_col1, _sig_col2, _sig_col3, _sig_col4 = st.columns(4)

            def _pill(col, label, value, color, bg, help_txt=""):
                _h = (
                    '<div style="background:BG;border:1px solid COLOR;border-radius:10px;'
                    'padding:14px 16px;text-align:center;">'
                    '<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;'
                    'letter-spacing:0.11em;margin-bottom:6px;">LABEL</div>'
                    '<div style="font-size:18px;font-weight:700;color:COLOR;">VALUE</div>'
                    '</div>'
                )
                _h = _h.replace("BG", bg).replace("COLOR", color).replace("LABEL", label).replace("VALUE", value)
                col.html(_h)

            _pill(_sig_col1, "Valuation",       _val_label,    _val_color,    "#F8FAFC")
            _pill(_sig_col2, "Company quality", _qual_label,   _qual_color,   "#F8FAFC")
            _pill(_sig_col3, "Growth",          _growth_label, _growth_color, "#F8FAFC")
            _pill(_sig_col4, "Risk level",      _risk_label,   _risk_color,   "#F8FAFC")

            # ══════════════════════════════════════════════════════════
            # LAYER 2b — YieldIQ Composite Score
            # ══════════════════════════════════════════════════════════
            _ys_pio    = enriched.get("piotroski_score") or int(fs.get("score", 50) / 100 * 9)
            _ys_pt     = (raw.get("finnhub_price_target") or {}) if raw else {}
            _ys_pt_mean = float(_ys_pt.get("mean", 0)) * fx if _ys_pt.get("mean") else 0
            _ys_upside  = ((_ys_pt_mean - price_d) / price_d * 100) if price_d and _ys_pt_mean else 0
            _ys = compute_yieldiq_score(
                mos_pct        = mos_pct,
                piotroski      = int(_ys_pio) if _ys_pio else 0,
                moat_grade     = moat_grade,
                rev_growth     = _rev_growth,
                analyst_upside = _ys_upside,
            )
            _grade_colors = {
                "A+": "#16a34a", "A": "#22c55e", "B+": "#65a30d", "B": "#84cc16",
                "C+": "#ca8a04", "C": "#f59e0b", "D": "#dc2626",
            }
            _gc = _grade_colors.get(_ys["grade"], "#94a3b8")

            st.html('<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;'
                    'letter-spacing:0.14em;margin:16px 0 8px;padding-left:2px;">YieldIQ Composite Score</div>')
            _ysc1, _ysc2 = st.columns([1, 2])
            with _ysc1:
                st.html(
                    f'<div style="text-align:center;padding:20px 16px;border-radius:12px;'
                    f'background:#0a1628;border:2px solid {_gc};">'
                    f'<div style="font-size:48px;font-weight:900;color:{_gc};line-height:1;">'
                    f'{_ys["score"]}</div>'
                    f'<div style="font-size:24px;font-weight:700;color:{_gc};margin-top:4px;">'
                    f'{_ys["grade"]}</div>'
                    f'<div style="font-size:11px;color:#94a3b8;margin-top:6px;'
                    f'text-transform:uppercase;letter-spacing:.08em;">YieldIQ Score</div>'
                    f'</div>'
                )
            with _ysc2:
                for _comp_lbl, _comp_pts in _ys["components"].items():
                    _comp_max = int(_comp_lbl.split("(")[1].replace("pts)", "")) or 1
                    st.progress(
                        _comp_pts / _comp_max,
                        text=f"{_comp_lbl}: **{_comp_pts}** / {_comp_max}",
                    )
            # Contextual note when valuation score is 0
            if _ys["components"].get("Valuation (40pts)", 0) == 0 and mos_pct is not None:
                _mos_str = f"{mos_pct:+.1f}%"
                st.caption(
                    f"ℹ️ Valuation: 0 / 40 pts — the stock trades {_mos_str} vs the model's fair value estimate. "
                    "Valuation points require the stock to trade at or below the model estimate (margin of safety ≥ −15%). "
                    "This reflects current pricing, not business quality."
                )
            st.caption(
                "⚠️ YieldIQ Score is a model output combining DCF, fundamentals, "
                "growth, and analyst data. It is not investment advice."
            )

            # ══════════════════════════════════════════════════════════
            # LAYER 3 — Why? (plain-language expandable, Simple mode only)
            # ══════════════════════════════════════════════════════════
            if not pro_mode:
                with st.expander("💡 Why These Signals? — Plain English Explanation"):
                    _why_val = (
                        f"Our model estimates this stock's fair value at **{fmts(iv_d, sym)}**. "
                        f"At the current price of **{fmts(price_d, sym)}**, it appears **{_val_label.lower()}** "
                        f"by around **{abs(mos_pct):.0f}%**."
                        if abs(mos_pct) > 2 else
                        f"The stock is trading very close to our estimated fair value of **{fmts(iv_d, sym)}**."
                    )
                    _why_qual = (
                        f"We rate this company's financial health as **{_qual_label}** "
                        f"(score: {fs.get('score',0)}/100). "
                        + ("It shows strong profitability, low debt, and consistent cash generation." if _fs_label == "STRONG" else
                           "It shows solid fundamentals with some areas to watch." if _fs_label == "GOOD" else
                           "It has average financial health — neither particularly strong nor weak." if _fs_label == "AVERAGE" else
                           "It shows some financial weaknesses worth monitoring.")
                    )
                    _why_growth = (
                        f"Revenue is growing at **{_rev_growth:.1f}%** and cash flows are "
                        + ("growing strongly" if _fcf_growth > 10 else
                           "roughly stable" if _fcf_growth > -5 else
                           "under pressure")
                        + f" ({_fcf_growth:.1f}% growth). This signals **{_growth_label.lower()}** momentum."
                    )
                    _why_moat = (
                        f"**Competitive advantage:** {_moat_plain}. "
                        + (f"This is supported by: {_moat_types}." if _moat_types and _moat_grade != "None" else "")
                    )
                    _why_risk = (
                        f"We rate the risk as **{_risk_label.lower()}**. "
                        + ("The company has healthy margins and a strong competitive position." if _risk_label == "Lower" else
                           "There are some risks to watch, but the fundamentals are broadly solid." if _risk_label == "Moderate" else
                           "The company has thin margins or unusual financial patterns — exercise caution.")
                    )

                    st.markdown("**📊 Valuation**  \n" + _why_val)
                    st.markdown("**🏢 Company quality**  \n" + _why_qual)
                    st.markdown("**📈 Growth**  \n" + _why_growth)
                    st.markdown("**🔰 Competitive strength**  \n" + _why_moat)
                    st.markdown("**⚡ Risk**  \n" + _why_risk)

                    if _moat_summary:
                        st.markdown(f"*{_moat_summary[:200]}{'...' if len(_moat_summary) > 200 else ''}*")

            # ══════════════════════════════════════════════════════════
            # LAYER 3b — Model Insight Card (target / downside / horizon)
            # ══════════════════════════════════════════════════════════
            if pt.get("buy_price") and forecast_result.get("reliable", True):
                _ins_tgt_d   = (pt.get("target_price") or 0) * fx
                _ins_sl_d    = (pt.get("stop_loss")    or 0) * fx
                _ins_price_d = price_d
                _ins_upside  = ((_ins_tgt_d - _ins_price_d) / _ins_price_d * 100) if _ins_price_d > 0 else 0
                _ins_down    = pt.get("sl_pct", 12)
                _ins_rr      = pt.get("rr_ratio", 0)
                _ins_hp      = (hp.get("label","Long-term") or "Long-term").replace("'","&#39;").replace('"','&quot;')
                _ins_summary = (inv_plan.get("summary","") or "")[:220].replace("'","&#39;").replace('"','&quot;').replace('<','&lt;').replace('>','&gt;')

                if _ins_mos_pct := mos_pct:
                    pass
                _ins_badge_color, _ins_badge_bg, _ins_badge_bd, _ins_gauge_color = (
                    ("#0D7A4E","#F0FDF4","#BBF7D0","#0D7A4E") if mos_pct >= 20 else
                    ("#B45309","#FFFBEB","#FDE68A","#B45309") if mos_pct >= 5  else
                    ("#1D4ED8","#EFF6FF","#BFDBFE","#1D4ED8") if mos_pct >= -5 else
                    ("#B91C1C","#FEF2F2","#FECACA","#B91C1C")
                )
                _ins_badge_txt = (
                    f"{mos_pct:.0f}% below estimated fair value"   if mos_pct >= 5  else
                    "Near estimated fair value"                      if mos_pct >= -5 else
                    f"{abs(mos_pct):.0f}% above estimated fair value"
                )
                _ins_gauge_w  = min(max(abs(mos_pct), 2), 100)
                _ins_ret_str  = ("+" if _ins_upside >= 0 else "") + f"{_ins_upside:.1f}%"
                _ins_ret_color = "#0D7A4E" if _ins_upside >= 0 else "#B91C1C"
                _ins_risk_label = "Lower risk" if _ins_down <= 8 else "Moderate risk" if _ins_down <= 15 else "Higher risk"
                _ins_risk_color = "#0D7A4E"    if _ins_down <= 8 else "#B45309"       if _ins_down <= 15 else "#B91C1C"
                _ins_rr_txt = (
                    "The model estimate is " + _ins_ret_str +
                    " above current price against a downside scenario of -" + f"{_ins_down:.0f}%" +
                    (f" (model upside/downside ratio: {_ins_rr:.1f}×)." if _ins_rr else ".")
                )

                _tpl = (
                    '<div style="background:#FFFFFF;border:1px solid #E2E8F0;border-radius:12px;'
                    'overflow:hidden;margin-bottom:6px;">'

                    '<div style="padding:16px 24px 12px;border-bottom:1px solid #F1F5F9;'
                    'background:linear-gradient(135deg,#FAFBFD,#F4F7FC);">'
                    '<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;'
                    'letter-spacing:0.14em;margin-bottom:8px;">📊 Model Output</div>'
                    '<div style="display:inline-block;padding:5px 14px;background:BADGE_BG;'
                    'border:1px solid BADGE_BD;border-radius:20px;margin-bottom:10px;">'
                    '<span style="font-size:13px;font-weight:700;color:BADGE_COLOR;">BADGE_TXT</span>'
                    '</div>'
                    '<div style="height:7px;background:#F1F5F9;border-radius:4px;overflow:hidden;">'
                    '<div style="height:100%;width:GAUGE_W%;background:GAUGE_COLOR;border-radius:4px;"></div>'
                    '</div>'
                    '</div>'

                    '<div style="display:grid;grid-template-columns:1fr 1fr 1fr;">'

                    '<div style="padding:16px 20px;border-right:1px solid #F1F5F9;">'
                    '<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;'
                    'letter-spacing:0.11em;margin-bottom:5px;">DCF Model Estimate</div>'
                    '<div style="font-size:21px;font-weight:700;color:RET_COLOR;">TGT_D</div>'
                    '<div style="font-size:12px;color:RET_COLOR;margin-top:3px;font-weight:600;">'
                    'RET_STR gap vs current price and model estimate</div>'
                    '</div>'

                    '<div style="padding:16px 20px;border-right:1px solid #F1F5F9;">'
                    '<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;'
                    'letter-spacing:0.11em;margin-bottom:5px;">Downside scenario</div>'
                    '<div style="font-size:21px;font-weight:700;color:RISK_COLOR;">SL_D</div>'
                    '<div style="font-size:12px;color:RISK_COLOR;margin-top:3px;font-weight:500;">'
                    'RISK_LABEL · -DOWN% range</div>'
                    '</div>'

                    '<div style="padding:16px 20px;">'
                    '<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;'
                    'letter-spacing:0.11em;margin-bottom:5px;">DCF projection horizon</div>'
                    '<div style="font-size:16px;font-weight:600;color:#334155;margin-top:4px;">HP_LABEL</div>'
                    '<div style="font-size:12px;color:#94A3B8;margin-top:3px;">Model forecast period</div>'
                    '</div>'

                    '</div>'

                    '<div style="padding:10px 24px 8px;background:#FAFBFD;border-top:1px solid #F1F5F9;">'
                    '<div style="font-size:12px;color:#64748B;line-height:1.6;">RR_TXT</div>'
                    '</div>'

                    '<div style="padding:8px 24px 12px;background:#FAFBFD;">'
                    '<div style="font-size:11px;color:#94A3B8;line-height:1.5;font-style:italic;'
                    'border-top:1px dashed #E2E8F0;padding-top:8px;">'
                    '⚠️ Model-based estimate for informational purposes only. '
                    'Not a price target, return projection, or investment advice. '
                    'YieldIQ is not a registered investment adviser.'
                    '</div>'
                    '</div>'
                    '</div>'
                )
                _tpl = (
                    _tpl
                    .replace("BADGE_BG",    _ins_badge_bg)
                    .replace("BADGE_BD",    _ins_badge_bd)
                    .replace("BADGE_COLOR", _ins_badge_color)
                    .replace("BADGE_TXT",   _ins_badge_txt)
                    .replace("GAUGE_W",     str(int(_ins_gauge_w)))
                    .replace("GAUGE_COLOR", _ins_gauge_color)
                    .replace("RET_COLOR",   _ins_ret_color)
                    .replace("TGT_D",       fmts(_ins_tgt_d, sym))
                    .replace("RET_STR",     _ins_ret_str)
                    .replace("RISK_COLOR",  _ins_risk_color)
                    .replace("SL_D",        fmts(_ins_sl_d, sym))
                    .replace("RISK_LABEL",  _ins_risk_label)
                    .replace("DOWN",        f"{_ins_down:.0f}")
                    .replace("HP_LABEL",    _ins_hp)
                    .replace("RR_TXT",      _ins_rr_txt)
                )
                st.html(_tpl)

                if _ins_summary:
                    with st.expander("🧮 Model Reasoning — Key Drivers of This Estimate"):
                        st.markdown("> " + _ins_summary.replace("&#39;","'").replace("&quot;",'"').replace("&lt;","<").replace("&gt;",">"))
                        st.caption("Model output based on public data. Not a recommendation.")

            # ══════════════════════════════════════════════════════════
            # LAYER 3c — AI Analysis Summary (Premium / Pro only)
            # ══════════════════════════════════════════════════════════
            if can("action_plan"):
                with st.expander("🤖 AI Analysis Summary — Plain English", expanded=False):
                    with st.spinner("Generating AI analysis…"):
                        _ai_pf_score = (
                            st.session_state.get("_last_pf_score") or
                            int(fs.get("score", 50) / 100 * 9)   # rough proxy from fund score
                        )
                        _ai_text = generate_ai_summary(
                            ticker          = ticker_input,
                            company_name    = company_name,
                            price           = price_d,
                            iv              = iv_d,
                            mos_pct         = mos_pct,
                            signal          = sig,
                            piotroski_score = _ai_pf_score,
                            wacc            = wacc,
                            rev_growth      = enriched.get("revenue_growth", 0),
                            fcf_growth      = enriched.get("fcf_growth", 0),
                            op_margin       = enriched.get("op_margin", 0),
                            moat_grade      = moat_grade,
                            sym             = sym,
                        )
                    _ai_safe = _ai_text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    # Split into paragraphs for nicer rendering
                    _ai_paras = [p.strip() for p in _ai_safe.split("\n\n") if p.strip()]
                    _ai_paras_html = "".join(
                        f'<p style="margin:0 0 12px;line-height:1.8;font-size:13px;color:#334155;">{p}</p>'
                        for p in _ai_paras
                    )
                    st.html(f"""
                    <div style="padding:16px 20px;
                                background:linear-gradient(135deg,#F8FAFC,#F4F6F9);
                                border-left:3px solid #7C3AED;
                                border-radius:0 8px 8px 0;">
                      <div style="font-size:11px;font-weight:700;color:#7C3AED;
                                  text-transform:uppercase;letter-spacing:0.12em;
                                  margin-bottom:12px;display:flex;align-items:center;gap:6px;">
                        <span>🤖</span>
                        <span>AI Analysis — Powered by Groq</span>
                      </div>
                      {_ai_paras_html}
                    </div>
                    """)
            else:
                st.html("""
                <div style="padding:10px 14px;background:#F8FAFC;
                            border:1px dashed #CBD5E1;border-radius:8px;
                            font-size:12px;color:#94A3B8;margin-bottom:4px;">
                  🤖 <strong style="color:#64748B;">AI Analysis Summary</strong>
                  — Available on Starter &amp; Pro plans
                  <a href="#" style="color:#7C3AED;margin-left:8px;font-weight:600;">
                    Upgrade →
                  </a>
                </div>
                """)

            # ══════════════════════════════════════════════════════════
            # LAYER 4 — Advanced Analysis (all hidden)
            # ══════════════════════════════════════════════════════════
            with st.expander("📊 Live Price, Analyst Views & Earnings — Detailed Data"):
                render_live_price_header(ticker=ticker_input, sym=sym, fx=fx, refresh_every=60)
                ccard("What do professional analysts say?", "#7C3AED")
                render_analyst_consensus(ticker=ticker_input, current_price=price_d, sym=sym, fx=fx, raw_data=raw)
                ccard_end()
                ccard("Upcoming earnings & past surprises", "#0891B2")
                render_earnings_calendar(ticker=ticker_input, sym=sym, raw_data=raw)
                ccard_end()

            with st.expander("📋 Detailed Financial Metrics & Key Ratios"):
                k1,k2,k3,k4 = st.columns(4)
                _dcf_iv_d = enriched.get("dcf_iv", 0) * fx
                _pe_iv_d  = enriched.get("pe_iv",  0) * fx
                k1.metric("Current price",        fmts(price_d, sym))
                k2.metric("Estimated fair value", fmts(iv_d, sym),
                          delta=f"DCF:{fmts(_dcf_iv_d,sym)} | PE:{fmts(_pe_iv_d,sym)}" if _pe_iv_d > 0 else None,
                          delta_color="off")
                k3.metric("Discount to fair value", f"{mos_pct:.1f}%",
                          help=FINANCIAL_TOOLTIPS["Margin of Safety"])
                k4.metric("Model confidence", f"{confidence['grade']} ({confidence['score']}/100)")
                k5,k6,k7,k8 = st.columns(4)
                k5.metric("Revenue growth",       f"{_rev_growth:.1f}%")
                k6.metric("Profit margin",        f"{_op_margin:.1f}%")
                k7.metric("Cash flow growth",     f"{_fcf_growth:.1f}%",
                          help=FINANCIAL_TOOLTIPS["FCF"])
                k8.metric("Required return rate", f"{wacc:.1%}",
                          help=FINANCIAL_TOOLTIPS["WACC"])
                if pro_mode:
                    _pq1, _pq2, _pq3, _pq4 = st.columns(4)
                    _pq1.metric("Enterprise Value",  fmt(dcf_res.get("enterprise_value", 0) * fx, sym))
                    _pq2.metric("PV Terminal Value", fmt(pv_tv_d, sym),
                                help=FINANCIAL_TOOLTIPS["Terminal Value"])
                    _pq3.metric("TV % of EV",        f"{dcf_res.get('tv_pct_of_ev', 0):.0%}")
                    _pq4.metric("Equity Value",      fmt(dcf_res.get("equity_value", 0) * fx, sym))

            # ── Legal compliance disclaimer banner ────────────────────
            st.html("""
<div style="margin:10px 0 6px;padding:8px 16px;background:#FFFBEB;
            border:1px solid #FDE68A;border-radius:6px;text-align:center;">
  <span style="font-size:11px;color:#92400E;font-family:'Inter',sans-serif;line-height:1.4;">
    ⚠️ <strong>Model output only — not investment advice.</strong>
    YieldIQ is not a registered investment adviser.
    All signals reflect mathematical model comparisons, not personalised model outputs.
    See full disclaimer in the <strong>Guide</strong> tab.
  </span>
</div>
""")

            if use_auto_wacc and wacc_data.get("auto_computed"):
                with st.expander(f"🔢 Required Return Rate ({wacc:.1%}) — How WACC Was Calculated"):
                    w1,w2,w3,w4,w5 = st.columns(5)
                    w1.metric("Required return (WACC)",    f"{wacc_data['wacc']:.1%}",
                              help=FINANCIAL_TOOLTIPS["WACC"])
                    w1.caption("Higher WACC = more conservative (lower) intrinsic value estimate")
                    w2.metric("Expected equity return",    f"{wacc_data['re']:.1%}")
                    w3.metric("Volatility vs market (Beta)", f"{wacc_data['beta']:.2f}")
                    w4.metric("Risk-free rate",             f"{wacc_data['rf']:.1%}")
                    w5.metric("Cost of debt",               f"{wacc_data['rd']:.1%}")

        if _active == "valuation":
            if not st.session_state.get("fin_enriched"):
                st.info("Run an analysis first to see this section.")
                st.stop()

            # ══════════════════════════════════════════════════════════
            # SECTION 1 — Valuation Summary (top card)
            # ══════════════════════════════════════════════════════════
            _vl_val_label  = (
                "Undervalued"   if mos_pct > 10 else
                "Fairly priced" if mos_pct > -10 else
                "Overvalued"
            )
            _vl_val_color  = (
                "#0D7A4E" if _vl_val_label == "Undervalued"   else
                "#1D4ED8" if _vl_val_label == "Fairly priced" else "#B91C1C"
            )
            _vl_val_bg     = (
                "#F0FDF4" if _vl_val_label == "Undervalued"   else
                "#EFF6FF" if _vl_val_label == "Fairly priced" else "#FEF2F2"
            )
            _vl_val_bd     = (
                "#BBF7D0" if _vl_val_label == "Undervalued"   else
                "#BFDBFE" if _vl_val_label == "Fairly priced" else "#FECACA"
            )
            _vl_gauge_w = min(max(abs(mos_pct), 2), 100)
            _vl_upside  = (("+" if mos_pct >= 0 else "") + f"{mos_pct:.1f}%")

            _summary_tpl = (
                '<div style="background:#FFFFFF;border-radius:12px;border:1px solid #E2E8F0;'
                'overflow:hidden;margin-bottom:8px;">'
                '<div style="height:4px;background:linear-gradient(90deg,VAL_COLOR,#06B6D4,transparent);"></div>'
                '<div style="padding:20px 24px;">'

                # label + upside
                '<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;">'
                '<div>'
                '<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;letter-spacing:0.14em;margin-bottom:6px;">Valuation</div>'
                '<div style="display:inline-flex;align-items:center;gap:8px;padding:6px 14px;'
                'background:VAL_BG;border:1px solid VAL_BD;border-radius:20px;">'
                '<span style="font-size:14px;font-weight:700;color:VAL_COLOR;">VAL_LABEL</span>'
                '</div>'
                '</div>'
                '<div style="text-align:right;">'
                '<div style="font-size:11px;color:#94A3B8;margin-bottom:3px;">vs estimated fair value</div>'
                '<div style="font-size:28px;font-weight:700;color:VAL_COLOR;">UPSIDE</div>'
                '</div>'
                '</div>'

                # 3 metrics row
                '<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:1px;'
                'background:#F1F5F9;border-radius:8px;overflow:hidden;margin-bottom:16px;">'

                '<div style="background:#FFFFFF;padding:14px 16px;">'
                '<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:4px;">Current price</div>'
                '<div style="font-size:20px;font-weight:700;color:#0F172A;">PRICE</div>'
                '</div>'

                '<div style="background:#FFFFFF;padding:14px 16px;">'
                '<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:4px;">Estimated fair value</div>'
                '<div style="font-size:20px;font-weight:700;color:#1D4ED8;">IV</div>'
                '</div>'

                '<div style="background:#FFFFFF;padding:14px 16px;">'
                '<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:4px;">Model confidence</div>'
                '<div style="font-size:20px;font-weight:700;color:#334155;">CONF</div>'
                '</div>'

                '</div>'

                # progress bar
                '<div style="display:flex;justify-content:space-between;margin-bottom:5px;">'
                '<span style="font-size:12px;color:#64748B;">Discount to fair value</span>'
                '<span style="font-size:12px;font-weight:700;color:VAL_COLOR;">MOS_PCT%</span>'
                '</div>'
                '<div style="height:7px;background:#F1F5F9;border-radius:4px;overflow:hidden;">'
                '<div style="height:100%;width:GAUGE_W%;background:VAL_COLOR;border-radius:4px;"></div>'
                '</div>'

                '</div>'  # end padding
                '</div>'  # end card
            )
            _summary_tpl = (
                _summary_tpl
                .replace("VAL_COLOR",  _vl_val_color)
                .replace("VAL_BG",     _vl_val_bg)
                .replace("VAL_BD",     _vl_val_bd)
                .replace("VAL_LABEL",  _vl_val_label)
                .replace("UPSIDE",     _vl_upside)
                .replace("PRICE",      fmts(price_d, sym))
                .replace("IV",         fmts(iv_d, sym))
                .replace("CONF",       f"{confidence.get('grade','N/A')} · {confidence.get('score',0)}/100")
                .replace("MOS_PCT",    f"{mos_pct:.1f}")
                .replace("GAUGE_W",    str(int(_vl_gauge_w)))
            )
            st.html(_summary_tpl)

            # ── Live RF Rate assumptions banner ────────────────────────
            _rf_ui = st.session_state.get("_rf_rate_info", {})
            if _rf_ui:
                _rf_src_icon = "🟢" if _rf_ui.get("source") == "live" else "🟡"
                _rf_mkt      = "US 10Y" if _rf_ui.get("market") == "us" else "India 10Y"
                _rf_adj_val  = _rf_ui.get("wacc_adj", 0.0)
                _rf_adj_s    = f"  ·  WACC adj {_rf_adj_val:+.2%}" if _rf_adj_val else ""
                _rf_env      = _rf_ui.get("environment", "")
                _rf_env_clr  = "#DC2626" if "Tight" in _rf_env else "#059669" if "Loose" in _rf_env else "#64748B"
                st.html(f"""
                <div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:8px;
                            padding:8px 16px;margin-bottom:12px;display:flex;align-items:center;
                            gap:18px;flex-wrap:wrap;font-size:12px;">
                  <span style="color:#64748B;font-weight:600;">⚙️ DCF Assumptions</span>
                  <span style="color:#475569;">{_rf_src_icon} Risk-Free Rate:
                    <b style="color:#0F172A;font-family:'IBM Plex Mono',monospace;">
                      {_rf_ui.get('rate_pct', 0):.2f}% ({_rf_mkt})
                    </b>
                  </span>
                  <span style="color:#475569;">WACC: <b style="color:#0F172A;font-family:'IBM Plex Mono',monospace;">{wacc:.2%}</b></span>
                  <span style="color:#475569;">Terminal g: <b style="color:#0F172A;font-family:'IBM Plex Mono',monospace;">{terminal_g:.2%}</b></span>
                  <span style="color:{_rf_env_clr};">{_rf_env}{_rf_adj_s}</span>
                </div>
                """)

            # ══════════════════════════════════════════════════════════
            # SECTION 2 — Scenario Cards (Bull / Base / Bear)
            # ══════════════════════════════════════════════════════════
            st.html('<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;'
                    'letter-spacing:0.14em;margin:14px 0 8px;padding-left:2px;">Scenarios</div>')

            _sc_cols = st.columns(3)
            _sc_cfg = [
                ("🐻 Bear case",    "Cautious outlook",        scenarios.get("Bear 🐻",  {}), "#B91C1C", "#FEF2F2", "#FECACA"),
                ("📊 Base case",    "Most likely outcome",     scenarios.get("Base 📊",  {}), "#1D4ED8", "#EFF6FF", "#BFDBFE"),
                ("🐂 Bull case",    "Optimistic outlook",      scenarios.get("Bull 🐂",  {}), "#0D7A4E", "#F0FDF4", "#BBF7D0"),
            ]
            for _col, (_sc_name, _sc_desc, _sc_data, _sc_c, _sc_bg, _sc_bd) in zip(_sc_cols, _sc_cfg):
                _sc_iv  = (_sc_data.get("iv", 0) or 0) * fx
                _sc_mos = _sc_data.get("mos_pct", 0) or 0
                _sc_lbl = (
                    "Undervalued"   if _sc_mos > 10 else
                    "Fairly priced" if _sc_mos > -10 else
                    "Overvalued"
                )
                _sc_card = (
                    '<div style="background:#FFFFFF;border:1px solid SCBD;border-top:3px solid SCC;'
                    'border-radius:10px;padding:16px 18px;height:100%;">'
                    '<div style="font-size:13px;font-weight:700;color:SCC;margin-bottom:4px;">SCNAME</div>'
                    '<div style="font-size:11px;color:#94A3B8;margin-bottom:12px;">SCDESC</div>'
                    '<div style="font-size:22px;font-weight:700;color:#0F172A;margin-bottom:4px;">SCIV</div>'
                    '<div style="font-size:12px;color:#64748B;margin-bottom:8px;">Estimated fair value</div>'
                    '<div style="display:inline-block;padding:3px 10px;background:SCBG;'
                    'border:1px solid SCBD;border-radius:12px;">'
                    '<span style="font-size:12px;font-weight:600;color:SCC;">SCLBL · SCMOS%</span>'
                    '</div>'
                    '</div>'
                )
                _sc_card = (
                    _sc_card
                    .replace("SCNAME", _sc_name)
                    .replace("SCDESC", _sc_desc)
                    .replace("SCIV",   fmts(_sc_iv, sym) if _sc_iv else "—")
                    .replace("SCMOS",  f"{_sc_mos:+.0f}")
                    .replace("SCLBL",  _sc_lbl)
                    .replace("SCC",    _sc_c)
                    .replace("SCBG",   _sc_bg)
                    .replace("SCBD",   _sc_bd)
                )
                _col.html(_sc_card)

            # ══════════════════════════════════════════════════════════
            # SECTION 3 — Key Insights (plain English, Simple mode only)
            # ══════════════════════════════════════════════════════════
            _insights = []

            if mos_pct > 20:
                _insights.append(("🟢", f"The stock appears meaningfully undervalued — our model estimates a {mos_pct:.0f}% discount to fair value."))
            elif mos_pct > 5:
                _insights.append(("🟡", f"The stock looks modestly undervalued by around {mos_pct:.0f}% based on our cash flow model."))
            elif mos_pct > -5:
                _insights.append(("🔵", "The stock is trading close to what our model considers fair value."))
            else:
                _insights.append(("🔴", f"The stock may be trading above fair value — our model suggests it is {abs(mos_pct):.0f}% above its estimated worth."))

            _bull_iv = (scenarios.get("Bull 🐂", {}).get("iv", 0) or 0) * fx
            _bear_iv = (scenarios.get("Bear 🐻", {}).get("iv", 0) or 0) * fx
            if _bull_iv > 0 and _bear_iv > 0:
                _range_pct = ((_bull_iv - _bear_iv) / price_d * 100) if price_d > 0 else 0
                if _range_pct > 50:
                    _insights.append(("⚡", "The valuation is sensitive to assumptions — bull and bear cases are far apart. Treat the estimate with appropriate caution."))
                else:
                    _insights.append(("✅", "Bull and bear scenarios are relatively close together, suggesting moderate uncertainty in the estimate."))

            _rev_g = enriched.get("revenue_growth", 0) * 100
            _fcf_g = enriched.get("fcf_growth", 0) * 100
            if _rev_g > 15:
                _insights.append(("📈", f"Revenue is growing strongly ({_rev_g:.0f}%), which supports a higher fair value estimate."))
            elif _rev_g > 0:
                _insights.append(("📊", f"Revenue is growing at a moderate pace ({_rev_g:.0f}%), reflected in the base case estimate."))
            else:
                _insights.append(("📉", "Revenue growth has been slow or negative, which limits upside in our model."))

            _moat_g = enriched.get("moat_grade", "None")
            if _moat_g == "Wide":
                _insights.append(("🔰", "The company has a strong competitive advantage, which adds a premium to the fair value estimate."))
            elif _moat_g == "Narrow":
                _insights.append(("🔰", "The company has some competitive advantages, providing modest support to the valuation."))

            if confidence.get("score", 0) < 50:
                _insights.append(("⚠️", "Model confidence is lower than usual — the estimate may be less reliable for this stock."))

            if not pro_mode:
                st.html('<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;'
                        'letter-spacing:0.14em;margin:16px 0 8px;padding-left:2px;">Key insights</div>')
                _ins_html = '<div style="background:#FFFFFF;border:1px solid #E2E8F0;border-radius:10px;padding:16px 20px;margin-bottom:8px;">'
                for _icon, _txt in _insights[:5]:
                    _ins_html += (
                        '<div style="display:flex;align-items:flex-start;gap:10px;padding:8px 0;'
                        'border-bottom:1px solid #F8FAFC;">'
                        '<span style="font-size:16px;flex-shrink:0;margin-top:1px;">' + _icon + '</span>'
                        '<span style="font-size:13px;color:#334155;line-height:1.6;">' + _txt + '</span>'
                        '</div>'
                    )
                _ins_html += '</div>'
                st.html(_ins_html)

            # ══════════════════════════════════════════════════════════
            # SECTION 4 — Price Chart (compact)
            # ══════════════════════════════════════════════════════════
            if not price_hist.empty:
                with st.expander("📈 Price History Chart with Fair Value Line", expanded=True):
                    phd = price_hist.copy()
                    for col in ["Open","High","Low","Close"]:
                        phd[col] = phd[col] * fx

                # Build candle + volume data for lightweight-charts
                _candles = []
                _volumes = []
                for _, row in phd.iterrows():
                    try:
                        _dt = row["Date"]
                        if hasattr(_dt, "strftime"):
                            _ts = _dt.strftime("%Y-%m-%d")
                        else:
                            _ts = str(_dt)[:10]
                        _o = float(row["Open"])
                        _h = float(row["High"])
                        _l = float(row["Low"])
                        _c = float(row["Close"])
                        _v = float(row.get("Volume", 0))
                        if _o > 0 and _h > 0:
                            _candles.append({"time":_ts,"open":round(_o,4),"high":round(_h,4),"low":round(_l,4),"close":round(_c,4)})
                            _volumes.append({"time":_ts,"value":_v,"color":"#10b981" if _c>=_o else "#ef4444"})
                    except Exception:
                        continue

                # MA20
                _ma20 = []
                if len(phd) >= 20:
                    phd["_ma"] = phd["Close"].rolling(20).mean()
                    for _, row in phd.dropna(subset=["_ma"]).iterrows():
                        try:
                            _dt = row["Date"]
                            _ts = _dt.strftime("%Y-%m-%d") if hasattr(_dt,"strftime") else str(_dt)[:10]
                            _ma20.append({"time":_ts,"value":round(float(row["_ma"]),4)})
                        except Exception:
                            continue

                # IV line + buy/sell markers
                _iv_val   = round(iv_d, 4) if iv_d > 0 else 0
                _iv_label = f"YIQ IV  {sym}{iv_d:,.2f}" if iv_d > 0 else ""

                # MA50 and MA200
                _ma50, _ma200 = [], []
                if len(phd) >= 50:
                    phd["_ma50"] = phd["Close"].rolling(50).mean()
                    for _, row in phd.dropna(subset=["_ma50"]).iterrows():
                        try:
                            _dt = row["Date"]
                            _ts = _dt.strftime("%Y-%m-%d") if hasattr(_dt,"strftime") else str(_dt)[:10]
                            _ma50.append({"time":_ts,"value":round(float(row["_ma50"]),4)})
                        except Exception: continue
                if len(phd) >= 200:
                    phd["_ma200"] = phd["Close"].rolling(200).mean()
                    for _, row in phd.dropna(subset=["_ma200"]).iterrows():
                        try:
                            _dt = row["Date"]
                            _ts = _dt.strftime("%Y-%m-%d") if hasattr(_dt,"strftime") else str(_dt)[:10]
                            _ma200.append({"time":_ts,"value":round(float(row["_ma200"]),4)})
                        except Exception: continue

                # 52-week high/low for shaded band
                _52w_high = float(phd["High"].tail(252).max()) if len(phd) >= 20 else 0
                _52w_low  = float(phd["Low"].tail(252).min())  if len(phd) >= 20 else 0

                # Serialize all data for JS
                import json as _json
                _candles_js  = _json.dumps(_candles)
                _volumes_js  = _json.dumps(_volumes)
                _ma20_js     = _json.dumps(_ma20)
                _ma50_js     = _json.dumps(_ma50)
                _ma200_js    = _json.dumps(_ma200)
                _markers_js  = "[]"
                _ticker_js   = _json.dumps(ticker_input)
                _sym_js      = _json.dumps(sym)

                # ── Clean chart: price + MA20 + IV fair value line ─────
                st.components.v1.html(f"""
<!DOCTYPE html>
<html>
<head>
<script src="https://unpkg.com/lightweight-charts@4.1.1/dist/lightweight-charts.standalone.production.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: #FFFFFF; font-family: 'Inter', system-ui, sans-serif; }}
  #wrap {{ position: relative; width: 100%; }}
  #main-chart {{ width: 100%; height: 320px; }}
  #vol-chart  {{ width: 100%; height: 60px; margin-top: 1px; }}

  /* Toolbar */
  #toolbar {{
    display: flex; align-items: center; gap: 6px;
    padding: 8px 12px 6px; background: #FFFFFF;
    border-bottom: 1px solid #F1F5F9;
    flex-wrap: wrap;
  }}
  .ticker-label {{
    font-size: 13px; font-weight: 700; color: #0F172A;
    margin-right: 6px; letter-spacing: -0.01em;
  }}
  .price-label {{
    font-size: 13px; font-weight: 600; color: #0F172A;
    font-family: 'IBM Plex Mono', monospace;
    margin-right: 10px;
  }}
  .tb-btn {{
    font-size: 11px; font-weight: 500; padding: 3px 8px;
    border: 1px solid #E2E8F0; border-radius: 4px;
    background: #FFFFFF; color: #475569; cursor: pointer;
    letter-spacing: 0.04em; transition: all 0.1s;
  }}
  .tb-btn:hover  {{ background: #EFF6FF; border-color: #BFDBFE; color: #1D4ED8; }}
  .tb-btn.active {{ background: #1D4ED8; color: #FFFFFF; border-color: #1D4ED8; }}
  .tb-sep {{ width: 1px; height: 14px; background: #E2E8F0; margin: 0 2px; }}

  /* OHLC legend — fixed top-right, no overlap */
  #ohlc-legend {{
    position: absolute; top: 44px; right: 12px;
    background: rgba(255,255,255,0.92);
    border: 1px solid #E2E8F0; border-radius: 6px;
    padding: 5px 10px; font-size: 11px; color: #475569;
    font-family: 'IBM Plex Mono', monospace;
    line-height: 1.7; pointer-events: none;
    display: none; z-index: 10;
    box-shadow: 0 2px 8px rgba(15,23,42,0.08);
  }}

  /* IV badge — bottom-left of chart */
  #iv-badge {{
    position: absolute; bottom: 68px; left: 12px;
    background: rgba(29,78,216,0.08);
    border: 1px solid rgba(29,78,216,0.25); border-radius: 4px;
    padding: 3px 8px; font-size: 11px; font-weight: 600;
    color: #1D4ED8; pointer-events: none; z-index: 10;
    font-family: 'IBM Plex Mono', monospace;
  }}

  /* Caption row */
  #caption {{
    padding: 5px 12px; font-size: 11px; color: #94A3B8;
    border-top: 1px solid #F1F5F9; background: #FAFBFD;
  }}
</style>
</head>
<body>
<div id="wrap">
  <div id="toolbar">
    <span class="ticker-label">{ticker_input}</span>
    <span class="price-label" id="cur-price">{fmts(price_d, sym)}</span>
    <div class="tb-sep"></div>
    <button class="tb-btn" onclick="setRange('1M',this)">1M</button>
    <button class="tb-btn" onclick="setRange('3M',this)">3M</button>
    <button class="tb-btn" onclick="setRange('6M',this)">6M</button>
    <button class="tb-btn active" onclick="setRange('1Y',this)">1Y</button>
    <div class="tb-sep"></div>
    <button class="tb-btn active" id="ma-btn" onclick="toggleMA(this)">MA20</button>
    <button class="tb-btn active" id="iv-btn" onclick="toggleIV(this)">Fair Value</button>
  </div>

  <div id="main-chart"></div>
  <div id="vol-chart"></div>
  <div id="ohlc-legend"></div>
  {'<div id="iv-badge">Fair value ' + fmts(iv_d, sym) + '</div>' if iv_d > 0 else ''}
  <div id="caption">Price data · MA20 (20-day average) · Fair value line · For informational purposes only</div>
</div>

<script>
const candleData  = {_candles_js};
const volumeData  = {_volumes_js};
const ma20Data    = {_ma20_js};
const markers     = [];
const ivVal       = {_iv_val};
const sym         = {_sym_js};

// ── Main chart ────────────────────────────────────────────────
const chart = LightweightCharts.createChart(
  document.getElementById('main-chart'), {{
    width:  document.getElementById('main-chart').offsetWidth,
    height: 320,
    layout: {{
      background: {{ color: '#FFFFFF' }},
      textColor:  '#64748B',
      fontSize:   11,
      fontFamily: "'Inter', system-ui, sans-serif",
    }},
    grid: {{
      vertLines: {{ color: '#F8FAFC' }},
      horzLines: {{ color: '#F1F5F9' }},
    }},
    crosshair: {{
      mode: LightweightCharts.CrosshairMode.Normal,
      vertLine: {{ color: '#94A3B8', width: 1, style: 3, labelBackgroundColor: '#334155' }},
      horzLine: {{ color: '#94A3B8', width: 1, style: 3, labelBackgroundColor: '#334155' }},
    }},
    rightPriceScale: {{
      borderColor: '#F1F5F9',
      scaleMargins: {{ top: 0.08, bottom: 0.08 }},
    }},
    timeScale: {{
      borderColor:     '#F1F5F9',
      timeVisible:     true,
      secondsVisible:  false,
      fixLeftEdge:     true,
      fixRightEdge:    true,
    }},
    handleScroll:   {{ mouseWheel: true, pressedMouseMove: true }},
    handleScale:    {{ mouseWheel: true, pinch: true }},
  }}
);

// ── Candlestick series ────────────────────────────────────────
const candleSeries = chart.addCandlestickSeries({{
  upColor:          '#10b981',
  downColor:        '#ef4444',
  borderUpColor:    '#059669',
  borderDownColor:  '#dc2626',
  wickUpColor:      '#10b981',
  wickDownColor:    '#ef4444',
  priceLineVisible: false,
}});
candleSeries.setData(candleData);

// ── MA20 ──────────────────────────────────────────────────────
const maSeries = chart.addLineSeries({{
  color:            '#64748b',
  lineWidth:        1,
  lineStyle:        LightweightCharts.LineStyle.Dashed,
  priceLineVisible: false,
  lastValueVisible: false,
  title:            'MA20',
}});
maSeries.setData(ma20Data);

// ── MA50 ──────────────────────────────────────────────────
const ma50Data = {_ma50_js};
const ma50Series = chart.addLineSeries({{
  color:            '#3b82f6',
  lineWidth:        1.5,
  lineStyle:        LightweightCharts.LineStyle.Solid,
  priceLineVisible: false,
  lastValueVisible: true,
  title:            'MA50',
}});
if (ma50Data.length > 0) ma50Series.setData(ma50Data);

// ── MA200 ─────────────────────────────────────────────────
const ma200Data = {_ma200_js};
const ma200Series = chart.addLineSeries({{
  color:            '#f59e0b',
  lineWidth:        1.5,
  lineStyle:        LightweightCharts.LineStyle.Solid,
  priceLineVisible: false,
  lastValueVisible: true,
  title:            'MA200',
}});
if (ma200Data.length > 0) ma200Series.setData(ma200Data);

// ── 52-week high/low shaded band ─────────────────────────
const hiVal = {_52w_high}; const loVal = {_52w_low};
if (hiVal > 0 && candleData.length > 0) {{
  const firstT = candleData[0].time;
  const lastT  = candleData[candleData.length-1].time;
  // high zone (top 5% of 52w range)
  const hiSeries = chart.addLineSeries({{
    color:'rgba(16,185,129,0.0)',lineWidth:0,priceLineVisible:false,lastValueVisible:false,
  }});
  hiSeries.setData([{{time:firstT,value:hiVal}},{{time:lastT,value:hiVal}}]);
  // low zone
  const loSeries = chart.addLineSeries({{
    color:'rgba(239,68,68,0.0)',lineWidth:0,priceLineVisible:false,lastValueVisible:false,
  }});
  loSeries.setData([{{time:firstT,value:loVal}},{{time:lastT,value:loVal}}]);
}}

// ── Fair value (IV) line ──────────────────────────────────────
let ivSeries = null;
if (ivVal > 0 && candleData.length > 0) {{
  ivSeries = chart.addLineSeries({{
    color:            '#1D4ED8',
    lineWidth:        2,
    lineStyle:        LightweightCharts.LineStyle.Dashed,
    priceLineVisible: false,
    lastValueVisible: true,
    title:            'Fair Value',
  }});
  ivSeries.setData([
    {{ time: candleData[0].time,                    value: ivVal }},
    {{ time: candleData[candleData.length - 1].time, value: ivVal }},
  ]);
}}

// ── Volume chart ──────────────────────────────────────────────
const volChart = LightweightCharts.createChart(
  document.getElementById('vol-chart'), {{
    width:  document.getElementById('vol-chart').offsetWidth,
    height: 60,
    layout: {{ background: {{ color: '#FFFFFF' }}, textColor: '#94A3B8', fontSize: 10 }},
    grid:   {{ vertLines: {{ color: '#F8FAFC' }}, horzLines: {{ visible: false }} }},
    rightPriceScale: {{
      borderColor:   '#F1F5F9',
      scaleMargins:  {{ top: 0.1, bottom: 0 }},
      drawTicks:     false,
    }},
    timeScale: {{
      borderColor:    '#F1F5F9',
      timeVisible:    false,
      fixLeftEdge:    true,
      fixRightEdge:   true,
    }},
    crosshair: {{ horzLine: {{ visible: false }} }},
    handleScroll: false,
    handleScale:  false,
  }}
);
const volSeries = volChart.addHistogramSeries({{
  priceFormat:  {{ type: 'volume' }},
  priceScaleId: 'volume',
}});
volSeries.setData(volumeData);

// ── Sync crosshair ────────────────────────────────────────────
chart.subscribeCrosshairMove(p => {{
  if (p.time) volChart.setCrosshairPosition(0, p.time, volSeries);
  else        volChart.clearCrosshairPosition();
}});
// ── Sync time scale ───────────────────────────────────────────
chart.timeScale().subscribeVisibleLogicalRangeChange(r => {{
  if (r) volChart.timeScale().setVisibleLogicalRange(r);
}});
volChart.timeScale().subscribeVisibleLogicalRangeChange(r => {{
  if (r) chart.timeScale().setVisibleLogicalRange(r);
}});

// ── OHLC Legend (fixed position, no overlap) ──────────────────
const legend = document.getElementById('ohlc-legend');
chart.subscribeCrosshairMove(p => {{
  if (!p.time || !p.seriesData) {{ legend.style.display = 'none'; return; }}
  const d = p.seriesData.get(candleSeries);
  if (!d) {{ legend.style.display = 'none'; return; }}
  legend.style.display = 'block';
  const chg    = d.close - d.open;
  const chgPct = (chg / d.open * 100).toFixed(2);
  const clr    = chg >= 0 ? '#10b981' : '#ef4444';
  legend.innerHTML =
    '<span style="color:#94A3B8;">' + p.time + '</span><br>' +
    'O <b>' + sym + d.open.toFixed(2)  + '</b>  ' +
    'H <b>' + sym + d.high.toFixed(2)  + '</b>  ' +
    'L <b>' + sym + d.low.toFixed(2)   + '</b>  ' +
    'C <b style="color:' + clr + '">'  + sym + d.close.toFixed(2) + '</b>  ' +
    '<span style="color:' + clr + '">(' + (chg >= 0 ? '+' : '') + chgPct + '%)</span>';
}});

// ── Time range buttons ────────────────────────────────────────
function setRange(r, btn) {{
  document.querySelectorAll('.tb-btn').forEach(b => {{
    if (['1M','3M','6M','1Y'].includes(b.textContent)) b.classList.remove('active');
  }});
  btn.classList.add('active');
  const last = candleData[candleData.length - 1].time;
  const d    = new Date(last);
  let from;
  if      (r === '1M') {{ d.setMonth(d.getMonth() - 1);          from = d.toISOString().slice(0,10); }}
  else if (r === '3M') {{ d.setMonth(d.getMonth() - 3);          from = d.toISOString().slice(0,10); }}
  else if (r === '6M') {{ d.setMonth(d.getMonth() - 6);          from = d.toISOString().slice(0,10); }}
  else                 {{ d.setFullYear(d.getFullYear() - 1);     from = d.toISOString().slice(0,10); }}
  chart.timeScale().setVisibleRange({{ from, to: last }});
}}

// ── Toggle MA20 ───────────────────────────────────────────────
let maVisible = true;
function toggleMA(btn) {{
  maVisible = !maVisible;
  maSeries.applyOptions({{ visible: maVisible }});
  btn.classList.toggle('active', maVisible);
}}

// ── Toggle IV line ────────────────────────────────────────────
let ivVisible = true;
function toggleIV(btn) {{
  ivVisible = !ivVisible;
  if (ivSeries) ivSeries.applyOptions({{ visible: ivVisible }});
  const badge = document.getElementById('iv-badge');
  if (badge) badge.style.display = ivVisible ? 'block' : 'none';
  btn.classList.toggle('active', ivVisible);
}}

// ── Fit content on load ───────────────────────────────────────
chart.timeScale().fitContent();

// ── Responsive resize ────────────────────────────────────────
const ro = new ResizeObserver(() => {{
  const w = document.getElementById('main-chart').offsetWidth;
  chart.applyOptions({{ width: w }});
  volChart.applyOptions({{ width: w }});
}});
ro.observe(document.getElementById('wrap'));
</script>
</body>
</html>
""", height=470, scrolling=False)
    
            # ── Historical + Projected FCF
            # ══════════════════════════════════════════════════════════
            # SECTION 5 — Advanced Analysis (all in expanders)
            # ══════════════════════════════════════════════════════════
            st.html('<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;'
                    'letter-spacing:0.14em;margin:16px 0 8px;padding-left:2px;">Advanced analysis</div>')

            with st.expander("📈 FCF History & 10-Year Cash Flow Projections"):
                cf_df_plot = enriched.get("cf_df", pd.DataFrame())
                hist_fcfs  = []
                hist_yrs   = []

                if not cf_df_plot.empty and "fcf" in cf_df_plot.columns:
                    valid_hist = cf_df_plot[cf_df_plot["fcf"].abs() > 1e6]
                    if not valid_hist.empty:
                        hist_fcfs = (valid_hist["fcf"] * fx / 1e9).tolist()
                        hist_yrs  = [str(int(y)) for y in valid_hist["year"].tolist()]

                # Use calendar year labels for projected FCF so x-axis is consistent
                import datetime as _dt
                _base_yr   = int(hist_yrs[-1]) + 1 if hist_yrs else _dt.datetime.now().year + 1
                proj_labels = [str(_base_yr + i) for i in range(forecast_yrs)]
                all_labels  = hist_yrs + proj_labels

                fig_fcf = go.Figure()

                # ── Shaded projected region ──────────────────────────────
                if hist_yrs and proj_labels and len(hist_yrs) + len(proj_labels) > 0:
                    _sep_x = len(hist_yrs) / (len(hist_yrs) + len(proj_labels))
                    fig_fcf.add_shape(
                        type="rect", xref="paper", yref="paper",
                        x0=_sep_x, x1=1.0, y0=0, y1=1,
                        fillcolor="rgba(16,185,129,0.04)",
                        line=dict(width=0), layer="below",
                    )

                # ── Historical bars ──────────────────────────────────────
                if hist_fcfs:
                    fig_fcf.add_trace(go.Bar(
                        x=hist_yrs, y=hist_fcfs,
                        marker=dict(
                            color=["#3b82f6" if v >= 0 else "#ef4444" for v in hist_fcfs],
                            opacity=0.88, line=dict(width=0),
                        ),
                        name="Historical FCF",
                        text=[f"${abs(v):.1f}B" for v in hist_fcfs],
                        textposition="outside",
                        textfont=dict(size=9, color="#8b949e", family="IBM Plex Mono"),
                        hovertemplate=f"<b>%{{x}}</b><br>FCF: {sym}%{{y:.2f}}B<extra>Historical</extra>",
                    ))

                # ── Projected bars ───────────────────────────────────────
                proj_vals = [v / 1e9 for v in proj_d]
                fig_fcf.add_trace(go.Bar(
                    x=proj_labels, y=proj_vals,
                    marker=dict(
                        color=["#10b981" if v >= 0 else "#ef4444" for v in proj_vals],
                        opacity=0.88, line=dict(width=0),
                    ),
                    name="Projected FCF",
                    text=[f"${abs(v):.1f}B" for v in proj_vals],
                    textposition="outside",
                    textfont=dict(size=9, color="#8b949e", family="IBM Plex Mono"),
                    hovertemplate=f"<b>%{{x}}</b><br>FCF: {sym}%{{y:.2f}}B<extra>Projected</extra>",
                ))

                # ── Growth rate overlay (secondary y-axis) ───────────────
                growth_pct = [g * 100 for g in growth_schedule]
                fig_fcf.add_trace(go.Scatter(
                    x=proj_labels, y=growth_pct,
                    yaxis="y2",
                    line=dict(color="#f59e0b", width=2, dash="dot"),
                    mode="lines+markers",
                    marker=dict(size=5, color="#f59e0b", symbol="circle"),
                    name="Growth %",
                    hovertemplate="<b>%{x}</b><br>Growth: %{y:.1f}%<extra>Growth Rate</extra>",
                ))

                # ── Dashed separator + dual annotation ───────────────────
                if hist_yrs and proj_labels:
                    _sep = len(hist_yrs) / (len(hist_yrs) + len(proj_labels))
                    fig_fcf.add_shape(
                        type="line", xref="paper", yref="paper",
                        x0=_sep, x1=_sep, y0=0, y1=0.92,
                        line=dict(color="#30363d", width=1.5, dash="dash"),
                    )
                    fig_fcf.add_annotation(
                        xref="paper", yref="paper",
                        x=_sep - 0.02, y=0.97,
                        text="← Historical",
                        showarrow=False,
                        font=dict(color="#8b949e", size=9, family="Inter"),
                        xanchor="right",
                    )
                    fig_fcf.add_annotation(
                        xref="paper", yref="paper",
                        x=_sep + 0.02, y=0.97,
                        text="Forecast →",
                        showarrow=False,
                        font=dict(color="#10b981", size=9, family="Inter"),
                        xanchor="left",
                    )

                apply_koyfin(fig_fcf, height=340, extra_kw=dict(
                    barmode="group",
                    showlegend=True,
                    yaxis=dict(
                        title=dict(text=f"{to_code}B", font=dict(size=11, color="#8b949e")),
                        gridcolor="#21262d", linecolor="#30363d", zeroline=True,
                        zerolinecolor="#30363d", tickfont=dict(color="#8b949e", size=10),
                    ),
                    yaxis2=dict(
                        title=dict(text="YoY Growth %", font=dict(size=11, color="#f59e0b")),
                        overlaying="y", side="right",
                        gridcolor="rgba(0,0,0,0)", ticksuffix="%",
                        tickfont=dict(color="#f59e0b", size=10), zeroline=False,
                    ),
                    xaxis=dict(
                        type="category", gridcolor="#21262d", linecolor="#30363d",
                        zeroline=False, tickfont=dict(color="#8b949e", size=10), tickangle=-30,
                    ),
                    legend=dict(
                        orientation="h", x=0.0, y=-0.18,
                        bgcolor="rgba(0,0,0,0)", borderwidth=0,
                        font=dict(color="#8b949e", size=11),
                    ),
                ))
                st.plotly_chart(fig_fcf, width="stretch", config={
                    "displayModeBar": True,
                    "modeBarButtonsToRemove": ["lasso2d", "select2d"],
                    "toImageButtonOptions": {"filename": f"FCF_{ticker_input}", "scale": 2},
                })

                # ── THREE SCENARIOS
            with st.expander("📊 Fair Value vs Market Price — Bull / Base / Bear Scenarios"):
                # ── IV vs Price ────────────────────────────────────────
                c1, c2 = st.columns(2)
                with c1:
                    ccard("Estimated fair value vs current price", "#10b981")
                    iv_color = "#10b981" if iv_d > price_d else "#ef4444"

                    # Show all 3 scenario IVs + current price
                    bar_x = ["Price"] + [s.split()[0] for s in scenarios.keys()]
                    bar_y = [price_d] + [sdata["iv"] * fx for sdata in scenarios.values()]
                    bar_c = ["#3b82f6"] + [sdata["color"] for sdata in scenarios.values()]

                    # Horizontal bullet gauge: Bear → Base → Bull + price marker
                    _bear_val = scenarios.get("Bear 🐻", {}).get("iv", 0) * fx or price_d * 0.7
                    _base_val = scenarios.get("Base 📊", {}).get("iv", 0) * fx or iv_d
                    _bull_val = scenarios.get("Bull 🐂", {}).get("iv", 0) * fx or price_d * 1.4
                    _min_v = min(_bear_val, price_d) * 0.92
                    _max_v = max(_bull_val, price_d) * 1.05

                    fig_iv = go.Figure()

                    # ── Continuous gradient zone (Bear → Bull) ───────────
                    # Use a filled scatter as the gradient bar background
                    _zone_xs = [_min_v, _bear_val, _base_val, _bull_val, _max_v]
                    _zone_cs = [
                        "rgba(239,68,68,0.35)",
                        "rgba(239,68,68,0.25)",
                        "rgba(245,158,11,0.25)",
                        "rgba(16,185,129,0.25)",
                        "rgba(16,185,129,0.35)",
                    ]
                    for i in range(len(_zone_xs) - 1):
                        fig_iv.add_shape(
                            type="rect", xref="x", yref="paper",
                            x0=_zone_xs[i], x1=_zone_xs[i+1], y0=0.15, y1=0.85,
                            fillcolor=_zone_cs[i], line=dict(width=0), layer="below",
                        )

                    # ── Bear / Base / Bull tick marks ────────────────────
                    for _sv, _slabel, _scolor in [
                        (_bear_val, f"Bear<br>{fmts(_bear_val,sym)}", "#ef4444"),
                        (_base_val, f"Base<br>{fmts(_base_val,sym)}", "#f59e0b"),
                        (_bull_val, f"Bull<br>{fmts(_bull_val,sym)}", "#10b981"),
                    ]:
                        fig_iv.add_shape(type="line", xref="x", yref="paper",
                            x0=_sv, x1=_sv, y0=0.15, y1=0.85,
                            line=dict(color=_scolor, width=1, dash="dot"),
                        )
                        fig_iv.add_annotation(
                            x=_sv, y=1.0, xref="x", yref="paper",
                            text=_slabel, showarrow=False,
                            font=dict(color=_scolor, size=9, family="IBM Plex Mono"),
                            align="center",
                        )

                    # ── Current price — thick gold marker ────────────────
                    fig_iv.add_shape(type="line", xref="x", yref="paper",
                        x0=price_d, x1=price_d, y0=0.0, y1=1.0,
                        line=dict(color="#f59e0b", width=4),
                        layer="above",
                    )
                    fig_iv.add_annotation(
                        x=price_d, y=0.05, xref="x", yref="paper",
                        text=f"▲ {fmts(price_d,sym)}",
                        showarrow=False,
                        font=dict(color="#f59e0b", size=11, family="IBM Plex Mono"),
                        align="center",
                    )

                    # ── MoS % as large overlay text ───────────────────────
                    _mos_txt = f"{mos_pct:+.1f}% vs base"
                    _mos_col = "#10b981" if mos_pct > 0 else "#ef4444"
                    fig_iv.add_annotation(
                        xref="paper", yref="paper", x=0.5, y=0.5,
                        text=f"<b>{_mos_txt}</b>",
                        showarrow=False,
                        font=dict(color=_mos_col, size=13, family="IBM Plex Mono"),
                        align="center",
                        bgcolor="rgba(13,17,23,0.7)",
                        bordercolor=_mos_col, borderwidth=1, borderpad=4,
                    )

                    # ── Invisible scatter for hover ───────────────────────
                    fig_iv.add_trace(go.Scatter(
                        x=[_bear_val, _base_val, _bull_val, price_d],
                        y=[0.5, 0.5, 0.5, 0.5],
                        mode="markers", marker=dict(color="rgba(0,0,0,0)", size=8),
                        text=[f"Bear: {fmts(_bear_val,sym)}", f"Base: {fmts(_base_val,sym)}",
                              f"Bull: {fmts(_bull_val,sym)}", f"Price: {fmts(price_d,sym)}"],
                        hovertemplate="%{text}<extra></extra>",
                    ))
                    apply_koyfin(fig_iv, height=200, extra_kw=dict(
                        showlegend=False,
                        xaxis=dict(
                            title=f"{to_code}/share", range=[_min_v, _max_v],
                            gridcolor="rgba(0,0,0,0)", tickfont=dict(color="#8b949e", size=10),
                            showgrid=False,
                        ),
                        yaxis=dict(visible=False, range=[0, 1]),
                        margin=dict(t=48, b=36, l=20, r=20),
                    ))
                    st.plotly_chart(fig_iv, width="stretch", config={"displayModeBar":True,"modeBarButtonsToRemove":["lasso2d","select2d"],"toImageButtonOptions":{"filename":"iv_vs_price","scale":2}})
                    ccard_end()

                with c2:
                    ccard("Where does the fair value come from?", "#f59e0b")
                    # ── True Waterfall chart ─────────────────────────────
                    _sum_pv_fcfs = sum(v/1e9 for v in pv_fcfs_d)
                    _pv_tv       = pv_tv_d / 1e9
                    _ev          = _sum_pv_fcfs + _pv_tv
                    _debt        = -(enriched.get("total_debt", 0) * fx / 1e9)
                    _cash        = enriched.get("total_cash", 0) * fx / 1e9
                    _eq_val      = _ev + _debt + _cash
                    _shares_b    = enriched.get("shares", 1) / 1e9
                    _iv_per_sh   = iv_d  # already in display currency

                    fig_wf = go.Figure(go.Waterfall(
                        name="DCF Build-up",
                        orientation="v",
                        measure=[
                            "relative", "relative", "total",
                            "relative", "relative", "total",
                        ],
                        x=["PV of FCFs", "Terminal Value", "Enterprise Value",
                           "Less Debt", "Plus Cash", "Equity Value"],
                        y=[_sum_pv_fcfs, _pv_tv, 0,
                           _debt, _cash, 0],
                        text=[
                            f"{sym}{_sum_pv_fcfs:.1f}B",
                            f"{sym}{_pv_tv:.1f}B",
                            f"{sym}{_ev:.1f}B",
                            f"{sym}{abs(_debt):.1f}B",
                            f"{sym}{_cash:.1f}B",
                            f"{sym}{_eq_val:.1f}B",
                        ],
                        textposition="inside",
                        textfont=dict(size=9, color="#e6edf3", family="IBM Plex Mono"),
                        hovertemplate="<b>%{x}</b><br>%{text}<extra></extra>",
                        increasing=dict(marker=dict(color="#10b981", line=dict(width=0))),
                        decreasing=dict(marker=dict(color="#ef4444", line=dict(width=0))),
                        totals=dict(marker=dict(color="#00b4d8", line=dict(width=0))),
                        connector=dict(
                            line=dict(color="#30363d", width=1, dash="dot"),
                            visible=True,
                        ),
                    ))
                    # ── Add IV/share annotation ───────────────────────────
                    fig_wf.add_annotation(
                        xref="paper", yref="paper", x=0.5, y=1.06,
                        text=f"Intrinsic Value / share: <b>{fmts(_iv_per_sh, sym)}</b>",
                        showarrow=False,
                        font=dict(color="#f59e0b", size=12, family="IBM Plex Mono"),
                    )
                    apply_koyfin(fig_wf, height=360, extra_kw=dict(
                        showlegend=False,
                        yaxis=dict(
                            title=f"{to_code}B",
                            gridcolor="#21262d", tickfont=dict(color="#8b949e", size=10),
                        ),
                        xaxis=dict(tickfont=dict(color="#8b949e", size=10)),
                        margin=dict(t=54, b=16, l=56, r=16),
                    ))
                    st.plotly_chart(fig_wf, width="stretch", config={"displayModeBar":True,"modeBarButtonsToRemove":["lasso2d","select2d"],"toImageButtonOptions":{"filename":"dcf_waterfall","scale":2}})
                    ccard_end()

                # ── FIXED Sensitivity Heatmap
            with st.expander("🎯 Sensitivity Analysis — How WACC & Growth Rate Affect Fair Value"):
                # ── FIXED Sensitivity Heatmap ──────────────────────────
                # ── TIER CHECK: sensitivity ───────────────────────────
                if not _show_sensitive:
                    show_upgrade_modal("Sensitivity & scenario analysis")
                else:
                    sa_df = sensitivity_analysis(
                    projected_fcfs=projected, terminal_fcf_norm=terminal_norm,
                    total_debt=enriched["total_debt"], total_cash=enriched["total_cash"],
                    shares_outstanding=enriched["shares"], current_price=price_n,
                )
                    sa_display = (sa_df * fx).round(0)

                    # Find base-case cell indices for highlight border
                    _sa_cols = sa_display.columns.tolist()
                    _sa_rows = sa_display.index.tolist()

                    def _parse_sa_val(v):
                        """Parse sensitivity header in ANY format → float.
                        Handles: 'g=1', 'wacc=9.5', '1%', '1.0', '9', '3.0%'
                        """
                        import re as _re
                        s = str(v).strip()
                        # Extract the numeric part after '=' if present
                        if '=' in s:
                            s = s.split('=')[-1].strip()
                        # Remove any non-numeric characters except '.' and '-'
                        s = s.replace('%', '').replace(' ', '')
                        # Extract first number found
                        m = _re.search(r'-?\d+\.?\d*', s)
                        if m:
                            try:
                                return float(m.group())
                            except ValueError:
                                return 0.0
                        return 0.0

                    _base_col_idx = min(range(len(_sa_cols)), key=lambda i: abs(_parse_sa_val(_sa_cols[i]) - terminal_g * 100)) if _sa_cols else 0
                    _base_row_idx = min(range(len(_sa_rows)), key=lambda i: abs(_parse_sa_val(_sa_rows[i]) - wacc * 100)) if _sa_rows else 0

                    fig_sa = go.Figure(go.Heatmap(
                        z=sa_display.values.astype(float),
                        x=[str(c) for c in _sa_cols],
                        y=[str(r) for r in _sa_rows],
                        colorscale=[
                            [0.0,  "#7f1d1d"],
                            [0.25, "#ef4444"],
                            [0.45, "#f59e0b"],
                            [0.55, "#fbbf24"],
                            [0.70, "#10b981"],
                            [1.0,  "#064e3b"],
                        ],
                        zmid=price_d * fx if iv_d > 0 else None,
                        text=[[f"{sym}{v:,.0f}" if not np.isnan(v) else "N/A"
                               for v in row] for row in sa_display.values],
                        texttemplate="%{text}",
                        textfont=dict(size=11, color="#e6edf3", family="IBM Plex Mono"),
                        hovertemplate="WACC: %{y}<br>Terminal g: %{x}<br>IV: %{text}<extra></extra>",
                        showscale=True,
                        colorbar=dict(
                            tickfont=dict(color="#8b949e", size=10),
                            title=dict(text=f"IV ({to_code})", font=dict(color="#8b949e", size=11)),
                            thickness=10, bgcolor="#161b22",
                            bordercolor="#30363d", borderwidth=1,
                        ),
                    ))
                    # Highlight base-case cell
                    fig_sa.add_shape(type="rect",
                        xref="x", yref="y",
                        x0=_base_col_idx - 0.5, x1=_base_col_idx + 0.5,
                        y0=_base_row_idx - 0.5, y1=_base_row_idx + 0.5,
                        line=dict(color="#ffffff", width=2),
                    )
                    fig_sa.add_annotation(
                        text=f"Base case  ·  Current price {fmts(price_d, sym)}",
                        xref="paper", yref="paper", x=0.0, y=1.06,
                        font=dict(color="#8b949e", size=11, family="IBM Plex Mono"),
                        showarrow=False, xanchor="left",
                    )
                    fig_sa.add_annotation(
                        text="Darker green = more undervalued  ·  Red = overvalued",
                        xref="paper", yref="paper", x=1.0, y=1.06,
                        font=dict(color="#8b949e", size=9, family="Inter"),
                        showarrow=False, xanchor="right",
                    )
                    apply_koyfin(fig_sa, height=300, extra_kw=dict(
                        margin=dict(t=44, b=44, l=64, r=90),
                        xaxis=dict(title="Terminal Growth Rate →",
                                   tickfont=dict(size=11, color="#8b949e")),
                        yaxis=dict(title="← WACC",
                                   tickfont=dict(size=11, color="#8b949e")),
                    ))
                    st.plotly_chart(fig_sa, width="stretch", config={"displayModeBar":True,"modeBarButtonsToRemove":["lasso2d","select2d"],"toImageButtonOptions":{"filename":"sensitivity","scale":2}})
                    st.caption(f"White border = base case  ·  Green = undervalued  ·  Red = overvalued  ·  Values in {to_code}")
                    ccard_end()

                # ── Monte Carlo
                if _show_mc:
                    with st.expander("🎲 Monte Carlo Simulation — Probability Range of 1,000 Outcomes"):
                        if run_mc and mc_result and "iv_values" in mc_result:
                            mc_arr = mc_result["iv_values"] * fx
                            fig_mc = go.Figure()
                            fig_mc.add_trace(go.Histogram(
                            x=mc_arr, nbinsx=60,
                            marker=dict(color="#3b82f6", opacity=0.85, line=dict(width=0.5, color="#1d4ed8")),
                            name="IV Distribution",
                            ))
                            fig_mc.add_vline(x=price_d, line=dict(color="#ef4444", width=2, dash="dash"),
                                     annotation_text=f"Price {fmts(price_d, sym)}", annotation_font_color="#ef4444")
                            fig_mc.add_vline(x=mc_result["median_iv"]*fx, line=dict(color="#10b981", width=2, dash="dot"),
                                     annotation_text=f"Median IV {fmts(mc_result['median_iv']*fx, sym)}",
                                     annotation_font_color="#10b981")
                            apply_koyfin(fig_mc, height=240, extra_kw=dict(
                                        showlegend=False,
                                        xaxis=dict(title=f"IV ({to_code})", gridcolor="#21262d"),
                                        yaxis=dict(title="Frequency", gridcolor="#21262d"),
                                    ))
                            st.plotly_chart(fig_mc, width="stretch", config={"displayModeBar":True,"modeBarButtonsToRemove":["lasso2d","select2d"],"toImageButtonOptions":{"filename":"monte_carlo","scale":2}})
                            mc1,mc2,mc3,mc4,mc5 = st.columns(5)
                            mc1.metric("Median IV",   fmts(mc_result["median_iv"]*fx, sym))
                            mc2.metric("Bear (P10)",  fmts(mc_result["p10"]*fx, sym))
                            mc3.metric("Bull (P90)",  fmts(mc_result["p90"]*fx, sym))
                            mc4.metric("Std Dev",     fmts(mc_result["std_iv"]*fx, sym))
                            mc5.metric("P(Undervalued)", f"{mc_result['prob_undervalued']:.0%}")

                # ── Reverse DCF
            with st.expander("🔍 Reverse DCF — What Growth Rate Does the Current Price Imply?"):
                reverse_dcf_tab.render(
                    enriched=enriched, price_n=price_n, wacc=wacc,
                    terminal_g=terminal_g, forecast_yrs=forecast_yrs,
                    fx=fx, sym=sym,
                )

                # ── EV/EBITDA Multiples
            with st.expander("⚖️ Peer Comparison — EV/EBITDA & P/E vs Similar Companies"):
                try:
                    ev_res = run_ev_ebitda_analysis(
                        enriched=enriched,
                        current_price=price_n,
                        fx=fx,
                    )

                    if not ev_res.get("applicable"):
                        st.info(f"ℹ️ {ev_res.get('reason', 'EV/EBITDA not applicable for this sector')}")
                    else:
                        ev_colour_map = {"green":("🟢","#0D7A4E","#ECFDF5","#BBF7D0"),
                                         "amber":("🟡","#B45309","#FFFBEB","#FDE68A"),
                                         "red":  ("🔴","#B91C1C","#FEF2F2","#FECACA")}
                        ev_emoji, ev_txt_c, ev_bg_c, ev_bd_c = ev_colour_map.get(
                            ev_res["verdict_colour"], ev_colour_map["amber"])

                        # Verdict banner
                        st.html(f"""
                        <div style="padding:14px 20px;background:{ev_bg_c};
                                    border:1.5px solid {ev_bd_c};border-radius:10px;
                                    margin-bottom:16px;">
                          <div style="font-size:13px;font-weight:700;color:{ev_txt_c};
                                      text-transform:uppercase;letter-spacing:.05em;margin-bottom:4px;">
                            {ev_emoji} {ev_res['verdict']}
                          </div>
                          <div style="font-size:13px;color:#0F172A;line-height:1.7;">
                            {ev_res['summary']}
                          </div>
                          {f'<div style="font-size:12px;color:#94A3B8;margin-top:6px;">{ev_res["sector_note"]}</div>' if ev_res.get("sector_note") else ""}
                        </div>
                        """)

                        # Metrics row
                        em1, em2, em3, em4, em5 = st.columns(5)
                        em1.metric(
                            "Current EV/EBITDA",
                            f"{ev_res['current_multiple']:.1f}×",
                            help="Enterprise Value ÷ EBITDA at today's market price"
                        )
                        _live_peer = ev_res.get("live_peer_median")
                        em2.metric(
                            "Live peer median",
                            f"{_live_peer:.1f}×" if _live_peer else "Fetching…",
                            delta=f"{(ev_res['current_multiple']/_live_peer - 1)*100:+.0f}% vs peers" if _live_peer else None,
                            delta_color="inverse",
                            help="Current EV/EBITDA of sector peers (live from Yahoo Finance)"
                        )
                        em3.metric(
                            "Damodaran median",
                            f"{ev_res['sector_median']}×",
                            delta=f"{ev_res['multiple_pct']*100:+.0f}% vs long-run",
                            delta_color="inverse",
                            help="Damodaran NYU Jan 2025 long-run sector median — mean reversion target"
                        )
                        _peer_iv = ev_res.get("peer_iv", 0)
                        em4.metric(
                            f"IV at peer median",
                            fmts(_peer_iv, sym) if _peer_iv else "—",
                            help=f"Fair value at current peer multiple ({_live_peer:.0f}×)" if _live_peer else "Peer IV"
                        )
                        em5.metric(
                            f"IV at Damodaran",
                            fmts(ev_res['median_iv'], sym),
                            delta=f"{ev_res['mos_at_median']*100:+.1f}% MoS",
                            delta_color="normal",
                            help=f"Fair value at long-run median ({ev_res['sector_median']}×) — mean reversion scenario"
                        )

                        # Multiple gauge bar chart
                        bear_m  = ev_res["sector_bear"]
                        med_m   = ev_res["sector_median"]
                        bull_m  = ev_res["sector_bull"]
                        curr_m  = ev_res["current_multiple"]

                        fig_ev = go.Figure()

                        # Shaded zones
                        fig_ev.add_vrect(x0=0,      x1=bear_m, fillcolor="#ECFDF5", opacity=0.4, line_width=0)
                        fig_ev.add_vrect(x0=bear_m, x1=med_m,  fillcolor="#FFFBEB", opacity=0.4, line_width=0)
                        fig_ev.add_vrect(x0=med_m,  x1=bull_m, fillcolor="#FEF2F2", opacity=0.4, line_width=0)

                        # Zone labels
                        mid_cheap  = bear_m / 2
                        mid_fair   = (bear_m + med_m) / 2
                        mid_exp    = (med_m + bull_m) / 2
                        for x, label in [(mid_cheap,"Cheap"),(mid_fair,"Fair"),(mid_exp,"Expensive")]:
                            fig_ev.add_annotation(x=x, y=0.9, text=label, showarrow=False,
                                font=dict(size=11, color="#94A3B8"),
                                yref="paper", xanchor="center")

                        # Sector lines
                        for xval, label, clr, dash in [
                            (bear_m, f"Bear {bear_m}×", "#10B981", "dot"),
                            (med_m,  f"Median {med_m}×", "#3B82F6", "solid"),
                            (bull_m, f"Bull {bull_m}×",  "#EF4444", "dot"),
                        ]:
                            fig_ev.add_vline(x=xval, line=dict(color=clr, width=2, dash=dash),
                                annotation_text=label, annotation_font=dict(color=clr, size=11))

                        # Current multiple marker
                        ev_marker_clr = "#0D7A4E" if curr_m <= med_m else "#B91C1C"
                        fig_ev.add_trace(go.Scatter(
                            x=[curr_m], y=[0.5], mode="markers+text",
                            marker=dict(size=18, color=ev_marker_clr, symbol="diamond"),
                            text=[f"{curr_m:.1f}×"], textposition="top center",
                            textfont=dict(size=12, color=ev_marker_clr, family="IBM Plex Mono"),
                            name=f"{ticker_input} today",
                        ))

                        # Bear/Median/Bull IV values
                        ev_iv_x = [ev_res["bear_iv"]/fx, ev_res["median_iv"]/fx, ev_res["bull_iv"]/fx]
                        ev_scenarios = [
                            (bear_m,  f"Bear: {fmts(ev_res['bear_iv'],sym)}",   "#10B981"),
                            (med_m,   f"Base: {fmts(ev_res['median_iv'],sym)}", "#3B82F6"),
                            (bull_m,  f"Bull: {fmts(ev_res['bull_iv'],sym)}",   "#EF4444"),
                        ]
                        for xv, lbl, clr in ev_scenarios:
                            fig_ev.add_annotation(x=xv, y=0.15, text=lbl, showarrow=False,
                                font=dict(size=10, color=clr, family="IBM Plex Mono"),
                                yref="paper", xanchor="center")

                        # Add live peer median line if available
                        _live_m = ev_res.get("live_peer_median")
                        if _live_m:
                            fig_ev.add_vline(
                                x=_live_m,
                                line=dict(color="#8B5CF6", width=2.5, dash="dashdot"),
                                annotation_text=f"Live peers {_live_m:.1f}×",
                                annotation_font=dict(color="#8B5CF6", size=11),
                            )

                        max_x = max(bull_m * 1.3, curr_m * 1.1)
                        apply_koyfin(fig_ev, height=200, extra_kw=dict(
                            margin=dict(t=44, b=20, l=30, r=30),
                            xaxis=dict(title="EV/EBITDA Multiple", range=[0, max_x],
                                       gridcolor="#21262d", ticksuffix="×",
                                       tickfont=dict(color="#8b949e")),
                            yaxis=dict(visible=False, range=[0, 1]),
                            showlegend=False,
                        ))
                        st.plotly_chart(fig_ev, width="stretch",
                                        config={"displayModeBar": False})

                        # Bear/Median/Bull table
                        # Build scenario table with peer IV
                        _piv   = ev_res.get("peer_iv", 0)
                        _lpm   = ev_res.get("live_peer_median")
                        _scenarios  = ["Bear (trough)",   "Live peers (now)", "Damodaran (long-run)", "Bull (premium)"]
                        _multiples  = [f"{bear_m}×",
                                       f"{_lpm:.1f}×" if _lpm else "—",
                                       f"{med_m}×",
                                       f"{bull_m}×"]
                        _ivs        = [
                            fmts(ev_res['bear_iv'], sym),
                            fmts(_piv, sym) if _piv else "—",
                            fmts(ev_res['median_iv'], sym),
                            fmts(ev_res['bull_iv'], sym),
                        ]
                        _vs_today   = [
                            f"{(ev_res['bear_iv']/price_d - 1)*100:+.1f}%",
                            f"{(_piv/price_d - 1)*100:+.1f}%" if _piv and price_d else "—",
                            f"{(ev_res['median_iv']/price_d - 1)*100:+.1f}%",
                            f"{(ev_res['bull_iv']/price_d - 1)*100:+.1f}%",
                        ]
                        _benchmark  = ["Damodaran bear", "Current market", "Damodaran median", "Damodaran bull"]
                        ev_tbl = pd.DataFrame({
                            "Scenario":      _scenarios,
                            "EV/EBITDA":     _multiples,
                            f"IV ({sym})":   _ivs,
                            "vs today":      _vs_today,
                            "Benchmark":     _benchmark,
                        })
                        st.dataframe(ev_tbl, width='stretch', hide_index=True)

                        # Peer breakdown
                        _peers = ev_res.get("peer_data", {}).get("peers", [])
                        _valid_peers = [p for p in _peers if p.get("valid")]
                        if _valid_peers:
                            with st.expander(f"📊 Peer Multiples — {len(_valid_peers)} Comparable Companies"):
                                peer_tbl = pd.DataFrame([
                                    {"Peer": p["ticker"],
                                     "EV/EBITDA": f"{p['multiple']:.1f}×" if p.get("multiple") else "N/A"}
                                    for p in _peers
                                ])
                                st.dataframe(peer_tbl, width='stretch', hide_index=True)
                                st.caption("Live data from Yahoo Finance · Refreshed every hour")

                        st.caption(
                            f"Damodaran NYU Jan 2025 long-run median · "
                            f"Live peers: Yahoo Finance · "
                            f"EBITDA: {ev_res['ebitda_method']}"
                        )

                except Exception as _ev_err:
                    st.warning(f"EV/EBITDA could not run: {_ev_err}")
                    st.exception(_ev_err)

                # ── Historical Fair Value Chart
            with st.expander("📅 Historical Fair Value vs Actual Price — Model Track Record"):
                try:
                    hist = compute_historical_iv(
                        enriched=enriched,
                        current_price=price_n,
                        current_iv=iv_n,
                        wacc=wacc,
                        terminal_g=terminal_g,
                        forecast_yrs=forecast_yrs,
                        fx=fx,
                    )

                    if not hist["available"]:
                        st.info(f"ℹ️ {hist.get('reason','Historical data not available')}")
                    else:
                        # Summary banner
                        mos = hist["current_mos"]
                        if mos > 0.15:
                            hv_bg, hv_bd, hv_tc = "#ECFDF5","#A7F3D0","#065F46"
                            hv_icon = "🟢"
                        elif mos < -0.15:
                            hv_bg, hv_bd, hv_tc = "#FEF2F2","#FECACA","#991B1B"
                            hv_icon = "🔴"
                        else:
                            hv_bg, hv_bd, hv_tc = "#FFFBEB","#FDE68A","#92400E"
                            hv_icon = "🟡"

                        st.html(f"""
                        <div style="padding:12px 18px;background:{hv_bg};
                                    border:1.5px solid {hv_bd};border-radius:10px;
                                    margin-bottom:16px;font-size:13px;
                                    color:{hv_tc};line-height:1.7;">
                          {hv_icon} {hist['summary']}
                        </div>
                        """)

                        # ── Main chart ────────────────────────────────────
                        labels   = hist["labels"]
                        iv_hist  = hist["iv_history"]
                        curr_px  = hist["current_price"]

                        fig_hv = go.Figure()

                        # ±20% band around fair value (buy/sell zones)
                        iv_upper = [v * 1.20 for v in iv_hist]
                        iv_lower = [v * 0.80 for v in iv_hist]

                        # Green zone (below fair value = buy zone)
                        fig_hv.add_trace(go.Scatter(
                            x=labels + labels[::-1],
                            y=iv_upper + iv_lower[::-1],
                            fill="toself",
                            fillcolor="rgba(5,150,105,0.08)",
                            line=dict(color="rgba(0,0,0,0)"),
                            showlegend=False,
                            hoverinfo="skip",
                            name="±20% band",
                        ))

                        # Upper band line (overvalued threshold)
                        fig_hv.add_trace(go.Scatter(
                            x=labels, y=iv_upper,
                            mode="lines",
                            line=dict(color="rgba(220,38,38,0.3)", width=1, dash="dot"),
                            name="+20% (overvalued)",
                            showlegend=True,
                        ))

                        # Lower band line (undervalued threshold)
                        fig_hv.add_trace(go.Scatter(
                            x=labels, y=iv_lower,
                            mode="lines",
                            line=dict(color="rgba(5,150,105,0.3)", width=1, dash="dot"),
                            name="−20% (undervalued)",
                            showlegend=True,
                        ))

                        # Fair value line (main model output)
                        fig_hv.add_trace(go.Scatter(
                            x=labels, y=iv_hist,
                            mode="lines+markers",
                            line=dict(color="#2563EB", width=2.5),
                            marker=dict(size=7, color="#2563EB",
                                        line=dict(color="#FFFFFF", width=2)),
                            name="Model fair value",
                            hovertemplate=f"Fair value: {sym}%{{y:,.0f}}<extra></extra>",
                        ))

                        # Current price horizontal line
                        fig_hv.add_hline(
                            y=curr_px,
                            line=dict(color="#EF4444", width=2, dash="solid"),
                            annotation_text=f"Today's price: {fmts(curr_px, sym)}",
                            annotation_font=dict(color="#EF4444", size=11),
                        )

                        # Current price dot on "Today" bar
                        if "Today" in labels:
                            today_idx = labels.index("Today")
                            fig_hv.add_trace(go.Scatter(
                                x=["Today"], y=[curr_px],
                                mode="markers",
                                marker=dict(size=14, color="#EF4444",
                                            symbol="diamond",
                                            line=dict(color="#FFFFFF", width=2)),
                                name="Current price",
                                hovertemplate=f"Current price: {sym}{curr_px:,.0f}<extra></extra>",
                            ))

                        # Shade over/under valued regions
                        apply_koyfin(fig_hv, height=260, extra_kw=dict(
                            margin=dict(t=44, b=40, l=60, r=80),
                            xaxis=dict(title="Year", gridcolor="#21262d",
                                       tickfont=dict(size=11, color="#8b949e")),
                            yaxis=dict(title=f"Price per share ({sym})",
                                       gridcolor="#21262d",
                                       tickfont=dict(size=11, color="#8b949e"),
                                       tickprefix=sym),
                            legend=dict(orientation="h", yanchor="bottom", y=1.02,
                                        xanchor="left", x=0, font=dict(size=11, color="#8b949e")),
                            hovermode="x unified",
                        ))
                        st.plotly_chart(fig_hv, width="stretch",
                                        config={"displayModeBar": True,
                                                "modeBarButtonsToRemove": ["lasso2d","select2d"],
                                                "toImageButtonOptions": {"filename": f"historical_iv_{ticker_input}", "scale": 2}})

                        # Summary metrics
                        hm1, hm2, hm3, hm4 = st.columns(4)
                        first_iv = iv_hist[0] if iv_hist else 0
                        curr_iv  = iv_hist[-1] if iv_hist else 0
                        iv_cagr  = ((curr_iv / first_iv) ** (1 / max(len(iv_hist)-1, 1)) - 1) if first_iv > 0 else 0

                        hm1.metric(
                            f"Fair value {hist['labels'][0]}",
                            fmts(first_iv, sym),
                            help="Model fair value in the earliest year available"
                        )
                        hm2.metric(
                            "Fair value today",
                            fmts(curr_iv, sym),
                            delta=f"{iv_cagr*100:+.1f}% CAGR",
                            help="Fair value CAGR reflects FCF growth over the period"
                        )
                        hm3.metric(
                            "Current price",
                            fmts(curr_px, sym),
                            delta=f"{mos*100:+.1f}% vs fair value",
                            delta_color="inverse",
                        )
                        hm4.metric(
                            "Model trend",
                            hist["model_trend"].title(),
                            help="Whether our fair value estimate has been rising or falling"
                        )

                        st.caption(
                            "Fair value computed using actual FCF for each historical year "
                            "with same WACC and terminal growth as today's model. "
                            "Not a backtest — shows what our model would have estimated at the time."
                        )

                except Exception as _hv_err:
                    st.warning(f"Historical fair value could not run: {_hv_err}")
                    st.exception(_hv_err)

                # ── Dividend Discount Model
            with st.expander("💰 DDM — Dividend-Based Valuation (for Income Investors)"):
                # Collect dividend yield from all possible sources
                _div_y1 = enriched.get("dividend_yield", 0) or 0
                _div_y2 = raw.get("dividend_yield", 0) or 0
                _div_r1 = enriched.get("dividend_rate", 0) or 0
                _div_r2 = raw.get("dividend_rate", 0) or 0
                _div_yield_raw = _div_y1 or _div_y2 or 0
                _div_rate_raw  = _div_r1 or _div_r2 or 0
                # Normalize: Yahoo sometimes returns yield as integer % (e.g. 276 = 2.76%)
                if _div_yield_raw > 1:
                    _div_yield_raw = _div_yield_raw / 100
                if _div_yield_raw == 0 and _div_rate_raw > 0 and price_n > 0:
                    _div_yield_raw = _div_rate_raw / price_n
                if _div_yield_raw > 0:
                    enriched["dividend_yield"] = _div_yield_raw
                if _div_rate_raw > 0:
                    enriched["dividend_rate"] = _div_rate_raw

                if _div_yield_raw >= 0.005:
                    try:
                        ddm = compute_ddm(
                            enriched=enriched,
                            current_price=price_n,
                            dcf_iv=iv_n,
                            wacc=wacc,
                            fx=fx,
                        )
                        if not ddm["applicable"]:
                            st.info(f"DDM not applicable: {ddm['not_applicable_reason']}")
                        else:
                            sc = ddm["sustainability_colour"]
                            sc_map = {"green":("#065F46","#ECFDF5","#A7F3D0","✅"),
                                      "amber":("#92400E","#FFFBEB","#FDE68A","⚠️"),
                                      "red":  ("#991B1B","#FEF2F2","#FECACA","🚨")}
                            sc_tc,sc_bg,sc_bd,sc_icon = sc_map.get(sc, sc_map["amber"])
                            st.html(
                                f'''<div style="padding:12px 18px;background:{sc_bg};border:1.5px solid {sc_bd};border-radius:10px;margin-bottom:16px;">
                                <div style="font-size:13px;font-weight:700;color:{sc_tc};margin-bottom:4px;">{sc_icon} {ddm["sustainability_msg"]}</div>
                                <div style="font-size:13px;color:#0F172A;line-height:1.7;">{ddm["summary"]}</div></div>''',
                            )
                            dm1,dm2,dm3,dm4,dm5 = st.columns(5)
                            dm1.metric("Dividend yield",  f"{ddm['div_yield']*100:.1f}%")
                            dm2.metric("Annual dividend", fmts(ddm['div_rate'], sym))
                            dm3.metric("DDM fair value",  fmts(ddm['ddm_iv'], sym), help=ddm['model_used'])
                            dm4.metric("Blended IV",      fmts(ddm['blended_iv'], sym),
                                       delta=f"{ddm['mos_blended']*100:+.1f}% MoS", delta_color="normal")
                            dm5.metric("Payout ratio",    f"{ddm['payout_ratio']*100:.0f}%")
                            st.html(
                                f'''<div style="display:flex;gap:24px;padding:10px 14px;background:#F8FAFC;
                                border:1px solid #E2E8F0;border-radius:8px;margin:8px 0;font-size:13px;">
                                <span>📈 <b>Near-term growth:</b> {ddm["g_high"]*100:.1f}%</span>
                                <span>→</span>
                                <span>📉 <b>Stable growth:</b> {ddm["g_stable"]*100:.1f}%</span>
                                <span style="margin-left:auto;color:#64748B;">{ddm["growth_method"][:45]}</span></div>''',
                            )
                            scen_rows = []
                            for sname, sdata in ddm["scenarios"].items():
                                scen_rows.append({
                                    "Scenario": sname,
                                    "Growth": f"{sdata['g_high']*100:.1f}% → {sdata['g_stable']*100:.1f}%",
                                    f"IV ({sym})": fmts(sdata["iv"], sym),
                                    "MoS": f"{sdata['mos']*100:+.1f}%",
                                })
                            scen_rows.append({"Scenario":"── DCF","Growth":"—",
                                f"IV ({sym})":fmts(ddm["dcf_iv"],sym),"MoS":f"{ddm['mos_dcf']*100:+.1f}%"})
                            scen_rows.append({"Scenario":"⭐ Blended",
                                "Growth":f"{ddm['ddm_weight']:.0%} DDM / {ddm['dcf_weight']:.0%} DCF",
                                f"IV ({sym})":fmts(ddm["blended_iv"],sym),"MoS":f"{ddm['mos_blended']*100:+.1f}%"})
                            st.dataframe(pd.DataFrame(scen_rows), width='stretch', hide_index=True)
                            st.caption(f"Model: {ddm['model_used']} · Required return: {ddm['re']*100:.1f}% · "
                                       f"Blend: {ddm['ddm_weight']:.0%} DDM / {ddm['dcf_weight']:.0%} DCF")
                    except Exception as _ddm_err:
                        st.warning(f"DDM could not run: {_ddm_err}")
                        import traceback; st.code(traceback.format_exc())

                        # ── FCF Yield vs Bond Yield
            with st.expander("🛡️ Risk-Adjusted Return — FCF Yield vs Risk-Free Bond Rate"):
                # ── FCF Yield vs Bond Yield ────────────────────────────
                try:
                    fy = compute_fcf_yield_analysis(
                        enriched=enriched,
                        current_price=price_n,
                        fx=fx,
                    )

                    YIELD_COLOURS = {
                        "🟢": ("#059669","#ECFDF5","#A7F3D0"),
                        "🔵": ("#2563EB","#EFF6FF","#BFDBFE"),
                        "🟡": ("#D97706","#FFFBEB","#FDE68A"),
                        "🟠": ("#EA580C","#FFF7ED","#FED7AA"),
                        "🔴": ("#DC2626","#FEF2F2","#FECACA"),
                        "⚪": ("#64748B","#F8FAFC","#E2E8F0"),
                    }
                    fy_emoji = fy["verdict_emoji"]
                    fy_txt_c, fy_bg_c, fy_bd_c = YIELD_COLOURS.get(fy_emoji, YIELD_COLOURS["⚪"])

                    # Verdict banner
                    st.html(f"""
                    <div style="padding:14px 20px;background:{fy_bg_c};
                                border:1.5px solid {fy_bd_c};border-radius:10px;
                                margin-bottom:16px;">
                      <div style="font-size:13px;font-weight:700;color:{fy_txt_c};
                                  text-transform:uppercase;letter-spacing:.05em;margin-bottom:4px;">
                        {fy_emoji} {fy['verdict']}
                      </div>
                      <div style="font-size:13px;color:#0F172A;line-height:1.7;">
                        {fy['summary']}
                      </div>
                    </div>
                    """)

                    # 5 key metrics
                    ym1,ym2,ym3,ym4,ym5 = st.columns(5)
                    ym1.metric(
                        "FCF Yield",
                        f"{fy['fcf_yield']*100:.1f}%",
                        help="Free Cash Flow per share ÷ Current price. "
                             "Higher = more cash earned per dollar invested."
                    )
                    ym2.metric(
                        "Bond Yield",
                        f"{fy['bond_yield']*100:.1f}%",
                        help=f"Risk-free rate: {fy['bond_source']}"
                    )
                    ym3.metric(
                        "Equity Risk Premium",
                        f"{fy['erp']*100:+.1f}%",
                        delta="vs bonds",
                        delta_color="normal" if fy['erp'] > 0 else "inverse",
                        help="FCF Yield minus Bond Yield. Positive = stock pays more than bonds. "
                             "Negative = bonds pay more — you're not being compensated for equity risk."
                    )
                    ym4.metric(
                        f"vs {fy.get('index_label','S&P 500')} avg",
                        f"{fy.get('vs_index', fy.get('vs_sp500',0))*100:+.1f}%",
                        help=f"{fy.get('index_label','S&P 500')} average FCF yield is ~{fy.get('index_avg_fcf_yield', fy['sp500_avg_fcf_yield'])*100:.1f}%. "
                             f"Positive = this stock is cheaper than the index average on FCF basis."
                    )
                    ym5.metric(
                        "Payback period",
                        f"{fy['payback_years']:.0f} yrs" if fy['payback_years'] and fy['payback_years'] < 100 else "100+ yrs",
                        help="Years of current FCF needed to equal the price you pay today. "
                             "Lower is better. S&P 500 average is ~28 years."
                    )

                    # Yield comparison bar chart
                    ctx = fy["context"]
                    fig_fy = go.Figure()

                    bar_colours = [c["colour"] for c in ctx]
                    fig_fy.add_trace(go.Bar(
                        x=[c["label"] for c in ctx],
                        y=[c["value"] * 100 for c in ctx],
                        marker_color=bar_colours,
                        text=[f"{c['value']*100:.1f}%" for c in ctx],
                        textposition="outside",
                        textfont=dict(size=12, family="IBM Plex Mono"),
                    ))

                    # Bond yield reference line
                    fig_fy.add_hline(
                        y=fy["bond_yield"] * 100,
                        line=dict(color="#EF4444", width=2, dash="dot"),
                        annotation_text=f"Bond yield {fy['bond_yield']*100:.1f}%",
                        annotation_font=dict(color="#EF4444", size=11),
                    )

                    # S&P 500 average line
                    fig_fy.add_hline(
                        y=fy["sp500_avg_fcf_yield"] * 100,
                        line=dict(color="#3B82F6", width=1.5, dash="dash"),
                        annotation_text=f"S&P 500 avg {fy['sp500_avg_fcf_yield']*100:.1f}%",
                        annotation_font=dict(color="#3B82F6", size=11),
                    )

                    fig_fy.update_layout(
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="#FFFFFF",
                        height=260, margin=dict(t=30, b=20, l=40, r=80),
                        yaxis=dict(title="Yield (%)", gridcolor="#F1F5F9",
                                   ticksuffix="%"),
                        xaxis=dict(gridcolor="#F1F5F9"),
                        showlegend=False,
                    )
                    st.plotly_chart(fig_fy, width="stretch",
                                    config={"displayModeBar": False})

                    # Shareholder yield breakdown
                    if fy["div_yield"] > 0 or fy["buyback_yield"] > 0:
                        st.html("""
                        <div style="font-size:12px;font-weight:700;color:#475569;
                                    text-transform:uppercase;letter-spacing:.07em;
                                    margin-bottom:8px;">Shareholder yield breakdown</div>
                        """)
                        sy1, sy2, sy3 = st.columns(3)
                        sy1.metric("FCF yield",      f"{fy['fcf_yield']*100:.1f}%")
                        sy2.metric("Dividend yield", f"{fy['div_yield']*100:.1f}%")
                        sy3.metric("Total sh. yield",f"{fy['total_sh_yield']*100:.1f}%",
                                   help="FCF yield + dividend yield — total cash return to shareholders")

                    # Key insight box
                    if fy["erp"] < 0:
                        st.html(f"""
                        <div style="padding:12px 16px;background:#FFF7ED;
                                    border:1px solid #FED7AA;border-radius:8px;
                                    font-size:13px;color:#9A3412;margin-top:8px;">
                          ⚠️ <strong>Important context:</strong> This metric does NOT mean the stock is
                          a bad investment. Growth stocks often have low FCF yields because the market
                          is paying for future growth. The question is: <em>is the growth assumption
                          priced into the stock actually going to materialise?</em>
                          Check the <strong>Reverse DCF</strong> above to see what growth rate is implied.
                        </div>
                        """)
                    else:
                        st.html(f"""
                        <div style="padding:12px 16px;background:#ECFDF5;
                                    border:1px solid #A7F3D0;border-radius:8px;
                                    font-size:13px;color:#065F46;margin-top:8px;">
                          ✅ <strong>Key insight:</strong> This stock earns more in free cash flow than
                          risk-free bonds. You are being compensated for taking equity risk.
                          This is the foundation of a value investment thesis.
                        </div>
                        """)

                    _idx_lbl = fy.get('index_label','S&P 500')
                    _idx_yld = fy.get('index_avg_fcf_yield', fy['sp500_avg_fcf_yield'])
                    st.caption(f"Bond yield: {fy['bond_source']} · "
                               f"{_idx_lbl} avg FCF yield: ~{_idx_yld*100:.1f}% (long-run) · "
                               f"FCF: Yahoo Finance")

                except Exception as _fy_err:
                    st.warning(f"FCF Yield analysis could not run: {_fy_err}")
                    st.exception(_fy_err)
                ccard_end()

                # ── Quality (Bloomberg-grade redesign)
        if _active == "fundamentals":
            if not enriched:
                st.warning("Analysis data unavailable. Please run a new analysis.")
            else:
                earnings_quality_tab.render(enriched)

        if _active == "consensus":
            if not st.session_state.get("fin_enriched"):
                st.info("Run an analysis first to see this section.")
                st.stop()

            # ══════════════════════════════════════════════════════════
            # SECTION — Analyst Price Target Distribution
            # ══════════════════════════════════════════════════════════
            _pt_data = raw.get("finnhub_price_target", {}) if raw else {}
            _rec_trend = raw.get("finnhub_rec_trend", []) if raw else []

            if _pt_data and _pt_data.get("mean"):
                ccard("🎯 Analyst Price Target Distribution", "#0f4c75")
                _pt_mean  = float(_pt_data.get("mean",   0)) * fx
                _pt_high  = float(_pt_data.get("high",   0)) * fx
                _pt_low   = float(_pt_data.get("low",    0)) * fx
                _pt_count = int(_pt_data.get("count",    0))
                _pt_src   = _pt_data.get("source", "")

                import plotly.graph_objects as _go_pt

                # ── Range bar: low → mean → high vs current price ──
                _pt_fig = _go_pt.Figure()

                # Range bar (low to high)
                _pt_fig.add_trace(_go_pt.Bar(
                    x=[_pt_high - _pt_low],
                    y=["Price Targets"],
                    base=[_pt_low],
                    orientation="h",
                    marker_color="rgba(59,130,246,0.18)",
                    marker_line=dict(color="rgba(59,130,246,0.5)", width=1.5),
                    name="PT Range",
                    hovertemplate=f"Low: {sym}{_pt_low:,.2f}<br>High: {sym}{_pt_high:,.2f}<extra></extra>",
                ))
                # Mean price target dot
                _pt_fig.add_trace(_go_pt.Scatter(
                    x=[_pt_mean],
                    y=["Price Targets"],
                    mode="markers+text",
                    marker=dict(size=14, color="#2563EB", symbol="diamond"),
                    text=[f"Mean {sym}{_pt_mean:,.2f}"],
                    textposition="top center",
                    textfont=dict(size=11, color="#1D4ED8"),
                    name="Consensus Mean",
                    hovertemplate=f"Consensus Mean: {sym}{_pt_mean:,.2f}<extra></extra>",
                ))
                # Current price line
                _pt_fig.add_vline(
                    x=price_d,
                    line_color="#059669",
                    line_width=2,
                    line_dash="dash",
                    annotation_text=f"Current {sym}{price_d:,.2f}",
                    annotation_position="top right",
                    annotation_font=dict(size=11, color="#059669"),
                )
                _pt_upside = (((_pt_mean - price_d) / price_d) * 100) if price_d else 0
                _pt_fig.update_layout(
                    title=dict(
                        text=f"Analyst PT: {sym}{_pt_low:,.2f} – {sym}{_pt_high:,.2f}"
                             f"  ·  {_pt_count} analysts"
                             + (f"  ·  {_pt_upside:+.1f}% upside to mean" if _pt_upside else ""),
                        font_size=12,
                        x=0,
                    ),
                    height=170,
                    margin=dict(t=40, b=20, l=10, r=20),
                    plot_bgcolor="#FAFAFA",
                    paper_bgcolor="#FFFFFF",
                    showlegend=False,
                    xaxis=dict(
                        title=f"Price ({sym})",
                        gridcolor="#E2E8F0",
                        tickformat=",.0f",
                    ),
                    yaxis=dict(showticklabels=False, showgrid=False),
                )
                st.plotly_chart(_pt_fig, width='stretch')

                # Upside callout badge
                if _pt_upside:
                    _up_c = "#065F46" if _pt_upside > 15 else "#1E40AF" if _pt_upside > 0 else "#7F1D1D"
                    _up_bg = "#D1FAE5" if _pt_upside > 15 else "#DBEAFE" if _pt_upside > 0 else "#FEE2E2"
                    st.html(
                        f'<div style="display:inline-block;background:{_up_bg};color:{_up_c};'
                        f'font-size:12px;font-weight:700;padding:5px 14px;border-radius:14px;">'
                        f'{"📈" if _pt_upside > 0 else "📉"} '
                        f'{_pt_upside:+.1f}% consensus upside to mean · {_pt_count} analyst{"s" if _pt_count != 1 else ""}'
                        f'</div>'
                    )

                # Recommendation trend (stacked bar — last 4 months)
                if _rec_trend:
                    _rt_periods  = [r.get("period", "") for r in _rec_trend[-4:]]
                    _rt_sb  = [r.get("strongBuy",  0) for r in _rec_trend[-4:]]
                    _rt_b   = [r.get("buy",        0) for r in _rec_trend[-4:]]
                    _rt_h   = [r.get("hold",       0) for r in _rec_trend[-4:]]
                    _rt_s   = [r.get("sell",       0) for r in _rec_trend[-4:]]
                    _rt_ss  = [r.get("strongSell", 0) for r in _rec_trend[-4:]]

                    _rt_fig = _go_pt.Figure(data=[
                        _go_pt.Bar(name="Strong Buy",  x=_rt_periods, y=_rt_sb,  marker_color="#059669"),
                        _go_pt.Bar(name="Buy",         x=_rt_periods, y=_rt_b,   marker_color="#34d399"),
                        _go_pt.Bar(name="Hold",        x=_rt_periods, y=_rt_h,   marker_color="#d1d5db"),
                        _go_pt.Bar(name="Sell",        x=_rt_periods, y=_rt_s,   marker_color="#fca5a5"),
                        _go_pt.Bar(name="Strong Sell", x=_rt_periods, y=_rt_ss,  marker_color="#dc2626"),
                    ])
                    _rt_fig.update_layout(
                        barmode="stack",
                        title=dict(text="Analyst Recommendation Trend", font_size=12, x=0),
                        height=220,
                        margin=dict(t=36, b=20, l=10, r=10),
                        plot_bgcolor="#FAFAFA",
                        paper_bgcolor="#FFFFFF",
                        legend=dict(orientation="h", y=-0.25, x=0),
                        yaxis=dict(title="# Analysts", gridcolor="#E2E8F0"),
                        xaxis=dict(gridcolor="#E2E8F0"),
                    )
                    st.plotly_chart(_rt_fig, width='stretch')

                ccard_end()
                st.markdown("---")

            # ══════════════════════════════════════════════════════════
            # SECTION — Insider Activity
            # ══════════════════════════════════════════════════════════
            ccard("🔍 Insider Activity", "#1a1a2e")

            _ins      = raw.get("finnhub_insider", {}) if raw else {}
            _ins_sent = _ins.get("sentiment", "NEUTRAL")
            _ins_net  = _ins.get("net_shares_90d", 0)
            _ins_txns = _ins.get("transactions", [])
            _ins_mo   = _ins.get("monthly_net", {})

            _SENT_CFG = {
                "STRONG BUY":  ("#065F46", "#D1FAE5", "🟢"),
                "BUY":         ("#1E40AF", "#DBEAFE", "🔵"),
                "NEUTRAL":     ("#374151", "#F3F4F6", "⚪"),
                "SELL":        ("#92400E", "#FEF3C7", "🟡"),
                "STRONG SELL": ("#7F1D1D", "#FEE2E2", "🔴"),
            }
            _s_fg, _s_bg, _s_ico = _SENT_CFG.get(_ins_sent, ("#374151", "#F3F4F6", "⚪"))
            _net_dir = "net bought" if _ins_net >= 0 else "net sold"
            _net_abs = abs(_ins_net)

            # Sentiment badge + 90-day net summary
            if _ins:
                _net_color   = '#059669' if _ins_net >= 0 else '#DC2626'
                _net_sign    = '+' if _ins_net >= 0 else ''
                _wacc_nudge  = f'&nbsp;·&nbsp;WACC nudge: {_insider_adj:+.2%}' if _insider_adj != 0 else ''
                st.html(
                    f'<div style="display:flex;align-items:center;gap:16px;flex-wrap:wrap;margin-bottom:16px;">'
                    f'<div style="background:{_s_bg};color:{_s_fg};font-size:13px;font-weight:700;'
                    f'padding:8px 18px;border-radius:20px;letter-spacing:.04em;">'
                    f'{_s_ico} Insider Sentiment: {_ins_sent}</div>'
                    f'<div style="font-size:13px;color:#475569;">'
                    f'90-day net: <b style="color:{_net_color};">'
                    f'{_net_sign}{_ins_net:,} shares {_net_dir}</b>'
                    f'{_wacc_nudge}</div></div>'
                )
            else:
                st.info("Insider transaction data unavailable (requires Finnhub API key).")

            if _ins_txns:
                # ── Last 5 transactions table ─────────────────────
                st.html("""
                <div style="font-size:12px;font-weight:700;color:#475569;
                            text-transform:uppercase;letter-spacing:.07em;
                            margin-bottom:6px;">Recent Transactions (Last 5)</div>
                """)
                _tbl_rows = ""
                for _t in _ins_txns[:5]:
                    _t_color = "#059669" if _t["type"] == "Buy" else "#DC2626"
                    _t_price = f"${_t['price']:,.2f}" if _t["price"] > 0 else "—"
                    _t_val   = f"${_t['value']/1e6:.2f}M" if _t["value"] >= 1e6 \
                                else f"${_t['value']:,.0f}" if _t["value"] > 0 else "—"
                    _t_name  = (_t["name"] or "Unknown")[:28]
                    _tbl_rows += f"""
                    <tr>
                      <td style="padding:6px 10px;border-bottom:1px solid #E2E8F0;
                                 font-size:12px;color:#0F172A;">{_t_name}</td>
                      <td style="padding:6px 10px;border-bottom:1px solid #E2E8F0;
                                 font-size:12px;color:#0F172A;">{_t['date']}</td>
                      <td style="padding:6px 10px;border-bottom:1px solid #E2E8F0;">
                        <span style="background:{'#D1FAE5' if _t['type']=='Buy' else '#FEE2E2'};
                                     color:{_t_color};font-size:11px;font-weight:700;
                                     padding:2px 8px;border-radius:10px;">{_t['type']}</span>
                      </td>
                      <td style="padding:6px 10px;border-bottom:1px solid #E2E8F0;
                                 font-size:12px;color:#0F172A;text-align:right;">
                        {_t['shares']:,}</td>
                      <td style="padding:6px 10px;border-bottom:1px solid #E2E8F0;
                                 font-size:12px;color:#0F172A;text-align:right;">
                        {_t_price}</td>
                      <td style="padding:6px 10px;border-bottom:1px solid #E2E8F0;
                                 font-size:12px;color:#475569;text-align:right;">
                        {_t_val}</td>
                    </tr>"""
                st.html(f"""
                <div style="overflow-x:auto;border-radius:8px;border:1px solid #E2E8F0;">
                  <table style="width:100%;border-collapse:collapse;">
                    <thead>
                      <tr style="background:#F8FAFC;">
                        <th style="padding:7px 10px;text-align:left;font-size:11px;
                                   color:#64748B;font-weight:600;white-space:nowrap;">
                          INSIDER</th>
                        <th style="padding:7px 10px;text-align:left;font-size:11px;
                                   color:#64748B;font-weight:600;">DATE</th>
                        <th style="padding:7px 10px;text-align:left;font-size:11px;
                                   color:#64748B;font-weight:600;">TYPE</th>
                        <th style="padding:7px 10px;text-align:right;font-size:11px;
                                   color:#64748B;font-weight:600;">SHARES</th>
                        <th style="padding:7px 10px;text-align:right;font-size:11px;
                                   color:#64748B;font-weight:600;">PRICE</th>
                        <th style="padding:7px 10px;text-align:right;font-size:11px;
                                   color:#64748B;font-weight:600;">VALUE</th>
                      </tr>
                    </thead>
                    <tbody>{_tbl_rows}</tbody>
                  </table>
                </div>
                """)

                # ── Net buying/selling bar chart (12 months) ──────
                if _ins_mo:
                    import plotly.graph_objects as _go_ins
                    _sorted_mo = sorted(_ins_mo.items())
                    _mo_labels = [m for m, _ in _sorted_mo]
                    _mo_values = [v for _, v in _sorted_mo]
                    _mo_colors = ["#059669" if v >= 0 else "#DC2626" for v in _mo_values]
                    _fig_ins = _go_ins.Figure(
                        _go_ins.Bar(
                            x=_mo_labels, y=_mo_values,
                            marker_color=_mo_colors,
                            text=[f"{v:+,.0f}" for v in _mo_values],
                            textposition="outside",
                            textfont_size=10,
                        )
                    )
                    _fig_ins.update_layout(
                        title="Monthly Net Insider Share Activity (12M)",
                        title_font_size=13,
                        height=260,
                        margin=dict(t=40, b=30, l=50, r=20),
                        plot_bgcolor="#FAFAFA",
                        paper_bgcolor="#FFFFFF",
                        yaxis=dict(
                            title="Net Shares",
                            gridcolor="#E2E8F0",
                            zerolinecolor="#94A3B8",
                        ),
                        xaxis=dict(gridcolor="#E2E8F0"),
                        showlegend=False,
                    )
                    _fig_ins.add_hline(y=0, line_color="#94A3B8", line_width=1)
                    st.plotly_chart(_fig_ins, width='stretch')

            st.markdown("---")

            # ══════════════════════════════════════════════════════════
            # SECTION — Smart Money (Institutional Ownership)
            # ══════════════════════════════════════════════════════════
            ccard("🏛️ Smart Money — Institutional Ownership", "#0f2942")

            _inst = raw.get("finnhub_institutional", {}) if raw else {}

            if _inst:
                # Persist snapshot to DB so history chart can be built over time
                try:
                    save_institutional_ownership(ticker_input, _inst)
                except Exception:
                    pass

                _it_pct      = _inst.get("total_pct", 0)
                _it_qoq      = _inst.get("qoq_change_pct", 0)
                _it_trend    = _inst.get("trend", "STABLE")
                _it_accum    = _inst.get("accumulation", False)
                _it_avg5     = _inst.get("avg_top5_chg", 0)
                _it_num      = _inst.get("num_holders", 0)
                _it_quarter  = _inst.get("quarter", "")
                _it_holders  = _inst.get("holders", [])

                _TREND_CFG = {
                    "ACCUMULATING": ("#065F46", "#D1FAE5", "📈"),
                    "DISTRIBUTING": ("#7F1D1D", "#FEE2E2", "📉"),
                    "STABLE":       ("#1E3A5F", "#DBEAFE", "➡️"),
                }
                _it_fg, _it_bg, _it_ico = _TREND_CFG.get(_it_trend, ("#374151", "#F3F4F6", "➡️"))
                _qoq_sign = "+" if _it_qoq >= 0 else ""

                # Metric row
                _mc1, _mc2, _mc3, _mc4 = st.columns(4)
                _mc1.metric(
                    "Institutional Ownership",
                    f"{_it_pct:.1f}%",
                    help=f"% of shares outstanding held by {_it_num} reporting institutions"
                )
                _mc2.metric(
                    "QoQ Change",
                    f"{_qoq_sign}{_it_qoq:.2f}%",
                    delta=f"{'▲ Accumulating' if _it_qoq > 0 else '▼ Distributing' if _it_qoq < 0 else 'Stable'}",
                    delta_color="normal" if _it_qoq > 0 else "inverse" if _it_qoq < 0 else "off",
                    help="QoQ change in total institutional share count"
                )
                _mc3.metric(
                    "Top-5 Avg Change",
                    f"{'+' if _it_avg5 >= 0 else ''}{_it_avg5:.1f}%",
                    help="Average QoQ position change across top-5 holders"
                )
                _mc4.metric(
                    "# Institutions",
                    f"{_it_num:,}",
                    help="Number of institutions reporting positions this quarter"
                )

                # Trend badge + optional Institutional Accumulation badge
                _badges_html = f"""
                <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;
                            margin:10px 0 14px;">
                  <span style="background:{_it_bg};color:{_it_fg};font-size:12px;
                               font-weight:700;padding:5px 14px;border-radius:14px;
                               letter-spacing:.03em;">
                    {_it_ico} {_it_trend}
                    {f' · as of {_it_quarter}' if _it_quarter else ''}
                  </span>
                """
                if _it_accum:
                    _badges_html += """
                  <span style="background:#FEF9C3;color:#713F12;font-size:12px;
                               font-weight:700;padding:5px 14px;border-radius:14px;
                               border:1px solid #FDE68A;letter-spacing:.03em;">
                    ⭐ Institutional Accumulation
                  </span>
                """
                _badges_html += "</div>"
                st.html(_badges_html)

                # Top-5 holders table
                if _it_holders:
                    st.html("""
                    <div style="font-size:12px;font-weight:700;color:#475569;
                                text-transform:uppercase;letter-spacing:.07em;
                                margin-bottom:6px;">Top-5 Institutional Holders</div>
                    """)
                    _h_rows = ""
                    for _h in _it_holders:
                        _h_chg = _h.get("change_pct", 0)
                        _h_chg_clr = "#059669" if _h_chg > 0 else "#DC2626" if _h_chg < 0 else "#64748B"
                        _h_chg_sym = "▲" if _h_chg > 0 else "▼" if _h_chg < 0 else "—"
                        _h_sh_chg  = _h.get("change_shares", 0)
                        _h_sh_sign = "+" if _h_sh_chg >= 0 else ""
                        _h_rows += f"""
                        <tr>
                          <td style="padding:6px 10px;border-bottom:1px solid #E2E8F0;
                                     font-size:12px;color:#0F172A;font-weight:500;">
                            {_h.get('name','')[:35]}</td>
                          <td style="padding:6px 10px;border-bottom:1px solid #E2E8F0;
                                     font-size:12px;color:#0F172A;text-align:right;">
                            {_h.get('shares', 0):,}</td>
                          <td style="padding:6px 10px;border-bottom:1px solid #E2E8F0;
                                     font-size:12px;color:#0F172A;text-align:right;">
                            {_h.get('share_pct', 0):.3f}%</td>
                          <td style="padding:6px 10px;border-bottom:1px solid #E2E8F0;
                                     font-size:12px;text-align:right;font-weight:600;
                                     color:{_h_chg_clr};">
                            {_h_chg_sym} {abs(_h_chg):.1f}%</td>
                          <td style="padding:6px 10px;border-bottom:1px solid #E2E8F0;
                                     font-size:12px;color:{_h_chg_clr};text-align:right;">
                            {_h_sh_sign}{_h_sh_chg:,}</td>
                          <td style="padding:6px 10px;border-bottom:1px solid #E2E8F0;
                                     font-size:11px;color:#94A3B8;text-align:right;">
                            {_h.get('filing_date','')}</td>
                        </tr>"""
                    st.html(f"""
                    <div style="overflow-x:auto;border-radius:8px;border:1px solid #E2E8F0;">
                      <table style="width:100%;border-collapse:collapse;">
                        <thead>
                          <tr style="background:#F8FAFC;">
                            <th style="padding:7px 10px;text-align:left;font-size:11px;
                                       color:#64748B;font-weight:600;">INSTITUTION</th>
                            <th style="padding:7px 10px;text-align:right;font-size:11px;
                                       color:#64748B;font-weight:600;">SHARES HELD</th>
                            <th style="padding:7px 10px;text-align:right;font-size:11px;
                                       color:#64748B;font-weight:600;">% FLOAT</th>
                            <th style="padding:7px 10px;text-align:right;font-size:11px;
                                       color:#64748B;font-weight:600;">QoQ CHG %</th>
                            <th style="padding:7px 10px;text-align:right;font-size:11px;
                                       color:#64748B;font-weight:600;">SHARES ΔQoQ</th>
                            <th style="padding:7px 10px;text-align:right;font-size:11px;
                                       color:#64748B;font-weight:600;">FILED</th>
                          </tr>
                        </thead>
                        <tbody>{_h_rows}</tbody>
                      </table>
                    </div>
                    """)

                # Ownership trend chart from DB history
                _hist = get_institutional_history(ticker_input, quarters=8)
                if len(_hist) >= 2:
                    import plotly.graph_objects as _go_inst
                    _hist_rev = list(reversed(_hist))   # oldest → newest for chart
                    _h_quarters = [r["quarter"] for r in _hist_rev]
                    _h_pcts     = [r["total_pct"] for r in _hist_rev]
                    _h_qoq      = [r["qoq_change"] for r in _hist_rev]

                    _fig_inst = _go_inst.Figure()
                    _fig_inst.add_trace(_go_inst.Scatter(
                        x=_h_quarters, y=_h_pcts,
                        mode="lines+markers+text",
                        name="Institutional %",
                        line=dict(color="#2563EB", width=2.5),
                        marker=dict(size=7, color="#2563EB"),
                        text=[f"{p:.1f}%" for p in _h_pcts],
                        textposition="top center",
                        textfont_size=10,
                        fill="tozeroy",
                        fillcolor="rgba(37,99,235,0.08)",
                        yaxis="y",
                    ))
                    _fig_inst.add_trace(_go_inst.Bar(
                        x=_h_quarters, y=_h_qoq,
                        name="QoQ Δ%",
                        marker_color=["#059669" if v >= 0 else "#DC2626" for v in _h_qoq],
                        opacity=0.7,
                        yaxis="y2",
                    ))
                    _fig_inst.update_layout(
                        title="Institutional Ownership Trend",
                        title_font_size=13,
                        height=270,
                        margin=dict(t=40, b=30, l=50, r=60),
                        plot_bgcolor="#FAFAFA",
                        paper_bgcolor="#FFFFFF",
                        legend=dict(orientation="h", y=-0.15, x=0),
                        yaxis=dict(
                            title="Total Ownership %",
                            gridcolor="#E2E8F0",
                            ticksuffix="%",
                        ),
                        yaxis2=dict(
                            title="QoQ Δ%",
                            overlaying="y",
                            side="right",
                            showgrid=False,
                            ticksuffix="%",
                            zeroline=True,
                            zerolinecolor="#94A3B8",
                        ),
                        xaxis=dict(gridcolor="#E2E8F0"),
                    )
                    st.plotly_chart(_fig_inst, width='stretch')
                elif len(_hist) == 1:
                    st.caption("📊 Ownership trend chart will appear after multiple quarters of data have been recorded.")

            else:
                st.info("Institutional ownership data unavailable (requires Finnhub API key or this endpoint may need a paid plan).")

            st.markdown("---")

            # ── Sector / Peer Comparison (existing content) ───────
            ccard("How does this stock compare to its peers?", "#0f4c75")
            try:
                sr = compute_sector_relative(
                    enriched=enriched,
                    current_price=price_n,
                    current_iv=iv_n,
                    current_mos=mos,
                    fx=fx,
                )

                SR_COLOURS = {
                    "#059669": ("#ECFDF5","#A7F3D0"),
                    "#2563EB": ("#EFF6FF","#BFDBFE"),
                    "#D97706": ("#FFFBEB","#FDE68A"),
                    "#EA580C": ("#FFF7ED","#FED7AA"),
                    "#DC2626": ("#FEF2F2","#FECACA"),
                }
                sr_tc  = sr["verdict_colour"]
                sr_bg, sr_bd = SR_COLOURS.get(sr_tc, ("#F8FAFC","#E2E8F0"))

                # Verdict banner
                st.html(f"""
                <div style="padding:12px 18px;background:{sr_bg};
                            border:1.5px solid {sr_bd};border-radius:10px;
                            margin-bottom:16px;">
                  <div style="font-size:13px;font-weight:700;color:{sr_tc};
                              text-transform:uppercase;letter-spacing:.05em;
                              margin-bottom:4px;">
                    {sr['verdict_emoji']} {sr['verdict']}
                  </div>
                  <div style="font-size:13px;color:#0F172A;line-height:1.7;">
                    {sr['summary']}
                  </div>
                </div>
                """)

                # ── Screener stats panel ──────────────────────────
                screen = sr.get("screener", {})
                if screen.get("available"):
                    sc1, sc2, sc3, sc4 = st.columns(4)
                    sc1.metric(
                        "Sector rank",
                        f"#{screen['rank']} of {screen['rank_of']}",
                        help=f"Ranked by margin of safety vs {screen['total_stocks']} stocks in {sr['sector_name']}"
                    )
                    sc2.metric(
                        "Percentile",
                        f"{screen['percentile']:.0f}th",
                        delta="cheaper than peers" if screen["percentile"] > 60 else
                              "pricier than peers" if screen["percentile"] < 40 else "in line",
                        delta_color="normal" if screen["percentile"] > 60 else
                                    "inverse" if screen["percentile"] < 40 else "off",
                        help="What % of sector peers are cheaper than this stock"
                    )
                    sc3.metric(
                        "Sector median MoS",
                        f"{screen['median_mos']:+.1f}%",
                        help="Median margin of safety across all stocks in this sector"
                    )
                    sc4.metric(
                        "Sector BUY signals",
                        f"{screen['buy_pct']:.0f}%",
                        help=f"% of sector with BUY signal. SELL: {screen['sell_pct']:.0f}%"
                    )

                    # Sector signal distribution bar
                    total_known = screen["buy_pct"] + screen["sell_pct"] + screen.get("watch_pct", 0)
                    st.html(f"""
                    <div style="margin:12px 0 8px;font-size:12px;font-weight:700;
                                color:#475569;text-transform:uppercase;letter-spacing:.07em;">
                      Sector signal distribution ({screen['total_stocks']} stocks)
                    </div>
                    <div style="display:flex;height:20px;border-radius:6px;overflow:hidden;gap:2px;">
                      <div style="width:{screen['buy_pct']}%;background:#059669;
                                  display:flex;align-items:center;justify-content:center;
                                  font-size:12px;color:#fff;font-weight:700;min-width:0;">
                        {'BUY ' + str(round(screen['buy_pct'])) + '%' if screen['buy_pct'] > 8 else ''}
                      </div>
                      <div style="width:{screen.get('watch_pct',0)}%;background:#2563EB;
                                  min-width:0;"></div>
                      <div style="flex:1;background:#E2E8F0;min-width:0;"></div>
                      <div style="width:{screen['sell_pct']}%;background:#DC2626;
                                  display:flex;align-items:center;justify-content:center;
                                  font-size:12px;color:#fff;font-weight:700;min-width:0;">
                        {'SELL ' + str(round(screen['sell_pct'])) + '%' if screen['sell_pct'] > 8 else ''}
                      </div>
                    </div>
                    <div style="display:flex;gap:16px;margin-top:4px;font-size:12px;color:#64748B;">
                      <span>🟢 BUY {screen['buy_pct']:.0f}%</span>
                      <span>🔴 SELL {screen['sell_pct']:.0f}%</span>
                      <span>⚪ Other {100-screen['buy_pct']-screen['sell_pct']:.0f}%</span>
                    </div>
                    """)

                    # Top picks in sector
                    top = screen.get("top_picks")
                    if top is not None and not top.empty:
                        with st.expander(f"🏆 Top Picks in {sr['sector_name']} — Ranked by Margin of Safety"):
                            st.dataframe(
                                top.rename(columns={
                                    "ticker":"Ticker",
                                    "price":"Price",
                                    "intrinsic_value":"Fair Value",
                                    "margin_of_safety":"MoS %",
                                    "signal":"Signal",
                                    "fundamental_grade":"Quality",
                                }),
                                width='stretch',
                                hide_index=True,
                            )

                # ── Live peer comparison ──────────────────────────
                peers            = sr.get("peer_metrics", [])
                curr             = sr.get("current_metrics", {})
                _peer_grp_label  = sr.get("peer_group_label", "sector peers")

                if peers or curr:
                    st.html(f"""
                    <div style="margin:16px 0 8px;font-size:12px;font-weight:700;
                                color:#475569;text-transform:uppercase;letter-spacing:.07em;">
                      Head-to-head — {_peer_grp_label}
                    </div>
                    """)

                    # Build comparison table
                    all_rows = []
                    # Current stock first
                    all_rows.append({
                        "Ticker": f"⭐ {curr['ticker']}",
                        "Price/Earnings": f"{curr['pe']:.1f}×" if curr.get("pe") else "—",
                        "Company Value Multiple": f"{curr['ev_ebitda']:.1f}×" if curr.get("ev_ebitda") else "—",
                        "Cash Yield": f"{curr['fcf_yield']:.1f}%" if curr.get("fcf_yield") else "—",
                        "Market Cap ($B)": f"${curr['mktcap_b']:.0f}B" if curr.get("mktcap_b") else "—",
                        "MoS% (analysed stock only)": f"{curr.get('mos_pct',0):+.1f}%",
                    })
                    for p in peers:
                        all_rows.append({
                            "Ticker": p["ticker"],
                            "Price/Earnings": f"{p['pe']:.1f}×" if p.get("pe") else "—",
                            "Company Value Multiple": f"{p['ev_ebitda']:.1f}×" if p.get("ev_ebitda") else "—",
                            "Cash Yield": f"{p['fcf_yield']:.1f}%" if p.get("fcf_yield") else "—",
                            "Market Cap ($B)": f"${p['mktcap_b']:.0f}B" if p.get("mktcap_b") else "—",
                            "MoS% (analysed stock only)": "n/a",
                        })

                    # Peer medians row
                    pm_pe  = sr.get("peer_median_pe")
                    pm_ev  = sr.get("peer_median_ev")
                    pm_fcf = sr.get("peer_median_fcf")
                    all_rows.append({
                        "Ticker": "── Peer median ──",
                        "Price/Earnings": f"{pm_pe:.1f}×" if pm_pe else "—",
                        "Company Value Multiple": f"{pm_ev:.1f}×" if pm_ev else "—",
                        "Cash Yield": f"{pm_fcf:.1f}%" if pm_fcf else "—",
                        "Market Cap ($B)": "—",
                        "MoS% (analysed stock only)": "n/a",
                    })

                    if all_rows:
                        peer_df = pd.DataFrame(all_rows)
                        st.dataframe(peer_df, width='stretch', hide_index=True)

                        # vs peers callout
                        pe_vs   = sr.get("pe_vs_peers")
                        ev_vs   = sr.get("ev_vs_peers")
                        # Guard: discard ratios that are clearly garbage (>500% diff = bad data)
                        if pe_vs is not None and abs(pe_vs) > 5:
                            pe_vs = None
                        if ev_vs is not None and abs(ev_vs) > 5:
                            ev_vs = None
                        if pe_vs is not None or ev_vs is not None:
                            msgs = []
                            if pe_vs is not None:
                                msgs.append(
                                    f"Price vs peers: {pe_vs*100:+.0f}%"
                                    f" — {'more expensive' if pe_vs > 0 else 'cheaper' } than similar companies"
                                )
                            if ev_vs is not None:
                                msgs.append(
                                    f"Company value vs peers: {ev_vs*100:+.0f}%"
                                    f" — {'more expensive' if ev_vs > 0 else 'cheaper'} than similar companies"
                                )
                            if msgs:
                                clr = "#065F46" if (pe_vs or 0) < -0.10 else "#991B1B" if (pe_vs or 0) > 0.20 else "#92400E"
                                bg  = "#ECFDF5" if (pe_vs or 0) < -0.10 else "#FEF2F2" if (pe_vs or 0) > 0.20 else "#FFFBEB"
                                st.html(f"""
                                <div style="padding:8px 14px;background:{bg};
                                            border-radius:8px;font-size:12px;color:{clr};">
                                  {" · ".join(msgs)}
                                </div>
                                """)

                        if not peers:
                            st.caption("💡 Live peer data unavailable — run with internet connection for head-to-head comparison")
                        else:
                            st.caption(
                                f"Peers selected by market cap and sector classification ({_peer_grp_label}). "
                                "Margin of Safety (MoS%) is only computed for the analysed stock."
                            )
                            st.caption("Live peer data: Yahoo Finance — MoS only computed for analysed stock")

            except Exception as _sr_err:
                st.warning(f"Sector relative valuation could not run: {_sr_err}")
            ccard_end()


        # ── Ask AI Analyst tab ─────────────────────────────────
        if _active == "ask_ai":
            try:
                import sys as _sys_ai
                from pathlib import Path as _Path_ai
                _dash_ai = str(_Path_ai(__file__).parent)
                if _dash_ai not in _sys_ai.path:
                    _sys_ai.path.insert(0, _dash_ai)
                from ai_chat import render_ai_chat as _render_ai_chat, build_stock_context as _build_ctx

                # Build analysis_data dict from variables in scope
                _ai_data = {
                    "ticker":                 ticker_input,
                    "company_name":           enriched.get("company_name", ticker_input),
                    "sector":                 enriched.get("sector",        ""),
                    "sym":                    sym,
                    "price":                  price_d,
                    "iv":                     iv_d,
                    "mos_pct":                mos_pct,
                    "signal":                 sig,
                    "wacc":                   wacc,
                    "terminal_g":             terminal_g,
                    "fcf_growth":             enriched.get("fcf_growth",      0),
                    "revenue_growth":         enriched.get("revenue_growth",  0),
                    "op_margin":              enriched.get("op_margin",       0),
                    "gross_margin":           enriched.get("gross_margin",    0),
                    "net_margin":             enriched.get("net_margin",      0),
                    "roe":                    enriched.get("roe",             0),
                    "roce":                   enriched.get("roce",            0),
                    "de_ratio":               enriched.get("de_ratio",        0),
                    "moat_score":             enriched.get("moat_score",      0),
                    "moat_grade":             enriched.get("moat_grade",      ""),
                    "moat_types":             enriched.get("moat_types",      []),
                    "piotroski_score":        enriched.get("piotroski_score", None),
                    "earnings_quality_grade": enriched.get("earnings_quality_grade", ""),
                    "earnings_quality_score": enriched.get("earnings_quality_score", None),
                    "forward_pe":             enriched.get("forward_pe",      0),
                    "ev_ebitda":              enriched.get("ev_ebitda",       0),
                    "fcf_yield":              enriched.get("fcf_yield",       0),
                    "scenarios":              scenarios,
                    "earnings_track_record":  raw.get("earnings_track_record", {}),
                }
                _render_ai_chat(_ai_data)
                st.caption(
                    "⚠️ Model output only — not investment advice. "
                    "YieldIQ is not a registered investment adviser. "
                    "Past model performance does not predict future results."
                )
            except Exception as _ai_err:
                st.error(f"AI chat error: {_ai_err}")

        # ── Download Section ───────────────────────────────────
        st.markdown("---")
        st.html("""
        <div style="font-size:13px;font-weight:700;color:#475569;text-transform:uppercase;
                    letter-spacing:0.04em;margin-bottom:12px;">⬇️ Export & Download</div>
        """)
        dl1, dl2, dl3 = st.columns(3)

        report_data = {
            "price": price_d, "iv": iv_d, "mos_pct": mos_pct,
            "signal": sig, "wacc": wacc, "term_g": terminal_g,
            "rev_growth": enriched.get("revenue_growth", 0),
            "fcf_growth": enriched.get("fcf_growth", 0),
            "op_margin":  enriched.get("op_margin", 0),
            "fund_grade": fs["grade"], "fund_score": fs["score"],
            "entry_signal": pt.get("entry_signal", ""),
            "buy_price": (pt.get("buy_price") or 0) * fx,
            "target_price": (pt.get("target_price") or 0) * fx,
            "stop_loss": (pt.get("stop_loss") or 0) * fx,
            "sl_pct": pt.get("sl_pct", 15),
            "rr_ratio": pt.get("rr_ratio", 0),
            "holding_period": hp.get("label", "N/A"),
            "sum_pv_fcfs": dcf_res.get("sum_pv_fcfs", 0) * fx,
            "pv_tv": pv_tv_d,
            "ev": dcf_res.get("enterprise_value", 0) * fx,
            "debt": enriched["total_debt"] * fx,
            "cash": enriched["total_cash"] * fx,
            "equity": dcf_res.get("equity_value", 0) * fx,
            "shares": enriched["shares"],
            "bear_iv": scenarios.get("Bear 🐻", {}).get("iv", 0) * fx,
            "bull_iv": scenarios.get("Bull 🐂", {}).get("iv", 0) * fx,
            "dcf_only_iv": enriched.get("dcf_iv", 0) * fx,
        }

        # Sensitivity table for Excel
        sa_df_for_excel = sensitivity_analysis(
            projected_fcfs=projected, terminal_fcf_norm=terminal_norm,
            total_debt=enriched["total_debt"], total_cash=enriched["total_cash"],
            shares_outstanding=enriched["shares"], current_price=price_n,
        )

        with dl1:
            report_bytes = generate_dcf_report(ticker_input, report_data, scenarios, sym)
            st.download_button(
                "📄 Download DCF Report (.txt)",
                data=report_bytes,
                file_name=f"DCF_{ticker_input}_{datetime.now().strftime('%Y%m%d')}.txt",
                mime="text/plain",
                width='stretch',
            )

        with dl2:
            # ── TIER CHECK: report download ────────────────────
            _can_dl, _dl_reason = can_download_report()
            if not _can_dl and limit("reports_per_month") == 0:
                # Free tier — offer purchase
                st.html(f"""
                <div style="text-align:center;padding:8px;">
                  <a href="https://buy.stripe.com/your_link_here" target="_blank"
                     style="background:#5046e4;color:#fff;font-size:13px;font-weight:600;
                            padding:9px 20px;border-radius:8px;text-decoration:none;
                            display:inline-block;width:100%;text-align:center;">
                    📄 Buy Report — $4.99
                  </a>
                </div>
                """)
            elif not _can_dl:
                st.warning(f"📄 {_dl_reason}")
                st.caption(f"Extra reports: ${limit('report_cost'):.2f} each")
                show_report_upsell()
            else:
                # ── Complete 14-Sheet Hedge Fund Model ─────────────
                st.caption(f"📄 {_dl_reason}")
            if _can_dl:
                try:
                    import sys as _sys
                    from pathlib import Path as _Path
                    _project_root = str(_Path(__file__).parent.parent)
                    if _project_root not in _sys.path:
                        _sys.path.insert(0, _project_root)
                    from generate_dcf_excel import generate_institutional_dcf
                    from generate_hf_excel import build_hedge_fund_sheets
                    from generate_portfolio_excel import build_portfolio_sheets
                    import io as _io
                    from openpyxl import load_workbook as _lwb

                    _hf_bytes = generate_institutional_dcf(
                    ticker=ticker_input, enriched=enriched, dcf_res=dcf_res,
                    forecast_result=forecast_result, scenarios=scenarios,
                    wacc_data=wacc_data, wacc=wacc, terminal_g=terminal_g,
                    forecast_yrs=forecast_yrs, sym=sym, to_code=to_code, fx=fx,
                    )
                    _wb = _lwb(filename=_io.BytesIO(_hf_bytes))
                    _wb = build_hedge_fund_sheets(
                    wb=_wb, ticker=ticker_input, enriched=enriched,
                    dcf_res=dcf_res, forecast_result=forecast_result,
                    scenarios=scenarios, wacc_data=wacc_data,
                    wacc=wacc, terminal_g=terminal_g,
                    forecast_yrs=forecast_yrs, sym=sym, fx=fx,
                    )
                    _port_capital = st.session_state.get("portfolio_capital", 10_000_000)
                    _wb = build_portfolio_sheets(
                    wb=_wb, ticker=ticker_input, enriched=enriched,
                    dcf_res=dcf_res, forecast_result=forecast_result,
                    scenarios=scenarios, wacc_data=wacc_data,
                    wacc=wacc, terminal_g=terminal_g,
                    forecast_yrs=forecast_yrs, sym=sym, fx=fx,
                    portfolio_size=_port_capital,
                    )
                    _buf = _io.BytesIO()
                    _wb.save(_buf)
                    if st.download_button(
                        "🏦 Download Complete HF Model (14 Sheets)",
                        data=_buf.getvalue(),
                        file_name=f"{ticker_input}_HedgeFund_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width='stretch',
                    ):
                        record_report()
                        track_event(st.session_state.get("auth_email",""), tier(), "report_download", {"ticker": ticker_input})
                except Exception as _e:
                    st.error(f"HF model error: {str(_e)}")
            st.html(f"""
            <div style="padding:14px 16px;background:#F8FAFC;border:1px solid #E2E8F0;
                        border-radius:8px;text-align:center;">
              <div style="font-size:12px;color:#475569;margin-bottom:4px;text-transform:uppercase;letter-spacing:0.04em;">Analysis as of</div>
              <div style="font-size:13px;font-weight:600;color:#475569;font-family:'IBM Plex Mono',monospace;">
                {datetime.now().strftime('%d %b %Y %H:%M')}
              </div>
              <div style="font-size:12px;color:#64748B;margin-top:6px;">14 sheets: DCF · WACC · FCFF<br>Scenarios · Monte Carlo · Reverse DCF<br>Risk-Reward · Portfolio System</div>
            </div>
            """)

        with dl3:
            # ── PDF Report download ─────────────────────────────
            _can_pdf, _pdf_reason = can_download_pdf()
            if not _can_pdf and limit("pdf_reports_per_month") == 0:
                st.html(f"""
                <div style="text-align:center;padding:8px;">
                  <a href="https://buy.stripe.com/your_pdf_link_here" target="_blank"
                     style="background:#0F2942;color:#fff;font-size:13px;font-weight:600;
                            padding:9px 20px;border-radius:8px;text-decoration:none;
                            display:inline-block;width:100%;text-align:center;">
                    📑 Buy PDF Report — $4.99
                  </a>
                </div>
                """)
            elif not _can_pdf:
                st.warning(f"📑 {_pdf_reason}")
                st.caption(f"Extra PDFs: ${limit('pdf_report_cost'):.2f} each")
            else:
                st.caption(f"📑 {_pdf_reason}")
            if _can_pdf:
                try:
                    import sys as _sys2
                    from pathlib import Path as _Path2
                    _dashboard_root = str(_Path2(__file__).parent)
                    if _dashboard_root not in _sys2.path:
                        _sys2.path.insert(0, _dashboard_root)
                    from pdf_report import generate_pdf_report as _gen_pdf
                    _pdf_bytes = _gen_pdf(
                        ticker=ticker_input,
                        enriched=enriched,
                        raw=raw,
                        dcf_res=dcf_res,
                        forecast_result=forecast_result,
                        scenarios=scenarios,
                        inv_plan=inv_plan,
                        wacc_data=wacc_data,
                        wacc=wacc,
                        terminal_g=terminal_g,
                        forecast_yrs=forecast_yrs,
                        sym=sym,
                        to_code=to_code,
                        fx=fx,
                        price_d=price_d,
                        iv_d=iv_d,
                        mos_pct=mos_pct,
                        sig=sig,
                    )
                    if st.download_button(
                        "📑 Download PDF Report (4 pages)",
                        data=_pdf_bytes,
                        file_name=f"{ticker_input}_YieldIQ_{datetime.now().strftime('%Y%m%d')}.pdf",
                        mime="application/pdf",
                        width='stretch',
                    ):
                        record_pdf_report()
                        track_event(st.session_state.get("auth_email",""), tier(), "pdf_download", {"ticker": ticker_input})
                except Exception as _pdf_e:
                    st.error(f"PDF error: {str(_pdf_e)}")
            st.html(f"""
            <div style="padding:14px 16px;background:#F8FAFC;border:1px solid #E2E8F0;
                        border-radius:8px;text-align:center;">
              <div style="font-size:12px;color:#475569;margin-bottom:4px;text-transform:uppercase;letter-spacing:0.04em;">Report contents</div>
              <div style="font-size:12px;color:#64748B;margin-top:4px;">
                4 pages: Overview · DCF Model<br>Quality Score · Recommendation<br>
                <span style="color:#0F2942;font-weight:600;">Starter 5/mo · Pro Unlimited</span>
              </div>
            </div>
            """)

        # ── Full DCF Audit ─────────────────────────────────────
        with st.expander("🔬 Full DCF Model Details — Assumptions & Technical Breakdown"):
            detail = {
                "── MODEL INPUTS ─────────────": "",
                "Required return rate (WACC)":  f"{wacc:.2%}",
                "Long-run growth rate":         f"{terminal_g:.2%}",
                "Starting cash flow growth":    f"{base_growth:.2%}",
                "Cash flow base method":        forecast_result.get("fcf_base_method",""),
                "── HOW WE BUILT THE VALUE ───": "",
                "Present value of cash flows":  fmt(dcf_res.get("sum_pv_fcfs",0)*fx, sym),
                "Terminal value (raw)":         fmt(dcf_res.get("terminal_value",0)*fx, sym),
                "Present value of terminal":    fmt(pv_tv_d, sym),
                "Terminal value % of total":    f"{dcf_res.get('tv_pct_of_ev',0):.0%}",
                "Total enterprise value":       fmt(dcf_res.get("enterprise_value",0)*fx, sym),
                "Less: total debt":             fmt(enriched["total_debt"]*fx, sym),
                "Plus: cash held":              fmt(enriched["total_cash"]*fx, sym),
                "Equity value":                 fmt(dcf_res.get("equity_value",0)*fx, sym),
                "Shares outstanding":           f"{enriched['shares']/1e9:.3f}B",
                "── RESULT ───────────────────": "",
                "Estimated fair value/share":   fmts(iv_d, sym),
                "Current price":                fmts(price_d, sym),
                "Margin of Safety":             f"{mos_pct:+.1f}%",
                "Model verdict": (
                    f"Undervalued by {mos_pct:.1f}%"   if mos_pct >= 5  else
                    f"Overvalued by {abs(mos_pct):.1f}%" if mos_pct <= -5 else
                    "Fairly valued (within ±5%)"
                ),
                "Signal (raw)":                 sig,
                "FX rate used":                 f"1 {native_ccy} = {fx:.6f} {to_code}  (live rate)",
                "Price in native currency":     f"{native_ccy} {price_n:,.2f}",
                "Fair value in native ccy":     f"{native_ccy} {iv_n:,.2f}",
            }
            detail_clean = {k: str(v) if not isinstance(v, str) else v
                           for k, v in detail.items()}
            st.dataframe(pd.DataFrame.from_dict(detail_clean, orient="index", columns=["Value"]),
                         width='stretch')


def mini_sparkline(values: list, width: int = 60, height: int = 22) -> str:
    """Returns an inline SVG sparkline for a list of values. Green if rising, red if falling."""
    if not values or len(values) < 2:
        return f'<svg width="{width}" height="{height}"></svg>'
    vals = [v for v in values if v is not None and not (isinstance(v, float) and v != v)]
    if len(vals) < 2:
        return f'<svg width="{width}" height="{height}"></svg>'
    mn, mx = min(vals), max(vals)
    rng = mx - mn if mx != mn else 1
    pts = []
    for i, v in enumerate(vals):
        x = i * (width - 6) / (len(vals) - 1) + 3
        y = height - 3 - ((v - mn) / rng) * (height - 6)
        pts.append(f"{x:.1f},{y:.1f}")
    polyline = " ".join(pts)
    trend_color = "#059669" if vals[-1] >= vals[0] else "#DC2626"
    last_x, last_y = pts[-1].split(",")
    return (
        f'<svg width="{width}" height="{height}" style="vertical-align:middle;display:block;">'
        f'<polyline points="{polyline}" fill="none" stroke="{trend_color}" stroke-width="1.5"'
        f' stroke-linecap="round" stroke-linejoin="round"/>'
        f'<circle cx="{last_x}" cy="{last_y}" r="2.2" fill="{trend_color}"/>'
        f'</svg>'
    )


def render_fin_table(df, title, rows_config, accent="#3b82f6"):
    """Render a financial statement as a Bloomberg/TIKR-quality HTML table.
    Upgrades: sparkline Trend column, YoY% sub-text per cell.
    """
    if df is None or df.empty:
        st.html(f"""
        <div style="padding:20px;background:#F8FAFC;border:1px solid #E2E8F0;border-radius:10px;
                    text-align:center;color:#475569;font-size:13px;">
          No data available for {title}
        </div>
        """)
        return

    years = []
    if "year" in df.columns:
        years = [str(int(y)) for y in df["year"].tolist()]
    elif df.index.name == "year" or df.index.dtype in [int, float]:
        years = [str(int(y)) for y in df.index.tolist()]
    else:
        years = [f"Period {i+1}" for i in range(len(df))]

    yr_headers = "".join([
        f'<th style="background:#EFF6FF;color:#475569;font-size:12px;font-weight:700;'
        f'padding:10px 14px;text-align:right;border:1px solid #F0F4F8;white-space:nowrap;">{yr}</th>'
        for yr in years
    ])
    # Trend column header
    yr_headers += (
        '<th style="background:#EFF6FF;color:#94a3b8;font-size:11px;font-weight:600;'
        'padding:10px 12px;text-align:center;border:1px solid #F0F4F8;white-space:nowrap;">Trend</th>'
    )

    rows_html = ""
    for label, col, is_pct, is_ratio, bold, is_section in rows_config:
        if is_section:
            span = len(years) + 2  # +1 for label col, +1 for Trend col
            rows_html += (
                f'<tr><td colspan="{span}" style="background:{accent}18;color:{accent};'
                f'font-size:11px;font-weight:700;padding:7px 16px;text-transform:uppercase;'
                f'letter-spacing:0.06em;border:1px solid #F0F4F8;">{label}</td></tr>'
            )
            continue

        row_bg   = "#FFFFFF" if bold else "#F8FAFC"
        lbl_style = (
            f'background:{row_bg};color:{"#0F172A" if bold else "#475569"};'
            f'font-size:{"13px" if bold else "12px"};font-weight:{"700" if bold else "400"};'
            f'padding:9px 16px;border:1px solid #F0F4F8;min-width:210px;white-space:nowrap;'
        )
        lbl_cell = f'<td style="{lbl_style}">{label}</td>'

        val_cells  = ""
        raw_values = []  # collect for sparkline

        if col and col in df.columns:
            col_vals = df[col].tolist()
            for i, val in enumerate(col_vals):
                _chg_pct = None  # YoY % change for cell background tinting
                if pd.isna(val) or val is None:
                    display    = "—"
                    color      = "#94a3b8"
                    yoy_html   = ""
                    raw_values.append(None)
                elif is_pct:
                    display    = f"{val * 100:.1f}%"
                    color      = "#059669" if val > 0 else ("#dc2626" if val < 0 else "#64748b")
                    raw_values.append(val)
                    # YoY for pct: absolute pp change
                    if i > 0 and col_vals[i-1] is not None and not pd.isna(col_vals[i-1]):
                        pp = (val - col_vals[i-1]) * 100
                        _chg_pct = pp
                        arrow = "▲" if pp >= 0 else "▼"
                        yoy_c = "#059669" if pp >= 0 else "#dc2626"
                        yoy_html = (
                            f'<div style="font-size:10px;color:{yoy_c};margin-top:2px;'
                            f'font-family:system-ui;">{arrow} {pp:+.1f}pp</div>'
                        )
                    else:
                        yoy_html = ""
                elif is_ratio:
                    display    = f"{val:.2f}x"
                    color      = "#059669" if val > 1 else "#f59e0b"
                    raw_values.append(val)
                    if i > 0 and col_vals[i-1] is not None and not pd.isna(col_vals[i-1]) and col_vals[i-1] != 0:
                        chg = (val - col_vals[i-1]) / abs(col_vals[i-1]) * 100
                        _chg_pct = chg
                        arrow = "▲" if chg >= 0 else "▼"
                        yoy_c = "#059669" if chg >= 0 else "#dc2626"
                        yoy_html = (
                            f'<div style="font-size:10px;color:{yoy_c};margin-top:2px;'
                            f'font-family:system-ui;">{arrow} {chg:+.1f}%</div>'
                        )
                    else:
                        yoy_html = ""
                else:
                    converted  = val * fin_fx / 1e9
                    display    = f"{fin_sym}{converted:,.2f}B"
                    color      = "#059669" if converted > 0 else ("#dc2626" if converted < 0 else "#64748b")
                    raw_values.append(converted)
                    if i > 0 and col_vals[i-1] is not None and not pd.isna(col_vals[i-1]) and col_vals[i-1] != 0:
                        prev_c = col_vals[i-1] * fin_fx / 1e9
                        if prev_c != 0:
                            chg = (converted - prev_c) / abs(prev_c) * 100
                            _chg_pct = chg
                            arrow = "▲" if chg >= 0 else "▼"
                            yoy_c = "#059669" if chg >= 0 else "#dc2626"
                            yoy_html = (
                                f'<div style="font-size:10px;color:{yoy_c};margin-top:2px;'
                                f'font-family:system-ui;">{arrow} {chg:+.1f}%</div>'
                            )
                        else:
                            yoy_html = ""
                    else:
                        yoy_html = ""

                # ── Cell background: tint by YoY growth direction/magnitude
                if _chg_pct is not None:
                    cell_bg = "#F0FDF4" if _chg_pct > 10 else "#FFFBEB" if _chg_pct >= 0 else "#FEF2F2"
                else:
                    cell_bg = row_bg

                val_cells += (
                    f'<td style="background:{cell_bg};color:{color};'
                    f'font-size:{"13px" if bold else "12px"};font-weight:{"700" if bold else "500"};'
                    f'padding:8px 14px;text-align:right;border:1px solid #F0F4F8;'
                    f'font-family:"IBM Plex Mono","Courier New",monospace;vertical-align:top;">'
                    f'{display}{yoy_html}</td>'
                )
        else:
            for _ in years:
                val_cells += (
                    f'<td style="background:{row_bg};color:#94a3b8;padding:9px 14px;'
                    f'text-align:right;border:1px solid #F0F4F8;">—</td>'
                )

        # ── Sparkline Trend cell ──────────────────────────────────
        spark_svg  = mini_sparkline(raw_values)
        spark_cell = (
            f'<td style="background:{row_bg};padding:8px 12px;text-align:center;'
            f'border:1px solid #F0F4F8;vertical-align:middle;">{spark_svg}</td>'
        )

        rows_html += f"<tr>{lbl_cell}{val_cells}{spark_cell}</tr>"

    ccard(title, accent)
    st.html(f"""
    <div style="overflow-x:auto;border-radius:8px;border:1px solid #E2E8F0;
                box-shadow:0 1px 4px rgba(0,0,0,0.06);">
      <table style="width:100%;border-collapse:collapse;">
        <thead>
          <tr>
            <th style="background:#EFF6FF;color:{accent};font-size:12px;font-weight:700;
                       padding:10px 16px;text-align:left;border:1px solid #F0F4F8;
                       min-width:210px;position:sticky;left:0;z-index:1;">Line Item</th>
            {yr_headers}
          </tr>
        </thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>
    """)
    ccard_end()


# ══════════════════════════════════════════════════════════════
# TAB 2 — FINANCIAL STATEMENTS
# ══════════════════════════════════════════════════════════════
if _active_main_tab == "financials":
    st.html("""
    <div style="margin-top:8px;margin-bottom:16px;padding:16px 20px;
                background:linear-gradient(135deg,#F8FAFC,#F4F6F9);
                border:1px solid #E2E8F0;border-radius:12px;">
      <div style="font-size:16px;font-weight:700;color:#0F172A;margin-bottom:4px;">📑 Financial Statements</div>
      <div style="font-size:13px;color:#475569;line-height:1.7;">
        Analyse a stock in the <b style="color:#3b82f6;">Single Stock Analysis</b> tab first, then return here to view
        the full historical P&amp;L, Cash Flow, and Balance Sheet data pulled from Yahoo Finance.
      </div>
    </div>
    """)

    # Check if analysis has been run (look for session state data)
    if "fin_enriched" not in st.session_state:
        st.html("""
        <div style="margin-top:32px;padding:48px;background:linear-gradient(135deg,#F8FAFC,#F4F6F9);
                    border:1px solid #E2E8F0;border-radius:16px;text-align:center;">
          <div style="font-size:40px;margin-bottom:14px;">📑</div>
          <div style="font-size:20px;font-weight:700;color:#0F172A;margin-bottom:8px;">No data loaded yet</div>
          <div style="font-size:13px;color:#475569;max-width:460px;margin:0 auto;line-height:1.8;">
            Go to <b style="color:#3b82f6;">Single Stock Analysis</b>, enter a ticker and click Analyse.<br>
            Financial Statements will appear here automatically.
          </div>
        </div>
        """)
    else:
        fin_enriched  = st.session_state.get("fin_enriched", {})
        fin_ticker    = st.session_state.get("fin_ticker", "")
        fin_fx        = st.session_state.get("fin_fx", 1.0)
        fin_to_code   = st.session_state.get("fin_to_code", "INR")
        fin_sym       = st.session_state.get("fin_sym", "₹")

        def _fmt_fin_val(v, is_pct=False, is_ratio=False):
            """Format a financial value for display."""
            if pd.isna(v) or v is None:
                return "—"
            if is_pct:
                return f"{v*100:.1f}%"
            if is_ratio:
                return f"{v:.2f}x"
            converted = v * fin_fx / 1e9
            color = "#10b981" if converted > 0 else ("#ef4444" if converted < 0 else "#64748b")
            return converted, color


        st.html(f"""
        <div style="padding:12px 18px;background:#F8FAFC;border:1px solid #E2E8F0;
                    border-radius:10px;margin-bottom:16px;display:flex;
                    align-items:center;justify-content:space-between;">
          <div>
            <span style="font-size:20px;font-weight:800;color:#0F172A;
                         font-family:'IBM Plex Mono',monospace;">{fin_ticker}</span>
            <span style="font-size:13px;color:#475569;margin-left:12px;">
              Values in {fin_to_code} Billions · Source: Yahoo Finance
            </span>
          </div>
          <div style="font-size:13px;color:#475569;">
            All figures converted at live FX rate
          </div>
        </div>
        """)

        income_df = fin_enriched.get("income_df", pd.DataFrame())
        cf_df     = fin_enriched.get("cf_df",     pd.DataFrame())
        bs_df     = fin_enriched.get("bs_df",     pd.DataFrame())

        # ── P&L ───────────────────────────────────────────────
        inc_config = [
            ("REVENUE",                   None,              False, False, True,  True),
            (f"Revenue ({fin_to_code}B)",         "revenue",         False, False, True,  False),
            (f"Gross Profit ({fin_to_code}B)",     "gross_profit",    False, False, False, False),
            ("PROFITABILITY",              None,              False, False, True,  True),
            (f"Operating Income ({fin_to_code}B)", "operating_income",False, False, True,  False),
            (f"Net Income ({fin_to_code}B)",        "net_income",      False, False, True,  False),
            ("MARGINS",                    None,              False, False, True,  True),
            ("Gross Margin",               "gross_margin",    True,  False, False, False),
            ("Operating Margin",           "op_margin",       True,  False, True,  False),
            ("Net Margin",                 "net_margin",      True,  False, False, False),
        ]
        render_fin_table(income_df, f"Income Statement (P&L) — {fin_ticker}", inc_config, "#3b82f6")

        # ── KEY METRICS CHART PANEL (Revenue / Op Margin / FCF) ───────
        if not income_df.empty and not cf_df.empty:
            try:
                _chart_years = [str(int(y)) for y in income_df["year"].tolist()] if "year" in income_df.columns else []
                _rev_vals  = [(v * fin_fx / 1e9) if v is not None and not pd.isna(v) else None
                              for v in income_df.get("revenue", pd.Series()).tolist()] if "revenue" in income_df.columns else []
                _opm_vals  = [(v * 100) if v is not None and not pd.isna(v) else None
                              for v in income_df.get("op_margin", pd.Series()).tolist()] if "op_margin" in income_df.columns else []
                _fcf_vals  = [(v * fin_fx / 1e9) if v is not None and not pd.isna(v) else None
                              for v in cf_df.get("fcf", pd.Series()).tolist()] if "fcf" in cf_df.columns else []

                _ch1, _ch2, _ch3 = st.columns(3)

                # Panel 1 — Revenue (bar)
                if _rev_vals and _chart_years:
                    with _ch1:
                        _fig_rev = go.Figure(go.Bar(
                            x=_chart_years, y=_rev_vals,
                            marker_color="#3b82f6",
                            marker_line_width=0,
                            hovertemplate="%{x}: " + fin_sym + "%{y:,.2f}B<extra></extra>",
                        ))
                        _fig_rev.update_layout(**CL(
                            height=190,
                            title=dict(text="Revenue", font=dict(size=12, color="#0F172A"), x=0.02),
                            margin=dict(t=36, b=32, l=40, r=8),
                            yaxis=dict(tickprefix=fin_sym, ticksuffix="B", gridcolor="#F1F5F9",
                                       linecolor="#E2E8F0", zeroline=False, tickfont=dict(color="#64748B", size=9)),
                            xaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0", zeroline=False,
                                       tickfont=dict(color="#64748B", size=9)),
                        ))
                        st.plotly_chart(_fig_rev, width="stretch",
                                        config={"displayModeBar": False})

                # Panel 2 — Operating Margin (line)
                if _opm_vals and _chart_years:
                    with _ch2:
                        _fig_opm = go.Figure(go.Scatter(
                            x=_chart_years, y=_opm_vals,
                            mode="lines+markers",
                            line=dict(color="#0d9488", width=2),
                            marker=dict(color="#0d9488", size=5),
                            fill="tozeroy",
                            fillcolor="rgba(13,148,136,0.08)",
                            hovertemplate="%{x}: %{y:.1f}%<extra></extra>",
                        ))
                        _fig_opm.update_layout(**CL(
                            height=190,
                            title=dict(text="Operating Margin", font=dict(size=12, color="#0F172A"), x=0.02),
                            margin=dict(t=36, b=32, l=40, r=8),
                            yaxis=dict(ticksuffix="%", gridcolor="#F1F5F9",
                                       linecolor="#E2E8F0", zeroline=False, tickfont=dict(color="#64748B", size=9)),
                            xaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0", zeroline=False,
                                       tickfont=dict(color="#64748B", size=9)),
                        ))
                        st.plotly_chart(_fig_opm, width="stretch",
                                        config={"displayModeBar": False})

                # Panel 3 — FCF (bar, green/red)
                if _fcf_vals and _chart_years:
                    with _ch3:
                        _fcf_colors = ["#059669" if (v or 0) >= 0 else "#dc2626" for v in _fcf_vals]
                        _fig_fcf2 = go.Figure(go.Bar(
                            x=_chart_years, y=_fcf_vals,
                            marker_color=_fcf_colors,
                            marker_line_width=0,
                            hovertemplate="%{x}: " + fin_sym + "%{y:,.2f}B<extra></extra>",
                        ))
                        _fig_fcf2.update_layout(**CL(
                            height=190,
                            title=dict(text="Free Cash Flow", font=dict(size=12, color="#0F172A"), x=0.02),
                            margin=dict(t=36, b=32, l=40, r=8),
                            yaxis=dict(tickprefix=fin_sym, ticksuffix="B", gridcolor="#F1F5F9",
                                       linecolor="#E2E8F0", zeroline=True, zeroline_color="#CBD5E1",
                                       tickfont=dict(color="#64748B", size=9)),
                            xaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0", zeroline=False,
                                       tickfont=dict(color="#64748B", size=9)),
                        ))
                        st.plotly_chart(_fig_fcf2, width="stretch",
                                        config={"displayModeBar": False})
            except Exception:
                pass  # Charts are non-critical; silently skip on data issues

        # ── Cash Flow ─────────────────────────────────────────
        cf_config = [
            ("OPERATING ACTIVITIES",       None,   False, False, True,  True),
            (f"Operating Cash Flow ({fin_to_code}B)", "cfo",  False, False, True,  False),
            (f"Capital Expenditure ({fin_to_code}B)", "capex",False, False, False, False),
            ("FREE CASH FLOW",             None,   False, False, True,  True),
            (f"Free Cash Flow ({fin_to_code}B)",      "fcf",  False, False, True,  False),
            ("GROWTH",                     None,   False, False, True,  True),
            ("FCF YoY Growth",             "fcf_growth", True, False, False, False),
        ]
        render_fin_table(cf_df, f"Cash Flow Statement — {fin_ticker}", cf_config, "#10b981")

        # ── Balance Sheet ─────────────────────────────────────
        bs_config_fallback = None
        if bs_df is not None and not bs_df.empty:
            bs_config = [
                ("ASSETS",                         None,              False, False, True,  True),
                (f"Total Assets ({fin_to_code}B)",         "total_assets",    False, False, True,  False),
                (f"Cash & Equivalents ({fin_to_code}B)",   "cash",            False, False, False, False),
                (f"Current Assets ({fin_to_code}B)",       "current_assets",  False, False, False, False),
                ("LIABILITIES",                    None,              False, False, True,  True),
                (f"Total Debt ({fin_to_code}B)",            "total_debt",      False, False, True,  False),
                (f"Current Liabilities ({fin_to_code}B)",  "current_liab",    False, False, False, False),
                ("EQUITY & SOLVENCY",              None,              False, False, True,  True),
                (f"Shareholders' Equity ({fin_to_code}B)", "equity",          False, False, True,  False),
                ("Debt / Equity",                  "de_ratio",        False, True,  False, False),
                ("Current Ratio",                  "current_ratio",   False, True,  False, False),
            ]
            render_fin_table(bs_df, f"Balance Sheet — {fin_ticker}", bs_config, "#06b6d4")
        else:
            # Snapshot from enriched data
            ccard(f"Balance Sheet Snapshot — {fin_ticker}", "#06b6d4")
            snap_rows = [
                ("Cash & Equivalents",    fin_enriched.get("total_cash", 0)),
                ("Total Debt",            fin_enriched.get("total_debt", 0)),
            ]
            snap_html = ""
            for label, raw_val in snap_rows:
                v = raw_val * fin_fx / 1e9
                color = "#10b981" if v >= 0 else "#ef4444"
                snap_html += f"""
                <tr>
                  <td style="background:#F4F6F9;color:#475569;font-size:13px;padding:10px 16px;border:1px solid #F8FAFC;">{label}</td>
                  <td style="background:#F4F6F9;color:{color};font-size:13px;font-weight:700;padding:10px 16px;text-align:right;border:1px solid #F8FAFC;font-family:'Courier New',monospace;">{fin_sym}{v:,.2f}B</td>
                </tr>"""
            st.html(f"""
            <div style="overflow-x:auto;border-radius:8px;border:1px solid #E2E8F0;">
              <table style="width:100%;border-collapse:collapse;">
                <thead><tr>
                  <th style="background:#EFF6FF;color:#06b6d4;font-size:13px;font-weight:700;padding:10px 16px;text-align:left;border:1px solid #F8FAFC;">Line Item</th>
                  <th style="background:#EFF6FF;color:#475569;font-size:13px;font-weight:700;padding:10px 16px;text-align:right;border:1px solid #F8FAFC;">Latest Available</th>
                </tr></thead>
                <tbody>{snap_html}</tbody>
              </table>
            </div>
            """)
            st.caption("Full multi-year balance sheet not available for this ticker via Yahoo Finance.")
            ccard_end()

        # ── EXCEL EXPORT ──────────────────────────────────────────
        st.html("<div style='height:8px'></div>")
        _can_dl_fin, _dl_fin_reason = can_download_report()
        if _can_dl_fin:
            try:
                import sys as _sys2
                from pathlib import Path as _Path2
                _proj2 = str(_Path2(__file__).parent.parent)
                if _proj2 not in _sys2.path:
                    _sys2.path.insert(0, _proj2)
                from generate_dcf_excel import generate_institutional_dcf
                from generate_hf_excel import build_hedge_fund_sheets
                from generate_portfolio_excel import build_portfolio_sheets
                import io as _io2
                from openpyxl import load_workbook as _lwb2
                _dcf_res_fin   = st.session_state.get("dcf_res", {})
                _fcst_fin      = st.session_state.get("forecast_result", {})
                _scen_fin      = st.session_state.get("scenarios", {})
                _wacc_data_fin = st.session_state.get("wacc_data", {})
                _wacc_fin      = st.session_state.get("wacc", 0.10)
                _tg_fin        = st.session_state.get("terminal_g", 0.03)
                _fy_fin        = st.session_state.get("forecast_yrs", 5)
                if _dcf_res_fin:
                    _hf2 = generate_institutional_dcf(
                        ticker=fin_ticker, enriched=fin_enriched, dcf_res=_dcf_res_fin,
                        forecast_result=_fcst_fin, scenarios=_scen_fin,
                        wacc_data=_wacc_data_fin, wacc=_wacc_fin, terminal_g=_tg_fin,
                        forecast_yrs=_fy_fin, sym=fin_sym, to_code=fin_to_code, fx=fin_fx,
                    )
                    _wb2 = _lwb2(filename=_io2.BytesIO(_hf2))
                    _wb2 = build_hedge_fund_sheets(
                        wb=_wb2, ticker=fin_ticker, enriched=fin_enriched, dcf_res=_dcf_res_fin,
                        forecast_result=_fcst_fin, scenarios=_scen_fin, wacc_data=_wacc_data_fin,
                        wacc=_wacc_fin, terminal_g=_tg_fin, forecast_yrs=_fy_fin,
                        sym=fin_sym, fx=fin_fx,
                    )
                    _wb2 = build_portfolio_sheets(
                        wb=_wb2, ticker=fin_ticker, enriched=fin_enriched, dcf_res=_dcf_res_fin,
                        forecast_result=_fcst_fin, scenarios=_scen_fin, wacc_data=_wacc_data_fin,
                        wacc=_wacc_fin, terminal_g=_tg_fin, forecast_yrs=_fy_fin,
                        sym=fin_sym, fx=fin_fx,
                        portfolio_size=st.session_state.get("portfolio_capital", 10_000_000),
                    )
                    _buf2 = _io2.BytesIO()
                    _wb2.save(_buf2)
                    st.download_button(
                        "📥 Download Full Financial Model (Excel)",
                        data=_buf2.getvalue(),
                        file_name=f"{fin_ticker}_FinancialModel_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width='content',
                    )
                else:
                    st.info("Run a full analysis in the Single Stock tab first to enable Excel export.", icon="ℹ️")
            except Exception as _ex_fin:
                st.info("Run a full analysis in the Single Stock tab first to enable Excel export.", icon="ℹ️")
        else:
            st.info(f"📥 Excel export requires a Pro account. {_dl_fin_reason}", icon="🔒")

        # ── Key Ratios Summary ─────────────────────────────────
        ccard("Key financial ratios at a glance", "#8b5cf6")
        r1,r2,r3,r4,r5,r6 = st.columns(6)
        r1.metric("Revenue growth",   f"{fin_enriched.get('revenue_growth', 0)*100:.1f}%")
        r2.metric("Cash flow growth",  f"{fin_enriched.get('fcf_growth', 0)*100:.1f}%")
        r3.metric("Profit margin",     f"{fin_enriched.get('op_margin', 0)*100:.1f}%")
        r4.metric("Free cash generated", fmt(fin_enriched.get("latest_fcf", 0) * fin_fx, fin_sym))
        r5.metric("Cash on hand",      fmt(fin_enriched.get("total_cash", 0) * fin_fx, fin_sym))
        r6.metric("Total debt",        fmt(fin_enriched.get("total_debt", 0) * fin_fx, fin_sym))
        ccard_end()


# ══════════════════════════════════════════════════════════════
# TAB 3 — SCREENER RESULTS
# ══════════════════════════════════════════════════════════════
if _active_main_tab == "screener":
    _screener_ok, _screener_reason = can_run_screener()
    if not _screener_ok:
        st.html("<br>")
        show_upgrade_modal("Batch stock screener")
        st.html("""
        <div style="margin-top:16px;padding:16px 20px;background:#f8fafc;
                    border:1px solid #e2e8f0;border-radius:10px;font-size:13px;color:#4a5568;">
          <strong>What the screener does:</strong> Runs our DCF model on 542+ US stocks
          and 2,270+ Indian stocks — ranked by margin of safety. Updated nightly.
        </div>""")
    else:
        df_screen = None
        if results_file is not None:
            df_screen = pd.read_csv(results_file)
        else:
            try:
                df_screen = pd.read_csv(RESULTS_PATH)
            except FileNotFoundError:
                st.info("No screener results found yet — run the screener above to generate them.")
        record_screener()
        track_event(st.session_state.get("auth_email",""), tier(), "screener_run")

    if _screener_ok and (df_screen is None or df_screen.empty):
        _, _ec, _ = st.columns([1, 3, 1])
        with _ec:
            st.html("""
            <div style="margin-top:60px;padding:56px 48px;
                        background:linear-gradient(135deg,#0d1117,#161b22);
                        border:1px solid #21262d;border-radius:16px;text-align:center;">
              <div style="font-size:48px;margin-bottom:16px;">&#128269;</div>
              <div style="font-size:22px;font-weight:700;color:#e6edf3;margin-bottom:12px;">
                Run your first screen</div>
              <div style="font-size:14px;color:#8b949e;max-width:480px;
                          margin:0 auto 28px;line-height:1.7;">
                Find undervalued stocks across US and Indian markets using
                institutional DCF methodology &#8212; 2,800+ stocks ranked by model estimate.
              </div>
              <div style="background:#161b22;border:1px solid #21262d;border-radius:8px;
                          padding:10px 20px;display:inline-block;">
                <code style="color:#00b4d8;font-size:12px;font-family:'IBM Plex Mono',monospace;">
                  python batch/nightly_precompute.py
                </code>
              </div>
            </div>""")

    elif _screener_ok and df_screen is not None and not df_screen.empty:
        # ── Batch run metadata banner ─────────────────────────
        try:
            import json as _json
            _status_path = Path("data/last_batch_run.json")
            if _status_path.exists():
                _bst = _json.loads(_status_path.read_text())
                _bts = _bst.get("timestamp", "")[:16].replace("T", " ")
                _bcomp = _bst.get("completed", "—")
                _bdur  = _bst.get("duration_min", "—")
                _bpick = _bst.get("top_pick", "—")
                st.html(f"""
                <div style="display:flex;align-items:center;gap:24px;padding:10px 16px;
                            background:#F0F9FF;border:1px solid #BAE6FD;border-radius:8px;
                            margin-bottom:12px;font-size:12px;color:#0369A1;">
                  <span>🕒 Last updated: <strong>{_bts}</strong></span>
                  <span>📊 Stocks analysed: <strong>{_bcomp:,}</strong></span>
                  <span>⏱ Runtime: <strong>{_bdur} min</strong></span>
                  <span>🏆 Top pick: <strong>{_bpick}</strong></span>
                </div>""")
        except Exception:
            pass
        _sc_clean = df_screen[~df_screen["signal"].astype(str).str.contains("Data Limited|N/A", na=False)]
        _sc_total = len(df_screen)
        _sc_buys  = len(_sc_clean[_sc_clean["signal"].astype(str).str.contains("Undervalued",    na=False)])
        _sc_watch = len(_sc_clean[_sc_clean["signal"].astype(str).str.contains("Near Fair Value", na=False)])
        _sc_sells = len(_sc_clean[_sc_clean["signal"].astype(str).str.contains("Overvalued",      na=False)])
        _sc_na    = len(df_screen[df_screen["signal"].astype(str).str.contains("N/A|Data Limited", na=False)])
        _sc_best  = (_sc_clean.loc[_sc_clean["margin_of_safety"].idxmax(), "ticker"]
                     if not _sc_clean.empty else "—")

        _k1, _k2, _k3, _k4, _k5, _k6 = st.columns(6)
        _k1.metric("Total",       _sc_total)
        _k2.metric("Undervalued", _sc_buys)
        _k3.metric("Discount",    _sc_watch)
        _k4.metric("Overvalued",  _sc_sells)
        _k5.metric("N/A",         _sc_na)
        _k6.metric("Top Pick",    _sc_best)

        st.caption(
            "⚠️ Screener results are model outputs only — not investment advice. "
            "Signals reflect DCF model estimates, not buy/sell recommendations. "
            "YieldIQ is not a registered investment adviser."
        )

        # ── Preset filter templates ───────────────────────────
        st.html("""<div style="font-size:11px;font-weight:600;color:#8b949e;
                    text-transform:uppercase;letter-spacing:.1em;margin-bottom:6px;">
                    Quick Presets</div>""")
        _pr1, _pr2, _pr3, _pr4 = st.columns(4)
        _PRESETS = {
            "🏆 Buffett Picks":    dict(sc_mos=20, sc_qual=60, sc_sig=["Undervalued 🟢"],            sc_sort="fundamental_score"),
            "🚀 Growth at Value":  dict(sc_mos=10, sc_qual=50, sc_sig=["Undervalued 🟢","Near Fair Value 🟡"], sc_sort="revenue_growth"),
            "💰 Deep Value":       dict(sc_mos=30, sc_qual=0,  sc_sig=["Undervalued 🟢"],            sc_sort="margin_of_safety"),
            "💎 Dividend Quality": dict(sc_mos=0,  sc_qual=55, sc_sig=["Undervalued 🟢","Near Fair Value 🟡","Fairly Valued 🔵"], sc_sort="fundamental_score"),
        }
        for _pcol, (_plabel, _pvals) in zip([_pr1, _pr2, _pr3, _pr4], _PRESETS.items()):
            with _pcol:
                if st.button(_plabel, key=f"preset_{_plabel}", width="stretch"):
                    for _pk, _pv in _pvals.items():
                        st.session_state[_pk] = _pv
                    st.rerun()

        # ── Filter bar
        st.html("""<div style="background:#161b22;border:1px solid #21262d;border-radius:10px;
                    padding:10px 14px;margin:10px 0 12px;">
          <span style="font-size:11px;font-weight:600;color:#8b949e;
                       text-transform:uppercase;letter-spacing:.1em;">Filters</span></div>""")

        _fb1, _fb2, _fb3, _fb4 = st.columns(4)
        with _fb1:
            if LAUNCH_REGION == "US":
                _mkt_sel = "US Only"   # locked — no market picker shown
                st.caption("🇺🇸 US markets")
            else:
                _mkt_sel = st.selectbox("Market", ["All Markets", "US Only", "India Only"], key="sc_mkt")
        with _fb2:
            _sig_filter = st.multiselect("Signal",
                ["Undervalued 🟢","Near Fair Value 🟡","Fairly Valued 🔵","Overvalued 🔴"],
                default=["Undervalued 🟢","Near Fair Value 🟡"], key="sc_sig")
        with _fb3:
            _sectors = ["All Sectors"]
            if "sector" in df_screen.columns:
                _sectors += sorted(df_screen["sector"].dropna().unique().tolist())
            _sec_sel = st.selectbox("Sector", _sectors, key="sc_sec")
        with _fb4:
            _sort_col = st.selectbox("Sort by",
                ["margin_of_safety","fundamental_score","rr_ratio","price"], key="sc_sort")

        _fb5, _fb6, _ = st.columns([2,2,2])
        with _fb5:
            _min_mos  = st.slider("Min MoS (%)", -50, 100, 0, key="sc_mos")
        with _fb6:
            _min_qual = st.slider("Min Quality", 0, 100, 0, key="sc_qual")

        # Apply filters
        _filtered = df_screen[~df_screen["signal"].astype(str).str.contains("CHECK|N/A", na=False)].copy()
        _filtered = _filtered[_filtered["margin_of_safety"] >= _min_mos]
        if _min_qual > 0 and "fundamental_score" in _filtered.columns:
            _filtered = _filtered[_filtered["fundamental_score"] >= _min_qual]
        if _sig_filter:
            _filtered = _filtered[_filtered["signal"].isin(_sig_filter)]
        if _mkt_sel == "US Only":
            _filtered = _filtered[~_filtered["ticker"].astype(str).str.endswith((".NS",".BO"))]
        elif _mkt_sel == "India Only":
            _filtered = _filtered[_filtered["ticker"].astype(str).str.endswith((".NS",".BO"))]
        if _sec_sel != "All Sectors" and "sector" in _filtered.columns:
            _filtered = _filtered[_filtered["sector"] == _sec_sel]
        if _sort_col in _filtered.columns:
            _filtered = _filtered.sort_values(_sort_col, ascending=False).reset_index(drop=True)

        _result_count = len(_filtered)
        _result_color = "#e6edf3" if _result_count > 0 else "#ef4444"
        st.html('<div style="font-size:12px;color:#8b949e;margin-bottom:6px;">Showing <strong style="color:'
                + _result_color + ';">' + str(_result_count) + '</strong> of ' + str(_sc_total) + ' stocks</div>')

        if _result_count == 0:
            st.html("""
            <div style="padding:28px;background:#161b22;border:1px solid #21262d;border-radius:10px;
                        text-align:center;margin:10px 0;">
              <div style="font-size:28px;margin-bottom:8px;">🔍</div>
              <div style="font-size:14px;font-weight:600;color:#e6edf3;margin-bottom:6px;">
                No stocks match these filters</div>
              <div style="font-size:12px;color:#8b949e;">Try loosening the Min MoS or Min Quality sliders,
                or select more signal types.</div>
            </div>
            """)

        # ── Styled HTML table
        _display_cols = ["ticker","price","intrinsic_value","margin_of_safety",
                         "signal","fundamental_grade","fundamental_score",
                         "revenue_growth","op_margin","rr_ratio"]
        _show_cols = [c for c in _display_cols if c in _filtered.columns]
        _col_labels = {"ticker":"Ticker","price":"Price","intrinsic_value":"Fair Value",
                       "margin_of_safety":"Discount","signal":"Signal","fundamental_grade":"Grade",
                       "fundamental_score":"Quality","revenue_growth":"Rev Gr","op_margin":"Margin",
                       "rr_ratio":"R/R"}
        _SIG_META = {"Undervalued 🟢":("#0D7A4E","#022c1d"),"Near Fair Value 🟡":("#B45309","#2d1f05"),
                     "Fairly Valued 🔵":("#1D4ED8","#07112e"),"Overvalued 🔴":("#B91C1C","#2d0606")}

        def _badge(s):
            fg,bg = _SIG_META.get(str(s), ("#475569","#161b22"))
            lbl = str(s).split()[0] if s else "—"
            return ('<span style="background:' + bg + ';color:' + fg + ';border:1px solid ' + fg
                    + '66;font-size:10px;font-weight:700;padding:2px 8px;border-radius:12px;">'
                    + lbl + '</span>')

        def _bar(v):
            try: v = float(v)
            except (TypeError, ValueError): return "—"
            clr = "#10b981" if v>20 else "#f59e0b" if v>0 else "#ef4444"
            pct = min(max(int(abs(v)),2),100)
            sign = "+" if v>=0 else ""
            return ('<div style="display:flex;align-items:center;gap:5px;">'
                    '<div style="width:48px;height:5px;background:#21262d;border-radius:3px;flex-shrink:0;">'
                    '<div style="height:100%;width:' + str(pct) + '%;background:' + clr + ';border-radius:3px;"></div></div>'
                    '<span style="font-size:11px;color:' + clr + ';font-family:IBM Plex Mono,monospace;">'
                    + sign + '{:.1f}%'.format(v) + '</span></div>')

        def _ring(v):
            try: v = float(v)
            except (TypeError, ValueError): return "—"
            clr = "#10b981" if v>=70 else "#f59e0b" if v>=40 else "#ef4444"
            pct = int(v)
            return ('<div style="display:flex;align-items:center;gap:4px;">'
                    '<div style="width:20px;height:20px;border-radius:50%;flex-shrink:0;background:'
                    'conic-gradient(' + clr + ' ' + str(pct) + '%, #21262d ' + str(pct) + '%);">'
                    '</div><span style="font-size:11px;color:' + clr + ';font-family:IBM Plex Mono,monospace;">'
                    + str(pct) + '</span></div>')

        _th = "".join('<th style="padding:8px 12px;background:#0d1117;color:#8b949e;font-size:11px;'
                      'font-weight:600;text-transform:uppercase;letter-spacing:.08em;'
                      'border-bottom:2px solid #21262d;white-space:nowrap;'
                      'position:sticky;top:0;z-index:1;">'
                      + _col_labels.get(c, c) + '</th>' for c in _show_cols)

        _tb = ""
        for _ri, _row in _filtered.head(100).iterrows():
            _bg = "#0d1117" if _ri % 2 == 0 else "#0f1318"
            _row_cells = ""
            for _col in _show_cols:
                _v = _row.get(_col, "")
                if _col == "ticker":
                    _c = '<span style="color:#00b4d8;font-weight:700;font-family:IBM Plex Mono,monospace;font-size:12px;">' + str(_v) + '</span>'
                elif _col == "signal":    _c = _badge(_v)
                elif _col == "margin_of_safety": _c = _bar(_v)
                elif _col == "fundamental_score": _c = _ring(_v)
                elif _col in ("price","intrinsic_value"):
                    try: _c = '<span style="font-family:IBM Plex Mono,monospace;font-size:11px;color:#e6edf3;">' + sym + '{:,.2f}'.format(float(_v)) + '</span>'
                    except (TypeError, ValueError): _c = str(_v)
                elif _col in ("revenue_growth","op_margin"):
                    try:
                        _fv = float(_v)*100 if abs(float(_v))<5 else float(_v)
                        _cc = "#10b981" if _fv>0 else "#ef4444"
                        _c = '<span style="font-family:IBM Plex Mono,monospace;font-size:11px;color:' + _cc + ';">' + '{:+.1f}%'.format(_fv) + '</span>'
                    except (TypeError, ValueError): _c = str(_v)
                elif _col == "rr_ratio":
                    try: _c = '<span style="font-family:IBM Plex Mono,monospace;font-size:11px;color:#8b949e;">' + '{:.1f}x'.format(float(_v)) + '</span>'
                    except (TypeError, ValueError): _c = str(_v)
                else:
                    _c = '<span style="font-size:11px;color:#8b949e;">' + str(_v) + '</span>'
                _row_cells += '<td style="padding:7px 12px;border-bottom:1px solid #161b22;background:' + _bg + ';">' + _c + '</td>'
            _tb += '<tr>' + _row_cells + '</tr>'

        st.html('<div style="overflow:auto;max-height:460px;border:1px solid #21262d;'
                'border-radius:10px;margin-bottom:14px;">'
                '<table style="width:100%;border-collapse:collapse;min-width:800px;">'
                '<thead><tr>' + _th + '</tr></thead>'
                '<tbody>' + _tb + '</tbody>'
                '</table></div>')

        # ── Top 15 chart
        ccard("Top 15 — highest model discount right now", "#10b981")
        _top15 = _filtered.head(15)
        if not _top15.empty:
            _bcolors = ["#10b981" if v>20 else "#f59e0b" if v>5 else "#3b82f6" if v>0 else "#ef4444"
                        for v in _top15["margin_of_safety"]]
            _fig_top = go.Figure(go.Bar(
                x=_top15["ticker"], y=_top15["margin_of_safety"],
                marker=dict(color=_bcolors, opacity=0.88, line=dict(width=0)),
                text=["{:.1f}%".format(v) for v in _top15["margin_of_safety"]],
                textposition="outside",
                textfont=dict(size=10, color="#8b949e", family="IBM Plex Mono"),
                hovertemplate="<b>%{x}</b><br>Discount: %{y:.1f}%<extra></extra>",
            ))
            apply_koyfin(_fig_top, height=300, extra_kw=dict(
                showlegend=False,
                yaxis=dict(title="Discount to model value (%)", gridcolor="#21262d",
                           tickfont=dict(color="#8b949e")),
                xaxis=dict(tickfont=dict(color="#e6edf3", size=11)),
                margin=dict(t=44, b=16, l=48, r=16),
            ))
            st.plotly_chart(_fig_top, width="stretch",
                config={"displayModeBar":True,"modeBarButtonsToRemove":["lasso2d","select2d"],
                        "toImageButtonOptions":{"filename":"screener_top15","scale":2}})
        ccard_end()

        st.caption(
            "⚠️ Model output only — not investment advice. "
            "YieldIQ is not a registered investment adviser. "
            "Past model performance does not predict future results."
        )

        # ── Export + Analyse top pick
        st.html('<div style="height:8px"></div>')
        _ex1, _ex2, _ex3 = st.columns(3)
        with _ex1:
            st.download_button("⬇️ Download filtered CSV",
                data=_filtered.to_csv(index=False).encode("utf-8"),
                file_name="screener_{}.csv".format(datetime.now().strftime("%Y%m%d")),
                mime="text/csv", width='stretch', type="primary")
        with _ex2:
            _buys_only = _filtered[_filtered["signal"].astype(str).str.contains("BUY", na=False)]
            st.download_button("🎯 BUY signals only",
                data=_buys_only.to_csv(index=False).encode("utf-8"),
                file_name="BUY_{}.csv".format(datetime.now().strftime("%Y%m%d")),
                mime="text/csv", width='stretch')
        with _ex3:
            if not _filtered.empty:
                if st.button("🚀 Analyse top pick: " + str(_sc_best),
                             width='stretch', key="sc_analyse_top"):
                    st.session_state["_prefill_ticker"] = _sc_best
                    st.rerun()

# ══════════════════════════════════════════════════════════════
# TAB — PORTFOLIO
# ══════════════════════════════════════════════════════════════
if _active_main_tab == "portfolio":
    _port_analysed = None
    if st.session_state.get("fin_ticker"):
        _port_analysed = {
            "entry_price":  st.session_state.get("fin_iv_d", 0) and
                            st.session_state.get("fin_enriched", {}).get("price", 0) *
                            st.session_state.get("fin_fx", 1),
            "iv":           st.session_state.get("fin_iv_d", 0),
            "mos_pct":      st.session_state.get("fin_mos_pct", 0),
            "signal":       st.session_state.get("fin_signal", ""),
            "wacc":         st.session_state.get("fin_enriched", {}).get("wacc_used",
                            st.session_state.get("fin_enriched", {}).get("wacc", 0)),
            "to_code":      st.session_state.get("fin_to_code", "USD"),
            "company_name": st.session_state.get("fin_raw", {}).get("company_name", ""),
            "sector":       st.session_state.get("fin_enriched", {}).get("sector_name", ""),
        }
        # Get actual entry price (price * fx)
        _raw_price = st.session_state.get("fin_enriched", {}).get("price", 0)
        _fx_val    = st.session_state.get("fin_fx", 1)
        if _raw_price and _fx_val:
            _port_analysed["entry_price"] = _raw_price * _fx_val

    render_portfolio_tab(
        sym              = sym,
        analysed_ticker  = st.session_state.get("fin_ticker", ""),
        analysed_data    = _port_analysed,
    )


if _active_main_tab == "backtest":
    render_backtest_tab()


# ══════════════════════════════════════════════════════════════
# TAB — EARNINGS CALENDAR
# ══════════════════════════════════════════════════════════════
if _active_main_tab == "earnings":
    render_earnings_tab(ticker=st.session_state.get("fin_ticker", ""))


if _active_main_tab == "sectors":
    render_sector_dashboard()


if _active_main_tab == "watchlist":

    # ── Live price helper (cached 2 min) ──────────────────────
    @st.cache_data(ttl=120, show_spinner=False)
    def _wl_fetch_price(ticker: str) -> tuple:
        """Returns (last_price, day_change_pct)."""
        try:
            import yfinance as yf
            info = yf.Ticker(ticker).fast_info
            price = float(getattr(info, "last_price", 0) or 0)
            prev  = float(getattr(info, "previous_close", 0) or 0)
            chg   = ((price - prev) / prev * 100) if prev > 0 else 0.0
            return price, chg
        except Exception as _wl_pe:
            print(f"[YieldIQ] Watchlist price fetch failed: {_wl_pe}")
            return 0.0, 0.0

    # ── Load watchlist items ──────────────────────────────────
    _wl_items = get_watchlist()
    _wl_now   = datetime.now().strftime("%H:%M:%S")

    # ── Enrich each item with live price + current MoS ───────
    _wl_enriched = []
    for _wl_item in _wl_items:
        _live_px, _day_chg = _wl_fetch_price(_wl_item["ticker"])
        _tgt = _wl_item["target_price"]
        _cur_mos = ((_tgt - _live_px) / _live_px * 100) if (_live_px > 0 and _tgt > 0) else 0.0
        _wl_enriched.append({
            **_wl_item,
            "live_price":  _live_px,
            "day_chg_pct": _day_chg,
            "current_mos": _cur_mos,
        })

    # ── 🔔 Alert banners — show BEFORE anything else ─────────
    for _wl_a in _wl_enriched:
        if _wl_a["live_price"] > 0 and _wl_a["current_mos"] >= _wl_a["alert_mos_threshold"]:
            st.warning(
                f"🔔 **Alert: {_wl_a['ticker']}** has crossed your "
                f"{_wl_a['alert_mos_threshold']:.0f}% MoS threshold — "
                f"currently at **{_wl_a['current_mos']:.1f}%** "
                f"(Target: {sym}{_wl_a['target_price']:,.2f} vs Live: {sym}{_wl_a['live_price']:,.2f})"
            )

    # ── Summary bar ───────────────────────────────────────────
    _wl_n_total = len(_wl_enriched)
    _wl_n_under = sum(1 for w in _wl_enriched if w["current_mos"] > 10)
    _wl_n_alert = sum(1 for w in _wl_enriched if w["current_mos"] >= w["alert_mos_threshold"])

    st.html(f"""
    <div style="display:flex;align-items:center;gap:0;
                background:#FFFFFF;border:1px solid #E2E8F0;border-radius:10px;
                overflow:hidden;margin-bottom:14px;">
      <div style="padding:12px 20px;border-right:1px solid #F1F5F9;flex:1;">
        <div style="font-size:10px;color:#94A3B8;text-transform:uppercase;
                    letter-spacing:0.12em;margin-bottom:3px;">Watching</div>
        <div style="font-size:20px;font-weight:700;color:#0F172A;
                    font-family:'IBM Plex Mono',monospace;">{_wl_n_total}</div>
      </div>
      <div style="padding:12px 20px;border-right:1px solid #F1F5F9;flex:1;">
        <div style="font-size:10px;color:#94A3B8;text-transform:uppercase;
                    letter-spacing:0.12em;margin-bottom:3px;">Undervalued</div>
        <div style="font-size:20px;font-weight:700;color:#059669;
                    font-family:'IBM Plex Mono',monospace;">{_wl_n_under}</div>
      </div>
      <div style="padding:12px 20px;border-right:1px solid #F1F5F9;flex:1;">
        <div style="font-size:10px;color:#94A3B8;text-transform:uppercase;
                    letter-spacing:0.12em;margin-bottom:3px;">Alerts Triggered</div>
        <div style="font-size:20px;font-weight:700;color:{'#DC2626' if _wl_n_alert else '#94A3B8'};
                    font-family:'IBM Plex Mono',monospace;">{_wl_n_alert}</div>
      </div>
      <div style="padding:12px 20px;flex:1;text-align:right;">
        <div style="font-size:10px;color:#94A3B8;text-transform:uppercase;
                    letter-spacing:0.12em;margin-bottom:3px;">Last updated</div>
        <div style="font-size:13px;font-weight:600;color:#64748B;
                    font-family:'IBM Plex Mono',monospace;">{_wl_now}</div>
      </div>
    </div>
    """)

    # ── Top controls ──────────────────────────────────────────
    _wl_ctrl1, _wl_ctrl2 = st.columns([5, 1])
    with _wl_ctrl2:
        if st.button("🔄 Refresh Prices", key="wl_refresh_all",
                     width='stretch'):
            _wl_fetch_price.clear()
            st.rerun()

    # ── Empty state ───────────────────────────────────────────
    if not _wl_enriched:
        st.html("""
        <div style="text-align:center;padding:48px 24px;background:#F8FAFC;
                    border:2px dashed #E2E8F0;border-radius:12px;margin-top:8px;">
          <div style="font-size:32px;margin-bottom:12px;">📌</div>
          <div style="font-size:15px;font-weight:600;color:#0F172A;margin-bottom:6px;">
            Your watchlist is empty
          </div>
          <div style="font-size:13px;color:#94A3B8;line-height:1.6;">
            Run a stock analysis on the <strong>🔍 Stock Analysis</strong> tab,<br>
            then click <strong>📌 Add to Watchlist</strong> to track it here.
          </div>
        </div>
        """)
    else:
        # ── Cards grid — 3 per row ────────────────────────────
        _WL_COLS = 3
        for _wl_row_start in range(0, len(_wl_enriched), _WL_COLS):
            _wl_row = _wl_enriched[_wl_row_start:_wl_row_start + _WL_COLS]
            _wl_cols = st.columns(_WL_COLS)

            for _wl_col, _w in zip(_wl_cols, _wl_row):
                _tk       = _w["ticker"]
                _mos      = _w["current_mos"]
                _live     = _w["live_price"]
                _added    = _w["added_price"]
                _tgt      = _w["target_price"]
                _thresh   = _w["alert_mos_threshold"]
                _notes_txt = (_w.get("notes") or "").strip()
                _since_added = ((_live - _added) / _added * 100) if _added > 0 else 0.0
                _day_c    = _w["day_chg_pct"]
                _co_name  = (_w.get("company_name") or _tk)[:28]

                # ── Colour scheme by MoS ──────────────────────────
                _alert_triggered = _mos >= _thresh
                if _alert_triggered:
                    _mos_col, _mos_bg, _top_bar = "#059669", "#ECFDF5", "#059669"
                elif _mos > 10:
                    _mos_col, _mos_bg, _top_bar = "#0D7A4E", "#F0FDF4", "#10B981"
                elif _mos > 0:
                    _mos_col, _mos_bg, _top_bar = "#D97706", "#FFFBEB", "#F59E0B"
                else:
                    _mos_col, _mos_bg, _top_bar = "#DC2626", "#FEF2F2", "#EF4444"

                _since_col = "#059669" if _since_added >= 0 else "#DC2626"
                _since_sym = "▲" if _since_added >= 0 else "▼"
                _day_col   = "#059669" if _day_c >= 0 else "#DC2626"
                _day_sym   = "▲" if _day_c >= 0 else "▼"

                # ── vs Target column ──────────────────────────────
                _vs_tgt_pct = ((_live - _added) / (_tgt - _added) * 100) if (_tgt > _added and _added > 0) else 0.0
                _vs_tgt_pct = max(0.0, min(100.0, _vs_tgt_pct))
                _vs_tgt_txt = f"{_vs_tgt_pct:.0f}% to target"

                # ── Progress bar fill ─────────────────────────────
                _prog_pct   = _vs_tgt_pct
                _prog_color = "#059669" if _prog_pct >= 50 else "#3B82F6"

                # ── Notes snippet ─────────────────────────────────
                _notes_html = (
                    f'<div style="font-size:11px;color:#94A3B8;font-style:italic;'
                    f'margin-top:8px;padding-top:8px;border-top:1px dashed #F1F5F9;'
                    f'line-height:1.5;">"{_notes_txt[:60]}{"…" if len(_notes_txt)>60 else ""}"</div>'
                ) if _notes_txt else ""

                # ── Alert border & pulse ──────────────────────────
                _card_border = "2px solid #F59E0B" if _alert_triggered else "1px solid #E2E8F0"
                _pulse_style = (
                    "animation:wl-pulse 2s ease-in-out infinite;"
                    if _alert_triggered else ""
                )
                _alert_chip = (
                    '<div style="display:inline-flex;align-items:center;gap:4px;'
                    'padding:2px 8px;background:#FEF3C7;border:1px solid #F59E0B;'
                    'border-radius:10px;font-size:10px;font-weight:700;color:#B45309;'
                    'margin-bottom:8px;">🔔 ALERT TRIGGERED</div>'
                    if _alert_triggered else ""
                )

                # ── Sparkline (use mini_sparkline if history available) ─
                _spark_html = ""
                try:
                    import yfinance as _yf_spark
                    _hist = _yf_spark.Ticker(_tk).history(period="7d", interval="1d")
                    if _hist is not None and not _hist.empty and len(_hist) >= 2:
                        _spark_vals = _hist["Close"].dropna().tolist()
                        _spark_html = (
                            '<div style="margin:8px 0 4px;">'
                            + mini_sparkline(_spark_vals, width=80, height=24)
                            + '</div>'
                        )
                except Exception:
                    pass

                _wl_col.html(f"""
                <style>
                @keyframes wl-pulse {{
                  0%,100% {{ box-shadow: 0 2px 8px rgba(245,158,11,0.15); }}
                  50%      {{ box-shadow: 0 4px 20px rgba(245,158,11,0.40); }}
                }}
                .wl-card:hover {{
                  transform: translateY(-2px);
                  box-shadow: 0 8px 24px rgba(15,23,42,0.10) !important;
                }}
                </style>

                <div class="wl-card" style="
                  background:#FFFFFF;
                  border:{_card_border};
                  border-radius:12px;
                  overflow:hidden;
                  margin-bottom:4px;
                  box-shadow:0 2px 8px rgba(15,23,42,0.06);
                  transition:transform .18s ease, box-shadow .18s ease;
                  {_pulse_style}
                ">
                  <!-- Top accent bar -->
                  <div style="height:3px;background:{_top_bar};"></div>

                  <div style="padding:14px 16px 12px;">
                    {_alert_chip}

                    <!-- Header row: Ticker + Price -->
                    <div style="display:flex;justify-content:space-between;
                                align-items:flex-start;margin-bottom:2px;">
                      <div>
                        <div style="font-family:'IBM Plex Mono',monospace;font-size:17px;
                                    font-weight:800;color:#0F172A;letter-spacing:-0.01em;">
                          {_tk}
                        </div>
                        <div style="font-size:11px;color:#94A3B8;margin-top:1px;
                                    white-space:nowrap;overflow:hidden;text-overflow:ellipsis;
                                    max-width:120px;">
                          {_co_name}
                        </div>
                      </div>
                      <div style="text-align:right;">
                        <div style="font-family:'IBM Plex Mono',monospace;font-size:16px;
                                    font-weight:700;color:#0F172A;">
                          {sym}{_live:,.2f}
                        </div>
                        <div style="font-size:11px;font-weight:600;color:{_day_col};">
                          {_day_sym} {abs(_day_c):.2f}%
                        </div>
                      </div>
                    </div>

                    <!-- Sparkline -->
                    {_spark_html}

                    <!-- Metrics row: MoS | Since Added | vs Target -->
                    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;
                                gap:5px;margin:10px 0;">
                      <div style="background:{_mos_bg};border-radius:7px;
                                  padding:7px 6px;text-align:center;">
                        <div style="font-size:9px;color:#94A3B8;text-transform:uppercase;
                                    letter-spacing:0.08em;margin-bottom:2px;">MoS</div>
                        <div style="font-family:'IBM Plex Mono',monospace;font-size:13px;
                                    font-weight:800;color:{_mos_col};">{_mos:+.1f}%</div>
                      </div>
                      <div style="background:#F8FAFC;border-radius:7px;
                                  padding:7px 6px;text-align:center;">
                        <div style="font-size:9px;color:#94A3B8;text-transform:uppercase;
                                    letter-spacing:0.08em;margin-bottom:2px;">Since Added</div>
                        <div style="font-family:'IBM Plex Mono',monospace;font-size:13px;
                                    font-weight:700;color:{_since_col};">
                          {_since_sym} {abs(_since_added):.1f}%
                        </div>
                      </div>
                      <div style="background:#EFF6FF;border-radius:7px;
                                  padding:7px 6px;text-align:center;">
                        <div style="font-size:9px;color:#3B82F6;text-transform:uppercase;
                                    letter-spacing:0.08em;margin-bottom:2px;">vs Target</div>
                        <div style="font-family:'IBM Plex Mono',monospace;font-size:11px;
                                    font-weight:700;color:#1D4ED8;">{_vs_tgt_txt}</div>
                      </div>
                    </div>

                    <!-- Progress bar: journey to target -->
                    <div style="margin-bottom:6px;">
                      <div style="display:flex;justify-content:space-between;
                                  font-size:9px;color:#94A3B8;margin-bottom:3px;">
                        <span>Added {sym}{_added:,.0f}</span>
                        <span>Target {sym}{_tgt:,.0f}</span>
                      </div>
                      <div style="height:5px;background:#F1F5F9;border-radius:3px;overflow:hidden;">
                        <div style="height:100%;width:{_prog_pct:.1f}%;
                                    background:{_prog_color};border-radius:3px;
                                    transition:width .4s ease;"></div>
                      </div>
                    </div>

                    <!-- Alert badge -->
                    <div style="font-size:10px;color:#94A3B8;margin-top:4px;">
                      <span style="padding:1px 7px;background:#F8FAFC;
                                   border:1px solid #E2E8F0;border-radius:8px;
                                   font-family:'IBM Plex Mono',monospace;">
                        Alert: {_thresh:.0f}%
                      </span>
                    </div>

                    <!-- Notes -->
                    {_notes_html}
                  </div>
                </div>
                """)

                # ── Action buttons ────────────────────────────────
                _btn_c1, _btn_c2 = _wl_col.columns(2)
                with _btn_c1:
                    if st.button(
                        "🔍 Analyse", key=f"wl_analyse_{_tk}",
                        width='stretch',
                        help=f"Switch to Stock Analysis and pre-fill {_tk}",
                    ):
                        st.session_state["_prefill_ticker"] = _tk
                        st.rerun()
                with _btn_c2:
                    if st.button(
                        "🗑 Remove", key=f"wl_remove_{_tk}",
                        width='stretch',
                    ):
                        remove_from_watchlist(_tk)
                        _wl_fetch_price.clear()
                        st.rerun()


if _active_main_tab == "about":
    a1, a2 = st.columns([3, 2])

    with a1:
        st.html("""
        <div style="background:linear-gradient(135deg,#F8FAFC,#F4F6F9);border:1px solid #E2E8F0;
                    border-radius:12px;padding:22px;margin-bottom:14px;">
          <div style="font-size:13px;font-weight:700;color:#0F172A;margin-bottom:14px;">⚙️ How it works</div>
        """)
        features = [
            ("Automatic return rate calculation",  "We use market data to figure out what return rate to expect from a stock (no guesswork needed)", "#3b82f6"),
            ("3 outcome scenarios",                "We model a pessimistic, a likely, and an optimistic case for every stock", "#f59e0b"),
            ("Realistic growth modelling",         "Growth rates gradually slow down over time — just like real businesses", "#8b5cf6"),
            ("Bank stocks handled separately",     "Cash-flow models don't work well for banks — we flag these automatically", "#10b981"),
            ("Quality filter",                     "Low-margin businesses are filtered out to avoid false 'good buy' signals", "#ef4444"),
            ("Currency auto-detection",            "Indian IT companies that report in USD (INFY, WIPRO, etc.) are automatically adjusted", "#06b6d4"),
            ("Model Price Levels",                 "Model-estimated DCF discount threshold, fair value estimate, and risk range — for research purposes only", "#10b981"),
            ("Download report",                    "Download a full analysis report as a text file for any stock you've analysed", "#3b82f6"),
        ]
        for title, desc, color in features:
            st.html(f"""
            <div style="display:flex;gap:12px;margin-bottom:10px;padding:12px 14px;
                        background:#F4F6F9;border:1px solid #E2E8F0;border-radius:8px;
                        border-left:3px solid {color};">
              <div>
                <div style="font-weight:600;color:#1E293B;font-size:12px;margin-bottom:2px;">{title}</div>
                <div style="color:#475569;font-size:12px;line-height:1.6;">{desc}</div>
              </div>
            </div>
            """)
        st.html("</div>")

    with a2:
        st.html("""
        <div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:10px;padding:16px;margin-bottom:12px;">
          <div style="font-size:12px;font-weight:700;color:#0F172A;margin-bottom:12px;">What our model signals mean</div>
        """)
        for human_label, internal, cond, color, bg in [
            ("📊 Trading below model estimate",  "Undervalued 🟢",    "Model price > market price by 20%+",  "#0D7A4E","#F0FDF4"),
            ("📉 Slight discount to model value","Slight Discount 🟡","Model price > market price by 5–20%", "#B45309","#FFFBEB"),
            ("⚖️ Near model fair value",          "Fairly Valued 🔵",  "Market price ≈ model fair value",      "#1D4ED8","#EFF6FF"),
            ("📈 Trading above model estimate",  "Overvalued 🔴",     "Market price > model price",           "#B91C1C","#FEF2F2"),
            ("⏳ Not applicable",                "N/A",               "Bank/NBFC or insufficient data",       "#64748b","#F8FAFC"),
        ]:
            st.html(f"""
            <div style="padding:10px 12px;background:{bg};border:1px solid {color}30;
                        border-radius:7px;margin-bottom:6px;">
              <div style="font-weight:700;color:{color};font-size:13px;">{human_label}</div>
              <div style="font-size:11px;color:{color};opacity:0.8;margin-top:2px;">{cond}</div>
            </div>
            """)
        st.html("</div>")

    st.html("""
    <div style="margin-top:8px;padding:18px 20px;background:#FFFBEB;border:1px solid #FDE68A;
                border-radius:10px;">
      <div style="font-weight:700;color:#f59e0b;font-size:13px;margin-bottom:8px;">⚠️ Important Disclosure</div>
      <div style="color:#B45309;font-size:12px;line-height:1.8;">
        <strong>For informational and educational purposes only. Not investment advice.</strong>
        This tool provides quantitative DCF analysis for research purposes. It does not constitute
        a recommendation to buy, sell, or hold any security. Past performance is not indicative of
        future results. All valuations are estimates based on publicly available data and model
        assumptions that may prove incorrect. Users should conduct their own due diligence and
        consult a registered investment advisor (RIA) or licensed financial professional before
        making any investment decisions. This platform is not registered as an investment advisor
        under the Investment Advisers Act of 1940 or any state securities law.
        Data sourced from Yahoo Finance — accuracy not guaranteed.
      </div>
      <div style="color:#92400E;font-size:11px;margin-top:10px;padding-top:8px;
                  border-top:1px solid #FDE68A;font-weight:600;">
        YieldIQ displays model outputs for informational purposes only.
        This is not investment advice. YieldIQ is not a registered investment adviser.
        Past model performance does not guarantee future results.
      </div>
      <div style="color:#92400E;font-size:12px;line-height:1.8;margin-top:10px;
                  padding-top:10px;border-top:1px solid #FDE68A;">
        <strong>Signal definitions:</strong> YieldIQ does not provide personalised investment
        model outputs. The &#39;Undervalued / Overvalued&#39; signals are purely mathematical
        outputs comparing the current market price to a model-estimated intrinsic value derived
        from a Discounted Cash Flow (DCF) analysis. They do <em>not</em> account for your
        personal financial situation, risk tolerance, tax position, investment objectives, or
        time horizon. The model relies on publicly available financial data that may be incomplete,
        delayed, or inaccurate. Intrinsic value estimates are highly sensitive to input assumptions
        (growth rates, discount rates) and can change materially with different inputs. No content
        on this platform should be construed as an offer or solicitation to buy or sell any
        financial instrument. Always consult a qualified financial adviser before making any
        investment decision.
      </div>
    </div>
    """)


# ══════════════════════════════════════════════════════════════
# TAB — ALERTS
# ══════════════════════════════════════════════════════════════
if _active_main_tab == "alerts":

    _tal_email    = st.session_state.get("auth_email", "")
    _tal_is_guest = not _tal_email or _tal_email == "guest"
    _tal_uid      = None if _tal_is_guest else _alerts_mod._get_user_id(_tal_email)

    if _tal_is_guest or _tal_uid is None:
        st.html("""
        <div style="padding:40px 32px;text-align:center;background:#F8FAFC;
                    border:1.5px solid #E2E8F0;border-radius:14px;margin:20px 0;">
          <div style="font-size:36px;margin-bottom:12px;">&#128276;</div>
          <div style="font-size:17px;font-weight:700;color:#0F172A;margin-bottom:8px;">
            Sign in to use Price Alerts
          </div>
          <div style="font-size:13px;color:#475569;max-width:380px;margin:0 auto 20px;line-height:1.7;">
            Get notified when a stock crosses your target price or reaches its
            intrinsic value &#8212; even while you&#8217;re away from the dashboard.
          </div>
        </div>""")
        upgrade_prompt("action_plan", compact=True)

    else:
        _tal_tier     = tier()
        _tal_cap      = _alerts_mod.get_alert_limit(_tal_tier)
        _tal_active   = _alerts_mod.get_active_alerts(_tal_uid)
        _tal_count    = len(_tal_active)
        _tal_cap_str  = "Unlimited" if _tal_cap >= 9_999 else str(_tal_cap)
        _tal_pct      = min(_tal_count / _tal_cap * 100, 100) if _tal_cap < 9_999 else 0
        _tal_bar_clr  = ("#dc2626" if _tal_pct >= 100
                         else "#d97706" if _tal_pct >= 66 else "#059669")

        # ── Header ────────────────────────────────────────────
        _tal_c1, _tal_c2 = st.columns([3, 1])
        with _tal_c1:
            st.html(f"""
            <div style="margin-bottom:18px;">
              <div style="font-size:22px;font-weight:800;color:#0F172A;
                          letter-spacing:-0.02em;margin-bottom:4px;">
                &#128276; Price Alerts
              </div>
              <div style="font-size:13px;color:#475569;">
                Get notified when stocks hit your target price or intrinsic value.
              </div>
            </div>""")
        with _tal_c2:
            _tal_tier_clr = {"free":"#8492a6","starter":"#5046e4","premium":"#5046e4","pro":"#059669"}.get(_tal_tier,"#8492a6")
            _tal_bar_html = (
                f'<div style="height:4px;background:#E2E8F0;border-radius:2px;">'
                f'<div style="height:100%;width:{int(_tal_pct)}%;background:{_tal_bar_clr};'
                f'border-radius:2px;"></div></div>'
                if _tal_cap < 9_999 else ""
            )
            st.html(f"""
            <div style="text-align:right;padding-top:12px;">
              <div style="font-size:11px;font-weight:700;color:{_tal_tier_clr};
                          text-transform:uppercase;letter-spacing:.1em;margin-bottom:4px;">
                {_tal_tier.capitalize()} &middot; {_tal_count} / {_tal_cap_str} alerts
              </div>
              {_tal_bar_html}
            </div>""")

        # ── Create new alert ──────────────────────────────────
        st.html('<div style="font-size:11px;font-weight:700;color:#64748B;'
                'text-transform:uppercase;letter-spacing:.1em;margin-bottom:10px;">'
                'New Alert</div>')

        if _tal_count >= _tal_cap:
            if _tal_tier == "free":
                upgrade_prompt("action_plan", compact=True)
            else:
                st.info(
                    f"You've reached the {_tal_cap}-alert limit for "
                    f"{_tal_tier.capitalize()}. Upgrade to Pro for unlimited alerts."
                )
        else:
            with st.form("_alerts_create_form", clear_on_submit=True):
                _fc1, _fc2, _fc3, _fc4 = st.columns([2, 2, 2, 1])
                with _fc1:
                    _new_ticker = st.text_input(
                        "Ticker", placeholder="e.g. AAPL",
                        help="Stock ticker symbol (e.g. AAPL, MSFT, GOOGL, NVDA)"
                    ).strip().upper()
                with _fc2:
                    _new_type = st.selectbox(
                        "Alert type",
                        options=list(_alerts_mod.ALERT_TYPE_LABELS.keys()),
                        format_func=lambda k: _alerts_mod.ALERT_TYPE_LABELS[k],
                        help=(
                            "above: triggers when price rises above target\n"
                            "below: triggers when price falls below target\n"
                            "iv_reached: triggers when price falls to your IV estimate"
                        ),
                    )
                with _fc3:
                    _new_price = st.number_input(
                        "Target price", min_value=0.01, value=100.00, step=1.0,
                        help="Price level that will trigger this alert"
                    )
                with _fc4:
                    st.html("<div style='height:28px'></div>")
                    _create_btn = st.form_submit_button(
                        "Add alert", width='stretch', type="primary"
                    )

                if _create_btn:
                    _al_res = _alerts_mod.create_alert(
                        _tal_uid, _new_ticker, _new_type, _new_price, _tal_tier
                    )
                    if _al_res["ok"]:
                        st.success(f"Alert created for {_new_ticker}.")
                        st.rerun()
                    else:
                        st.error(_al_res["error"])

        st.html('<div style="height:12px"></div>')

        # ── Active alerts list ────────────────────────────────
        st.html(
            '<div style="font-size:11px;font-weight:700;color:#64748B;'
            'text-transform:uppercase;letter-spacing:.1em;margin-bottom:10px;">'
            f'Active Alerts ({_tal_count})</div>'
        )
        if not _tal_active:
            st.html("""
            <div style="padding:20px;text-align:center;background:#F8FAFC;
                        border:1px solid #E2E8F0;border-radius:10px;
                        color:#94A3B8;font-size:13px;">
              No active alerts. Create one above to get started.
            </div>""")
        else:
            for _al in _tal_active:
                _al_type_lbl = _alerts_mod.ALERT_TYPE_LABELS.get(_al["alert_type"], _al["alert_type"])
                _al_clr = ("#059669" if _al["alert_type"] == "above"
                           else "#1D4ED8" if _al["alert_type"] == "iv_reached"
                           else "#DC2626")
                _al_col_l, _al_col_r = st.columns([10, 1])
                with _al_col_l:
                    st.html(f"""
                    <div style="display:flex;align-items:center;gap:14px;
                                padding:12px 16px;background:#FFFFFF;
                                border:1px solid #E2E8F0;border-radius:10px;
                                border-left:3px solid {_al_clr};">
                      <div style="font-size:15px;font-weight:800;color:#0F172A;
                                  font-family:'IBM Plex Mono',monospace;min-width:80px;">
                        {_al['ticker']}
                      </div>
                      <div style="width:1px;height:28px;background:#E2E8F0;"></div>
                      <div style="flex:1;">
                        <div style="font-size:12px;font-weight:600;color:{_al_clr};">
                          {_al_type_lbl}
                        </div>
                        <div style="font-size:13px;color:#0F172A;font-weight:700;
                                    font-family:'IBM Plex Mono',monospace;">
                          ${_al['target_price']:,.2f}
                        </div>
                      </div>
                      <div style="font-size:11px;color:#94A3B8;">
                        Created {_al['created_at'][:10]}
                      </div>
                    </div>""")
                with _al_col_r:
                    if st.button("✕", key=f"_del_alert_{_al['id']}",
                                 help="Delete this alert",
                                 width='stretch'):
                        _del_res = _alerts_mod.delete_alert(_al["id"], _tal_uid)
                        if _del_res["ok"]:
                            st.rerun()
                        else:
                            st.error(_del_res["error"])

        st.html('<div style="height:20px"></div>')

        # ── Recently triggered ────────────────────────────────
        _tal_triggered = _alerts_mod.get_triggered_alerts(_tal_uid, hours=24)
        if _tal_triggered:
            st.html(
                '<div style="font-size:11px;font-weight:700;color:#64748B;'
                'text-transform:uppercase;letter-spacing:.1em;margin-bottom:10px;">'
                f'Triggered in the last 24 h ({len(_tal_triggered)})</div>'
            )
            for _tr in _tal_triggered:
                _tr_lbl = _alerts_mod.ALERT_TYPE_LABELS.get(_tr["alert_type"], _tr["alert_type"])
                st.html(f"""
                <div style="display:flex;align-items:center;gap:14px;
                            padding:10px 16px;background:#FFFBEB;
                            border:1px solid #FDE68A;border-radius:10px;
                            margin-bottom:6px;">
                  <span style="font-size:15px;">&#128276;</span>
                  <div style="flex:1;">
                    <span style="font-weight:700;color:#0F172A;">{_tr['ticker']}</span>
                    <span style="color:#475569;font-size:13px;">
                      &mdash; {_tr_lbl} ${_tr['target_price']:,.2f}
                    </span>
                  </div>
                  <div style="font-size:11px;color:#92400E;">
                    {_tr['triggered_at'][:16].replace('T', ' ')} UTC
                  </div>
                </div>""")

            if st.button("Clear triggered alerts", key="_al_clear_triggered"):
                _alerts_mod.delete_all_triggered(_tal_uid)
                st.session_state["_al_fired"] = []
                st.rerun()

        # ── Manual re-check ───────────────────────────────────
        st.html('<div style="height:8px"></div>')
        if st.button("&#128260; Check alerts now", key="_al_check_now",
                     help="Fetch live prices and check all alerts immediately"):
            with st.spinner("Checking prices\u2026"):
                _manual_fired = _alerts_mod.check_alerts(_tal_uid)
            st.session_state["_al_last_check_ts"] = _time.time()
            st.session_state["_al_fired"] = (
                st.session_state.get("_al_fired", []) + _manual_fired
            )
            if _manual_fired:
                st.success(f"{len(_manual_fired)} alert(s) triggered!")
            else:
                st.success("All prices within range \u2014 no alerts triggered.")
            st.rerun()

# ══════════════════════════════════════════════════════════════
# TAB — ADMIN ANALYTICS  (YIELDIQ_ADMIN=1 only)
# ══════════════════════════════════════════════════════════════
if _ADMIN_MODE and _active_main_tab == "admin":
    render_admin_dashboard()
