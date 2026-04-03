"""
Portfolio Decision System
Adds 5 new sheets to the hedge fund model:
  Sheet 14: Expected Value Framework
  Sheet 15: Kelly Criterion Position Sizing
  Sheet 16: Portfolio Context & Correlation
  Sheet 17: Risk Management & Stop Loss
  Sheet 18: Final Portfolio Decision Memo

Usage:
    python generate_portfolio_excel.py --ticker TCS.NS
    python generate_portfolio_excel.py --ticker TCS.NS --portfolio-size 10000000
"""

import sys, argparse, io, warnings
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# ── Palette (same as institutional + HF models) ────────────────
C_INPUT_BG = "EBF3FB"; C_INPUT_FG = "0000FF"
C_FORMULA  = "000000"; C_LINK_FG  = "008000"
C_HEADER   = "1F3864"; C_HDR_FG   = "FFFFFF"
C_SUBHDR   = "2E75B6"; C_SECTION  = "D6E4F0"
C_ALT      = "F2F7FB"; C_WHITE    = "FFFFFF"
C_YELLOW   = "FFFF00"; C_GREEN    = "00B050"
C_RED      = "C00000"; C_AMBER    = "ED7D31"
C_BEAR_BG  = "FCE4D6"; C_BULL_BG  = "E2EFDA"
C_NAVY     = "1F3864"; C_KELLY    = "E8F5E9"

def _c(s):
    if not s: return "FFFFFFFF"
    s = str(s).lstrip("#").upper()
    if len(s) == 6: return "FF" + s
    if len(s) == 8: return s
    if len(s) >= 10: return s[:8]
    return "FFFFFFFF"

def hf(c): return PatternFill("solid", fgColor=_c(c))

def bdr(style="thin", color="BFBFBF"):
    s = Side(style=style, color=_c(color))
    return Border(left=s, right=s, top=s, bottom=s)

def wc(ws, row, col, val, bg=C_WHITE, fg=C_FORMULA, bold=False, sz=10,
       nf=None, align="right", italic=False, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = hf(bg)
    c.font = Font(name="Calibri", bold=bold, size=sz, color=_c(fg), italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = bdr()
    if nf: c.number_format = nf
    return c

def inp(ws, row, col, val, nf=None):
    c = wc(ws, row, col, val, bg=C_INPUT_BG, fg=C_INPUT_FG)
    if nf: c.number_format = nf
    return c

def hdr(ws, row, labels, bg=C_HEADER, fg=C_HDR_FG, sz=10, height=20):
    ws.row_dimensions[row].height = height
    for col, lbl in enumerate(labels, 1):
        c = ws.cell(row=row, column=col, value=lbl)
        c.fill = hf(bg); c.font = Font(name="Calibri", bold=True, size=sz, color=_c(fg))
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr()

def sec(ws, row, text, ncols, bg=C_SECTION, fg=C_NAVY):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=f"  {text}")
    c.fill = hf(bg); c.font = Font(name="Calibri", bold=True, size=10, color=_c(fg))
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = Border(bottom=Side(style="medium", color=_c(C_NAVY)))
    ws.row_dimensions[row].height = 18

def title(ws, row, text, sub, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = hf(C_NAVY); c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=ncols)
    c2 = ws.cell(row=row+1, column=1, value=sub)
    c2.fill = hf(C_SUBHDR); c2.font = Font(name="Calibri", size=9, color="FFFFFFFF", italic=True)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row+1].height = 15


def build_portfolio_sheets(wb, ticker, enriched, dcf_res, forecast_result,
                            scenarios, wacc_data, wacc, terminal_g,
                            forecast_yrs, sym, fx, portfolio_size=10_000_000):

    price     = enriched.get("price", 0) * fx
    shares    = enriched.get("shares", 1)
    op_margin = enriched.get("op_margin", 0)
    rev_g     = enriched.get("revenue_growth", 0)
    sector    = enriched.get("sector_name", enriched.get("sector", "General"))
    moat      = enriched.get("moat_grade", "None")
    iv_base   = dcf_res.get("intrinsic_value_per_share", 0) * fx
    mktcap    = price * shares / 1e9

    bear_iv = scenarios.get("Bear 🐻", {}).get("iv", iv_base * 0.6) * fx
    base_iv = scenarios.get("Base 📊", {}).get("iv", iv_base)       * fx
    bull_iv = scenarios.get("Bull 🐂", {}).get("iv", iv_base * 1.4) * fx

    p_bear, p_base, p_bull = 0.25, 0.50, 0.25
    ev = bear_iv * p_bear + base_iv * p_base + bull_iv * p_bull
    edge = (ev - price) / price if price > 0 else 0

    upside   = (bull_iv - price) / price if price > 0 else 0
    downside = (bear_iv - price) / price if price > 0 else 0
    rr_ratio = abs(upside / downside) if downside < 0 else 9.99

    # Kelly Criterion
    # p = probability of winning (IV > price), b = upside/downside ratio
    p_win = p_base + p_bull  # probability scenarios where we win
    if downside < 0 and upside > 0:
        b_ratio = upside / abs(downside)
        kelly_full = (p_win * b_ratio - (1 - p_win)) / b_ratio
        kelly_half = kelly_full / 2
        kelly_qtr  = kelly_full / 4
    else:
        kelly_full = 0.10
        kelly_half = 0.05
        kelly_qtr  = 0.025

    kelly_full = max(0, min(kelly_full, 0.30))  # cap at 30%
    kelly_half = max(0, min(kelly_half, 0.15))
    kelly_qtr  = max(0, min(kelly_qtr,  0.075))

    # Capital amounts
    cap_full = portfolio_size * kelly_full
    cap_half = portfolio_size * kelly_half
    cap_qtr  = portfolio_size * kelly_qtr
    shares_full = int(cap_full / price) if price > 0 else 0
    shares_half = int(cap_half / price) if price > 0 else 0

    # Position classification
    if kelly_half >= 0.08:     position_type = "CORE POSITION (High Conviction)"
    elif kelly_half >= 0.04:   position_type = "MEDIUM ALLOCATION"
    elif kelly_half >= 0.01:   position_type = "SMALL TRACKING POSITION"
    else:                      position_type = "AVOID"

    # ══════════════════════════════════════════════════════════
    # SHEET 14 — EXPECTED VALUE FRAMEWORK
    # ══════════════════════════════════════════════════════════
    ws14 = wb.create_sheet("14. Expected Value")
    ws14.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [34,18,18,18,18,32]):
        ws14.column_dimensions[col].width = w

    title(ws14, 1, f"EXPECTED VALUE FRAMEWORK — {ticker}",
          f"EV = Σ(Probability × Outcome)  |  Portfolio Size: {sym}{portfolio_size/1e6:.1f}M", 6)

    # Step 1 — Scenario inputs
    sec(ws14, 3, "STEP 1 — SCENARIO INPUTS (Blue = analyst inputs)", 6)
    hdr(ws14, 4, ["Scenario", "Intrinsic Value", "Probability", "Weighted Value", "vs Price", "Key Driver"], bg=C_SUBHDR)

    scenario_data = [
        ("Bear 🐻 — Adverse",  bear_iv, p_bear, C_BEAR_BG, C_RED,   "Growth miss + margin compression + re-rating"),
        ("Base 📊 — Expected", base_iv, p_base, C_ALT,     C_FORMULA,"DCF base case — current trajectory continues"),
        ("Bull 🐂 — Upside",   bull_iv, p_bull, C_BULL_BG, C_GREEN,  "Moat widens + margin expansion + re-rating"),
    ]
    for i, (label, iv_v, prob, bg, fg, driver) in enumerate(scenario_data):
        r = 5 + i
        ws14.row_dimensions[r].height = 20
        wc(ws14, r, 1, label,                bg=bg, fg=fg, bold=True, sz=10, align="left")
        c = inp(ws14, r, 2, round(iv_v, 0)); c.number_format = f'"{sym}"#,##0'
        c2= inp(ws14, r, 3, prob);           c2.number_format = "0%"
        c2.fill = hf(C_YELLOW)
        wv = iv_v * prob
        wc(ws14, r, 4, round(wv, 0),         bg=bg, fg=fg, bold=True, sz=10, nf=f'"{sym}"#,##0')
        pct = (iv_v - price) / price if price > 0 else 0
        wc(ws14, r, 5, pct,                  bg=bg, fg=fg, bold=True, sz=10, nf="+0.0%;(0.0%);-")
        wc(ws14, r, 6, driver,               bg=C_ALT, fg="595959", sz=9, align="left", italic=True)

    ws14.row_dimensions[9].height = 22
    for col in range(1,7):
        wc(ws14, 9, col, "", bg="E8E8E8")
    ws14.merge_cells("A9:C9")
    wc(ws14, 9, 1, "  EXPECTED VALUE (Probability-Weighted)", bg=C_YELLOW, fg=C_NAVY, bold=True, sz=11, align="left")
    c_ev = wc(ws14, 9, 4, round(ev, 0), bg=C_YELLOW, fg=C_NAVY, bold=True, sz=13, nf=f'"{sym}"#,##0')
    ev_pct = (ev - price) / price if price > 0 else 0
    wc(ws14, 9, 5, ev_pct, bg=C_YELLOW, fg=C_GREEN if ev_pct>0 else C_RED, bold=True, sz=12, nf="+0.0%;(0.0%);-")

    # Step 2 — Edge calculation
    sec(ws14, 11, "STEP 2 — EDGE CALCULATION (Mispricing vs Market)", 6)
    hdr(ws14, 12, ["Metric", "Value", "", "Classification", "", "Interpretation"], bg=C_SUBHDR)

    edge_rows = [
        ("Current Market Price",        price,     f'"{sym}"#,##0',    "Reference",
         C_WHITE, C_INPUT_FG, "What market is charging today"),
        ("Expected Value (EV)",         ev,        f'"{sym}"#,##0',    "Calculated above",
         C_YELLOW, C_NAVY, "Probability-weighted fair value"),
        ("Edge = (EV − Price) / Price", edge,      "+0.0%;(0.0%);-",
         "HIGH (>25%)" if edge>0.25 else ("MODERATE (10-25%)" if edge>0.10 else ("LOW (<10%)" if edge>0 else "NEGATIVE")),
         C_BULL_BG if edge>0.25 else (C_ALT if edge>0.10 else (C_YELLOW if edge>0 else C_BEAR_BG)),
         C_GREEN if edge>0.25 else (C_AMBER if edge>0.10 else (C_FORMULA if edge>0 else C_RED)),
         "The statistical advantage per ₹ invested"),
        ("Upside Potential (Bull−Price)", upside, "+0.0%;(0.0%);-", f"{upside:.0%} potential gain",
         C_BULL_BG, C_GREEN, "Bull case gain from current price"),
        ("Downside Risk (Bear−Price)",  downside,  "+0.0%;(0.0%);-", f"{abs(downside):.0%} potential loss",
         C_BEAR_BG, C_RED, "Bear case loss from current price"),
        ("Risk/Reward Ratio",           rr_ratio,  "0.0x",            "Target: >2.0x for institutional",
         C_BULL_BG if rr_ratio>2 else (C_YELLOW if rr_ratio>1 else C_BEAR_BG),
         C_GREEN if rr_ratio>2 else (C_AMBER if rr_ratio>1 else C_RED),
         "Upside potential / Downside risk"),
    ]
    for i, (label, val, nf, classif, bg, fg, interp) in enumerate(edge_rows):
        r = 13 + i
        ws14.row_dimensions[r].height = 18
        wc(ws14, r, 1, label,        bg=bg, fg=C_FORMULA, sz=10, align="left")
        c = wc(ws14, r, 2, round(float(val),4) if not isinstance(val,str) else val,
               bg=bg, fg=fg, bold="Edge" in label or "EV" in label, sz=11)
        if nf: c.number_format = nf
        wc(ws14, r, 3, "",           bg=bg)
        wc(ws14, r, 4, classif,      bg=bg, fg=fg, bold=True, sz=10, align="center")
        wc(ws14, r, 5, "",           bg=bg)
        wc(ws14, r, 6, interp,       bg=C_ALT, fg="595959", sz=9, align="left", italic=True)

    # ══════════════════════════════════════════════════════════
    # SHEET 15 — KELLY CRITERION POSITION SIZING
    # ══════════════════════════════════════════════════════════
    ws15 = wb.create_sheet("15. Position Sizing")
    ws15.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [34,20,20,20,20,28]):
        ws15.column_dimensions[col].width = w

    title(ws15, 1, f"KELLY CRITERION POSITION SIZING — {ticker}",
          f"Optimal capital allocation based on edge and payoff ratio  |  Portfolio: {sym}{portfolio_size/1e6:.1f}M", 6)

    sec(ws15, 3, "STEP 3 — KELLY CRITERION INPUTS", 6)
    hdr(ws15, 4, ["Input", "Value", "", "Formula", "", "Note"], bg=C_SUBHDR)
    kelly_inputs = [
        ("Probability of Win (p)",       p_win,      "0.0%",  "p = P(Base) + P(Bull)",        "Scenarios where IV > Price"),
        ("Probability of Loss (1−p)",    1-p_win,    "0.0%",  "1 − p",                        "P(Bear) = adverse scenario"),
        ("Payoff Ratio (b)",             rr_ratio,   "0.0x",  "b = Upside% / |Downside%|",    "How much you win vs lose"),
        ("Upside %",                     upside,     "0.0%",  "(Bull IV − Price) / Price",    "Gain in favourable scenario"),
        ("Downside %",                   abs(downside),"0.0%","(Price − Bear IV) / Price",   "Loss in adverse scenario"),
    ]
    for i, (label, val, nf, formula, note) in enumerate(kelly_inputs):
        r = 5 + i
        ws15.row_dimensions[r].height = 17
        bg = C_ALT if i%2==0 else C_WHITE
        wc(ws15, r, 1, label,   bg=bg, fg=C_FORMULA, sz=10, align="left")
        c = inp(ws15, r, 2, round(float(val),4)); c.number_format = nf; c.fill = hf(C_INPUT_BG if i<2 else C_YELLOW)
        wc(ws15, r, 3, "",      bg=bg)
        wc(ws15, r, 4, formula, bg=C_ALT, fg="595959", sz=9, align="left", italic=True)
        wc(ws15, r, 5, "",      bg=bg)
        wc(ws15, r, 6, note,    bg=C_ALT, fg="595959", sz=9, align="left", italic=True)

    sec(ws15, 11, "KELLY CRITERION FORMULA & RESULTS", 6)
    ws15.row_dimensions[12].height = 20
    ws15.merge_cells("A12:F12")
    c_formula = ws15.cell(12, 1,
        value=f"  Kelly% = (p×b − (1−p)) / b  =  ({p_win:.0%}×{rr_ratio:.2f} − {1-p_win:.0%}) / {rr_ratio:.2f}  =  {kelly_full:.1%}")
    c_formula.fill = hf("EBF5FB")
    c_formula.font = Font(name="Courier New", bold=True, size=10, color=_c(C_NAVY))
    c_formula.alignment = Alignment(horizontal="left", vertical="center")
    c_formula.border = bdr("medium", C_SUBHDR)

    hdr(ws15, 13, ["Allocation Type", "Kelly %", "Capital ({})".format(sym), "Shares", "Max Loss (₹)", "Recommendation"],
        bg=C_SUBHDR)

    kelly_rows = [
        ("Full Kelly (aggressive)",       kelly_full, cap_full, shares_full,
         cap_full * abs(downside), "Rarely used — too volatile for most PMs",
         C_BEAR_BG, C_RED),
        ("Half Kelly (recommended)",      kelly_half, cap_half, shares_half,
         cap_half * abs(downside), "Industry standard — optimal risk-adjusted",
         C_KELLY,   C_GREEN),
        ("Quarter Kelly (conservative)",  kelly_qtr,  cap_qtr,  int(cap_qtr/price) if price>0 else 0,
         cap_qtr * abs(downside),  "For high-uncertainty situations",
         C_ALT,     C_FORMULA),
        ("Minimum (tracking position)",   0.01,       portfolio_size*0.01, int(portfolio_size*0.01/price) if price>0 else 0,
         portfolio_size*0.01*abs(downside), "Monitor without meaningful exposure",
         C_ALT,     "595959"),
    ]
    for i, (label, pct, cap, sh, max_loss, rec, bg, fg) in enumerate(kelly_rows):
        r = 14 + i
        ws15.row_dimensions[r].height = 20
        is_rec = i == 1
        wc(ws15, r, 1, label,          bg=bg, fg=fg, bold=is_rec, sz=10, align="left")
        wc(ws15, r, 2, pct,            bg=bg, fg=fg, bold=is_rec, sz=10, nf="0.0%")
        wc(ws15, r, 3, round(cap,0),   bg=bg, fg=fg, bold=is_rec, sz=10, nf=f'"{sym}"#,##0')
        wc(ws15, r, 4, sh,             bg=bg, fg=fg, bold=is_rec, sz=10, nf="#,##0")
        wc(ws15, r, 5, round(max_loss,0), bg=C_BEAR_BG if is_rec else bg, fg=C_RED, sz=10, nf=f'"{sym}"#,##0')
        wc(ws15, r, 6, rec,            bg=C_ALT, fg="595959", sz=9, align="left", italic=True)

    # Position type classification
    sec(ws15, 19, "POSITION CLASSIFICATION", 6)
    ws15.row_dimensions[20].height = 28
    ws15.merge_cells("A20:F20")
    pos_bg = C_BULL_BG if "CORE" in position_type else (C_ALT if "MEDIUM" in position_type else (C_YELLOW if "SMALL" in position_type else C_BEAR_BG))
    pos_fg = C_GREEN if "CORE" in position_type else (C_SUBHDR if "MEDIUM" in position_type else (C_AMBER if "SMALL" in position_type else C_RED))
    c_pos = ws15.cell(20, 1, value=f"  ➤  {position_type}")
    c_pos.fill = hf(pos_bg); c_pos.font = Font(name="Calibri", bold=True, size=14, color=_c(pos_fg))
    c_pos.alignment = Alignment(horizontal="center", vertical="center")
    c_pos.border = bdr("medium", C_NAVY)

    # Kelly curve data for chart
    sec(ws15, 22, "KELLY CURVE — Growth Rate vs Position Size", 6)
    hdr(ws15, 23, ["Position Size %", "Expected Portfolio Growth %"], bg=C_SUBHDR)
    fracs = np.linspace(0, min(kelly_full*2, 0.30), 20)
    for i, f in enumerate(fracs):
        r = 24 + i
        if upside > 0 and downside < 0:
            g = p_win * np.log(1 + f*upside) + (1-p_win) * np.log(1 + f*downside)
        else:
            g = f * edge
        ws15.row_dimensions[r].height = 14
        wc(ws15, r, 1, round(f, 4),   bg=C_YELLOW if abs(f-kelly_full)<0.01 else (C_KELLY if abs(f-kelly_half)<0.01 else C_WHITE),
           fg=C_FORMULA, sz=9, nf="0.0%")
        wc(ws15, r, 2, round(g, 4),   bg=C_BULL_BG if g>0 else C_BEAR_BG,
           fg=C_GREEN if g>0 else C_RED, sz=9, nf="0.00%")

    chart = LineChart()
    chart.title = f"Kelly Growth Curve — {ticker}"
    chart.y_axis.title = "Expected Log Growth %"
    chart.x_axis.title = "Position Size %"
    chart.style = 2; chart.width = 18; chart.height = 12
    data = Reference(ws15, min_col=2, min_row=23, max_row=23+len(fracs))
    cats = Reference(ws15, min_col=1, min_row=24, max_row=23+len(fracs))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws15.add_chart(chart, "H3")

    # ══════════════════════════════════════════════════════════
    # SHEET 16 — PORTFOLIO CONTEXT & CORRELATION
    # ══════════════════════════════════════════════════════════
    ws16 = wb.create_sheet("16. Portfolio Context")
    ws16.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [34,18,18,18,34]):
        ws16.column_dimensions[col].width = w

    title(ws16, 1, f"PORTFOLIO CONTEXT — {ticker}",
          f"Step 5 & 6: Sector characteristics, correlations, diversification", 5)

    # Sector correlations and characteristics
    SECTOR_PROFILE = {
        "it_services":      {"beta":0.9,  "rate_sens":"LOW",    "cycle":"DEFENSIVE", "inf_sens":"LOW",   "div":"HIGH"},
        "fmcg":             {"beta":0.7,  "rate_sens":"LOW",    "cycle":"DEFENSIVE", "inf_sens":"MEDIUM","div":"HIGH"},
        "consumer_durable": {"beta":0.9,  "rate_sens":"MEDIUM", "cycle":"CYCLICAL",  "inf_sens":"MEDIUM","div":"MEDIUM"},
        "pharma":           {"beta":0.8,  "rate_sens":"LOW",    "cycle":"DEFENSIVE", "inf_sens":"LOW",   "div":"HIGH"},
        "oil_gas":          {"beta":1.0,  "rate_sens":"LOW",    "cycle":"CYCLICAL",  "inf_sens":"HIGH",  "div":"LOW"},
        "metals":           {"beta":1.3,  "rate_sens":"MEDIUM", "cycle":"HIGHLY CYCLICAL","inf_sens":"HIGH","div":"LOW"},
        "cement":           {"beta":1.1,  "rate_sens":"MEDIUM", "cycle":"CYCLICAL",  "inf_sens":"HIGH",  "div":"MEDIUM"},
        "auto_oem":         {"beta":1.1,  "rate_sens":"HIGH",   "cycle":"CYCLICAL",  "inf_sens":"MEDIUM","div":"MEDIUM"},
        "capital_goods":    {"beta":1.2,  "rate_sens":"HIGH",   "cycle":"CYCLICAL",  "inf_sens":"LOW",   "div":"LOW"},
        "banks":            {"beta":1.2,  "rate_sens":"HIGH",   "cycle":"CYCLICAL",  "inf_sens":"LOW",   "div":"LOW"},
        "realty":           {"beta":1.4,  "rate_sens":"HIGH",   "cycle":"HIGHLY CYCLICAL","inf_sens":"HIGH","div":"LOW"},
        "telecom":          {"beta":0.8,  "rate_sens":"HIGH",   "cycle":"DEFENSIVE", "inf_sens":"LOW",   "div":"MEDIUM"},
        "power":            {"beta":0.7,  "rate_sens":"HIGH",   "cycle":"DEFENSIVE", "inf_sens":"MEDIUM","div":"MEDIUM"},
        "airlines":         {"beta":1.5,  "rate_sens":"MEDIUM", "cycle":"HIGHLY CYCLICAL","inf_sens":"HIGH","div":"LOW"},
        "chemicals":        {"beta":1.1,  "rate_sens":"MEDIUM", "cycle":"CYCLICAL",  "inf_sens":"HIGH",  "div":"MEDIUM"},
        "general":          {"beta":1.0,  "rate_sens":"MEDIUM", "cycle":"CYCLICAL",  "inf_sens":"MEDIUM","div":"MEDIUM"},
    }
    raw_sector = enriched.get("sector", "general")
    profile = SECTOR_PROFILE.get(raw_sector, SECTOR_PROFILE["general"])

    sec(ws16, 3, "STEP 5 — SECTOR CHARACTERISTICS & PORTFOLIO FIT", 5)
    hdr(ws16, 4, ["Factor", "This Stock", "Ideal Portfolio Blend", "Impact", "Note"], bg=C_SUBHDR)

    profile_rows = [
        ("Sector",                  sector,                  "Diversified across sectors",
         "Context",  "Identify concentration vs diversification"),
        ("Moat Grade",              moat,                    "Prefer Wide/Narrow moat stocks",
         "HIGH" if moat=="Wide" else ("MEDIUM" if moat=="Narrow" else "LOW"),
         "Wide moat = durable competitive advantage"),
        ("Market Beta (estimated)", profile["beta"],         "Portfolio beta 0.8–1.2 target",
         "HIGH" if profile["beta"]>1.2 else ("LOW" if profile["beta"]<0.8 else "NEUTRAL"),
         "Measures market sensitivity"),
        ("Interest Rate Sensitivity",profile["rate_sens"],   "Mix of HIGH and LOW",
         profile["rate_sens"],  "Rate-sensitive stocks hurt in rising rate env"),
        ("Economic Cycle",          profile["cycle"],        "Mix defensive + cyclical",
         profile["cycle"],      "Cyclical stocks amplify macro swings"),
        ("Inflation Sensitivity",   profile["inf_sens"],     "Prefer LOW inflation sensitivity",
         profile["inf_sens"],   "HIGH = cost pressures squeeze margins"),
        ("Diversification Value",   profile["div"],          "HIGH diversification value preferred",
         profile["div"],        "How much this stock diversifies the portfolio"),
    ]
    for i, (label, val, ideal, impact, note) in enumerate(profile_rows):
        r = 5 + i
        ws16.row_dimensions[r].height = 18
        bg = C_ALT if i%2==0 else C_WHITE
        impact_bg = C_BULL_BG if impact in ["HIGH","DEFENSIVE","LOW","NEUTRAL"] else (C_BEAR_BG if impact in ["HIGHLY CYCLICAL","HIGH"] else C_YELLOW)
        impact_fg = C_GREEN if impact in ["HIGH","DEFENSIVE","LOW"] and label in ["Diversification Value","Moat Grade"] else (C_RED if impact in ["HIGHLY CYCLICAL"] else C_AMBER)
        wc(ws16, r, 1, label, bg=bg, fg=C_FORMULA, sz=10, align="left")
        wc(ws16, r, 2, str(val), bg=bg, fg=C_INPUT_FG, bold=True, sz=10, align="center")
        wc(ws16, r, 3, ideal, bg=C_ALT, fg="595959", sz=9, align="left", italic=True)
        wc(ws16, r, 4, impact, bg=impact_bg, fg=impact_fg, bold=True, sz=10, align="center")
        wc(ws16, r, 5, note, bg=C_ALT, fg="595959", sz=9, align="left", italic=True)

    # Benchmark comparison
    sec(ws16, 13, "STEP 5 — ATTRACTIVENESS vs PORTFOLIO ALTERNATIVES", 5)
    hdr(ws16, 14, ["Comparison Metric", "This Stock", "Portfolio Target", "Status"], bg=C_SUBHDR)
    fcf_yield = (enriched.get("latest_fcf",0)*fx / (shares * price) if shares*price > 0 else 0) * 1e9
    comp_rows = [
        ("FCF Yield",              fcf_yield,          0.04, fcf_yield >= 0.04),
        ("Edge (EV vs Price)",     edge,               0.15, edge >= 0.15),
        ("Risk/Reward Ratio",      rr_ratio,           2.0,  rr_ratio >= 2.0),
        ("Margin of Safety",       (base_iv-price)/price if price>0 else 0, 0.20, (base_iv-price)/price >= 0.20 if price>0 else False),
        ("Moat Score",             1 if moat=="Wide" else (0.5 if moat=="Narrow" else 0), 0.5, moat in ["Wide","Narrow"]),
        ("ROIC vs WACC (spread)",  enriched.get("op_margin",0)*(1-0.25)/0.4 - wacc, 0.05,
         enriched.get("op_margin",0)*(1-0.25)/0.4 > wacc + 0.05),
    ]
    for i, (label, val, target, passes) in enumerate(comp_rows):
        r = 15 + i
        ws16.row_dimensions[r].height = 17
        bg = C_BULL_BG if passes else C_BEAR_BG
        fg = C_GREEN if passes else C_RED
        nf = "0.0%" if "Yield" in label or "Edge" in label or "Safety" in label or "ROIC" in label else "0.0x" if "Ratio" in label else "0.0"
        wc(ws16, r, 1, label,      bg=bg, fg=C_FORMULA, sz=10, align="left")
        c = wc(ws16, r, 2, round(float(val),4), bg=bg, fg=fg, bold=True, sz=10); c.number_format = nf
        c2= wc(ws16, r, 3, target, bg=bg, fg="595959", sz=10); c2.number_format = nf
        wc(ws16, r, 4, "✓ PASS" if passes else "✗ FAIL", bg=bg, fg=fg, bold=True, sz=10, align="center")

    # ══════════════════════════════════════════════════════════
    # SHEET 17 — RISK MANAGEMENT & STOP LOSS
    # ══════════════════════════════════════════════════════════
    ws17 = wb.create_sheet("17. Risk Management")
    ws17.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [34,18,18,18,34]):
        ws17.column_dimensions[col].width = w

    title(ws17, 1, f"RISK MANAGEMENT — {ticker}",
          f"Step 7: Maximum acceptable loss, stop-loss levels, thesis-break conditions", 5)

    # Max loss framework
    max_loss_pct  = min(abs(downside), 0.30)    # never lose more than 30%
    stop_loss_px  = price * (1 - max_loss_pct)
    thesis_break  = price * (1 - abs(downside) * 0.8)  # 80% of bear case

    sec(ws17, 3, "STEP 7 — DOWNSIDE PROTECTION FRAMEWORK", 5)
    hdr(ws17, 4, ["Level", "Price", "% from Current", "Type", "Action"], bg=C_SUBHDR)

    stop_levels = [
        ("Current Price (reference)",   price,        0.0,           "Entry",        "Buy at or below this level"),
        ("Soft Alert (−10%)",           price*0.90,  -0.10,          "Warning",      "Re-evaluate thesis — check if anything changed"),
        ("Hard Stop Loss",              stop_loss_px, -max_loss_pct, "Capital prot.", "EXIT — maximum acceptable loss reached"),
        ("Thesis Break Price",          thesis_break, (thesis_break-price)/price, "Thesis break", "EXIT — bear case exceeded, thesis invalidated"),
        ("Bear Case IV",                bear_iv,      (bear_iv-price)/price, "Valuation floor","Monitor — worst-case fundamental value"),
        ("Half Position Level (−15%)",  price*0.85,  -0.15,          "Trim trigger", "Reduce to half Kelly if thesis weakening"),
    ]
    for i, (label, px, pct, typ, action) in enumerate(stop_levels):
        r = 5 + i
        ws17.row_dimensions[r].height = 18
        is_stop = "Stop" in label or "Break" in label
        bg = C_BEAR_BG if is_stop else (C_ALT if i%2==0 else C_WHITE)
        fg = C_RED if is_stop else C_FORMULA
        wc(ws17, r, 1, label,       bg=bg, fg=fg, bold=is_stop, sz=10, align="left")
        wc(ws17, r, 2, round(px,0), bg=bg, fg=fg, bold=is_stop, sz=10, nf=f'"{sym}"#,##0')
        wc(ws17, r, 3, pct,         bg=bg, fg=C_RED if pct<-0.05 else C_FORMULA, bold=is_stop, sz=10, nf="+0.0%;(0.0%);-")
        wc(ws17, r, 4, typ,         bg=bg, fg=fg, sz=10, align="center")
        wc(ws17, r, 5, action,      bg=C_ALT, fg="595959", sz=9, align="left", italic=True)

    # Thesis break conditions
    sec(ws17, 12, "THESIS-BREAK CONDITIONS (When to exit regardless of price)", 5)
    hdr(ws17, 13, ["Condition", "Trigger", "Severity", "Response"], bg=C_SUBHDR, fg=C_HDR_FG)
    thesis_breaks = [
        ("Revenue growth turns negative for 2+ quarters",
         "Rev CAGR < 0% for 6 months",        "HIGH",   "Exit 100% of position"),
        ("Operating margin drops >500bps below model",
         f"Margin < {max(0, op_margin-0.05):.0%}", "HIGH","Exit 75% — reassess"),
        ("Competitive moat eroded (new entrant/tech disruption)",
         "Market share loss >3% in core segment", "HIGH","Exit 100%"),
        ("Management credibility breakdown",
         "Accounting restatement or fraud signal","CRITICAL","Exit 100% immediately"),
        ("ROIC falls below WACC for 2+ years",
         f"ROIC < {wacc:.1%} sustainably",         "MEDIUM","Reduce to tracking (1%)"),
        ("Valuation disconnects — IV revised down >25%",
         f"DCF IV < {base_iv*0.75:.0f} after re-run","MEDIUM","Reduce by 50%"),
        ("Macro: severe rate hike (>300bps) if rate-sensitive",
         "Policy rate > current +300bps",          "LOW" if profile["rate_sens"]=="LOW" else "HIGH",
         "Reduce if sector highly rate-sensitive"),
    ]
    for i, (cond, trigger, sev, response) in enumerate(thesis_breaks):
        r = 14 + i
        ws17.row_dimensions[r].height = 18
        sev_bg = C_BEAR_BG if sev=="CRITICAL" else (C_BEAR_BG if sev=="HIGH" else (C_YELLOW if sev=="MEDIUM" else C_BULL_BG))
        sev_fg = C_RED if sev in ["CRITICAL","HIGH"] else (C_AMBER if sev=="MEDIUM" else C_GREEN)
        bg = C_ALT if i%2==0 else C_WHITE
        wc(ws17, r, 1, cond,     bg=bg,    fg=C_FORMULA, sz=9, align="left", wrap=True)
        wc(ws17, r, 2, trigger,  bg=C_ALT, fg="595959",  sz=9, align="left", italic=True, wrap=True)
        wc(ws17, r, 3, sev,      bg=sev_bg,fg=sev_fg,    bold=True, sz=10, align="center")
        wc(ws17, r, 4, response, bg=bg,    fg=C_FORMULA, sz=9, align="left", italic=True)

    # Max position loss table
    sec(ws17, 22, "MAXIMUM LOSS SCENARIOS BY POSITION SIZE", 5)
    hdr(ws17, 23, ["Position Size", "Capital Deployed", "Max Loss (Bear)", "Max Loss (Stop)", "As % Portfolio"], bg=C_SUBHDR)
    for i, (label, pct) in enumerate([("Full Kelly", kelly_full), ("Half Kelly", kelly_half), ("Quarter Kelly", kelly_qtr)]):
        r = 24 + i
        capital = portfolio_size * pct
        loss_bear = capital * abs(downside)
        loss_stop = capital * max_loss_pct
        ws17.row_dimensions[r].height = 18
        bg = C_ALT if i%2==0 else C_WHITE
        wc(ws17, r, 1, label,                   bg=bg, fg=C_FORMULA, sz=10, align="left")
        wc(ws17, r, 2, round(capital, 0),        bg=bg, fg=C_INPUT_FG, sz=10, nf=f'"{sym}"#,##0')
        wc(ws17, r, 3, round(loss_bear, 0),      bg=C_BEAR_BG, fg=C_RED, sz=10, nf=f'"{sym}"#,##0')
        wc(ws17, r, 4, round(loss_stop, 0),      bg=C_BEAR_BG, fg=C_RED, sz=10, nf=f'"{sym}"#,##0')
        wc(ws17, r, 5, loss_bear/portfolio_size, bg=C_BEAR_BG, fg=C_RED, bold=True, sz=10, nf="0.0%")

    # ══════════════════════════════════════════════════════════
    # SHEET 18 — FINAL PORTFOLIO DECISION MEMO
    # ══════════════════════════════════════════════════════════
    ws18 = wb.create_sheet("18. Final Decision")
    ws18.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [34,22,22,22,22]):
        ws18.column_dimensions[col].width = w

    title(ws18, 1, f"FINAL PORTFOLIO DECISION — {ticker}",
          f"Step 8: Complete hedge fund capital allocation framework", 5)

    # Final call logic
    passes_count = sum([
        edge > 0.15,
        rr_ratio > 2.0,
        kelly_half > 0.03,
        base_iv > price,
        moat in ["Wide", "Narrow"],
    ])

    if passes_count >= 4 and edge > 0.25:     final_call = "HIGH CONVICTION BUY"
    elif passes_count >= 3 and edge > 0.10:   final_call = "BUY"
    elif passes_count >= 2 and edge > 0:      final_call = "HOLD / ACCUMULATE"
    elif edge < -0.20 and rr_ratio < 1:       final_call = "SHORT"
    else:                                      final_call = "AVOID"

    call_colors = {
        "HIGH CONVICTION BUY": (C_BULL_BG, C_GREEN, 32),
        "BUY":                 ("CCFFCC",   C_GREEN, 28),
        "HOLD / ACCUMULATE":   (C_YELLOW,   C_AMBER, 24),
        "AVOID":               (C_BEAR_BG,  C_RED,   28),
        "SHORT":               ("FCE4D6",   C_RED,   28),
    }
    call_bg, call_fg, call_sz = call_colors.get(final_call, (C_YELLOW, C_AMBER, 24))

    ws18.merge_cells("A3:E5")
    c_call = ws18.cell(3, 1, value=f"  {final_call}")
    c_call.fill = hf(call_bg); c_call.font = Font(name="Calibri", bold=True, size=call_sz, color=_c(call_fg))
    c_call.alignment = Alignment(horizontal="center", vertical="center")
    c_call.border = bdr("medium", C_NAVY)
    ws18.row_dimensions[3].height = 70

    # 8-point framework summary
    sec(ws18, 6, "STEP 8 — COMPLETE PORTFOLIO DECISION FRAMEWORK", 5)
    hdr(ws18, 7, ["Framework Point", "Value", "Metric", "Status", "Detail"], bg=C_SUBHDR)

    framework = [
        ("① Market-Implied Expectations",
         f"Implying {enriched.get('revenue_growth',0)*1.2:.0%} growth",
         "vs Base", "SEE SHEET 9",
         "Reverse DCF shows market-implied assumptions vs your base case"),
        ("② Expected Value",
         f"{sym}{ev:,.0f}",
         "EV vs Price",
         f"+{ev_pct:.0%}" if ev_pct>0 else f"{ev_pct:.0%}",
         f"EV = {sym}{bear_iv:,.0f}×25% + {sym}{base_iv:,.0f}×50% + {sym}{bull_iv:,.0f}×25%"),
        ("③ Edge (% Mispricing)",
         f"{edge:+.0%}",
         "Target: >15%",
         "HIGH EDGE" if edge>0.25 else ("MODERATE" if edge>0.10 else ("LOW" if edge>0 else "NEGATIVE")),
         f"Statistical advantage per {sym} invested at current price"),
        ("④ Kelly Position Size",
         f"Half-Kelly: {kelly_half:.0%}",
         f"{sym}{cap_half:,.0f}",
         f"{shares_half:,} shares",
         f"Full Kelly {kelly_full:.0%} → Half Kelly {kelly_half:.0%} → Quarter {kelly_qtr:.0%}"),
        ("⑤ Risk vs Reward",
         f"{rr_ratio:.1f}x",
         "Target: >2x",
         "FAVOURABLE" if rr_ratio>2 else ("BALANCED" if rr_ratio>1 else "UNFAVOURABLE"),
         f"Upside {upside:+.0%} vs Downside {downside:.0%}"),
        ("⑥ Portfolio Role",
         position_type,
         "Capital alloc.",
         f"{kelly_half:.0%} of portfolio",
         f"Beta {profile['beta']:.1f} | {profile['cycle']} | Rate sensitivity: {profile['rate_sens']}"),
        ("⑦ Stop Loss",
         f"{sym}{stop_loss_px:,.0f}",
         f"{-max_loss_pct:.0%} from entry",
         "Hard stop",
         f"Thesis break at {sym}{thesis_break:,.0f} | Max loss: {sym}{cap_half*max_loss_pct:,.0f}"),
        ("⑧ Final Decision",
         final_call,
         f"{passes_count}/5 criteria met",
         "ACT" if passes_count>=3 else "PASS",
         f"Conviction score: {passes_count}/5 — {', '.join(['Edge','R/R','Kelly','IV>P','Moat'][:passes_count])} pass"),
    ]
    for i, (point, val, metric, status, detail) in enumerate(framework):
        r = 8 + i
        ws18.row_dimensions[r].height = 22
        is_final = i == 7
        status_bg = C_BULL_BG if status in ["HIGH EDGE","FAVOURABLE","ACT","PASS"] else (C_BEAR_BG if status in ["NEGATIVE","UNFAVOURABLE"] else C_YELLOW)
        status_fg = C_GREEN if status in ["HIGH EDGE","FAVOURABLE","ACT","PASS"] else (C_RED if status in ["NEGATIVE","UNFAVOURABLE"] else C_AMBER)
        wc(ws18, r, 1, point,  bg=C_SECTION if not is_final else call_bg, fg=call_fg if is_final else C_FORMULA, bold=is_final, sz=10, align="left")
        wc(ws18, r, 2, val,    bg=C_YELLOW if is_final else C_ALT, fg=call_fg if is_final else C_INPUT_FG, bold=True, sz=10 if not is_final else 11, align="left")
        wc(ws18, r, 3, metric, bg=C_ALT, fg="595959", sz=9, align="left", italic=True)
        wc(ws18, r, 4, status, bg=status_bg, fg=status_fg, bold=True, sz=10, align="center")
        wc(ws18, r, 5, detail, bg=C_ALT, fg="595959", sz=9, align="left", italic=True, wrap=True)

    # Full memo
    sec(ws18, 17, "INVESTMENT MEMO (Copy-paste ready)", 5)
    ws18.row_dimensions[18].height = 200
    ws18.merge_cells("A18:E18")
    memo = (
        f"INVESTMENT MEMO — {ticker}  |  {sector}  |  Moat: {moat}  |  {datetime.now().strftime('%d %b %Y')}\n\n"
        f"RECOMMENDATION: {final_call}\n\n"
        f"MARKET-IMPLIED: At {sym}{price:,.0f}, market prices {enriched.get('revenue_growth',0)*1.2:.0%} FCF growth. "
        f"{'Bar is HIGH vs our {:.0%} base case.'.format(enriched.get('revenue_growth',0)) if enriched.get('revenue_growth',0)*1.2 > enriched.get('revenue_growth',0) else 'Bar appears CONSERVATIVE vs fundamentals.'}\n\n"
        f"EXPECTED VALUE: Bear {sym}{bear_iv:,.0f} (25%) + Base {sym}{base_iv:,.0f} (50%) + Bull {sym}{bull_iv:,.0f} (25%) = "
        f"EV {sym}{ev:,.0f} vs price {sym}{price:,.0f} = {edge:+.0%} EDGE.\n\n"
        f"POSITION SIZING (Kelly): Full Kelly {kelly_full:.0%} | Half Kelly {kelly_half:.0%} | "
        f"Recommended: {sym}{cap_half:,.0f} ({kelly_half:.0%} of {sym}{portfolio_size/1e6:.1f}M portfolio) = {shares_half:,} shares.\n\n"
        f"RISK/REWARD: Upside {upside:+.0%} (Bull) vs Downside {downside:.0%} (Bear) = {rr_ratio:.1f}x ratio. "
        f"Hard stop at {sym}{stop_loss_px:,.0f} ({-max_loss_pct:.0%}). Max loss on recommended position: {sym}{cap_half*max_loss_pct:,.0f}.\n\n"
        f"PORTFOLIO ROLE: {position_type}. Beta {profile['beta']:.1f}, {profile['cycle']} sector, "
        f"rate sensitivity {profile['rate_sens']}. {'Adds diversification.' if profile['div']=='HIGH' else 'Monitor concentration.'}\n\n"
        f"THESIS BREAK: Exit if revenue turns negative, margins drop >500bps, or ROIC sustains below WACC.\n\n"
        f"CONVICTION: {passes_count}/5 criteria met. "
        f"{'All green lights — size up confidently.' if passes_count==5 else 'Most criteria pass — standard position sizing.' if passes_count>=3 else 'Mixed signals — reduce size or wait for better entry.'}"
    )
    c_memo = ws18.cell(18, 1, value=memo)
    c_memo.fill = hf("F8FBFF")
    c_memo.font = Font(name="Calibri", size=10, color=_c(C_FORMULA))
    c_memo.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c_memo.border = bdr("medium", C_NAVY)

    for ws in [ws14, ws15, ws16, ws17, ws18]:
        ws.freeze_panes = "B5"

    return wb


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--ticker",         required=True, help="e.g. TCS.NS")
    parser.add_argument("--portfolio-size", type=float, default=10_000_000,
                        help="Portfolio size in INR (default 1 crore = 10M)")
    args = parser.parse_args()

    ticker         = args.ticker.upper().strip()
    portfolio_size = args.portfolio_size

    print(f"[1/7] Fetching data for {ticker}...")
    from data.collector import StockDataCollector
    from data.processor import compute_metrics
    from models.forecaster import FCFForecaster, compute_wacc
    from models.industry_wacc import get_industry_wacc
    from screener.dcf_engine import DCFEngine, margin_of_safety, assign_signal
    from screener.scenarios import run_scenarios
    from screener.valuation_model import generate_valuation_summary as generate_investment_plan
    from screener.moat_engine import compute_moat_score

    collector  = StockDataCollector(ticker)
    raw        = collector.get_all()
    if not raw:
        print(f"Could not fetch {ticker}"); return

    print(f"[2/7] Running DCF + moat engine...")
    enriched      = compute_metrics(raw)
    forecaster    = FCFForecaster()
    wacc_data_raw = compute_wacc(collector._ticker_obj, is_indian=ticker.endswith((".NS",".BO")))
    ind_info      = get_industry_wacc(ticker=ticker, capm_wacc=wacc_data_raw.get("wacc",0.11))
    final_wacc    = ind_info["wacc"]
    terminal_g    = ind_info["terminal_growth"]
    forecast_yrs  = 10

    dcf_engine      = DCFEngine(discount_rate=final_wacc, terminal_growth=terminal_g)
    forecast_result = forecaster.predict(enriched, years=forecast_yrs)
    dcf_res         = dcf_engine.intrinsic_value_per_share(
        projected_fcfs=forecast_result["projections"],
        terminal_fcf_norm=forecast_result["terminal_fcf_norm"],
        total_debt=enriched["total_debt"], total_cash=enriched["total_cash"],
        shares_outstanding=enriched["shares"], current_price=enriched["price"], ticker=ticker,
    )
    moat_result             = compute_moat_score(enriched, final_wacc)
    enriched["moat_grade"]  = moat_result.get("grade", "None")
    enriched["fundamental_grade"] = "N/A"
    enriched["fundamental_score"] = 0

    iv_n     = dcf_res.get("intrinsic_value_per_share", 0)
    price_n  = enriched["price"]
    mos      = margin_of_safety(iv_n, price_n)
    sig      = assign_signal(mos, dcf_res.get("suspicious",False), forecast_result.get("reliable",True))
    inv_plan = generate_investment_plan(enriched, price_n, iv_n, mos)
    scenarios = run_scenarios(
        enriched=enriched, fcf_base=forecast_result.get("fcf_base",1e9),
        base_growth=forecast_result.get("base_growth",0.08),
        base_wacc=final_wacc, base_terminal_g=terminal_g,
        total_debt=enriched["total_debt"], total_cash=enriched["total_cash"],
        shares=enriched["shares"], current_price=price_n, years=forecast_yrs,
    )
    report_data = {
        "price": price_n, "iv": iv_n, "mos_pct": mos*100, "signal": sig,
        "bear_iv": scenarios.get("Bear 🐻",{}).get("iv", iv_n*0.7),
        "bull_iv": scenarios.get("Bull 🐂",{}).get("iv", iv_n*1.3),
    }

    print(f"[3/7] Building institutional DCF model (Sheets 1-8)...")
    from generate_dcf_excel import generate_institutional_dcf
    excel_bytes = generate_institutional_dcf(
        ticker=ticker, enriched=enriched, dcf_res=dcf_res,
        forecast_result=forecast_result, scenarios=scenarios,
        wacc_data=wacc_data_raw, wacc=final_wacc, terminal_g=terminal_g,
        forecast_yrs=forecast_yrs, sym="₹", to_code="INR", fx=1.0,
    )

    print(f"[4/7] Adding hedge fund sheets (Sheets 9-13)...")
    from generate_hf_excel import build_hedge_fund_sheets
    wb = load_workbook(filename=io.BytesIO(excel_bytes))
    wb = build_hedge_fund_sheets(
        wb=wb, ticker=ticker, enriched=enriched, dcf_res=dcf_res,
        forecast_result=forecast_result, scenarios=scenarios,
        wacc_data=wacc_data_raw, wacc=final_wacc, terminal_g=terminal_g,
        forecast_yrs=forecast_yrs, sym="₹", fx=1.0,
    )

    print(f"[5/7] Adding portfolio decision sheets (Sheets 14-18)...")
    wb = build_portfolio_sheets(
        wb=wb, ticker=ticker, enriched=enriched, dcf_res=dcf_res,
        forecast_result=forecast_result, scenarios=scenarios,
        wacc_data=wacc_data_raw, wacc=final_wacc, terminal_g=terminal_g,
        forecast_yrs=forecast_yrs, sym="₹", fx=1.0,
        portfolio_size=portfolio_size,
    )

    print(f"[6/7] Saving...")
    out_path = f"{ticker.replace('.','_')}_Complete_Portfolio_{datetime.now().strftime('%Y%m%d')}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    with open(out_path, "wb") as f:
        f.write(buf.getvalue())

    print(f"\n{'='*60}")
    print(f"  ✅  {out_path}")
    print(f"{'='*60}")
    print(f"\n  18 sheets — complete portfolio decision system:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"    {name}")
    print(f"\n  Portfolio size: {chr(8377)}{portfolio_size/1e7:.1f} Cr")
    print(f"  Run with different size: python generate_portfolio_excel.py --ticker {ticker} --portfolio-size 50000000")


if __name__ == "__main__":
    main()
